from collections import defaultdict
from flask import Flask, render_template, Response, make_response, request, redirect, session, url_for, jsonify, \
    send_file, flash, send_from_directory, logging
import folium
import requests
import hashlib  # Import hashlib for password hashing
from openpyxl import Workbook
from io import BytesIO
from fpdf import FPDF
import uuid
import base64
import mysql.connector
from openpyxl.packaging.manifest import mimetypes
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.protection import hash_password
import bcrypt
from werkzeug.utils import secure_filename
import re, os, csv
from flask_mail import Mail, Message
from itsdangerous import URLSafeTimedSerializer
import random
import json
from datetime import datetime, timedelta, date
import os
# from captcha.image import ImageCaptcha
import requests
from classes.caste import casteController
from tempfile import NamedTemporaryFile, gettempdir

# Middleware import
from authentication.middleware import auth, auth_admin

from classes.connection import HostConfig, ConfigPaths
# University Class
from classes.university import universityController

# ----------- FLASK APP CONFIGURATION -------------------
app = Flask(__name__)
app.config['SECRET_KEY'] = 'rootTanmay'
app.config['SESSION_TYPE'] = 'filesystem'
app.config['SESSION_PERMANENT'] = False
app.config['SESSION_USE_SIGNER'] = True
app.config['SESSION_KEY_PREFIX'] = 'session:'  # Optional, to prevent conflicts
# Session(app)
# ---------------------xxx-------------------------------

UPLOAD_PDF_FOLDER = os.path.join('/var/www/fellowship/fellowship/BartiFellowship/BartiFellowship/static', 'pdf_application_form')
app.config['UPLOAD_PDF_FOLDER'] = UPLOAD_PDF_FOLDER
# Test

UPLOAD_FOLDER = os.path.join('/static', 'Images')
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# -------------- Bilingual Configuration -------------
app.config['SUPPORTED_LANGUAGES'] = ['en', 'mr']

serializer = URLSafeTimedSerializer('SECRET_KEY')

# ------------ DATABASE CONFIGURATION -------------------
hostserver = '192.168.10.69'
localserver = '127.0.0.1'
host = hostserver
user = HostConfig.user
password = HostConfig.password
database = HostConfig.database

# For Host Server
if host == hostserver:
    app.config['USER_DOC_SEC_FIVE'] = '/var/www/fellowship/fellowship/BartiFellowship/BartiFellowship/static/uploads/user_doc_secfive/'
    app.config['RENT_AGREEMENT_REPORT'] = '/var/www/fellowship/fellowship/BartiFellowship/BartiFellowship/static/uploads/rent_agreement/'
    app.config['HALF_YEARLY_REPORTS'] = '/var/www/fellowship/fellowship/BartiFellowship/BartiFellowship/static/uploads/half_yearly/'
    app.config['PRESENTY_REPORTS'] = '/var/www/fellowship/fellowship/BartiFellowship/BartiFellowship/static/uploads/presenty_reports/'
    app.config['UPLOAD_PHOTO_SECTION1'] = '/var/www/fellowship/fellowship/BartiFellowship/BartiFellowship/static/uploads/image_retrive/'
    app.config['PDF_STORAGE_PATH'] = '/var/www/fellowship/fellowship/BartiFellowship/BartiFellowship/static/pdf_application_form/pdfform.pdf'
    app.config['AWARD_LETTER'] = '/var/www/fellowship/fellowship/BartiFellowship/BartiFellowship/static/pdf_application_form/award_letter.pdf'
    app.config['JOINING_REPORT'] = '/var/www/fellowship/fellowship/BartiFellowship/BartiFellowship/static/uploads/joining_reports/'
    app.config['PDF_CERTIFICATE'] = '/var/www/fellowship/fellowship/BartiFellowship/BartiFellowship/static/uploads/phd_certificate/'
    app.config['UPLOAD_THESIS'] = '/var/www/fellowship/fellowship/BartiFellowship/BartiFellowship/static/uploads/upload_thesis/'
    app.config['EMAIL_DOCS'] = '/var/www/fellowship/fellowship/BartiFellowship/BartiFellowship/static/uploads/sendbulkemails/'
    app.config['ASSESSMENT_REPORT'] = '/var/www/fellowship/fellowship/BartiFellowship/BartiFellowship/static/uploads/assessment_report/'
    app.config['SAVE_NEWS'] = '/var/www/fellowship/fellowship/BartiFellowship/BartiFellowship/static/uploads/save_news/'
    app.config['UNDERTAKING_REPORT'] = '/var/www/fellowship/fellowship/BartiFellowship/BartiFellowship/static/uploads/undertaking_doc/'
elif host == localserver:
    app.config['USER_DOC_SEC_FIVE'] = 'static/uploads/user_doc_secfive/'
    app.config['RENT_AGREEMENT_REPORT'] = 'static/uploads/rent_agreement/'
    app.config['HALF_YEARLY_REPORTS'] = 'static/uploads/half_yearly/'
    app.config['PRESENTY_REPORTS'] = 'static/uploads/presenty_reports/'
    app.config['UPLOAD_PHOTO_SECTION1'] = 'static/uploads/image_retrive/'
    app.config['PDF_STORAGE_PATH'] = 'static/pdf_application_form/pdfform.pdf'
    app.config['AWARD_LETTER'] = 'static/pdf_application_form/award_letter.pdf'
    app.config['JOINING_REPORT'] = 'static/uploads/joining_reports/'
    app.config['PDF_CERTIFICATE'] = 'static/uploads/phd_certificate/'
    app.config['UPLOAD_THESIS'] = 'static/uploads/upload_thesis/'
    app.config['EMAIL_DOCS'] = 'static/uploads/sendbulkemails/'
    app.config['ASSESSMENT_REPORT'] = 'static/uploads/assessment_report/'
    app.config['SAVE_NEWS'] = 'static/uploads/save_news/'
    app.config['UNDERTAKING_REPORT'] = 'static/uploads/undertaking_doc/'


cnx = mysql.connector.connect(user=user, password=password,       # --------  DATABASE CONNECTION
                              host=host,
                              database=database)
cursor = cnx.cursor()
# --------------- END DATABASE ---------------------------


# ------------ MAIL CONFIGURATION -------------------
# Outlook Mail
app.config['MAIL_SERVER'] = 'us2.smtp.mailhostbox.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USERNAME'] = 'helpdesk@trti-maha.in'                     # --------  E-MAIL CONNECTION
app.config['MAIL_PASSWORD'] = 'FOtIEzp9'
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USE_SSL'] = False
app.config['MAIL_DEBUG'] = True  # Enable debugging
mail = Mail(app)

# Zepto Mail
# app.config['MAIL_SERVER'] = 'smtp.zeptomail.in'
# app.config['MAIL_PORT'] = 587
# app.config['MAIL_USERNAME'] = 'noreply_fellowship@trti-maha.in'
# app.config['MAIL_PASSWORD'] = 'PHtE6r1YFuzp2TJ69BkFsfewQ8+iPI8v/7hvKABA5IxGCKRVGU0G/t4jkWS/rUsvAPNAFPaYz948tb6c4r2HIGa4N2pIWmqyqK3sx/VYSPOZsbq6x00Vt1gackbeVI/udNVt0S3Vud/fNA=='
# app.config['MAIL_USE_TLS'] = True
# app.config['MAIL_USE_SSL'] = False
# app.config['MAIL_DEBUG'] = True  # Enable debugging
# app.config['MAIL_DEFAULT_SENDER'] = ('Your Name', 'noreply_fellowship@trti-maha.in')
# mail = Mail(app)
# ---------------------xxx-------------------------------


# ---------------------- Language Array ------------------
multilingual_content = {
    'english': {
        # ---------- Title -------------
        'title_home' : 'Fellowship | Home',
        # ----- Top Menu -----
        'skip-to-main-content': 'Skip to Main Content',
        'mobile-no': '020-26333330, 26333339, 26343600',
        'a-': 'A-',
        'a': 'A',
        'a+': 'A+',
        'english': 'English',
        'marathi': 'Marathi',
        'trti_header':'Dr. Babasaheb Ambedkar Research and Training Institute (BARTI)',
        'admin_login_btn':'Admin Login',
        'fellowship': 'Fellowship',

        # -----------------------Leaders ----------------------
        'eknath_shinde': 'Shri. Eknath Shinde',
        'eknath_shinde_desgn': "Hon'ble Deputy Chief Minister, Maharashtra State",
        'dev_phad':'Shri. Devendra Fadnavis',
        'dev_phad_desgn':"Hon'ble Chief Minister, Maharashtra State",
        'ajit_pawar':'Shri. Ajit Pawar',
        'ajit_pawar_desgn':"Hon'ble Deputy Chief Minister, Maharashtra State",
        'vijay_gavit':'Dr. Vijaykumar Gavit',
        'vijay_gavit_desgn':"Hon'ble Minister, Tribal Development Department",
        'vijay_waghmare':"Dr. Harshdeep Kamble, I.A.S.",
        'vijay_waghmare_desgn':"Honb'le Principal secretary, Social Justice and Special Assistance Department",
        'sanjay_shirsat':"Mr. Sanjay Shirsat ",
        'sanjay_shirsat_design':"Hon'ble Minister of Social Justice, Government of Maharashtra",
        'madhuri_misal': "Mrs. Madhuri Misal",
        'madhuri_misal_design':"Hon'ble State Minister of Social Justice, Government of Maharashtra",
        'sunil_ware': "Mr. Sunil Ware, IRAS",
        'sunil_ware_design':"Director General, Dr. Babasaheb Ambedkar Research and Training Institute (BARTI), Pune",

        # --------------- Menubar -----------------
        'home' : 'Home',
        'about': "About",
        'government_regulation' : 'Circulars/Orders',
        'government_decision':'Government Decision',
        'circulars':'Circulars',
        'shcolarship_regulations':'Scholarship Regulations',
        'margadarshak':'Guidance Information',
        'video':'Videos - FAQ',
        'contact_us' : 'Contact Us',
        'charts': 'Reports',
        'login_sign_btn':'Login | Sign up',

        # ------------------ Banner Section ---------------------
        'total_applications' : "Total Application",
        'total_applications_2022' : "Total Application 2022",
        'total_applications_2021' : "Total Application 2021",
        'fellowship_awarded' : 'Fellowship Awarded',
        'fellowship_completed' : 'Fellowship Completed',
        'this_yrs_application' : 'This Year Applications',
        'get_your_fellowship_now' : 'Get Your Fellowship Now',
        'enroll_here' : 'Enroll Here',

        #------------------------ About --------------------------
        'about_the_institute' : 'About The Institution',
        'about_title' : 'Dr. Babasaheb Ambedkar Research and Training Institute (BARTI), Pune',
        'read_more' : "Read More",
        'criteria' : "Criteria",
        'scheme_benefits' : 'Scheme Benefits',
        'eligibility_criteria' : 'Eligibility Criteria',
        'updates' : 'Updates',
        'how_to_enroll' : 'How to enroll?',
        'step' : 'Step',
        'step_1_content' : 'Create Your Login or Sign on site',
        'step_2_content' : 'Fill Application Form',
        'step_3_content' : 'Get Selected by Admin',
        'step_4_content' : 'Upload Joining Letter of PhD',
        'step_5_content' : 'Get your Fellowship',

        'faq' : 'Frequently Asked Questions',
        'q1' : 'How many Fellows do you select?',
        'ans1': 'We hire 100 students every year',
        'q2' : 'I wasn’t selected, can I receive feedback on my application?',
        'ans2' : 'We will send you feedback through email, and it will be reflected on your dashboard as well.',
        'related_links' : 'Related Links',
        'security_brands' : 'SECURITY & BRANDS',
        'address': "28, Queen's Garden Rd, Near Old Circuit House, Camp, Pune, Maharashtra 411001" ,
        'phone' : 'Phone',
        'email' : 'Email',
        'national_portal_of_india' : 'National Portal of India',
        'tribal_development_department' : 'Tribal development Department',
        'tribal_commissionerate' : 'Tribal Commissionerate',
        'aaple_sarkar' : 'Aaple Sarkar',
        'ministry_of_tribal_affairs' : 'Ministry of Tribal Affairs- Govt of India',
        'my_gov' : 'My Gov',
        'security_brand' : 'SECURITY & BRAND',
        'hyperlink_policy' : 'Hyperlink Policy',
        'terms_and_conditions' : 'Terms and Conditions',
        'privacy_policy' : 'Privacy Policy',
        'copyright_policy' : 'Copyright Policy',
        'website_monitoring_policy' : 'Website Monitoring Policy',
        'sitemap' : 'Sitemap',
        'footer_text' : 'Website Content Owned & Managed by Dr. Babasaheb Ambedkar Research and Training Institute | Maharashtra, India',
        'desgn_dev' : 'Designed and Developed by',
        'ics' : 'Integrated Consultancy Services ',
        'you_name' : 'Your Name',
        'you_email' : 'Your Email',
        'subject' : 'Subject',
        'message' : 'Message',
        'send_message' : 'Send Message',
        'address' : 'Address',
        'address_text' : "28, QUEEN'S GARDEN RD, NEAR OLD CIRCUIT HOUSE, CAMP, PUNE, MAHARASHTRA 411001",
        'email_us' : 'Email us',
        'open_hours': 'Open hours',
        'monday_friday' : 'Monday - Friday',
        'year_wise_total':'Year Wise Total Application',
        'gender_ratio': 'Gender Ratio',
        'disability_count': 'Disability Count'

    },
    'marathi': {
        # ----------- Title -------------
        'title_home' : 'अधिछात्रवृत्ती | मुख्यपृष्ठ',

        # ----- Top Menu -----
        'skip-to-main-content': 'मुख्य सामग्रीवर जा',
        'mobile-no': '०२०-२६३३३३३०, २६३३३३३९, २६३४३६००',
        'a-': 'ए-',
        'a': 'ए',
        'a+': 'ए+',
        'english': 'इंग्रजी',
        'marathi': 'मराठी',
        'trti_header':'डॉ. बाबासाहेब आंबेडकर संशोधन व प्रशिक्षण संस्था (बार्टी)',
        'admin_login_btn':'प्रशासक लॉगिन',
        'fellowship': 'अधिछात्रवृत्ती',

        # -----------------------Leaders ----------------------
        'eknath_shinde': 'श्री. एकनाथ शिंदे',
        'eknath_shinde_desgn': "माननीय उपमुख्यमंत्री, महाराष्ट्र राज्य",
        'dev_phad':'श्री. देवेंद्र फडणवीस',
        'dev_phad_desgn':"माननीय मुख्यमंत्री, महाराष्ट्र राज्य",
        'ajit_pawar':'श्री. अजित पवार',
        'ajit_pawar_desgn':'माननीय उपमुख्यमंत्री, महाराष्ट्र राज्य',
        'vijay_gavit':'डॉ. विजयकुमार गावित',
        'vijay_gavit_desgn':"माननीय मंत्री, आदिवासी विकास विभाग",
        'vijay_waghmare':"डॉ. हर्षदीप कांबळे, IAS",
        'vijay_waghmare_desgn':"मा. प्रधान सचिव, सामाजिक न्याय आणि विशेष सहाय्य विभाग",
        'sanjay_shirsat':"श्री. संजय शिरसाट",
        'sanjay_shirsat_design':"मा. सामाजिक न्याय मंत्री, महाराष्ट्र राज्य",
        'madhuri_misal': "श्रीमती. माधुरी मिसाळ",
        'madhuri_misal_design':"मा. राज्य सामाजिक न्याय मंत्री, महाराष्ट्र राज्य",
        'sunil_ware': "श्री. सुनील वारे, IRAS ",
        'sunil_ware_design':"महासंचालक, डॉ. बाबासाहेब आंबेडकर संशोधन व प्रशिक्षण संस्था (बार्टी), पुणे",

        # --------------- Menubar -----------------
        'home' : 'मुख्यपृष्ठ',
        'about': "आमच्याविषयी",
        'government_regulation' : 'परिपत्रके/आदेश',

        'contact_us' : 'संपर्क',
        'charts' : 'अहवाल',
        'login_sign_btn':'लॉगिन | साइन अप करा',
        'government_decision':'सरकारी  निर्णय',
        'margadarshak':'मार्गदर्शक सूचना',
        'circulars':'परिपत्रके',
        'shcolarship_regulations':'अधिछात्रवृत्ती नियमावली',
        # ------------------ Banner Section ---------------------
        'total_applications' : "एकूण अर्ज",
        'total_applications_2022' : "एकूण अर्ज 2022",
        'total_applications_2021' : "एकूण अर्ज 2021",
        'fellowship_awarded' : 'अधिछात्रवृत्ती प्रदान',
        'fellowship_completed' : 'अधिछात्रवृत्ती पूर्ण केलेले',
        'this_yrs_application' : 'या वर्षाचे अर्ज',
        'enroll_here' : 'येथे नावनोंदणी करा',
        'get_your_fellowship_now' : 'आता आपली अधिछात्रवृत्ती मिळवा',


        # ----------------------------- About institute ---------------------
        'about_the_institute' : 'संस्थेबद्दल',
        'about_title' : 'डॉ. बाबासाहेब आंबेडकर संशोधन व प्रशिक्षण संस्था (बार्टी), पुणे',
        'read_more' : "पुढे वाचा",

        # ---------------------------
        'criteria' : "निकष",
        'scheme_benefits' : 'योजनेचे फायदे',
        'eligibility_criteria' : 'पात्रता निकष',
        'updates' : 'अपडेट्स',
        'how_to_enroll' : 'नोंदणी कशी करावी?',
        'step' : 'स्टेप',
        'step_1_content' : 'साइटवर आपले लॉगिन तयार करा किंवा साइन इन करा',
        'step_2_content' : 'अर्ज भरा',
        'step_3_content' : 'प्रशासकाद्वारे निवड मिळवा',
        'step_4_content' : 'पीएचडीचे सामील पत्र अपलोड करा',
        'step_5_content' : 'तुमची फेलोशिप मिळवा',

        'faq' : 'वारंवार विचारले जाणारे प्रश्न',
        'video' : 'व्हिडिओ - वारंवार विचारले जाणारे प्रश्न',
        'q1' : 'तुम्ही किती विद्यार्थी निवडता?',
        'ans1': 'आम्ही दरवर्षी 100 विद्यार्थी निवडतो',
        'q2' : 'मी निवडलो नाही, मला अभिप्राय मिळू शकेल का माझ्या अर्जावर?',
        'ans2' : 'आम्ही आपल्याला ईमेलद्वारे अभिप्राय पाठवू आणि ते आपल्या डॅशबोर्डवर देखील प्रतिबिंबित होईल.',
        'related_links' : 'संबंधित दुवे',
        'security_brands' : 'सुरक्षा आणि ब्रँड',

        'address': '२८, क्वीन्स गार्डन रोड, जवळ जुन्या सर्किट हाऊस, कॅम्प, पुणे, महाराष्ट्र ४११००१',
        'phone' : 'फोन',
        'email' : 'ईमेल',
        'national_portal_of_india' : 'भारताचा राष्ट्रीय पोर्टल',
        'tribal_development_department' : 'आदिवासी विकास विभाग',
        'tribal_commissionerate' : 'आदिवासी आयुक्त',
        'aaple_sarkar' : 'आपले सरकार',
        'ministry_of_tribal_affairs' : 'आदिवासी प्रकरण मंत्रालय- भारत सरकार',
        'my_gov' : 'माय गोव्ह',
        'security_brand' : 'सुरक्षा आणि ब्रँड',
        'hyperlink_policy' : 'हायपरलिंक धोरण',
        'terms_and_conditions' : 'नियम आणि अटी',
        'privacy_policy' : 'गोपनीयता धोरण',
        'copyright_policy' : 'कॉपीराइट धोरण',
        'website_monitoring_policy' : 'वेबसाइट देखरेख धोरण',
        'sitemap' : 'साइटमॅप',
        'footer_text' : 'वेबसाइट सामग्री डॉ. बाबासाहेब आंबेडकर संशोधन व प्रशिक्षण संस्था (बार्टी) संस्थेच्या मालकीची आणि व्यवस्थापित | महाराष्ट्र शासन, भारत',
        'desgn_dev' : ' डिझाइन आणि विकसित बाय',
        'ics' : 'इंटिग्रेटएड कन्सल्टन्सी सरवीसेस',
        'you_name' : 'तुमचे नाव',
        'you_email' : 'आपला ई - मेल',
        'subject' : 'विषय',
        'message' : 'संदेश',
        'send_message' : 'संदेश पाठवा',
        'address' : 'पत्ता',
        'call_us' : 'आम्हाला कॉल करा',
        'trti': 'आदिवासी संशोधन आणि प्रशिक्षण संस्था',
        'address_text' : '28, क्वीन्स गार्डन आरडी, ओल्ड  सर्किट हाऊसजवळ, कॅम्प, पुणे, महाराष्ट्र 411001',
        'email_us' : 'आम्हाला ईमेल करा',
        'open_hours': 'उघडे तास',
        'monday_friday' : 'सोमवार-शुक्रवार',
        'year_wise_total':'वर्षनिहाय एकूण अर्ज',
        'gender_ratio': 'लिंग गुणोत्तर',
        'disability_count': 'अपंगत्व संख्या'
    },
    # Add more languages and content as needed.
}


# --------------------- Set Session ---------------
@app.route('/set_session/<value>')
def set_session(value):
    session['language'] = value
    return redirect(request.referrer)


# ---------------- HOME PAGE ---------------------------


@app.route('/', methods=['GET', 'POST'])
def home_page():
    if 'language' in session:
        language = session['language']
    else:
        language = 'marathi'

    # --------------------------  HOME PAGE
    total_count = applications_today()
    fellow_awarded = fellow_awarded_count()
    total_appl_22 = total_appl_22_count()
    total_appl_23 = total_appl_23_count()

    # print("old user 2022",old_user_22)
    news_record = news_fetch()
    return render_template('home-page.html', total_count=total_count, fellow_awarded=fellow_awarded, total_appl_22=total_appl_22, total_appl_23=total_appl_23,
                           language=language, multilingual_content=multilingual_content, news_record=news_record)


def fellow_awarded_count():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute(" SELECT COUNT(*) FROM application_page where phd_registration_year='2023' and final_approval='accepted' ")
    result = cursor.fetchone()
    print(result)
    return result[0]


def total_appl_22_count():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute(" SELECT COUNT(*) FROM application_page where phd_registration_year='2022' ")
    result = cursor.fetchone()
    print(result)
    return result[0]


def total_appl_23_count():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute(" SELECT COUNT(*) FROM application_page where phd_registration_year='2023' ")
    result = cursor.fetchone()
    print(result)
    return result[0]



@app.route('/charts')
def reports():
    if 'language' in session:
        language = session['language']
    else:
        language = 'marathi'
    twentyone_count = year_twentyone_count()
    twentytwo_count = year_twentytwo_count()
    twentythree_count = year_twentythree_count()
    male_count = male_count_report()
    female_count = female_count_report()
    trans_count = trans_count_report()
    disability_yes = disability_yes_count_report()
    disability_no = disability_no_count_report()
    return render_template('report_homepage.html', twentyone_count=twentyone_count,
                           language=language, twentytwo_count=twentytwo_count, twentythree_count=twentythree_count,
                           male_count=male_count, female_count=female_count ,multilingual_content=multilingual_content,
                           trans_count=trans_count, disability_yes=disability_yes, disability_no=disability_no)

# ---------------- END HOME PAGE -----------------------


# ----------------- ABOUT US ---------------------------
@app.route('/aboutus')
def aboutus():
    if 'language' in session:
        language = session['language']
    else:
        language = 'marathi'

    return render_template('AboutUs.html', language = language, multilingual_content=multilingual_content)
# ----------------- END ABOUT US -----------------------


# ----------------- GOVERNMENT RELATIONS ---------------
@app.route('/gr')
def gr():                                          # --------------------------  GOVERNMENT RELATIONS PAGE
    return render_template('gr.html')
# ----------------- END GOVERNMENT RELATIONS ----------


# ----------------- CONTACT US ---------------------------
@app.route('/contact')
def contact_us():                                        # --------------------------  CONTACT US PAGE
    if 'language' in session:
        language = session['language']
    else:
        language = 'marathi'
    # Create a folium map centered at Pune, India
    m = folium.Map(location=[18.5204, 73.8567], zoom_start=12)
    # Add a marker at Pune, India
    folium.Marker(location=[18.5204, 73.8567], popup="Pune, India").add_to(m)
    # Render the map in the template
    map_html = m._repr_html_()
    # Pass the map HTML to the template
    return render_template('contact.html', map=map_html, language = language, multilingual_content = multilingual_content)


@app.route('/contact_submit', methods=['GET', 'POST'])
def contact_submit():
    if 'language' in session:
        language = session['language']
    else:
        language = 'marathi'
    cnx = mysql.connector.connect(user=user, password=password,  # --------  DATABASE CONNECTION
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    if request.method == 'POST':
        ticket = 'Contact Us'
        fullname = request.form['fullname']
        email = request.form['email']
        issue_subject = request.form['issue_subject']
        description = request.form['description']

        sql = "INSERT INTO issue_raised (ticket, fullname, email, issue_subject, description) " \
              "VALUES (%s, %s, %s, %s, %s)"
        # Execute the SQL statement with the data
        cursor.execute(sql, (ticket, fullname, email, issue_subject, description))
        cnx.commit()
        cursor.close()
        cnx.close()
        return render_template('contact.html', language = language, multilingual_content = multilingual_content)
    return redirect(url_for('contact_us'))
# ----------------- END CONTACT US -----------------------


# ----------------- LOGIN --------------------------------
def new_applicant_incomplete_form(email):                                     # -------------- CHECK IF USER HAS FILLED THE FORM
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    sql = """
            SELECT phd_registration_year, form_filled
            FROM application_page
            WHERE email = %s
              AND phd_registration_year >= 2023
              AND (form_filled = 0 OR form_filled IS NULL);
    """
    cursor.execute(sql, (email, ))
    result = cursor.fetchone()
    cursor.close()
    cnx.close()
    return result is not None


def old_user_incomplete_form(email, form_filled):                                     # -------------- CHECK IF USER HAS FILLED THE FORM
    email = session['email']
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    sql = "SELECT form_filled FROM application_page WHERE email = %s AND form_filled = %s"
    cursor.execute(sql, (email,form_filled))
    result = cursor.fetchone()
    cursor.close()
    cnx.close()
    return result is not None


def is_form_filled(email):                                     # -------------- CHECK IF USER HAS FILLED THE FORM
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    sql = "SELECT form_filled FROM application_page WHERE form_filled='1' and email=%s"
    cursor.execute(sql, (email,))
    result = cursor.fetchone()
    cursor.close()
    cnx.close()
    return result is not None


def check_final_approval(email):                                    # ----------- CHECK IF USER IS FINALLY APPROVED
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    sql = "SELECT final_approval FROM application_page WHERE email = %s AND final_approval = 'accepted'"
    cursor.execute(sql, (email,))
    result = cursor.fetchone()
    cursor.close()
    cnx.close()
    if result:
        flash('Please enter correct Email address', 'success')
    else:
        flash('Successfully Logged in', 'Error')
    return result is not None


def get_id_by_email(email):                                       # --------------- GET ID for EMAIL IN SESSION
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    sql = """ SELECT id FROM application_page WHERE email=%s"""
    cursor.execute(sql, (email,))
    result = cursor.fetchone()
    print(result[0])
    cursor.close()
    cnx.close()
    return result[0]


def is_withdrawn(email):
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    sql = """ SELECT fellowship_withdrawn='withdrawn' FROM signup WHERE email=%s"""
    cursor.execute(sql, (email,))
    result = cursor.fetchone()
    cursor.close()
    cnx.close()

    if result:
        flash('You cannot Login as the Fellowship has been withdrawn')
    else:
        flash('Successfully Logged in')
    return result[0]


def old_user(email):
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    # Corrected SQL query
    sql = """
          
          SELECT *
        FROM signup
        WHERE email = %s
          AND year IN ('2020', '2021', '2022') AND email NOT IN(SELECT email
        FROM application_page)
                
    """
    cursor.execute(sql, (email,))
    result = cursor.fetchone()
    if result:
        print(result)
    cursor.close()
    cnx.close()
    return result if result else None


# def send_sms(mobile_number, otp):
#     authkey = '413185AKlf5Kpy87NZ6597e17fP1'
#     sender = 'MHTRTI'
#     route = '4'
#     country = '91'
#     DLT_TE_ID = '1207171690915170968'
#
#     # Get the mobile number and OTP from the request
#     # mobile_number = request.args.get('mobile_number'
#
#     # Prepare the SMS message with the dynamic parameters
#     message = f"One Time Password for Fellowship Registration is ({otp}) use only once. Please do not share with anyone. MHTRTI Pune. I - C - O N COMPUTER"
#
#     # Construct the URL with dynamic parameters
#     sms_url = f"https://login.wishbysms.com/api/sendhttp.php?authkey={authkey}&mobiles={mobile_number}&message={message}&sender={sender}&route={route}&country={country}&DLT_TE_ID={DLT_TE_ID}"
#
#     # Send the SMS using the requests library
#     try:
#         response = requests.get(sms_url, verify=False)
#         if response.status_code == 200:
#             return 'SMS sent successfully!'
#         else:
#             return f"Error sending SMS. Status code: {response.status_code}"
#     except Exception as e:
#         return f"Error: {str(e)}"

# Suppress only the single InsecureRequestWarning from urllib3
# requests.packages.urllib3.disable_warnings(InsecureRequestWarning)
# Set up logging
# logging.basicConfig(level=logging.INFO)

def send_sms(mobile_number, otp):
    authkey = '413185AKlf5Kpy87NZ6597e17fP1'
    sender = 'MHTRTI'
    route = '4'
    country = '91'
    DLT_TE_ID = '1207171690915170968'

    message = f"One Time Password for Fellowship Registration is ({otp}) use only once. Please do not share with anyone. MHTRTI Pune."
    # encoded_message = urllib3.parse.quote(message)

    sms_url = f"https://login.wishbysms.com/api/sendhttp.php?authkey={authkey}&mobiles={mobile_number}&message=Your One Time Password for Fellowship Registration is {otp} use only once. Please do not share with anyone. MHTRTI Pune.I - C - O N COMPUTER&sender={sender}&route={route}&country={country}&DLT_TE_ID={DLT_TE_ID}"
    print(sms_url)


    try:
        response = requests.get(sms_url, verify=True)
        if response.status_code == 200:
            print(f"SMS sent successfully to {mobile_number}")
            return 'SMS sent successfully!'
        else:
            print(f"Error sending SMS to {mobile_number}. Status code: {response.status_code}")
            return f"Error sending SMS. Status code: {response.status_code}"
    except Exception as e:
        print(f"Error sending SMS: {str(e)}")
        return f"Error: {str(e)}"


@app.route('/login_closed_2023')
def login_closed_2023():
    return render_template('loginClosed2023.html')


@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'language' in session:
        language = session['language']
    else:
        language = 'marathi'

    try:
        # cnx = mysql.connector.connect(user=user, password=password,
        #                               host=host,
        #                               database=database)
        # cursor = cnx.cursor(dictionary=True)
        cnx = mysql.connector.connect(user='root1', password='Admin@#$123',
                                      host=host,
                                      database='BartiApplication')
        cursor = cnx.cursor(dictionary=True)

        if request.method == 'POST':
            email = request.form['email']
            password = request.form['password']
            print(email)
            print(password)

            # Validate that both email and password are provided
            if not email or not password:
                flash('Please enter username and password.', 'error')
                return redirect(url_for('login'))

            # Validate the email format
            email_regex = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
            if not re.match(email_regex, email):
                flash('Please enter a valid email address.', 'error')
                return redirect(url_for('login'))

            # Check if email and password are valid from signup table
            sql = "SELECT email, password, first_name, last_name, year FROM signup WHERE email=%s"
            cursor.execute(sql, (email,))
            user = cursor.fetchone()

            sql = "SELECT formfilled_again FROM application_page WHERE email=%s"
            cursor.execute(sql, (email,))
            user_result = cursor.fetchone()

            # special_case = ['valvibadal222@gmail.com', 'vishwaspadvi22@gmail.com',
            #                 'sunilarunnaik2014@gmail.com', 'tadaviakash62@gmail.com' ]
            # aadesh_email = ['manishapardhi33@gmail.com', 'gajanang996@gmail.com', 'Test@gmail.com',
            #                 'vidyashrihari4.vg@gmail.com', 'bdpatil642@gmail.com', 'niteenvasave2013@gmail.com']
            #
            # print(user['email'])
            
            # Check for valid emails
            # if user['email'].strip().lower() in [email.lower() for email in special_case] and user_result['formfilled_again'] == 0:
            #     if user['password '] == password:
            #         session['email'] = user['email']
            #         flash('Redirecting to form information.', 'info')
            #         print('Redirecting to section1 for', user['email'])
            #         return redirect(url_for('section1'))
            #     else:
            #         flash('Invalid Password', 'error')
            #         return redirect(url_for('login'))
            # elif user and user['email'] in aadesh_email and user_result['formfilled_again'] == 0:
            #     session['email'] = email
            #     flash('Redirecting to the accepted form for 2023.', 'info')
            #     return redirect(url_for('old_but_accepted_for_2023'))

            if user:
                user_password = user['password']
                # special_email = ['sveenashri@gmail.com']

                # If the password is not hashed
                if not user_password.startswith("$2b$"):
                    if password == user_password:
                        session['email'] = email
                        sql = "SELECT applicant_photo FROM application_page WHERE email=%s"
                        cursor.execute(sql, (email,))
                        user_image = cursor.fetchone()
                        session['applicant_photo'] = user_image['applicant_photo'] if user_image else '/static/assets/img/default_user.png'

                        if is_withdrawn(email):
                            flash('You have withdrawn from Fellowship. Please contact us.', 'error')
                            return redirect(url_for('login'))
                        # elif old_user(email):
                        #     session['logged_in_from_login'] = True
                        #     return redirect(url_for('old_user_preview'))
                        elif new_applicant_incomplete_form(email):
                            print('I am here')
                            flash('Your form is incomplete.', 'error')
                            return redirect(url_for('app_form_info'))
                        elif check_final_approval(email):
                            session['final_approval'] = "accepted"
                            session['logged_in_from_login'] = True
                            return redirect(url_for('main_page'))
                        elif is_form_filled(email):
                            session['final_approval'] = "pending"
                            id = get_id_by_email(email)
                            return redirect(url_for('viewform', id=id))
                        else:
                            flash('Redirecting to login closed page for 2023.', 'info')
                            return redirect(url_for('section1'))
                    else:
                        flash('Invalid password. Please try again.', 'error')
                        return redirect(url_for('login'))

                # For bcrypt hashed passwords
                elif bcrypt.checkpw(password.encode('utf-8'), user_password.encode('utf-8')):
                    session['email'] = email
                    sql = "SELECT applicant_photo FROM application_page WHERE email=%s"
                    cursor.execute(sql, (email,))
                    user_image = cursor.fetchone()
                    session['applicant_photo'] = user_image['applicant_photo'] if user_image else '/static/assets/img/default_user.png'

                    if is_withdrawn(email):
                        flash('You have withdrawn from Fellowship. Please contact us.', 'error')
                        return redirect(url_for('login'))
                    # elif old_user(email):
                    #     session['logged_in_from_login'] = True
                    #     return redirect(url_for('old_user_preview'))
                    elif new_applicant_incomplete_form(email):
                        print('I am here')
                        flash('Your form is incomplete.', 'error')
                        return redirect(url_for('section1'))
                    elif check_final_approval(email):
                        session['final_approval'] = "accepted"
                        session['logged_in_from_login'] = True
                        return redirect(url_for('main_page'))
                    elif is_form_filled(email):
                        session['final_approval'] = "pending"
                        id = get_id_by_email(email)
                        return redirect(url_for('viewform', id=id))
                    else:
                        flash('Redirecting to login closed page for 2023.', 'info')
                        print('Hrer mate')
                        return redirect(url_for('app_form_info'))
                else:
                    flash('Invalid password. Please try again.', 'error')
                    return redirect(url_for('login'))
            else:
                flash('Invalid Email or Password. Please enter valid credentials.', 'error')
                return redirect(url_for('login'))

    except mysql.connector.Error as err:
        flash(f'Database error: {err}', 'error')
        return redirect(url_for('login'))
    except Exception as e:
        flash(f'An error occurred: {e}', 'error')
        return redirect(url_for('login'))

    return render_template('login.html', language=language, multilingual_content=multilingual_content)


# @app.route('/refresh_login_captcha', methods=['GET'])
# def refresh_login_captcha():
#     captcha_number = random.randrange(100000, 999999)
#     # img = ImageCaptcha(width=280, height=90)
#     captcha_text = str(captcha_number)
#     img.write(captcha_text, 'static/Images/captcha/user_captcha.png')
# 
#     # Update the captcha number in the session
#     session['captcha_number'] = captcha_number
# 
#     # Redirect back to the login page after refreshing the Captcha
#     return redirect(url_for('login'))


# @app.route('/forgot_password', methods=['GET', 'POST'])
# def forgot_password():                                     # ---------------- RENDER APPLICATION FORM
#     if 'language' in session:
#         language = session['language']
#     else:
#         language = 'marathi'
#
#     if request.method == 'POST':
#         email = request.form.get('email')
#
#         # Check if the email exists in the database
#         cnx = mysql.connector.connect(user=user, password=password, host=host, database=database)
#         cursor = cnx.cursor(dictionary=True)
#         cursor.execute("SELECT * FROM signup WHERE email = %s", (email,))
#         user = cursor.fetchone()
#         cnx.close()
#
#         first_name = user['first_name']
#
#         if user:
#             # Generate OTP
#             otp = random.randint(100000, 999999)
#             session['otp'] = otp
#             session['email'] = email
#
#             # Send OTP to user's email
#             if send_otp_email(email, first_name, otp):
#                 flash('An OTP has been sent to your email. Please check your inbox.', 'success')
#                 return redirect(url_for('verify_otp'))
#             else:
#                 flash('Error sending OTP. Please try again.', 'danger')
#                 return redirect(url_for('forgot_password'))
#         else:
#             flash('Email is not in our database. Please use the Registered Mail ID', 'danger')
#             return redirect(url_for('forgot_password'))
#
#     return render_template('login.html', language=language, multilingual_content=multilingual_content)
#
#
# def send_otp_email(email, first_name, otp):
#     msg_body = f'''
#                <!DOCTYPE html>
# <html lang="en">
#
# <head>
#     <meta charset="UTF-8">
#     <meta name="viewport" content="width=device-width, initial-scale=1.0">
#     <title>OTP</title>
#     <link rel="preconnect" href="https://fonts.googleapis.com">
#     <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
#     <link
#         href="https://fonts.googleapis.com/css2?family=Poppins:ital,wght@0,100;0,200;0,300;0,400;0,500;0,600;0,700;0,800;1,100;1,200;1,300;1,400;1,500;1,600;1,700;1,800&display=swap"
#         rel="stylesheet">
# </head>
# <body style="font-family: 'Poppins', sans-serif;">
#     <link
#     href="https://fonts.googleapis.com/css2?family=Poppins:ital,wght@0,100;0,200;0,300;0,400;0,500;0,600;0,700;0,800;1,100;1,200;1,300;1,400;1,500;1,600;1,700;1,800&display=swap"
#     rel="stylesheet">
#
#     <table align="center" border="0" cellpadding="0" cellspacing="0" width="550" bgcolor="white"
#         style="border:2px solid  #660000; border-radius: 5px; box-shadow: 5px 15px 30px #6666; font-family: 'Poppins', sans-serif;">
#         <tbody>
#             <tr>
#                 <td align="center">
#                     <table align="center" border="0" cellpadding="0" cellspacing="0" class="col-550" width="550">
#                         <tbody>
#                             <tr>
#                                 <td align="center" style="background-color:  #660000;
#                                 height: 50px; border-bottom: 2px solid #660000; border-radius: 5px 5px 0px 0;">
#
#                                     <a href="#" style="text-decoration: none;">
#                                         <p style="color:#ffff;
#                                        font-weight:bold;font-size: 20px; text-transform: uppercase; ">
#                                             Verify Email
#                                         </p>
#                                     </a>
#                                 </td>
#                             </tr>
#                         </tbody>
#                     </table>
#                 </td>
#             </tr>
#             <tr style="height: 300px; background: #fff;">
#                 <td align="center" style="border: none;
#                 border-bottom: 2px solid #660000;
#                 padding-right: 20px;padding-left:20px;  padding: 30px;">
#                     <img src="https://fellowship.trti-maha.in/static/assets/img/tick_animation.gif" width="60px" height="60px" alt="Tick mark">
#                     <p style="font-weight: bolder;font-size: 22px;
#                    letter-spacing: 0.025em;
#                    color:#660000;">
#                         Dear, {first_name}
#                         <br>
#                         Thank you for your interest in creating a user account for Online Fellowship Portal.
#                         <br>
#                         To activate your account, please enter OTP to the portal.
#                     </p>
#                     <p
#                         style="border: 1px solid transparent; padding: 15px 35px; width: fit-content;  text-align: center; border-radius: 8px; font-weight: bold; background: #660000; color:#fff; letter-spacing: 10px;">
#                         {otp}
#                     </p>
#                 </td>
#             </tr>
#
#             <tr style="display: inline-block; width: 100%;">
#                 <td style="height: 150px;
#                 padding: 20px;
#                 border: none;
#                 width: 10%;
#                 border-bottom: 2px solid transparent;
#                 border-radius: 0px 0px 5px 5px;
#                 background-color: #ffff; ">
#
#                     <h2 style="text-align: left;
#                     align-items: center; color: #660000;">
#                         This OTP will expire in 10 minutes
#                     </h2>
#                     <p class="data" style="text-align: justify-all;
#                    align-items: center;
#                    font-size: 15px; color: #660000;">
#                         If you did not request a for sign up, no further action is required.
#                     </p>
#                     <p class="data" style="text-align: justify-all;
#                    align-items: center;
#                    font-size: 15px;
#                    padding-bottom: 12px; color: #660000;">
#                         Thank you,<br>
#                         Fellowship,
#                     </p>
#                 </td>
#             </tr>
#             <tr style="display: inline-block; width: 100%;">
#                 <td style="max-height: 150px;
#                 padding: 40px 20px;
#                 border: none;
#                 width: 10%;
#                 border-top: 1.5px solid #ffff;
#                 border-radius: 0px 0px 5px 5px;
#                 background-color: #660000; ">
#
#                     <p style="color: #fff; font-size: 13px;">
#                         In case of any technical issue while filling online application form, please contact us
#                     </p>
#                     <a href="#" style="text-decoration: none; color: #660000; padding: 10px 25px; box-shadow: 0 0 10px #fff; border-radius: 5px; background:#fff;">Contact Us</a>
#                 </td>
#             </tr>
#         </tbody>
#     </table>
#
#            '''
#     msg = Message('Verify Email', sender='helpdesk@trti-maha.in', recipients=[email])
#     msg.html = msg_body
#     mail.send(msg)


@app.route('/viewform_old_users/<int:id>', methods=['GET', 'POST'])
def viewform_old_users(id):                                                   # -------------- VIEW STUDENT FORM
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    sql = """SELECT * FROM application_page WHERE id = %s"""
    cursor.execute(sql, (id,))
    # Fetch all records matching the query
    records = cursor.fetchall()
    print(records)
    # Close the cursor and database connection
    cursor.close()
    cnx.close()
    return render_template('form-view-olduser.html', records=records)


# VIEW STUDENT FORM
@app.route('/viewform/<int:id>', methods=['GET', 'POST'])
def viewform(id):
    user = HostConfig.user
    password = HostConfig.password
    database = HostConfig.database

    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    sql = """SELECT * FROM application_page WHERE id = %s"""
    cursor.execute(sql, (id,))
    # Fetch all records matching the query
    records = cursor.fetchall()

    topic_of_phd_utf8_list = []

    # Iterate over each record and extract the value of 'topic_of_phd' column
    for record in records:
        # Assuming 'topic_of_phd' is the 42nd column in your query result

        topic_of_phd = record['topic_of_phd']
        # Encode the topic_of_phd as UTF-8 and append it to the list
        topic_of_phd_utf8 = topic_of_phd.encode('utf-8')
        topic_of_phd_utf8_list.append(topic_of_phd_utf8)

    if records:
        user = records[0]['first_name'] + ' ' + records[0]['last_name']
    else:
        user = "Admin"

    # Close the cursor and database connection
    cursor.close()
    cnx.close()
    return render_template('form-view.html', records=records, topic_of_phd_utf8_list=topic_of_phd_utf8_list, user=user)


@app.route('/mainpage')
@auth
def main_page():
    user = HostConfig.user
    password = HostConfig.password
    database = HostConfig.database
    # -------------- APPLICATION LIST PAGE
    if session.pop('logged_in_from_login', None):
        flash('Logged in Successfully', 'success')
    email = session['email']
    print(" User Email: " + email)
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    sql = """SELECT first_name, middle_name, last_name, email, applicant_photo, applicant_id, application_date, id, 
    phd_registration_date, phd_registration_year, adhaar_number FROM application_page WHERE email = %s"""
    cursor.execute(sql, (email,))
    # Fetch all records matching the query
    records = cursor.fetchall()
    # Close the cursor and database connection
    sql = """SELECT year FROM signup WHERE email = %s"""
    cursor.execute(sql, (email,))
    # Fetch all records matching the query
    result = cursor.fetchall()
    cursor.close()
    cnx.close()
    # Process each record
    for record in records:
        if not record['applicant_photo']:
            session['applicant_photo'] = '/static/assets/img/default_user.png'
        else:
            session['applicant_photo'] = record['applicant_photo']

    if records:
        user = records[0]['first_name'] + ' ' + records[0]['last_name']
    else:
        user = "Admin"
    return render_template('application-list.html', records=records, result=result, user=user)
# ----------------- END LOGIN -----------------------


# ----------------- SIGN UP --------------------------------
def is_user_registered(email):                              # ---------------- CHECK IF EMAIL IS IN THE DATABASE
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    sql = "SELECT verified FROM signup WHERE email = %s"
    cursor.execute(sql, (email,))
    result = cursor.fetchone()
    cursor.close()
    cnx.close()
    return result


def send_email_verification(email, first_name, otp):
    msg_body = f'''
               <!DOCTYPE html>
<html lang="en">

<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>OTP</title>
    <link rel="preconnect" href="https://fonts.googleapis.com">
    <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
    <link
        href="https://fonts.googleapis.com/css2?family=Poppins:ital,wght@0,100;0,200;0,300;0,400;0,500;0,600;0,700;0,800;1,100;1,200;1,300;1,400;1,500;1,600;1,700;1,800&display=swap"
        rel="stylesheet">
</head>
<body style="font-family: 'Poppins', sans-serif;">
    <link
    href="https://fonts.googleapis.com/css2?family=Poppins:ital,wght@0,100;0,200;0,300;0,400;0,500;0,600;0,700;0,800;1,100;1,200;1,300;1,400;1,500;1,600;1,700;1,800&display=swap"
    rel="stylesheet">

    <table align="center" border="0" cellpadding="0" cellspacing="0" width="550" bgcolor="white"
        style="border:2px solid  #660000; border-radius: 5px; box-shadow: 5px 15px 30px #6666; font-family: 'Poppins', sans-serif;">
        <tbody>
            <tr>
                <td align="center">
                    <table align="center" border="0" cellpadding="0" cellspacing="0" class="col-550" width="550">
                        <tbody>
                            <tr>
                                <td align="center" style="background-color:  #660000;
                                height: 50px; border-bottom: 2px solid #660000; border-radius: 5px 5px 0px 0;">

                                    <a href="#" style="text-decoration: none;">
                                        <p style="color:#ffff;
                                       font-weight:bold;font-size: 20px; text-transform: uppercase; ">
                                            Verify Email
                                        </p>
                                    </a>
                                </td>
                            </tr>
                        </tbody>
                    </table>
                </td>
            </tr>
            <tr style="height: 300px; background: #fff;">
                <td align="center" style="border: none;
                border-bottom: 2px solid #660000; 
                padding-right: 20px;padding-left:20px;  padding: 30px;">
                    <img src="https://fellowship.trti-maha.in/static/assets/img/tick_animation.gif" width="60px" height="60px" alt="Tick mark">
                    <p style="font-weight: bolder;font-size: 22px;
                   letter-spacing: 0.025em;
                   color:#660000;">
                        Dear, {first_name}
                        <br>
                        Thank you for your interest in creating a user account for Online Fellowship Portal.
                        <br>
                        To activate your account, please enter OTP to the portal.
                    </p>
                    <p
                        style="border: 1px solid transparent; padding: 15px 35px; width: fit-content;  text-align: center; border-radius: 8px; font-weight: bold; background: #660000; color:#fff; letter-spacing: 10px;">
                        {otp}
                    </p>
                </td>
            </tr>

            <tr style="display: inline-block; width: 100%;">
                <td style="height: 150px;
                padding: 20px;
                border: none; 
                width: 10%;
                border-bottom: 2px solid transparent;
                border-radius: 0px 0px 5px 5px;
                background-color: #ffff; ">

                    <h2 style="text-align: left;
                    align-items: center; color: #660000;">
                        This OTP will expire in 10 minutes
                    </h2>
                    <p class="data" style="text-align: justify-all;
                   align-items: center; 
                   font-size: 15px; color: #660000;">
                        If you did not request a for sign up, no further action is required.
                    </p>
                    <p class="data" style="text-align: justify-all;
                   align-items: center; 
                   font-size: 15px;
                   padding-bottom: 12px; color: #660000;">
                        Thank you,<br>
                        Fellowship,
                    </p>
                </td>
            </tr>
            <tr style="display: inline-block; width: 100%;">
                <td style="max-height: 150px;
                padding: 40px 20px;
                border: none; 
                width: 10%;
                border-top: 1.5px solid #ffff;
                border-radius: 0px 0px 5px 5px;
                background-color: #660000; ">

                    <p style="color: #fff; font-size: 13px;">
                        In case of any technical issue while filling online application form, please contact us
                    </p>
                    <a href="#" style="text-decoration: none; color: #660000; padding: 10px 25px; box-shadow: 0 0 10px #fff; border-radius: 5px; background:#fff;">Contact Us</a>
                </td>
            </tr>
        </tbody>
    </table>

           '''
    msg = Message('Verify Email', sender='noreply_fellowship@trti-maha.in', recipients=[email])
    msg.html = msg_body
    mail.send(msg)


def hash_password(password):
    # Generate a salt and hash the password
    salt = bcrypt.gensalt()
    hashed_password = bcrypt.hashpw(password.encode('utf-8'), salt)
    return hashed_password.decode('utf-8')


@app.route('/signup', methods=['GET', 'POST'])
def signup():                                               # --------------------------  SIGN UP PAGE
    if 'language' in session:
        language = session['language']
    else:
        language = 'marathi'

    if request.method == 'POST':
        first_name = request.form['first_name']
        middle_name = request.form['middle_name']
        last_name = request.form['last_name']
        email = request.form['email']
        password = request.form['password']
        confirm_password = request.form['confirm_password']
        year = request.form['year']
        mobile_number = request.form['mobile_number']

        if not email or not password:
            flash('Please enter email and password.', 'error')
            return redirect(url_for('login'))

        # Check if the passwords match
        if password != confirm_password:
            flash('Please enter the Password Correctly as it does not match', 'error')
            return redirect(url_for('login'))

        if old_user(email):
            error = ('Please Login with the registered email ID and Password for your login will be Fellowship123. '
                     'Please change the password after login. ')
            return render_template('login.html', error=error)

        if is_user_registered(email):
            error = ('This email is already registered. Please use a different email or log in with an existing one.')
            return render_template('login.html', error=error, language=language, multilingual_content=multilingual_content)
        else:
            # Encrypt the password before storing it
            hashed_password = hash_password(password)
            unique_id = random.randint(100000, 999999)
            global otp
            otp = random.randint(100000, 999999)

            # Store user registration data in a session for verification
            registration_data = {
                'first_name': first_name,
                'middle_name': middle_name,
                'last_name': last_name,
                'email': email,
                'password': hashed_password,
                'confirm_password': confirm_password,
                'year': year,
                'unique_id': unique_id,
                'mobile_number': mobile_number
            }
            print('Sending Email')
            # send_email_verification(email, first_name, otp)
            print('Sending SMS')
            # send_sms(mobile_number, otp)
            insert_user_data(registration_data)
            flash('Your email is verified and registration is successful.', 'success')
            # return render_template('email_verify.html', email=email)
            return redirect(url_for('login'))
    error = ('This email is already registered. Please use a different email or log in with an existing one.',
                 'error')
    return render_template('login.html', error=error, language=language, multilingual_content=multilingual_content)


# @app.route('/refresh_signup_captcha', methods=['GET'])
# def refresh_signup_captcha():
#     captcha_number = random.randrange(100000, 999999)
#     img = ImageCaptcha(width=280, height=90)
#     captcha_text = str(captcha_number)
#     img.write(captcha_text, 'static/Images/captcha/user_captcha.png')
# 
#     # Update the captcha number in the session
#     session['captcha_number'] = captcha_number
# 
#     # Redirect back to the login page after refreshing the Captcha
#     return redirect(url_for('signup'))


# Define a function to insert user registration data into the database
def insert_user_data(registration_data):
    try:
        cnx = mysql.connector.connect(user=user, password=password,
                                      host=host,
                                      database=database)
        cursor = cnx.cursor()

        # Define your INSERT SQL statement with %s placeholders
        sql = "INSERT INTO signup (first_name, middle_name, last_name, email, password, confirm_password, year, unique_id, mobile_number) " \
              "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)"

        # Extract the data from the 'registration_data' dictionary
        data = (
            registration_data['first_name'],
            registration_data['middle_name'],
            registration_data['last_name'],
            registration_data['email'],
            registration_data['password'],
            registration_data['confirm_password'],
            registration_data['year'],
            registration_data['mobile_number'],
            registration_data['unique_id']
        )

        # Execute the SQL statement with the data
        cursor.execute(sql, data)

        # Commit the changes to the database
        cnx.commit()

        # Close the cursor and database connection
        cursor.close()
        cnx.close()

        return True  # Return True to indicate a successful insertion

    except mysql.connector.Error as err:
        print("MySQL Error: {}".format(err))
        return False  # Return False to indicate an error occurred during insertion


@app.route('/email_verify', methods=['GET', 'POST'])
def email_verify():
    if 'registration_data' not in session:
        flash('Session data not found. Please sign up again.', 'error')
        return redirect(url_for('signup'))

    user_otp = request.form['otp']

    if otp == int(user_otp):
        registration_data = session.get('registration_data')
        insert_user_data(registration_data)
        flash('Your email is verified and registration is successful.')
        return redirect(url_for('login'))
    else:
        error = 'You have entered the wrong OTP. Please enter the OTP again sent to your email'
        return render_template('email_verify.html', error=error)
# ----------------- END SIGN UP -----------------------


# ----------------- ADMIN LOGIN -----------------------
@app.route('/adminlogin', methods=['GET', 'POST'])
def admin_login():
    user = HostConfig.user
    password = HostConfig.password
    database = HostConfig.database
    # ------------------ ADMIN LOGIN
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)

    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        # Validate that both username and password are provided
        if not username or not password:
            flash('Please enter username and password.', 'error')
            cursor.close()
            cnx.close()
            return redirect(url_for('admin_login'))

        # Check if the user is an admin
        sql = "SELECT * FROM admin WHERE username=%s AND password=%s"
        cursor.execute(sql, (username, password))
        user = cursor.fetchone()
        print(user)
        cnx.commit()

        if not user:
            flash('Please enter valid username and password.', 'error')
            cursor.close()
            cnx.close()
            return redirect(url_for('admin_login'))

        session['user'] = user['username']

        if user:
            # Set session variable to indicate user is logged in
            session['logged_in'] = True
            # Close the connection and cursor
            cursor.close()
            cnx.close()
            return redirect(url_for('index'))
        else:
            error = 'Invalid username or password'
            # Close the connection and cursor
            cursor.close()
            cnx.close()
            flash('Please enter Valid Details to Login', 'error')
            return render_template('adminlogin.html', error=error)
    return render_template('adminlogin.html')
# -------------------- END ADMIN LOGIN ------------------------------



# --------------------- USERS PROFILE USER SIDE PAGE ---------------
@app.route('/user_profile', methods=['GET', 'POST'])
def user_profile(): 
    flash_msg_profile = None      
    flash_msg = None                                            # ----------------  USERS PROFILE USER SIDE PAGE
    email = session['email']
    if request.method == 'POST':
        if 'edit_profile' in request.form:
            submit_edit_profile()  # Call the edit profile function
            flash_msg_profile = "success"
        elif 'change_password' in request.form:
            change_password_user()  # Call the change password function
            flash_msg = "success"  # Flash a success message

    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)

    sql = """SELECT * FROM application_page WHERE email = %s"""
    cursor.execute(sql, (email,))
    records = cursor.fetchall()
    cursor.close()
    cnx.close()

    return render_template('users-profile.html', records=records, flash_msg=flash_msg, flash_msg_profile=flash_msg_profile)


def submit_edit_profile():                                          # ------------- SUBMIT EDIT PROFILE FOR MY PROFILE
    if request.method == 'POST':
        first_name = request.form['first_name']
        address = request.form['add_1']
        phone = request.form['mobile_number']
        email = session['email']
        cnx = mysql.connector.connect(user=user, password=password,
                                      host=host,
                                      database=database)
        cursor = cnx.cursor(dictionary=True)
        # Construct the SQL update query
        sql = f"UPDATE application_page SET first_name = '{first_name}', add_1 = '{address}', mobile_number = '{phone}' WHERE email = '{email}'"
        cursor.execute(sql)
        cnx.commit()
        cursor.close()
        cnx.close()


# def change_password_user():                                            # -------------- CHANGE PASSWORD FOR USER
#     if request.method == 'POST':
#         current_password = request.form['current_password']
#         print(current_password)
#         new_password = request.form['new_password']
#         confirm_password = request.form['confirm_password']
#         email = session['email']
#         cnx = mysql.connector.connect(user=user, password=password,
#                                       host=host,
#                                       database=database)
#         cursor = cnx.cursor(dictionary=True)
#
#         query = 'SELECT password FROM signup WHERE email = %s'
#         print("SQL Query:", query)
#         cursor.execute(query, (email,))
#         result = cursor.fetchone()
#
#         if result:
#             # Get the stored hashed password from the result dictionary
#             stored_password = result['password']
#
#             if stored_password != :
#                 # Passwords match, update the password
#                 sql = f"UPDATE signup SET password = '{new_password}', confirm_password = '{confirm_password}' WHERE email = '{email}'"
#                 cursor.execute(sql)
#                 cnx.commit()
#         else:
#             print("User not found")
# ----------------- END  USERS PROFILE USER SIDE PAGE  -------------------


#                                                ---- FINALLY APPROVED USER FUNCTIONALITY
# ---------------- USER AFTER FINAL APPROVAL -----------------------------
@app.route('/manage_profile_AA')
def manage_profile():
    user = HostConfig.user
    password = HostConfig.password
    database = HostConfig.database
    # ---------- MANAGE PROFILE
    email = session['email']

    if request.method == 'POST':
        if 'edit_profile' in request.form:
            submit_edit_profile()  # Call the edit profile function
        elif 'change_password' in request.form:
            change_password_user()  # Call the change password function
            flash("Password changed successfully")  # Flash a success message

    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)

    sql = """SELECT * FROM application_page WHERE email = %s"""
    cursor.execute(sql, (email,))
    records = cursor.fetchall()

    if records:
        user = records[0]['first_name'] + ' ' + records[0]['last_name']
    else:
        user = 'Admin'

    cursor.close()
    cnx.close()

    return render_template('users-profile.html', records=records, user=user)


def submit_edit_profile():                                      # ------------- SUBMIT EDIT PROFILE FOR MY PROFILE
    if request.method == 'POST':
        first_name = request.form['first_name']
        address = request.form['add_1']
        phone = request.form['mobile_number']
        email = session['email']
        cnx = mysql.connector.connect(user=user, password=password,
                                      host=host,
                                      database=database)
        cursor = cnx.cursor(dictionary=True)
        # Construct the SQL update query
        sql = f"UPDATE application_page SET first_name = '{first_name}', add_1 = '{address}', mobile_number = '{phone}' WHERE email = '{email}'"
        cursor.execute(sql)
        cnx.commit()
        cursor.close()
        cnx.close()


def change_password_user():
    if request.method == 'POST':
        current_password = request.form['current_password']
        new_password = request.form['new_password']
        confirm_password = request.form['confirm_password']
        email = session['email']

        # Connect to the database
        cnx = mysql.connector.connect(user=user, password=password, host=host, database=database)
        cursor = cnx.cursor(dictionary=True)

        # Retrieve hashed password from the database
        query = 'SELECT password FROM signup WHERE email = %s'
        cursor.execute(query, (email,))
        result = cursor.fetchone()

        if result:
            stored_password = result['password']
            # Check if current password matches stored password
            if stored_password == current_password:
                # Update the password
                update_query = 'UPDATE signup SET password = %s, confirm_password = %s WHERE email = %s'
                cursor.execute(update_query, (new_password, confirm_password, email))
                cnx.commit()
                send_password_change_email(email)
                flash("Password updated successfully.")
            else:
                flash("Incorrect current password.")
        else:
            flash("User not found.")

        # Close database connections
        cursor.close()
        cnx.close()

    return result


def send_password_change_email(email):
    msg = Message('Password Changed', sender='helpdesk@trti-maha.in', recipients=[email])

    msg_body = f''' 
    <!DOCTYPE html>
    <html lang="en">

    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Document</title>
        <style>
            @import url('https://fonts.googleapis.com/css2?family=Montserrat:wght@100;200;300;400;500;600;700;800;900&display=swap');


        </style>
    </head>

    <body style="background: radial-gradient(rgb(235,236,240),rgb(235,236,240)); padding: 50px;  margin: 0;  font-family: 'Montserrat', sans-serif;">

        <table style="width: 90%; margin: auto; min-width: 480px; border-radius: 10px; overflow: hidden; width: 540px; border-spacing: 0;">
            <tr style="background: #F5F5F5; border-radius: 10px; ">
                <td style="text-align: center;">
                    <img src="https://fellowship.trti-maha.in/static/assets/img/logo/logo-new.png" style="width: 80px;"
                        alt="TRTI logo">

                </td>
                <td style="text-align: center;">
                    <img src="https://fellowship.trti-maha.in/static/assets/img/fellow_logo_1.png" style="width: 70px;"
                        alt="Fellowship Logo">
                </td>
                <td style="text-align: center;">
                    <h3 style="color: #175E97; font-weight: 700; ">FELLOWSHIP</h3>
                </td>
                <td style="text-align: center;">
                    <h3 style="color: #B71540; font-weight: 600; font-size: 15px;">HOME</h3>
                </td>
                <td style="text-align: center;">
                    <h3 style="color: #B71540; font-weight: 600; font-size: 15px;">CONTACT US</h3>
                </td>
            </tr>
            <tr>
                <td colspan="5"
                    style="background: linear-gradient(rgba(169,27,96,0.4), rgba(169,27,96,0.4)), url('https://fellowship.trti-maha.in/static/assets/img/banner_award.jpg'); width: 100%; height: 30vh; background-size: cover; background-repeat: no-repeat;">
                    <h2
                        style="text-transform: uppercase; text-align: center; font-size: 50px; color: #fff; width: 90%; letter-spacing: 5px; margin: auto; ">
                        Password is Changed
                    </h2>
                </td>
            </tr>
            <tr>
                <td colspan="5" style="background: #fff; padding: 40px;">
                    <h4
                        style="text-transform: uppercase; text-align: center; font-size: 20px; font-weight: 600; color: #A91B60;">
                        Hello, User - {email}</h4>
                    <h4
                        style="text-transform: uppercase; text-align: center; font-size: 18px; font-weight: 600; color: #A91B60; line-height: 28px;">
                        Your Password was changed successfully!!
                    </h4>
                    <h4
                        style="text-transform: uppercase; text-align: center; font-size: 18px; font-weight: 600; color: #A91B60; line-height: 28px;">
                    </h4>
                </td>
            </tr>
            <tr>
                <td colspan="5" style="background: #fff; padding: 40px; border-top: 1px solid rgb(235,236,240);">
                    <h4
                        style="text-transform: uppercase; text-align: center; font-size: 12px; font-weight: 600; color: #A91B60; line-height: 18px;">
                        In case of any technical issue while filling online application form, please contact on toll free
                        helpline number 18002330444 (From 09:45 AM to 06:30 PM </h4>
                    <p style="color:#A91B60; font-size: 11px; font-weight: 600; text-align: center;">
                        This is a system generated mail. Please do not reply to this Email ID
                    </p>
                </td>
            </tr>
            <tr>
                <td colspan="5" style="background: #A91B60; padding: 10px 40px; ">
                    <span style="color: #fff; font-size: 11px; ">Visit us at <a href="https://trti.maharashtra.gov.in"
                            target="_blank" style="color: #fff;">https://trti.maharashtra.gov.in</a> </span>
                    <img src="https://static.vecteezy.com/system/resources/thumbnails/027/395/710/small/twitter-brand-new-logo-3-d-with-new-x-shaped-graphic-of-the-world-s-most-popular-social-media-free-png.png" style="width: 32px; height: 32px; float: right; " alt="Twitter Logo">
                    <img src="https://cdn3.iconfinder.com/data/icons/social-network-30/512/social-06-512.png"
                        style="width: 32px;  height: 32px;  float: right; margin-right: 12px; background: transparent;"
                        alt="Youtube Logo">
                    <img src="https://freelogopng.com/images/all_img/1658030214facebook-logo-hd.png"
                        style="width: 32px; height: 32px; float: right; margin-right: 12px; " alt="Facebok Logo">
                </td>
            </tr>
        </table>

    </body>

    </html> 
    '''
    msg = Message('Password Changed', sender='noreply_fellowship@trti-maha.in', recipients=[email])
    msg.html = msg_body
    mail.send(msg)

# -------------------- END MANAGE PROFILE ---------------------------------

def get_base_url():
    base_url = request.url_root
    return base_url


# --------------------- AWARD LETTER ----------------------------------
def generate_award_letter_2022(data, filename):
    class PDF(FPDF):
        header_added = False  # To track whether the header is added to the first page

        def header(self):
            if not self.header_added:

                var = get_base_url()
                print(var)
                # Add a header
                self.set_font("Arial", "B", 12)
                self.image('/var/www/fellowship/fellowship/BartiFellowship/BartiFellowship/static/Images/satya.png', 94, 10, 20)  # Replace with the path to your small imag
                # Calculate the width of the image
                image_width = 100  # Assuming the width of the image is 100 (adjust if different)
                # Calculate the position for "Government of Maharashtra" text
                text_x_position = self.get_x()  # Get current X position
                text_y_position = self.get_y() + 20  # Set Y position below the image
                # Set cursor position
                self.set_xy(text_x_position, text_y_position)
                self.image('/var/www/fellowship/fellowship/BartiFellowship/BartiFellowship/static/Images/newtrtiImage.png', 10, 10, 45)  # Replace with the path to your symbol image
                self.image('/var/www/fellowship/fellowship/BartiFellowship/BartiFellowship/static/Images/mahashasn_new.png', 155, 10, 45)  # Replace with the path to your symbol image
                self.ln(5)
                self.ln(0)  # Reduce the space below the address
                self.cell(0, 5, "Government of Maharashtra", align="C", ln=True)

                self.cell(0, 5, "Tribal Research & Training Institute", align="C", ln=True)
                self.cell(0, 10, "28, Queens Garden, Pune - 411001", align="C", ln=True)
                self.dashed_line(10, self.get_y(), 200, self.get_y(), dash_length=3, space_length=1)

                self.ln(5)  # Adjust this value to control the space after the line
                self.set_font("Arial", "B", size=10)
                self.cell(0, 10,
                          " Fellowship Award Letter",
                          align="C", ln=True)
                self.ln(2)  # Adjust this value to control the space after the line

                self.rotate(45)  # Rotate the text by 45 degrees
                self.set_font('Arial', '', 45)
                self.set_text_color(192, 192, 192)
                self.text(-30, 195, "STRF FELLOWSHIP")  # Use text instead of rotated_text
                self.rotate(0)  # Reset the rotation to 0 degrees

                self.header_added = True  # Set to True after adding the header

        def to_name(self, data):
            # AWARD LETTER in the center

            # To, and Dear Candidate aligned to the left
            self.set_font("Arial", "", size=10)
            self.cell(0, 10, "To,", ln=True)
            self.set_font("Arial", "B", size=11)
            self.cell(0, 10, data['first_name'] + ' ' + data['middle_name'] + ' ' + data['last_name'], ln=True)


        def insert_static_data(self, data):
            # Insert your static data here
            self.set_font("Arial", "B", size=10)
            self.cell(0, 10, "Dear Candidate,", ln=True)
            self.set_font("Arial", "", 10)
            registration_year = data['phd_registration_year']
            fiscal_year = f"{registration_year} - {registration_year + 1}"
            self.multi_cell(0, 7, "         We are delighted to inform you that you have been selected for the award of "
                                  "a Fellowship for the year " + fiscal_year +
                            " for Ph.D. The Fellowship amount will be effective from the date of registration for Ph.D. Congratulations! "
                            )
            self.ln(3)  # Adjust this value to control the space before static data
            self.multi_cell(0, 7,
                            "       TRTI reserves all the rights to add terms and conditions as and when required, and "
                            "candidates are required to accept any changes in the terms and conditions of the fellowship."
                            )
            self.ln(3)  # Adjust this value to control the space before static data
            self.multi_cell(0, 7,
                            "       Attached with this letter is an undertaking stating that all the information provided "
                            "for the document verification is true to the best of my knowledge. Any discrepancy found "
                            "may result in the cancellation of the Fellowship. Please note that failure to submit the "
                            "undertaking will be assumed as non-acceptance of this offer, and the Fellowship will not "
                            "be  processed. "
                            )
            self.ln(3)  # Adjust this value to control the space before static data
            self.multi_cell(0, 7,
                            "       We believe this Fellowship will not only provide financial support but also contribute"
                            " to your academic growth. It will enable you to conduct research on your subject and "
                            "foster excellence in academia. Moreover, it will empower you to become an advocate for"
                            " equality, social justice, a contributor to peace, harmony and happiness within various"
                            " disadvantaged sections of society. "
                            )
            self.multi_cell(0, 20, "Wish you all the best. ")
            self.set_x(150)  # Adjust the x-coordinate as needed
            # self.image('static/Images/signature_awardletter.png', 20, 230, 30)
            self.image('/var/www/fellowship/fellowship/BartiFellowship/BartiFellowship/static/Images/sonanwanesir_signature.png', 125, 210, 50)
            self.ln(5)  # Adjust this value to control the space after static data

        def footer(self):
            # Add a footer
            self.set_y(-15)
            self.set_font("arial", "B", 8)
            self.cell(0, 10, f" {self.page_no()} ", align="C")

            # Center-align the "TRTI" text
            self.cell(0, 10, " TRTI  |  Fellowship | 2023 - 2024 ", align="R")

    pdf = PDF()
    pdf.add_page()
    pdf.header()
    pdf.to_name(data)
    # Insert static data
    pdf.insert_static_data(data)
    # Save the PDF to a file
    pdf.output(filename)


def generate_award_letter_2023(data, filename):
    class PDF(FPDF):
        header_added = False  # To track whether the header is added to the first page

        def header(self):
            if not self.header_added:

                var = get_base_url()
                print(var)
                # Add a header
                self.set_font("Arial", "B", 12)
                # self.image('static/Images/satya.png', 94, 10, 20)  # Replace with the path to your small imag
                self.image('/var/www/fellowship/fellowship/BartiFellowship/BartiFellowship/static/Images/satya.png', 94, 10, 20)  # Replace with the path to your small imag
                # Calculate the width of the image
                image_width = 100  # Assuming the width of the image is 100 (adjust if different)
                # Calculate the position for "Government of Maharashtra" text
                text_x_position = self.get_x()  # Get current X position
                text_y_position = self.get_y() + 20  # Set Y position below the image
                # Set cursor position
                self.set_xy(text_x_position, text_y_position)
                # self.image('static/Images/newtrtiImage.png', 10, 10, 45)  # Replace with the path to your symbol image
                self.image('/var/www/fellowship/fellowship/BartiFellowship/BartiFellowship/static/Images/newtrtiImage.png', 10, 10, 45)  # Replace with the path to your symbol image
                # self.image('static/Images/mahashasn_new.png', 155, 10, 45)  # Replace with the path to your symbol image
                self.image('/var/www/fellowship/fellowship/BartiFellowship/BartiFellowship/static/Images/mahashasn_new.png', 155, 10, 45)  # Replace with the path to your symbol image
                self.ln(5)
                self.ln(0)  # Reduce the space below the address
                self.cell(0, 5, "Government of Maharashtra", align="C", ln=True)

                self.cell(0, 5, "Tribal Research & Training Institute", align="C", ln=True)
                self.cell(0, 10, "28, Queens Garden, Pune - 411001", align="C", ln=True)
                self.dashed_line(10, self.get_y(), 200, self.get_y(), dash_length=3, space_length=1)

                self.ln(5)  # Adjust this value to control the space after the line
                self.set_font("Arial", size=10)
                self.cell(0, 10, "No.: Research-2024/Case.No 9/Desk-4/1832", ln=False)  # Add the number on the left without a line break

                # Move to the right for the date
                fellowship_year = data['fellowship_awarded_date']
                self.cell(0, 10, f"{fellowship_year}", align="R", ln=True)  # Add the date on the right with a line break

                self.set_font("Arial", "B", size=10)
                self.cell(0, 10,
                          " Fellowship Award Letter",
                          align="C", ln=True)
                self.ln(2)  # Adjust this value to control the space after the line

                self.rotate(45)  # Rotate the text by 45 degrees
                self.set_font('Arial', '', 45)
                self.set_text_color(192, 192, 192)
                self.text(-30, 195, "STRF FELLOWSHIP")  # Use text instead of rotated_text
                self.rotate(0)  # Reset the rotation to 0 degrees

                self.header_added = True  # Set to True after adding the header

        def to_name(self, data):
            # AWARD LETTER in the center

            # To, and Dear Candidate aligned to the left
            self.set_font("Arial", "", size=10)
            self.cell(0, 10, "To,", ln=True)
            self.set_font("Arial", "B", size=11)
            self.cell(0, 10, data['first_name'] + ' ' + data['middle_name'] + ' ' + data['last_name'], ln=True)


        def insert_static_data(self, data):
            # Insert your static data here
            self.set_font("Arial", "B", size=10)
            self.cell(0, 10, "Dear Candidate,", ln=True)
            self.set_font("Arial", "", 10)
            registration_year = data['phd_registration_year']
            fiscal_year = f"{registration_year} - {registration_year + 1}"
            self.multi_cell(0, 7, " We are delighted to inform you that you have been selected for the award of "
                                  "a Fellowship for the year " + fiscal_year +
                            " for Ph.D. The Fellowship amount will be effective from the date of registration for Ph.D. Congratulations! "
                            )
            self.ln(3)  # Adjust this value to control the space before static data
            self.multi_cell(0, 7,
                            "       TRTI reserves all the rights to add terms and conditions as and when required, and "
                            "candidates are required to accept any changes in the terms and conditions of the fellowship."
                            )
            self.ln(3)  # Adjust this value to control the space before static data
            self.multi_cell(0, 7,
                            "       Attached with this letter is an undertaking stating that all the information provided "
                            "for the document verification is true to the best of my knowledge. Any discrepancy found "
                            "may result in the cancellation of the Fellowship. Please note that failure to submit the "
                            "undertaking will be assumed as non-acceptance of this offer, and the Fellowship will not "
                            "be  processed. "
                            )
            self.ln(3)  # Adjust this value to control the space before static data
            self.multi_cell(0, 7,
                            "       We believe this Fellowship will not only provide financial support but also contribute"
                            " to your academic growth. It will enable you to conduct research on your subject and "
                            "foster excellence in academia. Moreover, it will empower you to become an advocate for"
                            " equality, social justice, a contributor to peace, harmony and happiness within various"
                            " disadvantaged sections of society. "
                            )
            self.multi_cell(0, 20, "Wish you all the best. ")
            self.set_x(150)  # Adjust the x-coordinate as needed
            # self.image('static/Images/signature_awardletter.png', 20, 230, 30)
            # self.image('static/Images/chanchalamam_signature.png', 125, 210, 50)
            self.image('/var/www/fellowship/fellowship/BartiFellowship/BartiFellowship/static/Images/chanchalamam_signature.png', 125, 210, 50)
            self.ln(5)  # Adjust this value to control the space after static data

        def footer(self):
            # Add a footer
            self.set_y(-15)
            self.set_font("arial", "B", 8)
            self.cell(0, 10, f" {self.page_no()} ", align="C")

            # Center-align the "TRTI" text
            self.cell(0, 10, " TRTI  |  Fellowship | 2023 - 2024 ", align="R")

    pdf = PDF()
    pdf.add_page()
    pdf.header()
    pdf.to_name(data)
    # Insert static data
    pdf.insert_static_data(data)
    # Save the PDF to a file
    pdf.output(filename)


def generate_award_letter_Aadesh15(data, filename):
    class PDF(FPDF):
        header_added = False  # To track whether the header is added to the first page

        def header(self):
            if not self.header_added:

                var = get_base_url()
                print(var)
                # Add a header
                self.set_font("Arial", "B", 12)
                # self.image('static/Images/satya.png', 94, 10, 20)  # Replace with the path to your small imag
                self.image('/var/www/fellowship/fellowship/BartiFellowship/BartiFellowship/static/Images/satya.png', 94, 10, 20)  # Replace with the path to your small imag
                # Calculate the width of the image
                image_width = 100  # Assuming the width of the image is 100 (adjust if different)
                # Calculate the position for "Government of Maharashtra" text
                text_x_position = self.get_x()  # Get current X position
                text_y_position = self.get_y() + 20  # Set Y position below the image
                # Set cursor position
                self.set_xy(text_x_position, text_y_position)
                # self.image('static/Images/newtrtiImage.png', 10, 10, 45)  # Replace with the path to your symbol image
                self.image('/var/www/fellowship/fellowship/BartiFellowship/BartiFellowship/static/Images/newtrtiImage.png', 10, 10, 45)  # Replace with the path to your symbol image
                # self.image('static/Images/mahashasn_new.png', 155, 10, 45)  # Replace with the path to your symbol image
                self.image('/var/www/fellowship/fellowship/BartiFellowship/BartiFellowship/static/Images/mahashasn_new.png', 155, 10, 45)  # Replace with the path to your symbol image
                self.ln(5)
                self.ln(0)  # Reduce the space below the address
                self.cell(0, 5, "Government of Maharashtra", align="C", ln=True)

                self.cell(0, 5, "Tribal Research & Training Institute", align="C", ln=True)
                self.cell(0, 10, "28, Queens Garden, Pune - 411001", align="C", ln=True)
                self.dashed_line(10, self.get_y(), 200, self.get_y(), dash_length=3, space_length=1)

                self.ln(5)  # Adjust this value to control the space after the line
                self.set_font("Arial", size=10)
                self.cell(0, 10, "No.: Research-2024/Case.No 9/Desk-4/2364", ln=False)  # Add the number on the left without a line break

                # Move to the right for the date
                self.cell(0, 10, "Date: 08-08-2024", align="R", ln=True)  # Add the date on the right with a line break

                self.set_font("Arial", "B", size=10)
                self.cell(0, 10,
                          " Fellowship Award Letter",
                          align="C", ln=True)
                self.ln(2)  # Adjust this value to control the space after the line

                self.rotate(45)  # Rotate the text by 45 degrees
                self.set_font('Arial', '', 45)
                self.set_text_color(192, 192, 192)
                self.text(-30, 195, "STRF FELLOWSHIP")  # Use text instead of rotated_text
                self.rotate(0)  # Reset the rotation to 0 degrees

                self.header_added = True  # Set to True after adding the header

        def to_name(self, data):
            # AWARD LETTER in the center

            # To, and Dear Candidate aligned to the left
            self.set_font("Arial", "", size=10)
            self.cell(0, 10, "To,", ln=True)
            self.set_font("Arial", "B", size=11)
            self.cell(0, 10, data['first_name'] + ' ' + data['middle_name'] + ' ' + data['last_name'], ln=True)


        def insert_static_data(self, data):
            # Insert your static data here
            self.set_font("Arial", "B", size=10)
            self.cell(0, 10, "Dear Candidate,", ln=True)
            self.set_font("Arial", "", 10)
            registration_year = data['phd_registration_year']
            fiscal_year = f"{registration_year} - {registration_year + 1}"
            self.multi_cell(0, 7, "         We are delighted to inform you that you have been selected for the award of "
                                  "a Fellowship for the year " + fiscal_year +
                            " for Ph.D. The Fellowship amount will be effective from the date of registration for Ph.D. Congratulations! "
                            )
            self.ln(3)  # Adjust this value to control the space before static data
            self.multi_cell(0, 7,
                            "       TRTI reserves all the rights to add terms and conditions as and when required, and "
                            "candidates are required to accept any changes in the terms and conditions of the fellowship."
                            )
            self.ln(3)  # Adjust this value to control the space before static data
            self.multi_cell(0, 7,
                            "       Attached with this letter is an undertaking stating that all the information provided "
                            "for the document verification is true to the best of my knowledge. Any discrepancy found "
                            "may result in the cancellation of the Fellowship. Please note that failure to submit the "
                            "undertaking will be assumed as non-acceptance of this offer, and the Fellowship will not "
                            "be  processed. "
                            )
            self.ln(3)  # Adjust this value to control the space before static data
            self.multi_cell(0, 7,
                            "       We believe this Fellowship will not only provide financial support but also contribute"
                            " to your academic growth. It will enable you to conduct research on your subject and "
                            "foster excellence in academia. Moreover, it will empower you to become an advocate for"
                            " equality, social justice, a contributor to peace, harmony and happiness within various"
                            " disadvantaged sections of society. "
                            )
            self.multi_cell(0, 20, "Wish you all the best. ")
            self.set_x(150)  # Adjust the x-coordinate as needed
            # self.image('static/Images/signature_awardletter.png', 20, 230, 30)
            # self.image('static/Images/chanchalamam_signature.png', 125, 210, 50)
            self.image('/var/www/fellowship/fellowship/BartiFellowship/BartiFellowship/static/Images/chanchalamam_signature.png', 125, 210, 50)
            self.ln(5)  # Adjust this value to control the space after static data

        def footer(self):
            # Add a footer
            self.set_y(-15)
            self.set_font("arial", "B", 8)
            self.cell(0, 10, f" {self.page_no()} ", align="C")

            # Center-align the "TRTI" text
            self.cell(0, 10, " TRTI  |  Fellowship | 2023 - 2024 ", align="R")

    pdf = PDF()
    pdf.add_page()
    pdf.header()
    pdf.to_name(data)
    # Insert static data
    pdf.insert_static_data(data)
    # Save the PDF to a file
    pdf.output(filename)


@app.route('/award_letter_AA')
def award_letter_AA():
    try:
        email = session['email']
        output_filename = '/var/www/fellowship/fellowship/BartiFellowship/BartiFellowship/static/pdf_application_form/award_letter.pdf'
        # output_filename = 'static/pdf_application_form/award_letter.pdf'
        cnx = mysql.connector.connect(user='icswebapp', password=password, host=host, database=database)
        cursor = cnx.cursor(dictionary=True)
        cursor.execute(
            "SELECT id, applicant_photo, applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number,"
            " email, date_of_birth, gender, age, caste, your_caste, marital_status, dependents, state, district,"
            " taluka, village, city, add_1, add_2, pincode, ssc_passing_year,"
            " ssc_percentage, ssc_school_name, hsc_passing_year, hsc_percentage, hsc_school_name,"
            " graduation_passing_year, graduation_percentage, graduation_school_name, phd_passing_year,"
            " phd_percentage, phd_school_name,have_you_qualified, name_of_college, name_of_guide, topic_of_phd,"
            " concerned_university, faculty, phd_registration_date, phd_registration_year,"
            " family_annual_income, "
            " income_certificate_number, issuing_authority, domicile, domicile_certificate, domicile_number,"
            " caste_certf, issuing_district, caste_issuing_authority, salaried, disability, type_of_disability,"
            " father_name, mother_name, work_in_government, bank_name, account_number, ifsc_code,"
            " account_holder_name, application_date FROM application_page WHERE email = %s", (email,))
        data = cursor.fetchone()

        cursor.execute(
            "SELECT id, phd_registration_year FROM application_page WHERE email = %s",
            (email,))
        result = cursor.fetchone()
        year = result['phd_registration_year']
        id = result['id']

        if(id in (310, 317, 330, 348, 365, 389, 408, 415, 420, 421, 426, 440, 442, 451, 480)):
            generate_award_letter_Aadesh15(data, output_filename)
        elif(year >= 2023):
            generate_award_letter_2023(data, output_filename)
        else:
            generate_award_letter_2022(data, output_filename)

        # Serve the generated PDF as a response
        with open(output_filename, "rb") as pdf_file:
            response = Response(pdf_file.read(), content_type="application/pdf")
            response.headers['Content-Disposition'] = 'inline; filename=award_letter.pdf'
    except BrokenPipeError:
        # Handle broken pipe error, e.g., log it
        pass

    return response


@app.route('/generate_award_letter_AA/<string:email>')
def generate_award_letter_AA(email):
    try:
        # email = session['email']
        output_filename = '/var/www/fellowship/fellowship/BartiFellowship/BartiFellowship/static/pdf_application_form/award_letter.pdf'
        # output_filename = 'static/pdf_application_form/award_letter.pdf'
        cnx = mysql.connector.connect(user='icswebapp', password=password, host=host, database=database)
        cursor = cnx.cursor(dictionary=True)
        cursor.execute(
            "SELECT id, applicant_photo, applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number,"
            " email, date_of_birth, gender, age, caste, your_caste, marital_status, dependents, state, district,"
            " taluka, village, city, add_1, add_2, pincode, ssc_passing_year,"
            " ssc_percentage, ssc_school_name, hsc_passing_year, hsc_percentage, hsc_school_name,"
            " graduation_passing_year, graduation_percentage, graduation_school_name, phd_passing_year,"
            " phd_percentage, phd_school_name,have_you_qualified, name_of_college, name_of_guide, topic_of_phd,"
            " concerned_university, faculty, phd_registration_date, phd_registration_year,"
            " family_annual_income, "
            " income_certificate_number, issuing_authority, domicile, domicile_certificate, domicile_number,"
            " caste_certf, issuing_district, caste_issuing_authority, salaried, disability, type_of_disability,"
            " father_name, mother_name, work_in_government, bank_name, account_number, ifsc_code,"
            " account_holder_name, application_date FROM application_page WHERE email = %s", (email,))
        data = cursor.fetchone()

        cursor.execute(
            "SELECT id, phd_registration_year FROM application_page WHERE email = %s",
            (email,))
        result = cursor.fetchone()
        year = result['phd_registration_year']
        id = result['id']

        if(id in (310, 317, 330, 348, 365, 389, 408, 415, 420, 421, 426, 440, 442, 451, 480)):
            generate_award_letter_Aadesh15(data, output_filename)
        elif(year >= 2023):
            generate_award_letter_2023(data, output_filename)
        else:
            generate_award_letter_2022(data, output_filename)

        # Serve the generated PDF as a response
        with open(output_filename, "rb") as pdf_file:
            response = Response(pdf_file.read(), content_type="application/pdf")
            response.headers['Content-Disposition'] = 'inline; filename=award_letter.pdf'
    except BrokenPipeError:
        # Handle broken pipe error, e.g., log it
        pass

    return response
# -------------------- END AWARD LETTER --------------------------------

  
# --------------------- JOINING LETTER ----------------------------------
@app.route('/joining_report_AA', methods=['GET', 'POST'])
def joining_report_AA():
    user = HostConfig.user
    password = HostConfig.password
    database = HostConfig.database
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    records = None
    joining_report = None
    joining_date = None  # Initialize with a default value
    email = session.get('email', None)
    existing_report = []
    # Fetch the joining date from the database

    email = session['email']
    cursor.execute("SELECT first_name, last_name, phd_registration_date, joining_date, joining_report FROM application_page WHERE email = %s", (email,))
    result = cursor.fetchone()
    print(result)
    if result:
        joining_date = result['phd_registration_date']
        existing_report = result['joining_report']
    if request.method == 'POST':
        first_name = result['first_name']
        last_name = result['last_name']
        existing_report = result['joining_report']
        joining_report = save_file_joining_report(request.files['joining_report'], first_name, last_name)
        print(joining_report)
        print(existing_report)

        update_query = "UPDATE application_page SET joining_report=%s, joining_date=%s  WHERE email = %s"
        cursor.execute(update_query, (joining_report, joining_date, email))
        cnx.commit()
        flash('Joining Report Uploaded Successfully', 'success')
    cursor.close()
    cnx.close()
    if result:
        user = result['first_name'] + ' ' + result['last_name']
    else:
        user = 'Admin'
    return render_template('joining_report_AA.html', records=records, joining_date=joining_date, joining_report=joining_report,
                           result=result, existing_report=existing_report, user=user)

# -------------------- END JOINING LETTER --------------------------------


# --------------------- PRESENTY LETTER ----------------------------------
@app.route('/presenty_AA', methods=['GET', 'POST'])
def presenty_AA():
    user = HostConfig.user
    password = HostConfig.password
    database = HostConfig.database
    if 'email' not in session:
        return redirect('/login')

    email = session['email']
    cnx = mysql.connector.connect(user=user, password=password, host=host, database=database)
    cursor = cnx.cursor(dictionary=True)
    existing_reports = []  # Initialize existing_reports here

    cursor.execute("SELECT first_name, last_name FROM application_page WHERE email = %s", (email,))
    result = cursor.fetchone()

    if result:
        user = result['first_name'] + ' ' + result['last_name']
    else:
        user = 'Admin'

    if request.method == 'POST':
        # Handle file uploads and save them to the database
        for i in range(1, 61):  # Assuming up to 60 months (5 years) of reports
            report = request.files.get(f'monthly_report{i}')
            if report:
                first_name = result['first_name']
                last_name = result['last_name']
                # Save the uploaded report to a directory
                report_path = save_file_presenty_report(report, first_name, last_name)
                # Update the database with the report path
                cursor.execute(f"UPDATE award_letter SET monthly_report{i} = %s WHERE email = %s",
                               (report_path, email))
        cnx.commit()

    # Fetch the saved reports for the user
    cursor.execute(
        f"SELECT monthly_report1, monthly_report2, monthly_report3, monthly_report4, monthly_report5, "
        f"monthly_report6, monthly_report7, monthly_report8, monthly_report9, monthly_report10, "
        f"monthly_report11, monthly_report12, monthly_report13, monthly_report14, monthly_report15, "
        f"monthly_report16, monthly_report17, monthly_report18, monthly_report19, monthly_report20, "
        f"monthly_report21, monthly_report22, monthly_report23, monthly_report24, monthly_report25, "
        f"monthly_report26, monthly_report27, monthly_report28, monthly_report29, monthly_report30, "
        f"monthly_report31, monthly_report32, monthly_report33, monthly_report34, monthly_report35, "
        f"monthly_report36, monthly_report37, monthly_report38, monthly_report39, monthly_report40, "
        f"monthly_report41, monthly_report42, monthly_report43, monthly_report44, monthly_report45, "
        f"monthly_report46, monthly_report47, monthly_report48, monthly_report49, monthly_report50, "
        f"monthly_report51, monthly_report52, monthly_report53, monthly_report54, monthly_report55, "
        f"monthly_report56, monthly_report57, monthly_report58, monthly_report59, monthly_report60 "
        f"FROM award_letter WHERE email = %s",
        (email,))
    reports = cursor.fetchone()
    # Count the number of submitted reports
    # Calculate the submitted count and get the list of submitted documents
    submitted_count = 0
    submitted_documents = []

    for i in range(1, 61):
        if reports[f'monthly_report{i}']:
            submitted_count += 1
            submitted_documents.append(f'monthly_report{i}')
    index = 1
    # Now you have the submitted count and the list of submitted documents
    print("Submitted Count:", submitted_count)
    print("Submitted Documents:", submitted_documents)
    # Fetch the joining date for the user
    cursor.execute("SELECT phd_registration_date FROM application_page WHERE email = %s", (email,))
    result = cursor.fetchone()

    i = submitted_count + 1

    # Use parameterized placeholders for the column names
    query = f"SELECT submission_date_report{i}, submission_day_report{i} FROM award_letter WHERE email = %s"
    cursor.execute(query, (email,))
    record = cursor.fetchone()
    
    query_award = f"SELECT * FROM award_letter WHERE email = %s"
    cursor.execute(query_award, (email,))
    award_letter_data = cursor.fetchone()
    
    if result:
        joining_date = result['phd_registration_date']
        print(joining_date)
        start_dates = [datetime.combine(joining_date, datetime.min.time()) + timedelta(days=i * 30) for i in range(60)]
        end_dates = [start_date + timedelta(days=30) for start_date in start_dates]

        if record:
            # Use existing submission date and day
            submitted_date = record[f'submission_date_report{i}']
            submitted_day = record[f'submission_day_report{i}']
        else:
            # Generate new submission date and day
            current_datetime = datetime.now()
            submitted_date = current_datetime.strftime('%Y-%m-%d')
            submitted_day = current_datetime.strftime('%A')  # Full weekday name (e.g., Monday)
    else:
        joining_date = None
        submitted_date = None
        submitted_day = None
        start_dates = []
        end_dates = []

    # Update the database
    update_query = f"""
        UPDATE award_letter
        SET submission_date_report{i} = %s,
            submission_day_report{i} = %s
        WHERE email = %s
    """
    cursor.execute(update_query, (submitted_date, submitted_day, email))
    cnx.commit()
        # Zip the start_dates and end_dates
    zipped_dates = list(zip(start_dates, end_dates))

    return render_template('presenty_AA.html', zipped_dates=zipped_dates, reports=reports,
                           start_dates=start_dates, end_dates=end_dates, joining_date=joining_date,
                           submitted_count=submitted_count, submitted_date=submitted_date, submitted_day=submitted_day,
                           submitted_documents=submitted_documents, index=index, user=user, award_letter_data=award_letter_data)


# -------------------- END PRESENTY LETTER --------------------------------


# --------------------- HALF YEARLY REPORT ----------------------------------
@app.route('/half_yearly_rep_AA', methods=['GET', 'POST'])
def half_yearly_rep_AA():
    if 'email' not in session:
        return redirect('/login')

    email = session['email']
    cnx = mysql.connector.connect(user=user, password=password, host=host, database=database)
    cursor = cnx.cursor(dictionary=True)
    existing_reports = []  # Initialize existing_reports here
    submitted_documents = []
    cursor.execute("SELECT first_name, last_name FROM application_page WHERE email = %s", (email,))
    result = cursor.fetchone()
    if result:
        user = result['first_name'] + ' ' + result['last_name']
    else:
        user = 'Admin'

    if request.method == 'POST':
        # Handle file uploads and save them to the database
        report_paths = []
        for i in range(1, 11):
            report = request.files.get(f'half_yearly_report{i}')
            if report:
                first_name = result['first_name']
                last_name = result['last_name']
                # Save the uploaded report to a directory
                # You can use your own logic to save the report and get the file path
                report_path = save_file_half_yearly(report, first_name, last_name)
                report_paths.append((f'half_yearly_report{i}', report_path))

        # Update the database with the report paths
        for report_field, report_path in report_paths:
            cursor.execute(f"UPDATE application_page SET {report_field} = %s WHERE email = %s",
                           (report_path, email))
        cnx.commit()

    # Fetch the saved reports for the user
    cursor.execute(
        f"SELECT half_yearly_report1, half_yearly_report2, half_yearly_report3, half_yearly_report4, half_yearly_report5, "
        f"half_yearly_report6, half_yearly_report7, half_yearly_report8, half_yearly_report9, half_yearly_report10 "
        f"FROM application_page WHERE email = %s",
        (email,))
    reports = cursor.fetchone()
    print(reports)
    # Count the number of submitted reports

    submitted_count = sum([1 for i in range(1, 11) if reports[f'half_yearly_report{i}']])
    for i in range(1, 11):
        if reports.get(f'half_yearly_report{i}'):
            submitted_documents.append(f'half_yearly_report{i}')

    # Fetch the joining date for the user
    cursor.execute("SELECT phd_registration_date FROM application_page WHERE email = %s", (email,))
    result = cursor.fetchone()

    if result:
        joining_date = result['phd_registration_date']
        start_dates = [joining_date + timedelta(days=i * 30 * 6) for i in range(10)]
        end_dates = [start_date + timedelta(days=30 * 6) for start_date in start_dates]
    else:
        joining_date = None
        start_dates = []
        end_dates = []

    cursor.close()
    cnx.close()
    return render_template('half_yearly_rep_AA.html', reports=reports, joining_date=joining_date,
                           start_dates=start_dates, end_dates=end_dates, submitted_count=submitted_count,
                           submitted_documents=submitted_documents, user=user)

# -------------------- END HALF YEARLY REPORT --------------------------------


# --------------------- HRA RENT AGREEMENT ----------------------------------
@app.route('/rent_agreement_AA', methods=['GET', 'POST'])
def rent_agreement_AA():
    if 'email' not in session:
        # Redirect to the login page if the user is not logged in
        return redirect('/login')

    email = session['email']
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    existing_reports = []  # Initialize existing_reports here
    submitted_documents = []
    cursor.execute("SELECT first_name, last_name FROM application_page WHERE email = %s", (email,))
    result = cursor.fetchone()
    
    if result:
        user = result['first_name'] + ' ' + result['last_name']
    else:
        user = 'Admin'
    
    if request.method == 'POST':
        # Handle file uploads and save them to the database
        report_paths = []
        for i in range(1, 6):
            report = request.files.get(f'rent_agreement{i}')
            if report:
                first_name = result['first_name']
                last_name = result['last_name']
                # Save the uploaded report to a directory
                # You can use your own logic to save the report and get the file path
                report_path = save_file_rent_agreement(report, first_name, last_name)
                report_paths.append(
                    (f'rent_agreement{i}', report_path))  # Store the field name along with the file path

        # Update the database with the report paths
        for report_field, report_path in report_paths:
            cursor.execute(f"UPDATE application_page SET {report_field} = %s WHERE email = %s",
                           (report_path, email))
        cnx.commit()

    # Fetch the saved reports for the user
    cursor.execute(
        f"SELECT rent_agreement1, rent_agreement2, rent_agreement3, rent_agreement4, rent_agreement5 FROM application_page WHERE email = %s",
        (email,))
    reports = cursor.fetchone()
    # Count the number of submitted reports
    submitted_count = sum([1 for i in range(1, 6) if reports[f'rent_agreement{i}']])
    for i in range(1, 11):
        if reports.get(f'rent_agreement{i}'):
            submitted_documents.append(f'rent_agreement{i}')
    # Fetch the joining date for the user
    cursor.execute("SELECT phd_registration_date FROM application_page WHERE email = %s", (email,))
    result = cursor.fetchone()

    if result:
        joining_date = result['phd_registration_date']
        start_dates = [joining_date + timedelta(days=i * 365) for i in range(5)]
        end_dates = [start_date + timedelta(days=365) for start_date in start_dates]
    else:
        joining_date = None
        start_dates = []
        end_dates = []

    cursor.close()
    cnx.close()

    return render_template('rent_agreement_AA.html', reports=reports, joining_date=joining_date,
                           start_dates=start_dates, end_dates=end_dates, submitted_count=submitted_count,
                           submitted_documents=submitted_documents, user=user)
# -------------------- END HRA RENT AGREEMENT --------------------------------


# --------------------- WITHDRAW FELLOWSHIP ----------------------------------
@app.route('/with_from_fellowship_AA', methods=['GET', 'POST'])
def with_from_fellowship_AA():
    email = session.get('email')
    if email is None:
        # Handle the case where the user is not logged in
        return redirect('/login')

    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)

    cursor.execute("SELECT first_name, last_name FROM application_page WHERE email = %s", (email,))
    result = cursor.fetchone()

    if result:
        user = result['first_name'] + ' ' + result['last_name']
    else:
        user = 'Admin'
    
    if request.method == 'POST':
        # Handle the form submission
        update_query = "UPDATE signup SET request_withdrawal = %s, withdrawal_request_date = %s WHERE email = %s"
        withdrawal_request_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        cursor.execute(update_query, (True, withdrawal_request_date, email))
        cnx.commit()

    # Fetch the relevant information
    select_query = "SELECT request_withdrawal, withdrawal_request_date FROM signup WHERE email = %s"
    cursor.execute(select_query, (email,))
    result = cursor.fetchone()

    cursor.close()
    cnx.close()
    return render_template('with_from_fellowship_AA.html', result=result, user=user)


@app.route('/withdrawed_application_admin', methods=['GET', 'POST'])
def withdrawed_application_admin():
    if not session.get('logged_in'):
        # Redirect to the admin login page if the user is not logged in
        return redirect(url_for('admin_login'))
    email = session.get('email')
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    # Fetch the relevant information
    select_query = """
        SELECT s.*, ps.*, ap.*
        FROM signup s
        LEFT JOIN payment_sheet ps ON s.email = ps.email
        LEFT JOIN application_page ap ON s.email = ap.email
        WHERE s.request_withdrawal = '1'
    """
    cursor.execute(select_query)
    result_set = cursor.fetchall()

    result_set_by_year = defaultdict(list)
    phd_date = None
    var = None
    for result in result_set:
        if result['request_withdrawal'] == 1:
            var = 'Raised for Withdrawal'
            registration_year = result['phd_registration_year']
            result_set_by_year[registration_year].append(result)
    sorted_years = sorted(result_set_by_year.keys(), reverse=True)

    application_2023 = """
                SELECT s.*, ps.*, ap.*
                FROM signup s
                LEFT JOIN payment_sheet ps ON s.email = ps.email
                LEFT JOIN application_page ap ON s.email = ap.email
                WHERE s.request_withdrawal = '1' and ap.phd_registration_year = '2023'
            """
    cursor.execute(application_2023)
    result2023 = cursor.fetchall()

    application_2022 = """
            SELECT s.*, ps.*, ap.*
            FROM signup s
            LEFT JOIN payment_sheet ps ON s.email = ps.email
            LEFT JOIN application_page ap ON s.email = ap.email
            WHERE s.request_withdrawal = '1' and ap.phd_registration_year = '2022'
        """
    cursor.execute(application_2022)
    result2022 = cursor.fetchall()

    application_2021 = """
                SELECT s.*, ps.*, ap.*
                FROM signup s
                LEFT JOIN payment_sheet ps ON s.email = ps.email
                LEFT JOIN application_page ap ON s.email = ap.email
                WHERE s.request_withdrawal = '1' and ap.phd_registration_year = '2021'
            """
    cursor.execute(application_2021)
    result2021 = cursor.fetchall()

    application_2020 = """
                SELECT s.*, ps.*, ap.*
                FROM signup s
                LEFT JOIN payment_sheet ps ON s.email = ps.email
                LEFT JOIN application_page ap ON s.email = ap.email
                WHERE s.request_withdrawal = '1' and ap.phd_registration_year = '2020'
            """
    cursor.execute(application_2020)
    result2020 = cursor.fetchall()

    cursor.close()
    cnx.close()
    return render_template('withdrawed_application_admin.html', result_set=result_set, var=var,
                           sorted_years=sorted_years, result_set_by_year=result_set_by_year,
                           result2023=result2023, result2022=result2022, result2021=result2021,
                           result2020=result2020)


# -------------------- END WITHDRAW FELLOWSHIP --------------------------------


# --------------------- CHANGE GUIDE ----------------------------------
@app.route('/change_guide_AA', methods=['GET', 'POST'])
def change_guide_AA():
    if 'email' not in session:
        # Redirect to the login page if the user is not logged in
        return redirect('/login')

    email = session.get('email')
        # Update the database to change the name of the guide
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("SELECT first_name, last_name FROM application_page WHERE email = %s", (email,))
    result = cursor.fetchone()

    if result:
        user = result[0] + ' ' + result[1]
    else:
        user = 'Admin'

    if request.method == 'POST':
        # Get the user's email from the session
        email = session['email']

        new_guide_name = request.form.get('new_guide_name')
        cursor = cnx.cursor()

        # Execute an SQL query to update the giude's name
        update_query = "UPDATE application_page SET name_of_guide = %s WHERE email = %s"
        cursor.execute(update_query, (new_guide_name, email))

        # Commit the transaction and close the cursor and connection
        cnx.commit()
        cursor.close()
        return redirect('/change_guide_AA')

    cursor = cnx.cursor()

    # SQL query to fetch the present guide name from database
    # cursor.execute("SELECT name_of_guide FROM application_page")
    email = session['email']
    cursor.execute("SELECT name_of_guide FROM application_page WHERE email = %s", (email,))
    result = cursor.fetchone()

    if result is not None:
        guide_name = result[0]
    else:
        guide_name = "No Guide Found"

    # cursor.close()
    cnx.close()

    return render_template('change_guide_AA.html', guide_name=guide_name, user=user)
# -------------------- CHANGE GUIDE  --------------------------------


# --------------------- CHANGE CENTER ----------------------------------
@app.route('/change_center_AA', methods=['GET', 'POST'])
def change_center_AA():
    if 'email' not in session:
        # Redirect to the login page if the user is not logged in

        return redirect('/login')

        # Update the database to change the name of the center
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)

    if request.method == 'POST':
        # Get the user's email from the session
        email = session['email']
        new_center_name = request.form.get('new_center_name')
        cursor = cnx.cursor()
        update_query = "UPDATE application_page SET name_of_college = %s WHERE email = %s"
        cursor.execute(update_query, (new_center_name, email))

        # Commit the transaction and close the cursor and connection
        cnx.commit()
        cursor.close()
        return redirect('/change_center_AA')
    cursor = cnx.cursor()

    # SQL query to fetch the present center name from database
    email = session['email']
    cursor.execute("SELECT name_of_college FROM application_page WHERE email = %s", (email,))
    result = cursor.fetchone()

    if result is not None:
        center_name = result[0]
    else:
        center_name = "No Center Found"

    # cursor.close()
    cnx.close()
    return render_template('change_center_AA.html', center_name=center_name)
# -------------------- CHANGE CENTER  ---------------------------------


#--------------------- PHD AWARD ----------------------------------
@app.route('/phd_award_AA', methods=['GET', 'POST'])
def phd_award_AA():
    if 'email' not in session:
        return redirect('/login')  # Redirect to login if user not logged in
    email = session['email']
    # Establish database connection
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT first_name, last_name FROM application_page WHERE email = %s", (email,))
    result = cursor.fetchone()

    if result:
        user = result['first_name'] + ' ' + result['last_name']
    else:
        user = 'Admin'

    if request.method == 'POST':
        if 'phd_award' in request.files:
            file = request.files['phd_award']
            if file.filename != '':
                file_paths = []  # Create an empty list to store tuples

                first_name = result['first_name']
                last_name = result['last_name']

                # Save the uploaded report to a directory and get the file path
                file_path = save_file_pdf_cert(file, first_name, last_name)

                # Append a tuple to the list
                file_paths.append((f'phd_award', file_path))


                # Get user's email from session
                email = session['email']

                cursor = cnx.cursor()
                update_query = "UPDATE application_page SET phd_award = %s WHERE email = %s"
                cursor.execute(update_query, (file_path, email))

                cnx.commit()
                cursor.close()
                return redirect('/phd_award_AA')

    email = session['email']
    cursor = cnx.cursor()
    cursor.execute("SELECT phd_award FROM application_page WHERE email = %s", (email,))
    result = cursor.fetchone()

    if result is not None:
        phd_award = result[0]
    else:
        phd_award = None

    cnx.close()
    return render_template('phd_award_AA.html',phd_award=phd_award, user=user)
# -------------------- END PHD AWARD -----------------------------


#--------------------- UPLOAD THESIS ----------------------------------
@app.route('/upload_thesis_AA', methods=['GET', 'POST'])
def upload_thesis_AA():
    if 'email' not in session:
        return redirect('/login')  # Redirect to login if user not logged in
    email = session['email']
    # Establish database connection
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT first_name, last_name FROM application_page WHERE email = %s", (email,))
    result = cursor.fetchone()
    
    if result:
        user = result['first_name'] + ' ' + result['last_name']
    else:
        user = 'Admin'
    
    if request.method == 'POST':
        if 'phd_thesis' in request.files:
            file = request.files['phd_thesis']
            if file.filename != '':
                file_paths = []  # Create an empty list to store tuples

                first_name = result['first_name']
                last_name = result['last_name']

                # Save the uploaded report to a directory and get the file path
                file_path = save_file_uplaod_thesis(file, first_name, last_name)

                # Append a tuple to the list
                file_paths.append((f'phd_award', file_path))

                # Get user's email from session
                email = session['email']

                cursor = cnx.cursor()
                update_query = "UPDATE application_page SET phd_thesis = %s WHERE email = %s"
                cursor.execute(update_query, (file_path, email))

                cnx.commit()
                cursor.close()
                return redirect('/upload_thesis_AA')

    email = session['email']
    cursor = cnx.cursor()
    cursor.execute("SELECT phd_thesis FROM application_page WHERE email = %s", (email,))
    result = cursor.fetchone()

    if result is not None:
        phd_thesis = result[0]
    else:
        phd_thesis = None

    cnx.close()
    return render_template('upload_thesis_AA.html', phd_thesis=phd_thesis, user=user)
# -------------------- END UPLOAD THESIS -----------------------------


# def save_file(file):
#     if file and allowed_file(file.filename):
#         filename = secure_filename(file.filename)
#         file_path = os.path.join(UPLOAD_FOLDER, filename)
#         file.save(file_path)


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'png', 'jpg', 'jpeg', 'gif', 'bmp', 'pdf'}


#                                               ---- ADMIN LEVEL 1, 2, 3
# ------------------ ADMIN LEVEL 1 ----------------
@app.route('/adminPage', methods=['GET', 'POST'])
def admin_level_1():
    if not session.get('logged_in'):
        # Redirect to the admin login page if the user is not logged in
        return redirect(url_for('admin_login'))
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)

    if request.method == 'POST':
        # Check if a form was submitted
        if 'accept' in request.form:
            # Handle accepting an application
            applicant_id = request.form['accept']
            update_status_admin(applicant_id, 'accepted')
        elif 'reject' in request.form:
            # Handle rejecting an application
            applicant_id = request.form['reject']
            update_status_admin(applicant_id, 'rejected')
            cursor.execute("SELECT email, first_name, status FROM application_page WHERE applicant_id = %s",
                           (applicant_id,))
            user_data = cursor.fetchone()

            if user_data:
                email = user_data[0]
                full_name = f"{user_data[1]} {user_data[2]}"

                send_email_rejection(email, full_name, 'Rejected', applicant_id)
        # Commit the changes to the database
        cnx.commit()

    cursor.execute("SELECT * FROM application_page where form_filled='1' and phd_registration_year>='2023' ")
    data = cursor.fetchall()
    print(data)
    cursor.close()
    cnx.close()
    return render_template('admin_main.html', data=data)
# ------------------ END ADMIN LEVEL 1 ----------------


#----------------------  EXPORT TO EXCEL Level One Admin-------------------------------------------
@app.route('/export_level_one_applications', methods=['GET', 'POST'])
def export_level_one_applications():
    if not session.get('logged_in'):
        # Redirect to the admin login page if the user is not logged in
        return redirect(url_for('admin_login'))
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute(
        "  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
        "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status,"
        "dependents, state, district, taluka, village, city, add_1, add_2, pincode,"
        "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total,"
        "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total,"
        "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream,"
        "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
        "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
        "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
        "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income,"
        "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka,"
        "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number,"
        "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
        "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
        "type_of_disability, father_name, mother_name, work_in_government, gov_department,"
        "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
        " where phd_registration_year = '2023' and status = 'accepted' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    ws.append(
        ['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
         'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
         'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
         'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
         'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts',
         'HSC Total',
         'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream',
         'Graduation Attempts',
         'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts',
         'PG Total',
         'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
         'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
         'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
         'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
         'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
         'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District',
         'Caste Issuing Authority',
         'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government',
         'Government Department',
         'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR'])

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=New_Applications.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response


# ----------------- ADMIN LEVEL 2 ----------------------
@app.route('/level_two_admin', methods=['GET', 'POST'])
def level_two_admin():
    if not session.get('logged_in'):
        # Redirect to the admin login page if the user is not logged in
        return redirect(url_for('admin_login'))
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    if request.method == 'POST':
        # Check if a form was submitted
        if 'accept' in request.form:
            # Handle accepting an application
            applicant_id = request.form['accept']
            update_scrutiny_admin(applicant_id, 'accepted')
        elif 'reject' in request.form:
            # Handle rejecting an application
            applicant_id = request.form['reject']
            update_scrutiny_admin(applicant_id, 'rejected')
            cursor.execute("SELECT email, first_name, last_name, scrutiny_status FROM application_page WHERE applicant_id = %s",
                           (applicant_id,))
            user_data = cursor.fetchone()

            if user_data:
                email = user_data['email']
                full_name = user_data['first_name'] + ' ' + user_data['last_name']

                # Send an email to the user
                send_email_rejection(email, full_name, 'Rejected', applicant_id)
        # Commit the changes to the database
        cnx.commit()

    cursor.execute("SELECT * FROM application_page WHERE status='accepted' and phd_registration_year>='2023' ")
    data = cursor.fetchall()
    print(data)
    cursor.close()
    cnx.close()
    return render_template('level_two_admin.html', data=data)
# ------------------- END ADMIN LEVEL 2 ------------------


#----------------------  EXPORT TO EXCEL Level Two Admin-------------------------------------------
@app.route('/export_level_two_applications', methods=['GET', 'POST'])
def export_level_two_applications():
    if not session.get('logged_in'):
        # Redirect to the admin login page if the user is not logged in
        return redirect(url_for('admin_login'))
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute(
        "  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
        "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status,"
        "dependents, state, district, taluka, village, city, add_1, add_2, pincode,"
        "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total,"
        "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total,"
        "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream,"
        "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
        "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
        "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
        "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income,"
        "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka,"
        "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number,"
        "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
        "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
        "type_of_disability, father_name, mother_name, work_in_government, gov_department,"
        "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
        " where phd_registration_year = '2023' and scrutiny_status = 'accepted' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    # ws.append(['applicant_id','email','first_name','last_name','application_date'])

    ws.append(
        ['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
         'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
         'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
         'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
         'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts',
         'HSC Total',
         'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream',
         'Graduation Attempts',
         'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts',
         'PG Total',
         'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
         'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
         'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
         'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
         'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
         'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District',
         'Caste Issuing Authority',
         'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government',
         'Government Department',
         'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR'])

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Level_Two_Applications.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response


#-------------------- ADMIN LEVEL 3 ----------------------
@app.route('/level_three_admin', methods=['GET', 'POST'])
def level_three_admin():
    if not session.get('logged_in'):
        # Redirect to the admin login page if the user is not logged in
        return redirect(url_for('admin_login'))
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    final_approval = request.get_data()
    if request.method == 'POST':
        # Check if a form was submitted
        if 'accept' in request.form:
            # Handle accepting an application
            applicant_id = request.form['accept']
            today = date.today()
            day = today.day
            month = today.month
            year = today.year
            print("Today's date:", today)
            print(f"Day: {day}, Month: {month}, Year: {year}")
            update_final_appr_admin(applicant_id, 'accepted', day, month, year)
            # Fetch the applicant's email and full name from the database
            cursor.execute("SELECT email, first_name, last_name, final_approval FROM application_page WHERE applicant_id = %s",
                           (applicant_id,))
            user_data = cursor.fetchone()

            if user_data:
                email = user_data['email']
                full_name = f"{user_data['first_name']} {user_data['last_name']}"

                # Send an email to the user
                send_email_approval(email, full_name, 'Accepted', applicant_id)

        elif 'reject' in request.form:
            # Handle rejecting an application
            applicant_id = request.form['reject']
            today = date.today()
            day = today.day
            month = today.month
            year = today.year
            update_final_appr_admin(applicant_id, 'rejected', day, month, year)
            cursor.execute("SELECT email, first_name, last_name, final_approval FROM application_page WHERE applicant_id = %s",
                           (applicant_id,))
            user_data = cursor.fetchone()
            if user_data:
                email = user_data['email']
                full_name =user_data['first_name'] + " " + user_data['last_name']

                # Send an email to the user
                send_email_rejection(email, full_name, 'Rejected', applicant_id)
        # Commit the changes to the database
        cnx.commit()

    cursor.execute("SELECT * FROM application_page WHERE scrutiny_status='accepted' and fellowship_awarded_year>='2023'")
    data = cursor.fetchall()
    cursor.close()
    cnx.close()
    return render_template('level_three_admin.html', data=data)


#----------------------  EXPORT TO EXCEL Level Three Admin-------------------------------------------
@app.route('/export_level_three_admin', methods=['GET', 'POST'])
def export_level_three_admin():
    if not session.get('logged_in'):
        # Redirect to the admin login page if the user is not logged in
        return redirect(url_for('admin_login'))
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute(
        "  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
        "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status,"
        "dependents, state, district, taluka, village, city, add_1, add_2, pincode,"
        "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total,"
        "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total,"
        "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream,"
        "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
        "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
        "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
        "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income,"
        "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka,"
        "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number,"
        "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
        "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
        "type_of_disability, father_name, mother_name, work_in_government, gov_department,"
        "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
        " where phd_registration_year = '2023' and final_approval = 'accepted' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    # ws.append(['applicant_id','email','first_name','last_name','application_date'])

    ws.append(
        ['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
         'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
         'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
         'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
         'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts',
         'HSC Total',
         'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream',
         'Graduation Attempts',
         'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts',
         'PG Total',
         'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
         'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
         'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
         'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
         'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
         'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District',
         'Caste Issuing Authority',
         'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government',
         'Government Department',
         'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR'])

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Level_Three_Applications.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response


#---------------By Aditya---Rejected applications at level 1--------------------------
@app.route('/rejected_applications_l1', methods=['GET', 'POST'])
def rejected_at_level1():
    if not session.get('logged_in'):
    # Redirect to the admin login page if the user is not logged in
        return redirect(url_for('admin_login'))
    cnx = mysql.connector.connect(user=user, password=password,
                                    host=host,
                                    database=database)
    cursor = cnx.cursor(dictionary=True)
    status = request.get_data()

    cursor.execute("SELECT * FROM application_page WHERE status='rejected'")
    data = cursor.fetchall()
    print(data)
    cursor.close()
    cnx.close()
    return render_template('rejected_applications_l1.html', data=data)

#---------------By Aditya---Rejected applications at level 2--------------------------
@app.route('/rejected_applications_l2', methods=['GET', 'POST'])
def rejected_at_level2():
    if not session.get('logged_in'):
    # Redirect to the admin login page if the user is not logged in
        return redirect(url_for('admin_login'))
    cnx = mysql.connector.connect(user=user, password=password,
                                    host=host,
                                    database=database)
    cursor = cnx.cursor(dictionary=True)
    scrutiny_status = request.get_data()

    cursor.execute("SELECT * FROM application_page WHERE scrutiny_status='rejected'")
    data = cursor.fetchall()
    print(data)
    cursor.close()
    cnx.close()
    return render_template('rejected_applications_l2.html', data=data)

# ------------------ END ADMIN LEVEL 3 -------------------


# ----------------- SEND EMAIL REJECTION & ACCEPTANCE -------------------
def send_email_rejection(email, full_name, status, applicant_id):
    # Email content in HTML format
    msg = Message('Application Status Changed', sender='helpdesk@trti-maha.in', recipients=[email])
    # msg.body = msg.body = "Hi, " + full_name + "\n Your Status for Fellowship : " + status + \
    #                       "\n Unfortunately the status of your application has changed to Rejected!!" + \
    #                       "\n Please login to view the status as Accepted for Fellowship" + \
    #                       "\n Your Application ID = " + applicant_id

    msg_body = f''' 
<!DOCTYPE html>
<html lang="en">

<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Document</title>
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Montserrat:wght@100;200;300;400;500;600;700;800;900&display=swap');

      
    </style>
</head>

<body style="background: radial-gradient(rgb(235,236,240),rgb(235,236,240)); padding: 50px;  margin: 0;  font-family: 'Montserrat', sans-serif;">

    <table style="width: 90%; margin: auto; min-width: 480px; border-radius: 10px; overflow: hidden; width: 540px; border-spacing: 0;">
        <tr style="background: #F5F5F5; border-radius: 10px; ">
            <td style="text-align: center;">
                <img src="https://fellowship.trti-maha.in/static/assets/img/logo/logo-new.png" style="width: 80px;"
                    alt="TRTI logo">

            </td>
            <td style="text-align: center;">
                <img src="https://fellowship.trti-maha.in/static/assets/img/fellow_logo_1.png" style="width: 70px;"
                    alt="Fellowship Logo">
            </td>
            <td style="text-align: center;">
                <h3 style="color: #175E97; font-weight: 700; ">FELLOWSHIP</h3>
            </td>
            <td style="text-align: center;">
                <h3 style="color: #B71540; font-weight: 600; font-size: 15px;">HOME</h3>
            </td>
            <td style="text-align: center;">
                <h3 style="color: #B71540; font-weight: 600; font-size: 15px;">CONTACT US</h3>
            </td>
        </tr>
        <tr>
            <td colspan="5"
                style="background: linear-gradient(rgba(169,27,96,0.4), rgba(169,27,96,0.4)), url('https://fellowship.trti-maha.in/static/assets/img/banner_award.jpg'); width: 100%; height: 30vh; background-size: cover; background-repeat: no-repeat;">
                <h2
                    style="text-transform: uppercase; text-align: center; font-size: 50px; color: #fff; width: 90%; letter-spacing: 5px; margin: auto; ">
                    Thanks For Applying
                </h2>
            </td>
        </tr>
        <tr>
            <td colspan="5" style="background: #fff; padding: 40px;">
                <h4
                    style="text-transform: uppercase; text-align: center; font-size: 20px; font-weight: 600; color: #A91B60;">
                    Hello, { full_name }</h4>
                <h4
                    style="text-transform: uppercase; text-align: center; font-size: 18px; font-weight: 600; color: #A91B60; line-height: 28px;">
                    Your Status for Fellowship : {status}
                    Unfortunately the status of your application has changed to Rejected!!
                </h4>
                <h4
                    style="text-transform: uppercase; text-align: center; font-size: 18px; font-weight: 600; color: #A91B60; line-height: 28px;">
                    Please Login To View The Status As Rejected For Fellowship
                </h4>
                <h4
                    style="text-transform: uppercase; text-align: center; font-size: 18px; font-weight: 600; color: #A91B60; line-height: 28px;">
                    Your Application ID:</h4>
                <p
                    style="text-align: center; padding: 25px; border: 3px solid #ECB322; color: #ECB322; font-weight: 700; letter-spacing: 10px; font-size: 20px;">{ applicant_id }</p>
            </td>
        </tr>
        <tr>
            <td colspan="5" style="background: #fff; padding: 40px; border-top: 1px solid rgb(235,236,240);">
                <h4
                    style="text-transform: uppercase; text-align: center; font-size: 12px; font-weight: 600; color: #A91B60; line-height: 18px;">
                    In case of any technical issue while filling online application form, please contact on toll free
                    helpline number 18002330444 (From 09:45 AM to 06:30 PM </h4>
                <p style="color:#A91B60; font-size: 11px; font-weight: 600; text-align: center;">
                    This is a system generated mail. Please do not reply to this Email ID
                </p>
            </td>
        </tr>
        <tr>
            <td colspan="5" style="background: #A91B60; padding: 10px 40px; ">
                <span style="color: #fff; font-size: 11px; ">Visit us at <a href="https://trti.maharashtra.gov.in"
                        target="_blank" style="color: #fff;">https://trti.maharashtra.gov.in</a> </span>
                <img src="https://static.vecteezy.com/system/resources/thumbnails/027/395/710/small/twitter-brand-new-logo-3-d-with-new-x-shaped-graphic-of-the-world-s-most-popular-social-media-free-png.png" style="width: 32px; height: 32px; float: right; " alt="Twitter Logo">
                <img src="https://cdn3.iconfinder.com/data/icons/social-network-30/512/social-06-512.png"
                    style="width: 32px;  height: 32px;  float: right; margin-right: 12px; background: transparent;"
                    alt="Youtube Logo">
                <img src="https://freelogopng.com/images/all_img/1658030214facebook-logo-hd.png"
                    style="width: 32px; height: 32px; float: right; margin-right: 12px; " alt="Facebok Logo">
            </td>
        </tr>
    </table>

</body>

</html> 
'''
    msg = Message('Application Status Changed', sender='noreply_fellowship@trti-maha.in', recipients=[email])
    msg.html = msg_body
    mail.send(msg)


def send_email_approval(email, full_name, status, applicant_id):
    base_url = request.url_root
# email_body = render_template('email_template.html', full_name=full_name, status=status,  applicant_id=applicant_id)
    # Construct the HTML email body
    msg_body = f'''
       <!DOCTYPE html>
<html lang="en">

<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Document</title>
    <style>
        @import url('https:                                                                                                                    f  fv  //fonts.googleapis.com/css2?family=Montserrat:wght@100;200;300;400;500;600;700;800;900&display=swap');
    </style>
</head>

<body style="background: radial-gradient(rgb(235,236,240),rgb(235,236,240));  margin: 0; font-family: 'Montserrat', sans-serif;  overflow: auto; padding:50px; width:100%;">

    <table style="width: 90%; margin: auto; min-width: 480px; width: 540px;  border-spacing: 0;">
        <tr style="background: #F5F5F5; border-radius: 10px; overflow: hidden;">
            <td style="text-align: center;">
                <img src="https://fellowship.trti-maha.in/static/assets/img/logo/logo-new.png" style="width: 80px;" alt="TRTI logo">

            </td>
            <td style="text-align: center;">
                <img src="https://fellowship.trti-maha.in/static/assets/img/fellow_logo_1.png" style="width: 70px;" alt="Fellowship Logo">
            </td>
            <td style="text-align: center;">
                <h3 style="color: #175E97; font-weight: 700; ">FELLOWSHIP</h3>
            </td>
            <td style="text-align: center;">
                <h3 style="color: #B71540; font-weight: 600; font-size: 15px;">HOME</h3>
            </td>
            <td style="text-align: center;">
                <h3 style="color: #B71540; font-weight: 600; font-size: 15px;">CONTACT US</h3>
            </td>
        </tr>
        <tr>
            <td colspan="5"
                style="background: linear-gradient(rgba(169,27,96,0.4), rgba(169,27,96,0.4)), url('https://fellowship.trti-maha.in/static/assets/img/banner_award.jpg'); width: 100%; height: 30vh; background-size: cover; background-repeat: no-repeat;">
                <h2 style="text-transform: uppercase; text-align: center; font-size: 50px; color: #fff;">Congratulations
                </h2>
            </td>
        </tr>
        <tr>
            <td colspan="5" style="background: #fff; padding: 40px;">
                <h4
                    style="text-transform: uppercase; text-align: center; font-size: 20px; font-weight: 600; color: #A91B60;">
                    Hello, {full_name}</h4>
                <h4
                    style="text-transform: uppercase; text-align: center; font-size: 18px; font-weight: 600; color: #A91B60; line-height: 28px;">
                    Congratulations The Status Of Your Application Has Changed To Accepted!! Please Login To View The
                    Status As Accepted For Fellowship</h4>
                <h4
                    style="text-transform: uppercase; text-align: center; font-size: 18px; font-weight: 600; color: #A91B60; line-height: 28px;">
                    Your Application ID:</h4>
                <p
                    style="text-align: center; padding: 25px; border: 3px solid #ECB322; color: #ECB322; font-weight: 700; letter-spacing: 10px; font-size: 20px;">
                    {applicant_id}</p>
            </td>
        </tr>
        <tr>
            <td colspan="5" style="background: #A91B60; padding: 40px;">
                <h4
                    style="text-transform: uppercase; text-align: center; font-size: 18px; font-weight: 600; color: #fff; line-height: 28px;">
                    Please Upload Your Joing Report as soon as you get it signed by concerned authority</h4>
            </td>
        </tr>
        <tr>
            <td colspan="5" style="background: #fff; padding: 40px;">
                <h4
                    style="text-transform: uppercase; text-align: center; font-size: 12px; font-weight: 600; color: #A91B60; line-height: 18px;">In case of any technical issue while filling online application form, please contact on toll free helpline number No. (From 09:45 AM to 06:30) PM </h4>
                <p style="color:#A91B60; font-size: 11px; font-weight: 600; text-align: center;">
                    This is a system generated mail. Please do not reply to this Email ID
                </p>
            </td>
        </tr>
        <tr>
            <td colspan="5" style="background: #A91B60; padding: 10px 40px; ">
               <span style="color: #fff; font-size: 11px; ">Visit us at <a href="https://trti.maharashtra.gov.in" target="_blank" style="color: #fff;">https://trti.maharashtra.gov.in</a> </span>
                <img src="https://static.vecteezy.com/system/resources/thumbnails/027/395/710/small/twitter-brand-new-logo-3-d-with-new-x-shaped-graphic-of-the-world-s-most-popular-social-media-free-png.png" style="width: 32px; height: 32px; float: right; " alt="Twitter Logo">
                <img src="https://cdn3.iconfinder.com/data/icons/social-network-30/512/social-06-512.png" style="width: 32px;  height: 32px;  float: right; margin-right: 12px; background: transparent;" alt="Youtube Logo">
               <img src="https://freelogopng.com/images/all_img/1658030214facebook-logo-hd.png" style="width: 32px; height: 32px; float: right; margin-right: 12px; " alt="Facebok Logo">
            </td>
        </tr>
    </table>

</body>

</html>

    '''

    # Email content in HTML format
    msg = Message('Application Status Changed', sender='noreply_fellowship@trti-maha.in', recipients=[email])
    msg.html = msg_body
    mail.send(msg)
# --------------------- END SEND EMAIL --------------------------


# --------------------- UPDATE STATUS ADMIN -------------------------
# STATUS UPDATE ON STUDENT RECORDS FUNCTIONALITY
# ----------------------------------------------
def update_status_admin(applicant_id, status):
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()

    # Update the status for the specified applicant ID
    update_query = "UPDATE application_page SET status = %s WHERE applicant_id = %s"
    cursor.execute(update_query, (status, applicant_id))

    # Commit the changes to the database
    cnx.commit()

    # Close the cursor and database connection
    cursor.close()
    cnx.close()


def update_scrutiny_admin(applicant_id, scrutiny_status):
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()

    # Update the status for the specified applicant ID
    update_query = "UPDATE application_page SET scrutiny_status = %s WHERE applicant_id = %s"
    cursor.execute(update_query, (scrutiny_status, applicant_id))

    # Commit the changes to the database
    cnx.commit()

    # Close the cursor and database connection
    cursor.close()
    cnx.close()


def update_final_appr_admin(applicant_id, final_approval, day, month, year):
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()

    # Update the status and date components for the specified applicant ID
    update_query = "UPDATE application_page SET final_approval = %s, final_approval_day = %s, final_approval_month = %s, final_approval_year = %s WHERE applicant_id = %s"
    print(update_query)
    cursor.execute(update_query, (final_approval, day, month, year, applicant_id))

    # Commit the changes to the database
    cnx.commit()

    # Close the cursor and database connection
    cursor.close()
    cnx.close()

# ------------------------- END UPDATE STATUS ADMIN ----------------------------


@app.route('/export_data', methods=['GET', 'POST'])
def export_data():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("SELECT * FROM application_page")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['Applicant Photo', 'Applicant Id', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email', 
               'Date Of Birth', 'Gender', 'Age', 'Caste', 'Your Caste', 'Marital Status', 'Dependents', 'Add 1', 'Add 2', 
               'Pincode', 'Village', 'Taluka', 'District', 'State', 'Phd Registration Date', 'Concerned University', 
               'Topic Of Phd', 'Name Of Guide', 'Name Of College', 'Stream', 'Board University', 'Admission Year', 
               'Passing Year', 'Percentage', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
               'Domicile', 'Domicile Certificate', 'Relation', 'Domicile Number', 'Caste', 'Caste Category', 
               'Caste Certificate Number', 'Issuing District', 'Caste Applicant Name', 'Caste Issuing Authority',
               'Salaried', 'Disability', 'Father Alive', 'Father Name', 'Mother Alive', 'Mother Name', 'Work In Government', 
               'Bank Name', 'Account Number', 'Ifsc Code', 'Account Holder Name', 'Documentfile1', 'Documentfile2',
               'Documentfile3', 'Documentfile4', 'Documentfile5', 'Documentfile6', 'Documentfile7', 'Documentfile8',
               'Documentfile9', 'Current Date', 'Formatted Datetime'])

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=ApplicantData.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response


@app.route('/export')
def export():
    export_data()


#                                           ------- ADMIN PAYMENT SHEET
@app.route('/payment_sheet', methods=['GET', 'POST'])
def payment_sheet():
    if not session.get('logged_in'):
        # Redirect to the admin login page if the user is not logged in
        return redirect(url_for('admin_login'))
    user_records = []
    if request.method == 'GET':
        # Establish a database connection
        cnx = mysql.connector.connect(user=user, password=password,
                                      host=host,
                                      database=database)
        cursor = cnx.cursor(dictionary=True)
        print('I have made connection')

        # Fetch user data based on the email
        cursor.execute("""
        
                    SELECT * 
                    FROM application_page 
                    WHERE final_approval = 'accepted' 
                      AND phd_registration_year >= '2023'
                    
                    UNION
                    
                    SELECT * 
                    FROM application_page 
                    WHERE phd_registration_year > '2020' 
                      AND aadesh = 1;
            
        """)
        user_data = cursor.fetchall()  # Use fetchall to retrieve all rows
        # print('user data:', user_data)

        for row in user_data:
            # Calculate values based on user data
            applicant_id = row['applicant_id']
            faculty = row["faculty"]
            print('faculty', faculty)
            fellowship_awarded_year = row['fellowship_awarded_year']
            fellowship_awarded_date = row['fellowship_awarded_date']
            joining_date = row["phd_registration_date"]
            city = row['city']
            bank_name = row['bank_name']
            account_number = row['account_number']
            ifsc = row['ifsc_code']
            year = '2023'
            print(joining_date)

            # Calculate Count Yearly
            if faculty == "Arts":
                count_yearly = 20500
            elif faculty == "Law":
                count_yearly = 20500
            elif faculty == "Commerce":
                count_yearly = 20500
            elif faculty == "Other":
                count_yearly = 20500
            elif faculty == "Science":
                count_yearly = 25000
            else:
                count_yearly = 0  # Handle other faculty values as needed

            if city in ['Ahmedabad', 'Bengaluru', 'Chennai', 'Delhi', 'Hyderabad', 'Kolkata', 'Mumbai', 'Pune']:
                rate = '30%'
            elif city in ['Agra', 'Ajmer', 'Aligarh', 'Amravati', 'Amritsar', 'Anand', 'Asansol', 'Aurangabad',
                          'Bareilly',
                          'Belagavi', 'Brahmapur', 'Bhavnagar', 'Bhiwandi', 'Bhopal', 'Bhubaneswar', 'Bikaner',
                          'Bilaspur',
                          'Bokaro Steel City', 'Burdwan', 'Chandigarh', 'Coimbatore', 'Cuttack', 'Dahod', 'Dehradun',
                          'Dombivli',
                          'Dhanbad', 'Bhilai', 'Durgapur', 'Erode', 'Faridabad', 'Ghaziabad', 'Gorakhpur', 'Guntur',
                          'Gurgaon',
                          'Guwahati', 'Gwalior', 'Hamirpur', 'Hubballi–Dharwad', 'Indore', 'Jabalpur', 'Jaipur',
                          'Jalandhar',
                          'Jalgaon', 'Jammu', 'Jamshedpur', 'Jhansi', 'Jodhpur', 'Kalaburagi', 'Kakinada', 'Kannur',
                          'Kanpur',
                          'Karnal', 'Kochi', 'Kolhapur', 'Kollam', 'Kota', 'Kozhikode', 'Kumbakonam', 'Kurnool',
                          'Ludhiana',
                          'Lucknow', 'Madurai', 'Malappuram', 'Mathura', 'Mangaluru', 'Meerut', 'Moradabad', 'Mysuru',
                          'Nagpur',
                          'Nanded', 'Nadiad', 'Nashik', 'Nellore', 'Noida', 'Patna', 'Puducherry', 'Purlia',
                          'Prayagraj', 'Raipur',
                          'Rajkot', 'Rajamahendravaram', 'Ranchi', 'Rourkela', 'Ratlam', 'Saharanpur', 'Salem',
                          'Sangli', 'Shimla',
                          'Siliguri', 'Solapur', 'Srinagar', 'Surat', 'Thanjavur', 'Thiruvananthapuram', 'Thrissur',
                          'Tiruchirappalli', 'Tirunelveli', 'Tiruvannamalai', 'Ujjain', 'Vijayapura', 'Vadodara',
                          'Varanasi',
                          'Vasai-Virar', 'Vijayawada', 'Visakhapatnam', 'Vellore', 'Warangal']:
                rate = '20%'
            else:
                rate = '10%'

            print("Rate:", rate)

            # Initialize the "from" and "to" date to empty strings
            duration_date_from = ""
            duration_date_to = ""

            if joining_date:  # Check if joining_date is not None
                # Calculate Duration Date (adding 3 months to joining date)
                duration_date_from = fellowship_awarded_date  # Assuming this is a datetime object
                duration_date_to = fellowship_awarded_date + timedelta(days=90)  # Adding 90 days to joining date
                # Extract day, month, and year
                day = duration_date_to.day
                month = duration_date_to.month
                year = duration_date_to.year

                # Print the stripped day, month, and year
                print(f"Day: {day}, Month: {month}, Year: {year}")
                # Format the dates for display in the desired format
                duration_date_from_str = duration_date_from.strftime('%d/%m/%Y')  # "17 Aug 2023"
                duration_date_to_str = duration_date_to.strftime('%d/%m/%Y')  # "15 Nov 2023"


            # Calculate Total Months
            total_months = 3

            # Calculate Fellowship
            fellowship = 37000 # Fixed value for 3 months

            # Calculate Total Fellowship
            total_fellowship = fellowship * total_months

            rate_str = float(rate.rstrip('%'))
            convert_rate = (rate_str/100)
            hra_amount = convert_rate * fellowship

            months = total_months

            total_hra = hra_amount * months

            total = total_fellowship + total_hra

            # Calculate the date 2 years after duration_date_from
            two_years_later = duration_date_from + timedelta(days=730)  # 2 years = 730 days

            # Assuming 'phd_registration_date' is already a datetime object
            if 'phd_registration_date' in row and row['phd_registration_date']:
                joiningDate = row['phd_registration_date'].strftime('%d/%m/%Y')  # This is for display or logs

                # Format the date to 'YYYY-MM-DD' for database insertion
                formattedDate = row['phd_registration_date'].strftime('%Y-%m-%d')

            # Get the current date
            current_date = datetime.now().date()
            current_year = current_date.year

            # Check the category based on 2 years difference
            if current_year >= fellowship_awarded_year+2:
                category = "SRF"  # Senior Research Fellowship
            else:
                category = "JRF"  # Junior Research Fellowship

            # Create a record dictionary for the user
            record = {
                "applicant_id": row['applicant_id'],
                "full_name": str(row['first_name']) + ' ' + str(row['middle_name']) + ' ' + str(row['last_name']),
                "first_name": row['first_name'],
                "last_name": row['last_name'],
                "middle_name": row['middle_name'],
                "email": row["email"],
                "faculty": row['faculty'],
                "fellowship_awarded_date": fellowship_awarded_date,
                "joining_date": formattedDate,
                "city":row['city'],
                "duration": f"{duration_date_from_str} <span class='fw-bold'>to</span> {duration_date_to_str}",
                "rate": rate,
                "count": count_yearly,
                "amount": hra_amount,
                "months": months,
                "total_hra":total_hra,
                "total": total,
                "duration_date_from": duration_date_from,
                "duration_date_to": duration_date_to,
                "duration_day": day,
                "duration_month": month,
                "duration_year": year,
                "total_months": total_months,
                "fellowship": fellowship,
                "to_fellowship": total_fellowship,
                "phd_registration_year": row['phd_registration_year'],
                "id": row['id'],
                "account_number": account_number,
                "ifsc": ifsc,
                "bank_name": bank_name,
                "year": year,
                "jrf_srf": category
            }

            user_records.append(record)

            email = record['email']

            cnx = mysql.connector.connect(user=user, password=password,
                                          host=host,
                                          database=database)
            cursor = cnx.cursor()

            cursor.execute(" SELECT * FROM payment_sheet where email=%s", (email,))
            result = cursor.fetchone()


            if result:
                print("Existing Record:", email)
                # Record already exists, do not insert again
            else:
                print("Record not found, proceeding with the INSERT query")
                # Insert values into the payment_sheet table
                cnx = mysql.connector.connect(user=user, password=password,
                                              host=host,
                                              database=database)
                cursor = cnx.cursor()

                insert_query = """
                    INSERT INTO payment_sheet (
                        full_name, faculty, fellowship_awarded_date, city, date, jrf_srf, duration_date_from, duration_date_to, duration_day, duration_month, duration_year, 
                        rate, count, amount, months, total_hra, total,
                        total_months, fellowship,
                        to_fellowship, bank_name, ifsc_code, account_number, fellowship_awarded_year, email
                    )
                    VALUES (%(full_name)s, %(faculty)s, %(fellowship_awarded_date)s, %(city)s, %(joining_date)s, %(jrf_srf)s, %(duration_date_from)s, %(duration_date_to)s,
                            %(duration_day)s, %(duration_month)s, %(duration_year)s,
                            %(rate)s, %(count)s, %(amount)s, %(months)s, %(total_hra)s, %(total)s, 
                            %(total_months)s, %(fellowship)s, %(to_fellowship)s, %(bank_name)s, %(ifsc)s,
                            %(account_number)s, %(year)s, %(email)s)         
                """
                # Execute the INSERT query
                cursor.execute(insert_query, record)

                # Commit the changes to the database
                cnx.commit()

            # Close the database cursor and connection
            cursor.close()
            cnx.close()
        # Close the database cursor and connection
        cursor.close()
        cnx.close()

    return render_template('payment_sheet.html', user_records=user_records)


@app.route('/submit_installments_admin', methods=['GET', 'POST'])
def submit_installments_admin():
        email = session['email']

        cnx = mysql.connector.connect(user=user, password=password,
                                      host=host,
                                      database=database)
        cursor = cnx.cursor(dictionary=True)

        if request.method == 'POST':
            email = session.get('email')  # Use .get() to avoid KeyError if 'email' is not in session
            start_period = request.form.get('start_period')
            end_period = request.form.get('end_period')
            recieved_pay = request.form.get('recieved_pay')
            recieved_date = request.form.get('recieved_date')

            try:
                # Check if the email already exists in the installments table
                check_query = "SELECT * FROM installments WHERE email = %s"
                cursor.execute(check_query, (email,))
                result = cursor.fetchone()

                if result:
                    # Check if each installment slot is completely filled
                    installment_2_filled = (result['start_period_2'] and
                                            result['end_period_2'] and
                                            result['recieved_pay_2'] and
                                            result['recieved_date_2'])
                    installment_3_filled = (result['start_period_3'] and
                                            result['end_period_3'] and
                                            result['recieved_pay_3'] and
                                            result['recieved_date_3'])
                    installment_4_filled = (result['start_period_4'] and
                                            result['end_period_4'] and
                                            result['recieved_pay_4'] and
                                            result['recieved_date_4'])

                    if not installment_2_filled:
                        update_query = """
                                UPDATE installments
                                SET inst_num_2 = %s, start_period_2 = %s, end_period_2 = %s, 
                                    recieved_pay_2 = %s, recieved_date_2 = %s, status_paid_2 = %s
                                WHERE email = %s
                            """
                        values = (2, start_period, end_period, recieved_pay, recieved_date, 'Paid', email)
                    elif not installment_3_filled:
                        update_query = """
                                UPDATE installments
                                SET inst_num_3 = %s, start_period_3 = %s, end_period_3 = %s, 
                                    recieved_pay_3 = %s, recieved_date_3 = %s, status_paid_3 = %s
                                WHERE email = %s
                            """
                        values = (3, start_period, end_period, recieved_pay, recieved_date, 'Paid', email)
                    elif not installment_4_filled:
                        update_query = """
                                UPDATE installments
                                SET inst_num_4 = %s, start_period_4 = %s, end_period_4 = %s, 
                                    recieved_pay_4 = %s, recieved_date_4 = %s, status_paid_4 = %s
                                WHERE email = %s
                            """
                        values = (4, start_period, end_period, recieved_pay, recieved_date, 'Paid', email)
                    else:
                        return "Maximum installments reached", 400

                    cursor.execute(update_query, values)
                    message = "Installment updated successfully"
                    flash(message, 'success')
                else:
                    # If email does not exist, insert a new record
                    insert_query = """
                            INSERT INTO installments (email, inst_num, start_period, end_period, 
                                                      recieved_pay, recieved_date, status_paid)
                            VALUES (%s, %s, %s, %s, %s, %s, %s)
                        """
                    values = (email, 1, start_period, end_period, recieved_pay, recieved_date, 'Paid')
                    cursor.execute(insert_query, values)
                    message = "Installment submitted successfully"
                    flash(message, 'success')

                cnx.commit()
                return redirect(url_for('fellowship_details', email=email))
            except Exception as e:
                cnx.rollback()
                return f"An error occurred: {str(e)}", 500
            finally:
                cursor.close()
                cnx.close()

        return redirect(url_for('fellowship_details', email=email))

# #----------------------  EXPORT TO EXCEL Payment Sheet (2023-2024) -------------------------------------------

@app.route('/export_payment_sheet')
def export_payment_sheet():
    cnx = mysql.connector.connect(user=user, password=password,
                                    host=host,
                                    database=database)
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT number, full_name, email, faculty, fellowship_awarded_date, date, duration_date_from, duration_date_to, "
                   "total_months, fellowship, to_fellowship, rate, amount, months, total_hra, count, pwd, total,"
                   "city, bank_name, ifsc_code, account_number FROM payment_sheet")

    data = cursor.fetchall()
    print(data)

    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    header = 4
    date_column_index = 1
    date_column_index_date = 8
    current_date = datetime.now().strftime('%B %Y')
    # Add the bold row before the header
    ws.cell(row=1, column=header, value='Appendix')  # Place "Date:" in column B
    ws.cell(row=2, column=date_column_index, value='Number:')  # Place "Date:" in column B
    ws.cell(row=2, column=date_column_index_date, value=f'Date:{current_date}')  # Place "Date:" in column B
    # ws.cell(row=1, column=date_column_index_date + 1, value=current_date)  # Place the date in column C
    # Add the bold row before the header
    # ws.append(['Date:', current_date])  # Format as "Date: September 2024"
    bold_row = ws[1]  # Get the last added row (the one we just added)
    bold_row_2 = ws[2]  # Get the last added row (the one we just added)
    for cell in bold_row and bold_row_2:
        cell.font = Font(bold=True)  # Make the text bold

    # Add header row
    ws.append(['Sr. No.', 'Name of Student', 'Date of PHD Registration', 'Fellowship Awarded Date',
           'Duration', 'Bank Name', 'Account Number', 'IFSC', 'Fellowship Amount'])

    # Add data to the worksheet with formatting
    for index, row in enumerate(data, start=1):
        full_name = row['full_name']

        # Joining date
        joining_date = row['date']
        fellowship_awarded_date = row['fellowship_awarded_date']
        # Duration dates
        duration_date_from = row['duration_date_from']
        duration_date_to = row['duration_date_to']

        # Convert string dates to datetime objects if they are not already
        if isinstance(duration_date_from, str):
            try:
                duration_date_from = datetime.strptime(duration_date_from, '%Y-%m-%d')  # Adjust format as needed
            except ValueError:
                duration_date_from = None

        if isinstance(duration_date_to, str):
            try:
                duration_date_to = datetime.strptime(duration_date_to, '%Y-%m-%d')  # Adjust format as needed
            except ValueError:
                duration_date_to = None

        if isinstance(duration_date_from, datetime) and isinstance(duration_date_to, datetime):
            duration_date_from_str = duration_date_from.strftime('%d %b %Y')  # Format as "17 Aug 2023"
            duration_date_to_str = duration_date_to.strftime('%d %b %Y')  # Format as "15 Nov 2023"
            duration = f"{duration_date_from_str} to {duration_date_to_str}"
        else:
            duration = "N/A"

        # Other fields
        bank_name = row['bank_name']
        account_number = row['account_number']
        ifsc = row['ifsc_code']
        fellowship_amount = row['total']

        # Append the formatted data
        ws.append([index, full_name, joining_date, fellowship_awarded_date, duration,
                   bank_name, account_number, ifsc, fellowship_amount])

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Payment_Sheet_2023_2024.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response


@app.route('/export_track_payments')
def export_track_payments():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    cursor.execute(
        "SELECT number, full_name, email, faculty, jrf_srf, fellowship_awarded_date, date, duration_date_from, duration_date_to, "
        "total_months, fellowship, to_fellowship, rate, amount, months, total_hra, count, pwd, total,"
        "city, bank_name, ifsc_code, account_number FROM payment_sheet")

    data = cursor.fetchall()
    print(data)

    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    header = 4
    date_column_index = 1
    date_column_index_date = 8
    current_date = datetime.now().strftime('%B %Y')
    # Add the bold row before the header
    ws.cell(row=1, column=header, value='Appendix')  # Place "Date:" in column B
    ws.cell(row=2, column=date_column_index, value='Number:')  # Place "Date:" in column B
    ws.cell(row=2, column=date_column_index_date, value=f'Date:{current_date}')  # Place "Date:" in column B
    # ws.cell(row=1, column=date_column_index_date + 1, value=current_date)  # Place the date in column C
    # Add the bold row before the header
    # ws.append(['Date:', current_date])  # Format as "Date: September 2024"
    bold_row = ws[1]  # Get the last added row (the one we just added)
    bold_row_2 = ws[2]  # Get the last added row (the one we just added)
    for cell in bold_row and bold_row_2:
        cell.font = Font(bold=True)  # Make the text bold

    # Add header row
    ws.append([ 'Sr. No.', 'Name of Student', 'Faculty', 'JRF/SRF', 'Date of PHD Registration',
               'Fellowship Awarded Date', 'Duration', 'Total Months', 'Fellowship', 'Total Fellowship', 'H.R.A Rate',
               'H.R.A Amount', 'Months', 'Total H.R.A', 'Contingency Yearly', 'PWD', 'Total Amount', 'City' ])

    # Add data to the worksheet with formatting
    for index, row in enumerate(data, start=1):
        full_name = row['full_name']

        # Joining date
        joining_date = row['date']
        fellowship_awarded_date = row['fellowship_awarded_date']
        # Duration dates
        duration_date_from = row['duration_date_from']
        duration_date_to = row['duration_date_to']

        # Convert string dates to datetime objects if they are not already
        if isinstance(duration_date_from, str):
            try:
                duration_date_from = datetime.strptime(duration_date_from, '%Y-%m-%d')  # Adjust format as needed
            except ValueError:
                duration_date_from = None

        if isinstance(duration_date_to, str):
            try:
                duration_date_to = datetime.strptime(duration_date_to, '%Y-%m-%d')  # Adjust format as needed
            except ValueError:
                duration_date_to = None

        if isinstance(duration_date_from, datetime) and isinstance(duration_date_to, datetime):
            duration_date_from_str = duration_date_from.strftime('%d %b %Y')  # Format as "17 Aug 2023"
            duration_date_to_str = duration_date_to.strftime('%d %b %Y')  # Format as "15 Nov 2023"
            duration = f"{duration_date_from_str} to {duration_date_to_str}"
        else:
            duration = "N/A"

        # Other fields
        faculty = row['faculty']
        jrf_srf = row['jrf_srf']
        total_months = row['total_months']
        fellowship = row['fellowship']
        total_felowship = row['to_fellowship']
        hra_rate = row['rate']
        hra_amount = row['amount']
        months = row['months']
        total_hra = row['total_hra']
        count_yearly = row['count']
        pwd = row['pwd']
        total = row['total']
        city = row['city']


        # Append the formatted data
        ws.append([index, full_name, faculty, jrf_srf, joining_date, fellowship_awarded_date, duration,
                   total_months, fellowship, total_felowship, hra_rate, hra_amount, months, total_hra,
                   count_yearly, pwd, total, city ])

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Track Payment Sheet 2023-2024.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response


class PDF(FPDF):
    def __init__(self):
        super().__init__(orientation='P', unit='mm', format='A3')  # Set format to A3

    header_added = False  # To track whether the header is added to the first page

    def header(self):
        if not self.header_added:
            var = get_base_url()
            print(var)

            # Adjusted dimensions for A3 format
            self.set_font("Arial", "B", 12)

            # Adjust the X, Y, and image size to fit A3 format
            # for LOCALSERVER
            self.image('/var/www/fellowship/fellowship/BartiFellowship/BartiFellowship/static/Images/satya.png', 140, 10, 30)  # Adjusted position and size for A3
            image_width = 140  # Updated to a size appropriate for A3
            text_x_position = self.get_x()
            text_y_position = self.get_y() + 25  # Adjusted for A3 format
            self.set_xy(text_x_position, text_y_position)
            # Adjusted positions for A3 format
            self.image('/var/www/fellowship/fellowship/BartiFellowship/BartiFellowship/static/Images/newtrtiImage.png', 20, 10, 60)  # Adjust size for larger format
            self.image('/var/www/fellowship/fellowship/BartiFellowship/BartiFellowship/static/Images/mahashasn_new.png', 215, 10, 60)  # Adjust size and position for A3

            # For HOSTSERVER
            # self.image('static/Images/satya.png', 140, 10, 30)  # Adjusted position and size for A3
            # image_width = 140  # Updated to a size appropriate for A3
            # text_x_position = self.get_x()
            # text_y_position = self.get_y() + 25  # Adjusted for A3 format
            # self.set_xy(text_x_position, text_y_position)
            # # Adjusted positions for A3 format
            # self.image('static/Images/newtrtiImage.png', 20, 10, 60)  # Adjust size for larger format
            # self.image('static/Images/mahashasn_new.png', 215, 10, 60)  # Adjust size and position for A3

            # Centered text for A3 format
            self.ln(10)
            self.cell(0, 10, "Government of Maharashtra", align="C", ln=True)
            self.cell(0, 10, "Tribal Research & Training Institute", align="C", ln=True)
            self.cell(0, 10, "28, Queens Garden, Pune - 411001", align="C", ln=True)

            # Adjust the dashed line width for A3
            self.dashed_line(10, self.get_y(), 290, self.get_y(), dash_length=3, space_length=1)  # Adjust for A3 width

            self.ln(5)  # Adjust space after the line
            self.set_font("Arial", size=10)

            current_date = datetime.now().strftime('%B %Y')
            # Left-aligned and right-aligned text for A3 format
            self.cell(0, 10, "No.: Research-_____/Case.No ____/Desk- __/_____ ", ln=False)
            self.cell(0, 10, f"Date: {current_date}", align="R", ln=True)

            self.set_font("Arial", "B", size=12)
            self.cell(0, 10, "Appendix", align="C", ln=True)

            # Adjust rotation for A3
            # self.ln(2)
            # self.rotate(45)
            # self.set_font('Arial', '', 65)  # Adjust font size for A3
            # self.set_text_color(192, 192, 192)

            # Adjust the position for the watermark on A3
            # self.text(-50, 280, "STRF FELLOWSHIP")  # Adjust position for A3 layout
            # self.rotate(0)

            self.header_added = True  # Set to True after adding the header

    def footer(self):
        self.set_y(-15)
        self.set_font("Arial", 'I', 8)
        self.cell(0, 10, f'Page {self.page_no()}', 0, 0, 'C')


@app.route('/export_payment_sheet_pdf')
def export_payment_sheet_pdf():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT number, full_name, email, faculty, fellowship_awarded_date, date, duration_date_from, duration_date_to, "
                   "total_months, fellowship, to_fellowship, rate, amount, months, total_hra, count, pwd, total,"
                   "city, bank_name, ifsc_code, account_number FROM payment_sheet")

    data = cursor.fetchall()
    pdf = PDF()
    pdf.add_page()

    # Set margins
    pdf.set_margins(5, 5, 10)  # Left, Top, Right margins

    current_date = datetime.now().strftime('%B %Y')

    # Add date information
    pdf.set_font("Arial", 'I', 12)
    # pdf.cell(0, 10, f'Number:', ln=True, align='L')
    # pdf.cell(0, 10, f'Date: {current_date}', ln=True, align='L')
    pdf.ln(5)  # Small line break

    # Set header
    pdf.set_font("Arial", 'B', 10)

    # Define fixed column widths
    column_widths = [15, 60, 30, 45, 40, 40, 30, 30]  # Set fixed widths for each column
    headers = ['Sr. No.', 'Name of Student', 'Date of PHD Reg.', 'Duration', 'Bank Name',
               'Account Number', 'IFSC', 'Amount']

    # Add header row with multi-cell for text wrapping
    x_start = pdf.get_x()  # Get the starting x position

    for i, header in enumerate(headers):
        # Store the current x and y position before adding the multi_cell
        x = pdf.get_x()
        y = pdf.get_y()

        # Add the multi_cell for the header
        pdf.cell(column_widths[i], 10, header, border=1, align='C')

        # Set the position back to where it was before adding multi_cell for the next cell
        pdf.set_xy(x + column_widths[i], y)

    pdf.ln()  # Move to the next line after adding the header

    # Set font for data
    pdf.set_font("Arial", '', 10)

    # Add data to the PDF
    for index, row in enumerate(data, start=1):
        pdf.cell(column_widths[0], 10, str(index), 1, align='C')

        # Add the name cell with multi_cell
        full_name = row['full_name']
        pdf.cell(column_widths[1], 10, full_name, 1, align='C')

        # Move cursor back to the right for the next column
        # pdf.cell(column_widths[0], 10, '', 0)  # Placeholder for Sr. No.

        # Add the remaining cells
        pdf.cell(column_widths[2], 10, str(row['date']), 1, align='C')
        #pdf.cell(column_widths[2], 10, str(row['fellowship_awarded_date']), 1, align='C')
        duration_text = f"{row['duration_date_from']} to {row['duration_date_to']}"
        pdf.cell(column_widths[3], 10, duration_text, 1, align='C')
        pdf.cell(column_widths[4], 10, row['bank_name'] if row['bank_name'] else 'N/A', 1, align='C')
        pdf.cell(column_widths[5], 10, row['account_number'] if row['account_number'] else 'N/A', 1, align='C')
        pdf.cell(column_widths[6], 10, row['ifsc_code'] if row['ifsc_code'] else 'N/A', 1, align='C')
        pdf.cell(column_widths[7], 10, str(row['total']), 1, align='C')

        # Move to the next line
        pdf.ln()

    # Finalize PDF output
    response = make_response(pdf.output(dest='S').encode('latin1'))
    response.headers['Content-Disposition'] = 'attachment; filename=Payment_Sheet_2023_2024.pdf'
    response.headers['Content-Type'] = 'application/pdf'

    return response


@app.route('/budget', methods=['GET', 'POST'])
def budget():
    if not session.get('logged_in'):
        # Redirect to the admin login page if the user is not logged in
        return redirect(url_for('admin_login'))
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)

    if request.method == 'POST':
        allocated_budget = float(request.form['allocated_budget'])  # Convert to float
        spent_budget = float(request.form['spent_budget'])  # Convert to float
        year = request.form['year']
        inhand_budget = allocated_budget - spent_budget
        print(inhand_budget)
        cursor.execute("INSERT INTO budget (allocated, spent, inhand, year) VALUES (%s, %s, %s, %s)",
                       (allocated_budget, spent_budget, inhand_budget, year))
        cnx.commit()

    sql = "SELECT * FROM budget"
    cursor.execute(sql)  # Corrected syntax for executing the SQL query
    budget = cursor.fetchall()  # Use fetchall to retrieve all rows
    print(budget)
    cursor.close()
    cnx.close()
    return render_template('budget.html', budget=budget)


@app.route('/payment_tracking', methods=['GET', 'POST'])
def payment_tracking():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    print('Connected')
    records_display = []
    print('Post method')

    sql = """
        SELECT * 
        FROM application_page 
        WHERE final_approval = 'accepted' 
          AND phd_registration_year >= '2023'

        UNION

        SELECT * 
        FROM application_page 
        WHERE phd_registration_year > '2020' 
          AND aadesh = 1;
    """
    cursor.execute(sql)
    records = cursor.fetchall()

    # Assuming you want to fetch payment records for each result from 'records'
    payment_records = []
    for record in records:
        email = record['email']  # Extract email from each record
        sql = "SELECT * FROM payment_sheet WHERE email=%s"
        cursor.execute(sql, (email,))
        result = cursor.fetchall()

        # Assuming 'duration_date_from' is a date field in the result, format it
        for payment_record in result:
            if 'duration_date_from' in payment_record:
                try:
                    # Convert the date string to a datetime object
                    date_obj = datetime.strptime(payment_record['duration_date_from'], '%Y-%m-%d')
                    date_objj = datetime.strptime(payment_record['duration_date_to'], '%Y-%m-%d')
                    # Format the date as day/month/year
                    payment_record['duration_date_from'] = date_obj.strftime('%d/%m/%Y')
                    payment_record['duration_date_to'] = date_objj.strftime('%d/%m/%Y')
                except ValueError:
                    pass  # Handle invalid date formats gracefully

        payment_records.append(result)  # Append formatted result to the list
    # print(payment_records)

    if request.method == 'POST':
        start_date = request.form.get('start_date')
        end_date = request.form.get('end_date')
        year = request.form.get('year')
        month = request.form.get('month')
        quarters = request.form.get('quarters')
        print(quarters)
        # Base SQL query without the trailing 'AND'
        sql = """
            SELECT ap.*, ps.*
            FROM application_page ap
            JOIN payment_sheet ps ON ap.email = ps.email
            WHERE ap.final_approval = 'accepted'          
        """

        # List to store query conditions
        conditions = []
        params = []

        # Adding conditions dynamically based on input
        if start_date and end_date:
            conditions.append("ps.duration_date_to BETWEEN %s AND %s")
            params.extend([start_date, end_date])
        if year:
            conditions.append("ps.fellowship_awarded_year = %s")
            params.append(year)
        if month:
            conditions.append("ps.duration_month = %s")
            params.append(month)
        if quarters:
            conditions.append("JSON_CONTAINS(ps.quarters, %s, '$')")
            params.append(str(quarters))  # Convert to string to pass the integer value

        # You can keep json.dumps() if it's for a JSON array like [1, 2]

        # Join the conditions with 'AND' if there are any
        if conditions:
            sql += " AND " + " AND ".join(conditions)

        # Execute the query
        cursor.execute(sql, params)
        records_display = cursor.fetchall()
        print(sql)
        flash('Payment information retrieved successfully', 'success')
    flattened_records = [record for sublist in payment_records for record in sublist]
    # No else block is needed in this case
    # print(records_display)
    return render_template('payment_tracking.html', records_display=records_display, records=records,
                           payment_records=flattened_records)


#----------------------  EXPORT TO EXCEL Payment Tracking Sheet-------------------------------------------
@app.route('/export_payment_tracking_sheet', methods=['GET', 'POST'])
def export_payment_tracking_sheet():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute(" SELECT number, full_name, faculty, jrf_srf, duration_date_from, duration_date_from, "
                   "total_months, fellowship, to_fellowship, rate, amount, months, total_hra, count, pwd, total, city"
                   " FROM payment_sheet WHERE fellowship_awarded_year = 2023 ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    ws.append(['Sr No', 'Full Name', 'Faculty', 'JRF/SRF', 'Date of PhD Registration', 'Duration', 'Total Months',
               'Fellowship', 'Total Fellowship', 'H.R.A Rate', 'H.R.A Amount', 'Months', 'Total H.R.A',
               'Cont. Yearly', 'Difference', 'Total', 'City'])

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Payment_Tracking_Sheet.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response


# @app.route('/installment_userpage', methods=['GET', 'POST'])
# def installment_userpage():
#     email = session['email']
#     cnx = mysql.connector.connect(user=user, password=password,
#                                   host=host,
#                                   database=database)
#     cursor = cnx.cursor(dictionary=True)
#
#     cursor.execute("SELECT * FROM application_page where email=%s", (email,))
#     result = cursor.fetchall()
#     print("result" + str(result))
#
#     cursor.execute("SELECT * FROM payment_sheet where email=%s", (email,))
#     record = cursor.fetchall()
#     print("record" + str(record))
#     table_data = []
#
#     for row in record:
#         total_months = int(row['total_months'])
#         print("Total Months" + str(total_months))
#         start_date = datetime.strptime(row['duration_date_from'], '%Y-%m-%d')
#         print("Start Date" + str(start_date))
#         end_date = datetime.strptime(row['duration_date_to'], '%Y-%m-%d')
#         print("End Date" + str(end_date))
#
#         # First installment
#         table_data.append({
#             'sr_no': 1,
#             'period': total_months,
#             'start_period': start_date.strftime('%Y-%m-%d'),
#             'end_period': end_date.strftime('%Y-%m-%d'),
#             'due_date': (end_date + timedelta(days=60)).strftime('%Y-%m-%d'),
#             'balance': 31000,
#             'installment_number': 1,
#             'paid': 'Not Paid'  # Adjust this based on your payment status logic
#         })
#
#         # Generate next two installments
#         for i in range(2, 4):
#             next_start_date = end_date + timedelta(days=30 * (i - 1))
#             next_end_date = next_start_date + timedelta(days=90)
#             table_data.append({
#                 'sr_no': i,
#                 'period': total_months,
#                 'start_period': next_start_date.strftime('%Y-%m-%d'),
#                 'end_period': next_end_date.strftime('%Y-%m-%d'),
#                 'due_date': (next_end_date + timedelta(days=60)).strftime('%Y-%m-%d'),
#                 'balance': 31000,
#                 'installment_number': i,
#                 'paid': row['paid_or_not_installment_1']  # Adjust this based on your payment status logic
#             })
#
#         print(table_data)
#         total_period = sum(int(row['period']) for row in table_data)
#         total_balance = sum(int(row['balance']) for row in table_data)
#         print("Total Period" + str(total_period))
#
#     approve_pay = approve_payment(email)
#
#     cursor.execute("SELECT * FROM award_letter where email=%s ", (email,))
#     solution = cursor.fetchall()
#     print("record" + str(solution))
#
#     cursor.execute("SELECT fellowship_withdrawn FROM signup where email=%s ", (email,))
#     output = cursor.fetchall()
#     print("record" + str(output))
#
#     cnx.commit()
#     cursor.close()
#     cnx.close()
#     return render_template('installment_userpage.html', result=result, record=record, output=output, solution=solution,
#                            table_data=table_data, total_period=total_period, total_balance=total_balance,
#                            approve_pay=approve_pay)


@app.route('/budget_report/<string:email>', methods=['GET', 'POST'])
def budget_report(email):
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)

    cursor.execute("SELECT * FROM application_page where email=%s ", (email,))
    result = cursor.fetchall()
    print("result" + str(result))

    cursor.execute("SELECT * FROM payment_sheet where email=%s ", (email,))
    record = cursor.fetchall()
    print("record" + str(record))
    table_data = []

    for row in record:
        total_months = int(row['total_months'])
        print("Total Months" + str(total_months))
        start_date = datetime.strptime(row['duration_date_from'], '%Y-%m-%d')
        print("Start Date" + str(start_date))
        end_date = datetime.strptime(row['duration_date_to'], '%Y-%m-%d')
        print("End Date" + str(end_date))

        # First installment
        table_data.append({
            'sr_no': 1,
            'period': total_months,
            'start_period': start_date.strftime('%Y-%m-%d'),
            'end_period': end_date.strftime('%Y-%m-%d'),
            'due_date': (end_date + timedelta(days=60)).strftime('%Y-%m-%d'),
            'balance': 31000,
            'installment_number': 1,
            'paid': 'Not Paid'  # Adjust this based on your payment status logic
        })

        # Generate next two installments
        for i in range(2, 4):
            next_start_date = end_date + timedelta(days=30 * (i - 1))
            next_end_date = next_start_date + timedelta(days=90)
            table_data.append({
                'sr_no': i,
                'period': total_months,
                'start_period': next_start_date.strftime('%Y-%m-%d'),
                'end_period': next_end_date.strftime('%Y-%m-%d'),
                'due_date': (next_end_date + timedelta(days=60)).strftime('%Y-%m-%d'),
                'balance': 31000,
                'installment_number': i,
            })

        print(table_data)
        total_period = sum(int(row['period']) for row in table_data)
        total_balance = sum(int(row['balance']) for row in table_data)
        print("Total Period" + str(total_period))


   #approve_pay = approve_payment(email)

    cursor.execute("SELECT * FROM award_letter where email=%s ", (email,))
    solution = cursor.fetchall()
    print("record" + str(solution))

    cursor.execute("SELECT fellowship_withdrawn FROM signup where email=%s ", (email,))
    output = cursor.fetchall()
    print("record" + str(output))

    cnx.commit()
    cursor.close()
    cnx.close()
    return render_template('budget_report.html', result=result, record=record, output=output, solution=solution,
                           table_data=table_data, total_period=total_period, total_balance=total_balance)


def approve_payment(email):
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    # Update the 'paid' column for the specified email
    cursor.execute("UPDATE payment_sheet SET paid_or_not_installment_1='Paid' WHERE email=%s", (email,))
    cnx.commit()
    cursor.close()
    cnx.close()


@app.route('/fellowship_details/<string:email>', methods=['GET', 'POST'])
def fellowship_details(email):
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)

    today = datetime.today().date()

    cursor.execute("SELECT * FROM application_page where email=%s ", (email,))
    result = cursor.fetchall()
    startDate = result[0]['final_approved_date']

    cursor.execute("SELECT * FROM installments where email=%s", (email,))
    installments = cursor.fetchall()
    print('installments', installments)

    cursor.execute("SELECT * FROM payment_sheet WHERE email=%s", (email,))
    record = cursor.fetchall()

    # Assuming only one row in record
    for row in record:
        total_months = int(row['total_months'])
        start_date = startDate
        installment_list = []

        # Loop to create 15 installments (5 years * 3 installments per year)
        for i in range(1, 16):
            # Set start and end dates for each installment
            if i == 1:
                current_start_date = start_date
            else:
                current_start_date = previous_end_date + timedelta(days=30)

            current_end_date = current_start_date + timedelta(days=90)

            # Create installment dictionary
            installment = {
                'sr_no': i,
                'period': total_months,
                'start_period': current_start_date.strftime('%Y-%m-%d'),
                'end_period': current_end_date.strftime('%Y-%m-%d'),
                'due_date': (current_end_date + timedelta(days=60)).strftime('%Y-%m-%d'),
                'balance': 42000,  # Adjust balance if necessary
                'installment_number': i,
                'paid': row.get(f'paid_or_not_installment_{i}', 'Not Available')
                # Assuming the field changes per installment
            }

            # Append to installment list
            installment_list.append(installment)
            # print(installments)
            # Update previous_end_date for the next iteration
            previous_end_date = current_end_date

        # Calculate total period and total balance for all installments
        total_period = sum(inst['period'] for inst in installment_list)
        total_balance = sum(inst['balance'] for inst in installment_list)

        for installment in installment_list:
            # Convert and format the dates
            start_period = datetime.strptime(installment['start_period'], '%Y-%m-%d').strftime('%d %B %Y')
            end_period = datetime.strptime(installment['end_period'], '%Y-%m-%d').strftime('%d %B %Y')
            installment['formatted_start_period'] = start_period
            installment['formatted_end_period'] = end_period

    cursor.execute("SELECT fellowship_withdrawn FROM signup where email=%s", (email,))
    output = cursor.fetchall()

    installment_button_status = []
    previously_paid = False  # Track if the previous installment was paid

    # Loop through 15 installments directly
    for current_installment_number in range(1, 16):
        status_paid = None

        # Check the status of the current installment
        for installment in installments:
            if installment.get(f'inst_num_{current_installment_number}') == current_installment_number:
                status_paid = installment.get(f'status_paid_{current_installment_number}')

        # Ensure the first installment can be paid if not already paid
        if current_installment_number == 1:
            if status_paid == 'Paid':
                installment_button_status.append('paid')
                previously_paid = True
            else:
                installment_button_status.append('pay_enabled')  # First installment always enabled
                previously_paid = False
        else:
            # Handle installments after the first
            if status_paid == 'Paid':
                installment_button_status.append('paid')
                previously_paid = True
            elif previously_paid:
                installment_button_status.append('pay_enabled')  # Enable pay if previous is paid
            else:
                installment_button_status.append('disabled')  # Disable if previous is unpaid

    # Example to print out installment statuses
    for i, button_status in enumerate(installment_button_status):
        print(f"Installment {i + 1}: {button_status}")

    cnx.commit()
    cursor.close()
    cnx.close()

    return render_template(
        'fellowship_details.html',
        result=result,
        record=record,
        output=output,
        installment_list=installment_list,
        total_period=total_period,
        total_balance=total_balance,
        today=today,
        installments=installments,
        installment_button_status=installment_button_status
    )


@app.route('/get_year_count', methods=['GET', 'POST'])
def get_year_count():
    if request.method == 'POST':
        data = request.form['selected_year']
        print(data)
        # selected_year = data.get('selected_year')
        count = f""" SELECT count(*) FROM application_page where phd_registration_year = {data} """

        count_accept = f""" SELECT count(*) FROM application_page where phd_registration_year = {data} and final_approval = 'accepted' """
        count_reject = f""" SELECT count(*) FROM application_page where phd_registration_year = {data} and final_approval = 'rejected' """
        # Queries to get district data for the map
        district_query = """SELECT district, COUNT(*) AS student_count 
                                    FROM application_page 
                                    WHERE phd_registration_year = %s 
                                    GROUP BY district;"""
        print(district_query)
        # Execute the queries
        cursor.execute(count)
        total_result = cursor.fetchone()
        total_count = total_result[0] if total_result else 0

        cursor.execute(count_accept)
        accept_result = cursor.fetchone()
        accept_count = accept_result[0] if accept_result else 0

        cursor.execute(count_reject)
        reject_result = cursor.fetchone()
        reject_count = reject_result[0] if reject_result else 0
        result = cursor.fetchone()  # Fetch the result of the query
        count = result[0] if result else 0  # Extract the count from the result

        # Execute district count query
        cursor.execute(district_query, (data,))
        district_results = cursor.fetchall()

        # Construct district data dictionary
        district_data = {row[0]: row[1] for row in district_results}  # row[0] is district, row[1] is student_count
        print(f"District Data: {district_data}")

        cnx.commit()
        return jsonify({'total_count': total_count, 'accept_count': accept_count, 'reject_count': reject_count,
                        'district_data': district_data})
    return 'INAVALID REQUEST'

#                                                     ---- ADMIN DASHBOARD

@app.route('/index', methods=['GET', 'POST'])
def index():
    if not session.get('logged_in'):
        # Redirect to the admin login page if the user is not logged in
        return redirect(url_for('admin_login'))

    selected_year = None  # Initialize selected_year

    session['user_name'] = 'Admin'
    total_count = applications_today()
    accepted_count = accepted_applications()
    form_filled = form_filled_applications()
    form_incomplete = form_incomplete_applications()


    #-------------------------------------------------------
    rejected_count = rejected_applications()
    male_count = male_count_report()
    male_count21 = male_count2021()
    male_count22 = male_count2022()
    male_count23 = male_count2023()
    female_count = female_count_report()
    female_count21 = female_count2021()
    female_count22 = female_count2022()
    female_count23 = female_count2023()
    trans_count = trans_count_report()

    # -------------------------------------------------------
    katkari = katkari_count_report()
    katkari21 = katkari_count21()
    katkari22 = katkari_count22()
    katkari23 = katkari_count23()
    kolam = kolam_count_report()
    kolam21 = kolam_count21()
    kolam22 = kolam_count22()
    kolam23 = kolam_count23()
    madia = madia_count_report()
    madia21 = madia_count21()
    madia22 = madia_count22()
    madia23 = madia_count23()


    twentyone_count = year_twentyone_count()
    twentytwo_count = year_twentytwo_count()
    twentythree_count = year_twentythree_count()
    accept_23 = accept_twentythree_count()
    reject_23 = reject_twentythree_count()
    olduser_count = old_users_count_2021()
    olduser22_count = old_users_count_2022()
    old_user_accepted_22 = old_users_count_2022_accepted()
    old_user_accepted_21 = old_users_count_2021_accepted()

    disability_yes = disability_yes_count_report()
    disability_no = disability_no_count_report()
    physical_disability = type_disability_physically()
    visually_disability = type_disability_visually()
    hearing_disability = type_disability_hearing()
    male_report = male_record_report()
    female_report = female_record_report()
    priority_people_caste = priority_by_caste()
    # map_data = maps()
    current_year = datetime.now().year

    return render_template('index.html', katkari=katkari, kolam=kolam, madia=madia,
                           form_incomplete=form_incomplete, form_filled=form_filled, accept_23=accept_23 ,reject_23=reject_23,
                           selected_year=selected_year,
                           total_count=total_count, accepted_count=accepted_count, rejected_count=rejected_count,
                           male_count=male_count, female_count=female_count, trans_count=trans_count, disability_yes=disability_yes,
                           disability_no=disability_no, male_report=male_report, female_report=female_report,
                           priority_people_caste=priority_people_caste,
                           twentyone_count=twentyone_count, twentytwo_count=twentytwo_count, olduser_count=olduser_count,
                           olduser22_count=olduser22_count, twentythree_count=twentythree_count,
                           physical_disability=physical_disability, hearing_disability=hearing_disability,
                           visually_disability=visually_disability, old_user_accepted_22=old_user_accepted_22,
                           old_user_accepted_21=old_user_accepted_21 , male_count21=male_count21, male_count22=male_count22,
                           male_count23=male_count23, female_count21=female_count21, female_count22=female_count22,
                           female_count23=female_count23, katkari21=katkari21, katkari22=katkari22, katkari23=katkari23,
                           madia21=madia21, madia22=madia22, madia23=madia23, kolam21=kolam21, kolam22=kolam22, kolam23=kolam23,
                           current_year=current_year)


# Assuming you're using a web framework to serve the data
@app.route('/get_district_data', methods=['POST'])
def get_district_data():
    # Get the year from the request
    selected_year = request.form['selected_year']

    # Queries to get district data for the map
    district_query = """SELECT district, COUNT(*) AS student_count 
                                        FROM application_page 
                                        WHERE phd_registration_year = %s 
                                        GROUP BY district;"""
    # Execute district count query
    cursor.execute(district_query, (selected_year,))
    district_results = cursor.fetchall()

    # Construct district data dictionary
    district_data = {row[0]: row[1] for row in district_results}  # row[0] is district, row[1] is student_count

    return jsonify(district_data=district_data)


# ---------------------  DISABILITY FUNCTIONALITY ---------------------------------
# ---------------------------------------------------------------------------------
@app.route('/disability_report_yes', methods=['GET', 'POST'])
def disability_report_yes():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT * FROM application_page WHERE disability='Yes'")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    return render_template('disability_report_yes.html', result=result)


#----------------------  EXPORT TO EXCEL Applicants with Disability -------------------------------------------
@app.route('/export_disability_report_yes', methods=['GET', 'POST'])
def export_disability_report_yes():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("SELECT * from application_page WHERE disability ='Yes' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['Applicant Id', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email', 'Date Of Birth', 
                'Gender', 'Age', 'Caste', 'Your Caste', 'Marital Status', 'Dependents', 'Add 1', 'Add 2', 'Pincode', 'Village', 
                'Taluka', 'District', 'State', 'Phd Registration Date', 'Concerned University', 'Topic Of Phd', 'Name Of Guide', 
                'Name Of College', 'Stream', 'Board University', 'Admission Year', 'Passing Year', 'Percentage', 'Family Annual Income', 
                'Income Certificate Number', 'Issuing Authority', 'Domicile', 'Domicile Certificate', 'Domicile Number', 'Caste Certf', 
                'Issuing District', 'Caste Issuing Authority', 'Salaried', 'Disability', 'Father Name', 'Mother Name', 'Work In Government', 
                'Bank Name', 'Account Number', 'Ifsc Code', 'Account Holder Name']	)

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Applicants_with_disability.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response    


@app.route('/disability_report_no', methods=['GET', 'POST'])
def disability_report_no():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT * FROM application_page WHERE disability='No'")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    print(result)
    return render_template('disability_report_no.html', result=result)


#----------------------  EXPORT TO EXCEL Applicants WITHOUT Disability -------------------------------------------
@app.route('/export_disability_report_no', methods=['GET', 'POST'])
def export_disability_report_no():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("SELECT * from application_page WHERE disability ='No' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['Applicant Id', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email', 'Date Of Birth', 
                'Gender', 'Age', 'Caste', 'Your Caste', 'Marital Status', 'Dependents', 'Add 1', 'Add 2', 'Pincode', 'Village', 
                'Taluka', 'District', 'State', 'Phd Registration Date', 'Concerned University', 'Topic Of Phd', 'Name Of Guide', 
                'Name Of College', 'Stream', 'Board University', 'Admission Year', 'Passing Year', 'Percentage', 'Family Annual Income', 
                'Income Certificate Number', 'Issuing Authority', 'Domicile', 'Domicile Certificate', 'Domicile Number', 'Caste Certf', 
                'Issuing District', 'Caste Issuing Authority', 'Salaried', 'Disability', 'Father Name', 'Mother Name', 'Work In Government', 
                'Bank Name', 'Account Number', 'Ifsc Code', 'Account Holder Name']	)

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Applicants_without_disability.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response


def disability_yes_count_report():       # ----- To count users with NO disability  ------
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("SELECT COUNT(*) FROM application_page WHERE disability='Yes'")
    result = cursor.fetchone()
    print(result)
    return result[0]


def disability_no_count_report():       # ----- To count users with NO disability  ------
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("SELECT COUNT(*) FROM application_page WHERE disability='no'")
    result = cursor.fetchone()
    print(result)
    return result[0]


def type_disability_physically():       # ----- To count users with NO disability  ------
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("SELECT COUNT(*) FROM application_page WHERE type_of_disability='Physically Handicapped'")
    result = cursor.fetchone()
    print(result)
    return result[0]    


def type_disability_visually():       # ----- To count users with NO disability  ------
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("SELECT COUNT(*) FROM application_page WHERE type_of_disability='Visually Impaired'")
    result = cursor.fetchone()
    print(result)
    return result[0]


def type_disability_hearing():       # ----- To count users with NO disability  ------
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("SELECT COUNT(*) FROM application_page WHERE type_of_disability='Hearing Impaired'")
    result = cursor.fetchone()
    print(result)
    return result[0]


@app.route('/disabilty_report_physical', methods=['GET', 'POST'])
def disabilty_report_physical():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT * FROM application_page WHERE type_of_disability='Physically Handicapped'")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    print(result)
    return render_template('disabilty_report_physical.html', result=result)


#----------------------  EXPORT TO EXCEL Physically Handicapped Applicants -------------------------------------------
@app.route('/export_physically_handicapped_report', methods=['GET', 'POST'])
def export_physically_handicapped_report():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("SELECT applicant_id, first_name, middle_name, last_name, mobile_number, email, date_of_birth,"
                   "gender, age, caste, your_caste, marital_status, add_1, add_2, pincode, village, taluka, district,"
                   "state, city,  phd_registration_date, concerned_university, topic_of_phd, "
                   "name_of_guide, name_of_college, stream, board_university, admission_year, passing_year, percentage,"
                   "family_annual_income, income_certificate_number, issuing_authority, domicile, domicile_certificate,"
                   "domicile_number, caste_certf, issuing_district, caste_issuing_authority, salaried, disability,"
                   "father_name, mother_name, work_in_government, bank_name, account_number, ifsc_code," 
                   "account_holder_name FROM application_page WHERE type_of_disability='Physically Handicapped'")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['Applicant Id', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email', 'Date Of Birth', 
                'Gender', 'Age', 'Caste', 'Your Caste', 'Marital Status', 'Dependents', 'Add 1', 'Add 2', 'Pincode', 'Village', 
                'Taluka', 'District', 'State', 'Phd Registration Date', 'Concerned University', 'Topic Of Phd', 'Name Of Guide', 
                'Name Of College', 'Stream', 'Board University', 'Admission Year', 'Passing Year', 'Percentage', 'Family Annual Income', 
                'Income Certificate Number', 'Issuing Authority', 'Domicile', 'Domicile Certificate', 'Domicile Number', 'Caste Certf', 
                'Issuing District', 'Caste Issuing Authority', 'Salaried', 'Disability', 'Father Name', 'Mother Name', 'Work In Government', 
                'Bank Name', 'Account Number', 'Ifsc Code', 'Account Holder Name']	)

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Physically_handicapped_applicants.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response      


@app.route('/disabilty_report_visual', methods=['GET', 'POST'])
def disabilty_report_visual():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT * FROM application_page WHERE type_of_disability='Visually Impaired'")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    print(result)
    return render_template('disabilty_report_visual.html', result=result)


#----------------------  EXPORT TO EXCEL Visually Impaired Applicants -------------------------------------------
@app.route('/export_visually_impaired_report', methods=['GET', 'POST'])
def export_visually_impaired_report():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("SELECT applicant_id, first_name, middle_name, last_name, mobile_number, email, date_of_birth,"
                   "gender, age, caste, your_caste, marital_status, add_1, add_2, pincode, village, taluka, district,"
                   "state, city,  phd_registration_date, concerned_university, topic_of_phd, "
                   "name_of_guide, name_of_college, stream, board_university, admission_year, passing_year, percentage,"
                   "family_annual_income, income_certificate_number, issuing_authority, domicile, domicile_certificate,"
                   "domicile_number, caste_certf, issuing_district, caste_issuing_authority, salaried, disability,"
                   "father_name, mother_name, work_in_government, bank_name, account_number, ifsc_code," 
                   "account_holder_name FROM application_page WHERE type_of_disability='Visually Impaired'")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['Applicant Id', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email', 'Date Of Birth', 
                'Gender', 'Age', 'Caste', 'Your Caste', 'Marital Status', 'Dependents', 'Add 1', 'Add 2', 'Pincode', 'Village', 
                'Taluka', 'District', 'State', 'Phd Registration Date', 'Concerned University', 'Topic Of Phd', 'Name Of Guide', 
                'Name Of College', 'Stream', 'Board University', 'Admission Year', 'Passing Year', 'Percentage', 'Family Annual Income', 
                'Income Certificate Number', 'Issuing Authority', 'Domicile', 'Domicile Certificate', 'Domicile Number', 'Caste Certf', 
                'Issuing District', 'Caste Issuing Authority', 'Salaried', 'Disability', 'Father Name', 'Mother Name', 'Work In Government', 
                'Bank Name', 'Account Number', 'Ifsc Code', 'Account Holder Name']	)

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Visually_impaired_applicants.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response         


@app.route('/disabilty_report_hearing', methods=['GET', 'POST'])
def disabilty_report_hearing():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT * FROM application_page WHERE type_of_disability='Hearing Impaired'")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    print(result)
    return render_template('disabilty_report_hearing.html', result=result)


#----------------------  EXPORT TO EXCEL Hearing Impaired Applicants -------------------------------------------
@app.route('/export_hearing_impaired_report', methods=['GET', 'POST'])
def export_hearing_impaired_report():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("SELECT applicant_id, first_name, middle_name, last_name, mobile_number, email, date_of_birth,"
                   "gender, age, caste, your_caste, marital_status, add_1, add_2, pincode, village, taluka, district,"
                   "state, city,  phd_registration_date, concerned_university, topic_of_phd, "
                   "name_of_guide, name_of_college, stream, board_university, admission_year, passing_year, percentage,"
                   "family_annual_income, income_certificate_number, issuing_authority, domicile, domicile_certificate,"
                   "domicile_number, caste_certf, issuing_district, caste_issuing_authority, salaried, disability,"
                   "father_name, mother_name, work_in_government, bank_name, account_number, ifsc_code," 
                   "account_holder_name FROM application_page WHERE type_of_disability='Hearing Impaired'")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['Applicant Id', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email', 'Date Of Birth', 
                'Gender', 'Age', 'Caste', 'Your Caste', 'Marital Status', 'Dependents', 'Add 1', 'Add 2', 'Pincode', 'Village', 
                'Taluka', 'District', 'State', 'Phd Registration Date', 'Concerned University', 'Topic Of Phd', 'Name Of Guide', 
                'Name Of College', 'Stream', 'Board University', 'Admission Year', 'Passing Year', 'Percentage', 'Family Annual Income', 
                'Income Certificate Number', 'Issuing Authority', 'Domicile', 'Domicile Certificate', 'Domicile Number', 'Caste Certf', 
                'Issuing District', 'Caste Issuing Authority', 'Salaried', 'Disability', 'Father Name', 'Mother Name', 'Work In Government', 
                'Bank Name', 'Account Number', 'Ifsc Code', 'Account Holder Name']	)
    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Hearing_impaired_applicants.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response     


# ------------------------------- END DISABILITY -----------------------------------------
# ----------------------------------------------------------------------------------------


# ------------------------------- MALE AND FEMALE FUNCTIONALITY --------------------------
# ----------------------------------------------------------------------------------------
@app.route('/male_record_report', methods=['GET', 'POST'])
def male_record_report():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT * FROM application_page WHERE gender='male'")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    #print(result)
    return render_template('male_record_report.html', result=result)

#----------------------  EXPORT TO EXCEL Male Applications -------------------------------------------
@app.route('/export_male_applications', methods=['GET', 'POST'])
def export_male_applications():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("SELECT applicant_id, first_name, middle_name, last_name, mobile_number, email, date_of_birth,"
                   "gender, age, caste, your_caste, marital_status, add_1, add_2, pincode, village, taluka, district,"
                   "state, city,  phd_registration_date, concerned_university, topic_of_phd, "
                   "name_of_guide, name_of_college, stream, board_university, admission_year, passing_year, percentage,"
                   "family_annual_income, income_certificate_number, issuing_authority, domicile, domicile_certificate,"
                   "domicile_number, caste_certf, issuing_district, caste_issuing_authority, salaried, disability,"
                   "father_name, mother_name, work_in_government, bank_name, account_number, ifsc_code," 
                   "account_holder_name FROM application_page WHERE gender ='male' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['Applicant Id', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email', 'Date Of Birth', 
                'Gender', 'Age', 'Caste', 'Your Caste', 'Marital Status', 'Dependents', 'Add 1', 'Add 2', 'Pincode', 'Village', 
                'Taluka', 'District', 'State', 'Phd Registration Date', 'Concerned University', 'Topic Of Phd', 'Name Of Guide', 
                'Name Of College', 'Stream', 'Board University', 'Admission Year', 'Passing Year', 'Percentage', 'Family Annual Income', 
                'Income Certificate Number', 'Issuing Authority', 'Domicile', 'Domicile Certificate', 'Domicile Number', 'Caste Certf', 
                'Issuing District', 'Caste Issuing Authority', 'Salaried', 'Disability', 'Father Name', 'Mother Name', 'Work In Government', 
                'Bank Name', 'Account Number', 'Ifsc Code', 'Account Holder Name']	)

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Male_Applicants_Data.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response    


@app.route('/female_record_report', methods=['GET', 'POST'])
def female_record_report():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT * FROM application_page WHERE gender='female'")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    print(result)
    return render_template('female_record_report.html', result=result)


#----------------------  EXPORT TO EXCEL Female Applications -------------------------------------------
@app.route('/export_female_applications', methods=['GET', 'POST'])
def export_female_applications():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("SELECT applicant_id, first_name, middle_name, last_name, mobile_number, email, date_of_birth,"
                   "gender, age, caste, your_caste, marital_status, add_1, add_2, pincode, village, taluka, district,"
                   "state, city,  phd_registration_date, concerned_university, topic_of_phd, "
                   "name_of_guide, name_of_college, stream, board_university, admission_year, passing_year, percentage,"
                   "family_annual_income, income_certificate_number, issuing_authority, domicile, domicile_certificate,"
                   "domicile_number, caste_certf, issuing_district, caste_issuing_authority, salaried, disability,"
                   "father_name, mother_name, work_in_government, bank_name, account_number, ifsc_code," 
                   "account_holder_name FROM application_page WHERE gender ='female' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['Applicant Id', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email', 'Date Of Birth', 
                'Gender', 'Age', 'Caste', 'Your Caste', 'Marital Status', 'Dependents', 'Add 1', 'Add 2', 'Pincode', 'Village', 
                'Taluka', 'District', 'State', 'Phd Registration Date', 'Concerned University', 'Topic Of Phd', 'Name Of Guide', 
                'Name Of College', 'Stream', 'Board University', 'Admission Year', 'Passing Year', 'Percentage', 'Family Annual Income', 
                'Income Certificate Number', 'Issuing Authority', 'Domicile', 'Domicile Certificate', 'Domicile Number', 'Caste Certf', 
                'Issuing District', 'Caste Issuing Authority', 'Salaried', 'Disability', 'Father Name', 'Mother Name', 'Work In Government', 
                'Bank Name', 'Account Number', 'Ifsc Code', 'Account Holder Name']	)

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Female_Applicants_Data.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response     


def applications_today():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute(" SELECT COUNT(*) FROM application_page ")
    result = cursor.fetchone()
    return result[0]


def form_filled_applications():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute(" SELECT COUNT(*) FROM application_page where form_filled='1' and phd_registration_year>='2023' ")
    result = cursor.fetchone()
    return result[0]


def form_incomplete_applications():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute(" SELECT COUNT(*) FROM application_page where form_filled='0' and phd_registration_year>='2023' ")
    result = cursor.fetchone()
    return result[0]



def accepted_applications():    # ----- To count accepted applications  ------
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute(" SELECT COUNT(*) FROM application_page WHERE final_approval='accepted' ")
    result = cursor.fetchone()
    return result[0]


def rejected_applications():    # ----- To count accepted applications  ------
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute(" SELECT COUNT(*) FROM application_page WHERE final_approval='rejected' ")
    result = cursor.fetchone()
    return result[0]


def male_count_report():       # ----- To count male users  ------
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("SELECT COUNT(*) FROM application_page WHERE gender='male' ")
    result = cursor.fetchone()
    return result[0]


def male_count2021():       # ----- To count male users  ------
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute(" SELECT count(*) FROM application_page where phd_registration_year='2021' and gender='male' ")
    result = cursor.fetchone()
    print(result)
    return result[0]


def male_count2022():       # ----- To count male users  ------
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute(" SELECT COUNT(*) FROM application_page WHERE gender='male' and phd_registration_year='2022' ")
    result = cursor.fetchone()
    return result[0]


def male_count2023():       # ----- To count male users  ------
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute(" SELECT COUNT(*) FROM application_page WHERE gender='male' and phd_registration_year='2023' ")
    result = cursor.fetchone()
    return result[0]


def female_count_report():     # ----- To count female users  ------
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("SELECT COUNT(*) FROM application_page WHERE gender='female'")
    result = cursor.fetchone()
    return result[0]


def female_count2021():     # ----- To count female users  ------
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("SELECT COUNT(*) FROM application_page WHERE gender='female' and phd_registration_year='2021' ")
    result = cursor.fetchone()
    return result[0]


def female_count2022():     # ----- To count female users  ------
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("SELECT COUNT(*) FROM application_page WHERE gender='female' and phd_registration_year='2022' ")
    result = cursor.fetchone()
    return result[0]


def female_count2023():     # ----- To count female users  ------
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("SELECT COUNT(*) FROM application_page WHERE gender='female' and phd_registration_year='2023' ")
    result = cursor.fetchone()
    print(result)
    return result[0]


def trans_count_report():       # ----- To count trans_gender users  ------
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("SELECT COUNT(*) FROM application_page WHERE gender='other'")
    result = cursor.fetchone()
    print(result)
    return result[0]


def katkari_count_report():       # ----- To count male users  ------
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute(" SELECT COUNT(*) FROM application_page WHERE your_caste='katkari' ")
    result = cursor.fetchone()
    print(result)
    return result[0]

def katkari_count21():       # ----- To count male users  ------
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute(" SELECT COUNT(*) FROM application_page WHERE your_caste='katkari' and phd_registration_year='2021' ")
    result = cursor.fetchone()
    print(result)
    return result[0]

def katkari_count22():       # ----- To count male users  ------
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute(" SELECT COUNT(*) FROM application_page WHERE your_caste='katkari' and phd_registration_year='2022' ")
    result = cursor.fetchone()
    print(result)
    return result[0]

def katkari_count23():       # ----- To count male users  ------
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute(" SELECT COUNT(*) FROM application_page WHERE your_caste='katkari' and phd_registration_year='2023' ")
    result = cursor.fetchone()
    print(result)
    return result[0]


def kolam_count_report():     # ----- To count female users  ------
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute(" SELECT COUNT(*) FROM application_page WHERE your_caste='kolam' ")
    result = cursor.fetchone()
    print(result)
    return result[0]


def kolam_count21():     # ----- To count female users  ------
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute(" SELECT COUNT(*) FROM application_page WHERE your_caste='kolam' and phd_registration_year='2021' ")
    result = cursor.fetchone()
    print(result)
    return result[0]


def kolam_count22():     # ----- To count female users  ------
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute(" SELECT COUNT(*) FROM application_page WHERE your_caste='kolam' and phd_registration_year='2022' ")
    result = cursor.fetchone()
    print(result)
    return result[0]


def kolam_count23():     # ----- To count female users  ------
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute(" SELECT COUNT(*) FROM application_page WHERE your_caste='kolam' and phd_registration_year='2023' ")
    result = cursor.fetchone()
    print(result)
    return result[0]


def madia_count_report():       # ----- To count trans_gender users  ------
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute(" SELECT COUNT(*) FROM application_page WHERE your_caste='madia' ")
    result = cursor.fetchone()
    print(result)
    return result[0]


def madia_count21():       # ----- To count trans_gender users  ------
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute(" SELECT COUNT(*) FROM application_page WHERE your_caste='madia' and phd_registration_year='2021' ")
    result = cursor.fetchone()
    print(result)
    return result[0]


def madia_count22():       # ----- To count trans_gender users  ------
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute(" SELECT COUNT(*) FROM application_page WHERE your_caste='madia' and phd_registration_year='2022' ")
    result = cursor.fetchone()
    print(result)
    return result[0]


def madia_count23():       # ----- To count trans_gender users  ------
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute(" SELECT COUNT(*) FROM application_page WHERE your_caste='madia' and phd_registration_year='2023' ")
    result = cursor.fetchone()
    print(result)
    return result[0]


def year_twentyone_count():      #----- To count users from 2021
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("SELECT COUNT(*) FROM application_page WHERE phd_registration_year='2021'")
    result = cursor.fetchone()
    print(result)
    return result[0]


def year_twentytwo_count():       #----- To count users from 2022
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("SELECT COUNT(*) FROM application_page WHERE phd_registration_year='2022'")
    result = cursor.fetchone()
    print(result)
    return result[0]


def year_twentythree_count():     #----- To count users from 2023
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("SELECT COUNT(*) FROM application_page WHERE phd_registration_year='2023'")
    result = cursor.fetchone()
    print(result)
    return result[0]

def accept_twentythree_count():     #----- To count users from 2023
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("SELECT COUNT(*) FROM application_page WHERE phd_registration_year='2023' and final_approval='accepted' ")
    result = cursor.fetchone()
    print(result)
    return result[0]

def reject_twentythree_count():     #----- To count users from 2023
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("SELECT COUNT(*) FROM application_page WHERE phd_registration_year='2023' and final_approval='rejected' ")
    result = cursor.fetchone()
    print(result)
    return result[0]


def old_users_count_2021():                                                  #----- To count old users
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("SELECT COUNT(*) FROM old_users where phd_registration_year='2021'")
    result = cursor.fetchone()
    print(result)
    return result[0]


def old_users_count_2022():                                                  #----- To count old users
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("SELECT COUNT(*) FROM old_users where phd_registration_year='2022'")
    result = cursor.fetchone()
    print(result)
    return result[0]

def old_users_count_2022_accepted():                                                  #----- To count old users
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("SELECT COUNT(*) FROM application_page where final_approval='accepted' AND phd_registration_year='2022'")
    result = cursor.fetchone()
    print(result)
    return result[0]


def old_users_count_2021_accepted():                                                  #----- To count old users
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("SELECT COUNT(*) FROM application_page where final_approval='accepted' AND phd_registration_year='2021'")
    result = cursor.fetchone()
    print(result)
    return result[0]


def total_appl_for_all_years():                                                  #----- To count old users
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("SELECT COUNT(*) AS TotalRows FROM ( SELECT email FROM old_users UNION SELECT email FROM application_page ) AS CombinedResult")
    result = cursor.fetchone()
    print(result)
    return result[0]


def priority_by_caste():              # ----- To count users with caste "Katkari , Kolam , madia"------
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT * FROM application_page WHERE your_caste IN ('katkari', 'kolam', 'madia')")
    result = cursor.fetchall()
    return result


def news_fetch():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    cursor.execute(" SELECT * FROM news_and_updates ORDER BY date DESC LIMIT 6 ")
    result = cursor.fetchall()
    return result

# ----------------------------------------------------------------------------
# ------------------- BEGIN APPLICATIONS 2023 REPORTS ------------------------------
# ----------------------------------------------------------------------------
@app.route('/total_application_report', methods=['GET', 'POST'])
def total_application_report():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    cursor.execute(" SELECT * FROM application_page where phd_registration_year = 2023 ")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    # print(result)
    return render_template('total_application_report.html', result=result)


@app.route('/completed_form', methods=['GET', 'POST'])
def completed_form():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    cursor.execute(" SELECT * FROM application_page where phd_registration_year = 2023 and form_filled='1' ")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    # print(result)
    return render_template('completed_form.html', result=result)


@app.route('/incompleted_form', methods=['GET', 'POST'])
def incompleted_form():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    cursor.execute(" SELECT * FROM application_page where phd_registration_year = 2023 and form_filled='0' ")
    # cursor.execute(" SELECT * FROM application_page where email = 'tupotbhare@gmail.com' ")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    # print(result)
    return render_template('incompleted_form.html', result=result)


#----------------------  EXPORT TO EXCEL Total applications 2023 -------------------------------------------
@app.route('/export_total_applications_23', methods=['GET', 'POST'])
def export_total_applications_23():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
                   "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status," 
                    "dependents, state, district, taluka, village, city, add_1, add_2, pincode," 
                    "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total," 
                    "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total," 
                    "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream," 
                    "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
                    "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
                    "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
                    "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income," 
                    "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka," 
                    "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number," 
                    "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
                    "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
                    "type_of_disability, father_name, mother_name, work_in_government, gov_department," 
                    "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
                   " where phd_registration_year = '2023' ")

    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    # ws.append(['applicant_id','email','first_name','last_name','application_date'])

    ws.append(['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
               'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
               'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
               'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
                'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts', 'HSC Total',
                'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream', 'Graduation Attempts',
                'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts', 'PG Total',
                'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
                'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
                'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
                'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
                'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
                'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District', 'Caste Issuing Authority',
                'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government', 'Government Department',
                'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR' ])

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Total_Applications_2023.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response


@app.route('/completed_form_export', methods=['GET', 'POST'])
def completed_form_export():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
                   "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status," 
                    "dependents, state, district, taluka, village, city, add_1, add_2, pincode," 
                    "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total," 
                    "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total," 
                    "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream," 
                    "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
                    "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
                    "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
                    "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income," 
                    "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka," 
                    "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number," 
                    "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
                    "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
                    "type_of_disability, father_name, mother_name, work_in_government, gov_department," 
                    "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
                   " where phd_registration_year = '2023' and form_filled='1' ")

    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    # ws.append(['applicant_id','email','first_name','last_name','application_date'])

    ws.append(['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
               'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Your Caste', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
               'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
               'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
                'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts', 'HSC Total',
                'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream', 'Graduation Attempts',
                'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts', 'PG Total',
                'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
                'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
                'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
                'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
                'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
                'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District', 'Caste Issuing Authority',
                'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government', 'Government Department',
                'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR' ])

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Completed Form Fellowship 2023.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response


@app.route('/incompleted_form_export', methods=['GET', 'POST'])
def incompleted_form_export():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
                   "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status," 
                    "dependents, state, district, taluka, village, city, add_1, add_2, pincode," 
                    "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total," 
                    "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total," 
                    "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream," 
                    "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
                    "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
                    "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
                    "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income," 
                    "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka," 
                    "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number," 
                    "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
                    "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
                    "type_of_disability, father_name, mother_name, work_in_government, gov_department," 
                    "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
                   " where phd_registration_year = '2023' and form_filled='0' ")

    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    # ws.append(['applicant_id','email','first_name','last_name','application_date'])

    ws.append(['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
               'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
               'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
               'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
                'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts', 'HSC Total',
                'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream', 'Graduation Attempts',
                'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts', 'PG Total',
                'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
                'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
                'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
                'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
                'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
                'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District', 'Caste Issuing Authority',
                'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government', 'Government Department',
                'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR' ])

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Total_Applications_2023.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response


@app.route('/total_accepted_report', methods=['GET', 'POST'])
def total_accepted_report():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    cursor.execute(" SELECT * FROM application_page WHERE final_approval='accepted' and phd_registration_year='2023' ")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    # print(result)
    return render_template('total_accepted_report.html', result=result)


#----------------------  EXPORT TO EXCEL Accepted Applications 2023 -------------------------------------------
@app.route('/export_accepted_applications_23', methods=['GET', 'POST'])
def export_accepted_applications_23():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
                   "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status," 
                    "dependents, state, district, taluka, village, city, add_1, add_2, pincode," 
                    "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total," 
                    "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total," 
                    "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream," 
                    "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
                    "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
                    "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
                    "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income," 
                    "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka," 
                    "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number," 
                    "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
                    "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
                    "type_of_disability, father_name, mother_name, work_in_government, gov_department," 
                    "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page "
                    " WHERE final_approval = 'accepted' and phd_registration_year='2023' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
               'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
               'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
               'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
                'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts', 'HSC Total',
                'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream', 'Graduation Attempts',
                'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts', 'PG Total',
                'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
                'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
                'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
                'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
                'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
                'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District', 'Caste Issuing Authority',
                'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government', 'Government Department',
                'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR' ])

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Accepted_Applications_2023.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response    


@app.route('/total_rejected_report', methods=['GET', 'POST'])
def total_rejected_report():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT * FROM application_page WHERE final_approval='rejected' and phd_registration_year='2023'")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    # print(result)
    return render_template('total_rejected_report.html', result=result)


#----------------------  EXPORT TO EXCEL Rejected Applications 2023 -------------------------------------------
@app.route('/export_rejected_applications_23', methods=['GET', 'POST'])
def export_rejected_applications_23():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
                   "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status," 
                    "dependents, state, district, taluka, village, city, add_1, add_2, pincode," 
                    "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total," 
                    "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total," 
                    "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream," 
                    "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
                    "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
                    "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
                    "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income," 
                    "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka," 
                    "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number," 
                    "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
                    "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
                    "type_of_disability, father_name, mother_name, work_in_government, gov_department," 
                    "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page "
                    " WHERE final_approval = 'rejected' and phd_registration_year='2023' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
               'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
               'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
               'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
                'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts', 'HSC Total',
                'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream', 'Graduation Attempts',
                'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts', 'PG Total',
                'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
                'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
                'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
                'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
                'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
                'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District', 'Caste Issuing Authority',
                'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government', 'Government Department',
                'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR' ])

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Rejected_Applications_2023.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response    

# ----------------------------------------------------------------------------
# ------------------- END APPLICATIONS 2023 REPORTS ------------------------------
# ----------------------------------------------------------------------------

# ----------------------------------------------------------------------------
# ------------------- BEGIN APPLICATIONS 2022 REPORTS ------------------------------
# ----------------------------------------------------------------------------
@app.route('/total_application_report_22', methods=['GET', 'POST'])
def total_application_report_22():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT * FROM application_page WHERE phd_registration_year='2022' ")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    # print(result)
    return render_template('total_application_report_22.html', result=result)


#----------------------  EXPORT TO EXCEL Total applications 2022 -------------------------------------------
@app.route('/export_total_applications_22', methods=['GET', 'POST'])
def export_total_applications_22():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
                   "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status," 
                    "dependents, state, district, taluka, village, city, add_1, add_2, pincode," 
                    "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total," 
                    "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total," 
                    "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream," 
                    "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
                    "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
                    "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
                    "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income," 
                    "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka," 
                    "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number," 
                    "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
                    "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
                    "type_of_disability, father_name, mother_name, work_in_government, gov_department," 
                    "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
                   " WHERE phd_registration_year='2022' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
               'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe','Your Caste', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
               'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
               'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
                'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts', 'HSC Total',
                'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream', 'Graduation Attempts',
                'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts', 'PG Total',
                'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
                'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
                'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
                'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
                'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
                'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District', 'Caste Issuing Authority',
                'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government', 'Government Department',
                'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR' ])

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Total_Applications_2022.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response    


@app.route('/total_accepted_report_22', methods=['GET', 'POST'])
def total_accepted_report_22():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT * FROM application_page WHERE final_approval='accepted' and phd_registration_year='2022' ")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    # print(result)
    return render_template('total_accepted_report_22.html', result=result)


#----------------------  EXPORT TO EXCEL Accepted Applications 2022 -------------------------------------------
@app.route('/export_accepted_applications_22', methods=['GET', 'POST'])
def export_accepted_applications_22():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
                   "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status," 
                    "dependents, state, district, taluka, village, city, add_1, add_2, pincode," 
                    "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total," 
                    "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total," 
                    "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream," 
                    "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
                    "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
                    "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
                    "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income," 
                    "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka," 
                    "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number," 
                    "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
                    "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
                    "type_of_disability, father_name, mother_name, work_in_government, gov_department," 
                    "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
                   " WHERE phd_registration_year='2022' and final_approval='accepted' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
               'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Your Caste', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
               'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
               'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
                'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts', 'HSC Total',
                'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream', 'Graduation Attempts',
                'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts', 'PG Total',
                'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
                'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
                'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
                'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
                'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
                'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District', 'Caste Issuing Authority',
                'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government', 'Government Department',
                'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR' ])

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Accepted_Applications_2022.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response     


@app.route('/total_rejected_report_22', methods=['GET', 'POST'])
def total_rejected_report_22():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT * FROM application_page WHERE phd_registration_year='2022' and final_approval='rejected' ")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    # print(result)
    return render_template('total_rejected_report_22.html', result=result)


#----------------------  EXPORT TO EXCEL Rejected Applications 2022 -------------------------------------------
@app.route('/export_rejected_applications_22', methods=['GET', 'POST'])
def export_rejected_applications_22():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
                   "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status," 
                    "dependents, state, district, taluka, village, city, add_1, add_2, pincode," 
                    "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total," 
                    "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total," 
                    "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream," 
                    "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
                    "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
                    "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
                    "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income," 
                    "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka," 
                    "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number," 
                    "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
                    "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
                    "type_of_disability, father_name, mother_name, work_in_government, gov_department," 
                    "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
                   " WHERE phd_registration_year='2022' and final_approval='rejected' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
               'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Your Caste', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
               'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
               'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
                'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts', 'HSC Total',
                'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream', 'Graduation Attempts',
                'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts', 'PG Total',
                'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
                'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
                'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
                'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
                'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
                'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District', 'Caste Issuing Authority',
                'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government', 'Government Department',
                'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR' ])

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Rejected_Applications_2022.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response    
         

# ----------------------------------------------------------------------------
# ------------------- END APPLICATIONS 2022 REPORTS ------------------------------
# ----------------------------------------------------------------------------

# ----------------------------------------------------------------------------
# ------------------- BEGIN APPLICATIONS 2021 REPORTS ------------------------------
# ----------------------------------------------------------------------------
@app.route('/total_application_report_21', methods=['GET', 'POST'])
def total_application_report_21():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT * FROM application_page WHERE phd_registration_year='2021'")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    # print(result)
    return render_template('total_application_report_21.html', result=result)


#----------------------  EXPORT TO EXCEL Total applications 2021 ----------------
@app.route('/export_total_applications_21', methods=['GET', 'POST'])
def export_total_applications_21():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
                   "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status," 
                    "dependents, state, district, taluka, village, city, add_1, add_2, pincode," 
                    "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total," 
                    "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total," 
                    "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream," 
                    "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
                    "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
                    "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
                    "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income," 
                    "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka," 
                    "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number," 
                    "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
                    "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
                    "type_of_disability, father_name, mother_name, work_in_government, gov_department," 
                    "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
                   " WHERE phd_registration_year='2021' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
               'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Your Caste', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
               'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
               'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
                'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts', 'HSC Total',
                'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream', 'Graduation Attempts',
                'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts', 'PG Total',
                'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
                'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
                'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
                'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
                'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
                'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District', 'Caste Issuing Authority',
                'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government', 'Government Department',
                'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR' ])

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Total_Applications_2021.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response     


@app.route('/total_accepted_report_21', methods=['GET', 'POST'])
def total_accepted_report_21():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT * FROM application_page WHERE phd_registration_year='2021' and final_approval='accepted' ")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    # print(result)
    return render_template('total_accepted_report_21.html', result=result)


#----------------------  EXPORT TO EXCEL Accepted Applications 2021 -------------------------------------------
@app.route('/export_accepted_applications_21', methods=['GET', 'POST'])
def export_accepted_applications_21():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
                   "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status," 
                    "dependents, state, district, taluka, village, city, add_1, add_2, pincode," 
                    "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total," 
                    "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total," 
                    "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream," 
                    "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
                    "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
                    "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
                    "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income," 
                    "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka," 
                    "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number," 
                    "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
                    "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
                    "type_of_disability, father_name, mother_name, work_in_government, gov_department," 
                    "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
                   " WHERE phd_registration_year='2021' and final_approval='accepted' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
               'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Your Caste', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
               'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
               'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
                'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts', 'HSC Total',
                'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream', 'Graduation Attempts',
                'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts', 'PG Total',
                'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
                'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
                'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
                'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
                'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
                'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District', 'Caste Issuing Authority',
                'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government', 'Government Department',
                'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR' ])

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Accepted_Applications_2021.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response      


@app.route('/total_rejected_report_21', methods=['GET', 'POST'])
def total_rejected_report_21():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT * FROM application_page WHERE phd_registration_year='2021' and final_approval='rejected'")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    # print(result)
    return render_template('total_rejected_report_21.html', result=result)


#----------------------  EXPORT TO EXCEL Rejected Applications 2021 -------------------------------------------
@app.route('/export_rejected_applications_21', methods=['GET', 'POST'])
def export_rejected_applications_21():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
                   "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status," 
                    "dependents, state, district, taluka, village, city, add_1, add_2, pincode," 
                    "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total," 
                    "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total," 
                    "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream," 
                    "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
                    "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
                    "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
                    "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income," 
                    "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka," 
                    "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number," 
                    "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
                    "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
                    "type_of_disability, father_name, mother_name, work_in_government, gov_department," 
                    "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
                   " WHERE phd_registration_year='2021' and final_approval='rejected' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
               'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Your Caste', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
               'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
               'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
                'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts', 'HSC Total',
                'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream', 'Graduation Attempts',
                'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts', 'PG Total',
                'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
                'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
                'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
                'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
                'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
                'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District', 'Caste Issuing Authority',
                'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government', 'Government Department',
                'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR' ])

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Rejected_Applications_2021.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response     

# ----------------------------------------------------------------------------
# ------------------- END APPLICATIONS 2021 REPORTS ------------------------------
# ----------------------------------------------------------------------------

@app.route('/priority_people_report', methods=['GET', 'POST'])
def priority_people_report():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT * FROM application_page WHERE your_caste IN ('katkari', 'kolam', 'madia')")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    # print(result)
    return render_template('priority_people_report.html', result=result)


# ---------------------  Export Disability Report Yes list to Excel  -----------------------
@app.route('/export_priority_caste_report', methods=['GET', 'POST'])
def export_priority_caste_report():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("SELECT applicant_id, first_name, middle_name, last_name, mobile_number, email, date_of_birth,"
                   "gender, age, caste, your_caste, marital_status, add_1, add_2, pincode, village, taluka, district,"
                   "state, city,  phd_registration_date, concerned_university, topic_of_phd, "
                   "name_of_guide, name_of_college, stream, board_university, admission_year, passing_year, percentage,"
                   "family_annual_income, income_certificate_number, issuing_authority, domicile, domicile_certificate,"
                   "domicile_number, caste_certf, issuing_district, caste_issuing_authority, salaried, disability,"
                   "father_name, mother_name, work_in_government, bank_name, account_number, ifsc_code," 
                   "account_holder_name FROM application_page WHERE your_caste IN ('katkari', 'kolam', 'madia')")    
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['Applicant Id', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email', 'Date Of Birth', 
                'Gender', 'Age', 'Caste', 'Your Caste', 'Marital Status', 'Dependents', 'Add 1', 'Add 2', 'Pincode', 'Village', 
                'Taluka', 'District', 'State', 'Phd Registration Date', 'Concerned University', 'Topic Of Phd', 'Name Of Guide', 
                'Name Of College', 'Stream', 'Board University', 'Admission Year', 'Passing Year', 'Percentage', 'Family Annual Income', 
                'Income Certificate Number', 'Issuing Authority', 'Domicile', 'Domicile Certificate', 'Domicile Number', 'Caste Certf', 
                'Issuing District', 'Caste Issuing Authority', 'Salaried', 'Disability', 'Father Name', 'Mother Name', 'Work In Government', 
                'Bank Name', 'Account Number', 'Ifsc Code', 'Account Holder Name'])

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Priority_Caste_Report.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response


@app.route('/all_news', methods=['GET', 'POST'])
def all_news():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT * FROM news_and_updates")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    # print(result)
    return render_template('all_news_updates.html', result=result)


#---------------------  Export News in Excel  -----------------------
@app.route('/export_news', methods=['GET', 'POST'])
def export_news():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("SELECT id, user, title, subtitle FROM news_and_updates")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['Id','User','Title','Subtitle'])

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=News_and_Updates.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response

#---------------------  Export Rejected Users list in Excel  -----------------------
@app.route('/export_rejected_applications', methods=['GET', 'POST'])
def export_rejected_applications():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("SELECT first_name,middle_name,last_name, mobile_number, email FROM application_page WHERE final_approval = 'rejected'")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['first_name','middle_name','last_name','mobile_number','email','day','month','year','age'])

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Rejected_Users_Data.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response

# --------------------------------------------------------------------------------------------------------------------
# --------------------------------------------------------------------------------------------------------------------
#                        ----------------------- APPLICATION FORM -------------------------------
# --------------------------------------------------------------------------------------------------------------------
# --------------------------------------------------------------------------------------------------------------------
@app.route('/get_subcastes/<int:unique_number>', methods=['GET'])
def get_subcastes(unique_number):
    caste_class = casteController(host)
    subcastes = caste_class.get_subcastes_by_unique_number(unique_number)
    return jsonify({'subcastes': subcastes})


@app.route('/section1', methods=['GET', 'POST'])
@auth
def section1():
    email = session['email']
    print('I am in section 1:' + email)
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)

    # Check if a record already exists for this user
    cursor.execute("SELECT applicant_photo, adhaar_number, adhaar_seeding, first_name, "
                   "middle_name, last_name, mobile_number, email, date_of_birth, gender, age, caste, your_caste,"
                   "marital_status, add_1, add_2, pincode, village, taluka, district, state, city"
                   " FROM application_page WHERE email = %s", (email,))
    record = cursor.fetchone()
    print(record)
    caste_class = casteController(host)
    all_caste = caste_class.get_all_caste_details()
    print('ALL Castes', all_caste)
    # Initialize an empty dictionary if no record is found
    if record is None:
        record = {}

    if request.method == 'POST':
        #
        adhaar_number = request.form.get('aadhaar_refnum')
        adhaar_refnum = request.form.get('aadhaar_refnum')
        adhaar_seeding = request.form['adhaar_seeding']
        first_name = request.form['first_name']
        middle_name = request.form['middle_name']
        last_name = request.form['last_name']
        mobile_number = request.form['mobile_number']
        email = session['email']
        date_of_birth = request.form['date_of_birth']
        gender = request.form['gender']
        age = request.form['age']
        caste = request.form['caste']
        your_caste = request.form['your_caste']
        # pvtg = request.form['pvtg']
        # pvtg_caste = request.form['pvtg_caste']
        marital_status = request.form['marital_status']
        add_1 = request.form['add_1']
        add_2 = request.form['add_2']
        pincode = request.form['pincode']
        village = request.form['village']
        taluka = request.form['taluka']
        district = request.form['district']
        state = request.form['state']
        city = request.form['city']
        # subcaste = request.form['subcaste']
        # Access other fields in a similar manner
        # Handle file upload (applicant's photo)
        photo = request.files['applicant_photo']
        photo_path = save_applicant_photo(photo, first_name, last_name)

        if not record:
            # Save the form data to the database
            print('Inserting new record for:' + email)
            cursor.execute("INSERT INTO application_page (applicant_photo, adhaar_number, adhaar_refnum, adhaar_seeding,  first_name, "
                           "middle_name, last_name, mobile_number, email, date_of_birth, gender, age, caste, your_caste,"
                           "marital_status, add_1, add_2, pincode, village, taluka, district, state, city) "
                           "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,%s, %s, %s)",
                           (photo_path, adhaar_number, adhaar_refnum, adhaar_seeding, first_name, middle_name, last_name, mobile_number,
                            email, date_of_birth, gender, age, caste, your_caste, marital_status, add_1, add_2,
                            pincode, village, taluka, district, state, city))
            cnx.commit()
            return redirect(url_for('section2'))
    return render_template('AForm_section1.html', record=record,all_caste = all_caste)


@app.route('/section2', methods=['GET', 'POST'])
@auth
def section2():
    email = session['email']
    print('I am in section 2:' + email)
    cnx = mysql.connector.connect(user=user, password=password,
                                     host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    university_data = universityController(host)
    university_names = university_data.get_all_university()
    # Check if a record already exists for this user
    cursor.execute("SELECT ssc_passing_year, ssc_percentage, ssc_school_name, "
               "hsc_passing_year, hsc_percentage, hsc_school_name, graduation_passing_year, "
               "graduation_percentage, graduation_school_name, phd_passing_year, phd_percentage, "
               "phd_school_name, have_you_qualified, phd_registration_year, "
               "concerned_university, department_name, topic_of_phd, name_of_guide, name_of_college, "
               " faculty, family_annual_income, phd_registration_date,"
               "income_certificate_number, issuing_authority,ssc_stream,ssc_attempts,ssc_total,hsc_stream,hsc_attempts,"
                "hsc_total,grad_stream,grad_attempts,grad_total,pg_stream,pg_attempts,pg_total, income_issuing_district "
                "FROM application_page WHERE email = %s", (email,))
    existing_record = cursor.fetchone()
    print('existing_record', existing_record)
    cursor.execute('SELECT year FROM signup where email = %s',(email,))
    signup_year = cursor.fetchone()
    cursor.execute('SELECT YEAR(date_of_birth) AS birth_year FROM application_page where email = %s', (email,))
    dob_year = cursor.fetchone()

    if existing_record:
        # Determine whether the user is in "view" or "edit" mode
        if not any(value is None or value == '' for value in existing_record.values()):
            view_mode = 'view'
        else:
            view_mode = 'edit'
    else:
        # Handle the case when existing_record is None or empty
        view_mode = 'edit'  # or 'view', depending on your logic for this case

    # print(existing_record.values())

    if request.method == 'POST':
        ssc_passing_year = request.form['ssc_passing_year']
        ssc_percentage = request.form['ssc_percentage']
        ssc_school_name = request.form['ssc_school_name']
        hsc_passing_year = request.form['hsc_passing_year']
        hsc_percentage = request.form['hsc_percentage']
        hsc_school_name = request.form['hsc_school_name']
        graduation_passing_year = request.form['graduation_passing_year']
        graduation_percentage = request.form['graduation_percentage']
        graduation_school_name = request.form['graduation_school_name']
        phd_passing_year = request.form['phd_passing_year']
        phd_percentage = request.form['phd_percentage']
        phd_school_name = request.form['phd_school_name']
        have_you_qualified = request.form['have_you_qualified']
        cet_other = request.form['cet_other']
        phd_registration_date = request.form['phd_registration_date']
        phd_registration_year = request.form['phd_registration_year']
        phd_registration_age = request.form['phd_registration_age']
        concerned_university = request.form['concerned_university']
        topic_of_phd = request.form['topic_of_phd']
        name_of_guide = request.form['name_of_guide']
        name_of_college = request.form['name_of_college']
        print(name_of_college)
        faculty = request.form['faculty']
        family_annual_income = request.form['family_annual_income']
        income_certificate_number = request.form['income_certificate_number']
        issuing_authority = request.form['issuing_authority']
        # Code added by Akash /*-------------- Starts here ------------*/
        # for ssc changes percentage
        ssc_stream = request.form['ssc_stream']
        ssc_attempts = request.form['ssc_attempts']
        ssc_total = request.form['ssc_total']

        # for hsc percentage changes
        hsc_stream = request.form['hsc_stream']
        hsc_attempts = request.form['hsc_attempts']
        hsc_total = request.form['hsc_total']

        # For Graduation section
        grad_stream = request.form['grad_stream']
        grad_attempts = request.form['grad_attempts']
        grad_total = request.form['grad_total']

        # Post Graduation Section
        pg_stream = request.form['pg_stream']
        pg_attempts = request.form['pg_attempts']
        pg_total = request.form['pg_total']


        # Department section of college
        department_name = request.form['department_name']
        income_issuing_district = request.form['income_issuing_district']
        income_issuing_taluka = request.form['income_issuing_taluka']

        print("Values in existing_record:", existing_record.values())
        if all(value is None or value == '' for value in existing_record.values()):
            print('Inserting new record for:' + email)
            if view_mode == 'edit':
                cursor.execute( "UPDATE application_page SET ssc_passing_year = %s, ssc_percentage = %s, ssc_school_name = %s, "
                                "hsc_passing_year = %s, hsc_percentage = %s, hsc_school_name = %s, graduation_passing_year = %s, "
                                "graduation_percentage = %s, graduation_school_name = %s, phd_passing_year = %s, phd_percentage = %s, "
                                "phd_school_name = %s, have_you_qualified = %s, phd_registration_date = %s, phd_registration_year = %s, "
                                "phd_registration_age = %s, concerned_university = %s, topic_of_phd = %s, name_of_guide = %s, name_of_college = %s, "
                                " faculty = %s, family_annual_income = %s, "
                                "income_certificate_number = %s, issuing_authority = %s,ssc_stream = %s, ssc_attempts =%s,"
                                "ssc_total = %s, hsc_stream = %s,hsc_attempts = %s,hsc_total = %s,grad_stream = %s,"
                                "grad_attempts =%s, grad_total = %s, pg_stream = %s, pg_attempts = %s, pg_total = %s,"
                                "department_name = %s,income_issuing_district = %s, income_issuing_taluka =%s, have_you_qualified_other = %s  "
                                "WHERE email = %s",
                                (ssc_passing_year, ssc_percentage, ssc_school_name, hsc_passing_year, hsc_percentage, hsc_school_name,
                                 graduation_passing_year, graduation_percentage, graduation_school_name, phd_passing_year, phd_percentage,
                                 phd_school_name, have_you_qualified, phd_registration_date, phd_registration_year,phd_registration_age, concerned_university,
                                 topic_of_phd, name_of_guide, name_of_college,  faculty,
                                 family_annual_income, income_certificate_number, issuing_authority, ssc_stream,
                                 ssc_attempts,ssc_total,hsc_stream,hsc_attempts,hsc_total ,grad_stream, grad_attempts,
                                 grad_total,pg_stream, pg_attempts, pg_total, department_name, income_issuing_district,
                                 income_issuing_taluka, cet_other, email))
                cnx.commit()
            return redirect(url_for('section3'))
    # Select records again to verify the insertion
    cursor.execute("SELECT ssc_passing_year, ssc_percentage, ssc_school_name, "
               "hsc_passing_year, hsc_percentage, hsc_school_name, graduation_passing_year, "
               "graduation_percentage, graduation_school_name, phd_passing_year, phd_percentage, "
               "phd_school_name, have_you_qualified, phd_registration_year, "
               "concerned_university, department_name, topic_of_phd, name_of_guide, name_of_college, "
               "other_college_name,  faculty, family_annual_income, phd_registration_date,"
               "income_certificate_number, issuing_authority,ssc_stream,ssc_attempts,ssc_total,hsc_stream,hsc_attempts,"
                "hsc_total,grad_stream,grad_attempts,grad_total,pg_stream,pg_attempts,pg_total, income_issuing_district,"
               "diploma_percentage, diploma_passing_year, diploma_stream, diploma_school_name, diploma_attempts, diploma_total "
                "FROM application_page WHERE email = %s", (email,))
    new_record = cursor.fetchone()
    print('new_record', new_record)
    # Initialize view_mode with a default value

    return render_template('AForm_section2.html', existing_record=existing_record, view_mode=view_mode,
                           university_data = university_names, signup_year = signup_year, dob_year = dob_year,
                           new_record=new_record)



@app.route('/section3', methods=['GET', 'POST'])
@auth
def section3():
    email = session['email']
    print('I am in section 3:' + email)
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)

    # Check if a record already exists for this user
    cursor.execute("SELECT domicile, domicile_certificate, domicile_number, caste_certf, issuing_district, "
                   "caste_issuing_authority, salaried, disability, type_of_disability, caste_certf_number,validity_certificate, validity_cert_number, validity_issuing_district, validity_issuing_authority  "
                   "FROM application_page WHERE email = %s", (email,))
    existing_record = cursor.fetchone()
    print("existing_record", existing_record)
    # print(existing_record)
    # Check if existing_record is not None
    if existing_record is not None:
        # Determine whether the user is in "view" or "edit" mode
        if not any(existing_record.values()):
            view_mode = 'edit'
        else:
            view_mode = 'view'
    else:
        # Handle the case when existing_record is None
        view_mode = 'edit'  # or 'view', depending on your logic for this case

    if request.method == 'POST':
        # Save form data in the session for this section
        if request.method == 'POST':
            domicile = request.form['domicile']
            domicile_certificate = request.form['domicile_certificate']
            domicile_number = request.form['domicile_number']
            caste_certf = request.form['caste_certf']
            issuing_district = request.form['issuing_district']
            caste_issuing_authority = request.form['caste_issuing_authority']
            caste_cert_number = request.form['caste_cert_number']
            validity_certificate = request.form['validity_certificate']
            validity_num = request.form['validity_num']
            validity_issuing_district = request.form['validity_issuing_district']
            validity_issuing_authority = request.form['validity_issuing_authority']
            salaried = request.form['salaried']
            disability = request.form['disability']
            type_of_disability = request.form['type_of_disability']

            print("I am breaking before the if condition")
            if all(value is None for value in existing_record.values()):
                # Save the form data to the database
                print('I am breaking while insertion')
                print('Inserting new record for:' + email)
                if view_mode == 'edit':
                    cursor.execute("UPDATE application_page SET domicile = %s, domicile_certificate = %s, domicile_number = %s, "
                                   "caste_certf = %s, issuing_district = %s, caste_issuing_authority = %s, salaried = %s, "
                                   "disability = %s, type_of_disability = %s ," 
                                   "caste_certf_number = %s, validity_certificate = %s,validity_cert_number = %s, validity_issuing_district =%s,validity_issuing_authority =%s  WHERE email = %s",
                                   (domicile, domicile_certificate, domicile_number, caste_certf, issuing_district,
                                    caste_issuing_authority, salaried, disability, type_of_disability,caste_cert_number, validity_certificate, validity_num, validity_issuing_district,validity_issuing_authority,
                                    email))
                    cnx.commit()
                    # cursor.execute("SELECT domicile, domicile_certificate, domicile_number, "
                    #                "caste_certf, issuing_district, caste_issuing_authority, salaried, disability,"
                    #                " type_of_disability FROM application_page WHERE email = %s", (email,))
                    # existing_record = cursor.fetchone()
                    # if existing_record:
                    #     return redirect(url_for('section4'))
            return redirect(url_for('section4'))
        # Select records again to verify the insertion
        cursor.execute("SELECT domicile, domicile_certificate, domicile_number, caste_certf, issuing_district, "
                   "caste_issuing_authority, salaried, disability, type_of_disability, caste_certf_number,validity_certificate, validity_cert_number, validity_issuing_district, validity_issuing_authority  "
                   "FROM application_page WHERE email = %s", (email,))
        existing_record = cursor.fetchone()
        print(existing_record)
    return render_template('AForm_section3.html', existing_record=existing_record, view_mode=view_mode)


@app.route('/section4', methods=['GET', 'POST'])
def section4():
    email = session['email']
    print('I am in section 4:' + email)
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)

    # Check if a record already exists for this user
    cursor.execute("SELECT father_name, mother_name, work_in_government, bank_name, account_number, "
                   "ifsc_code, account_holder_name, micr FROM application_page WHERE email = %s", (email,))
    existing_record = cursor.fetchone()
    print(existing_record)

    if existing_record is not None:
        # Determine whether the user is in "view" or "edit" mode
        if not any(existing_record.values()):
            view_mode = 'edit'
        else:
            view_mode = 'view'
    else:
        # Handle the case when existing_record is None
        view_mode = 'edit'  # or 'view', depending on your logic for this case

    if request.method == 'POST':
        # Save form data in the session for this section
        father_name = request.form['father_name']
        mother_name = request.form['mother_name']
        work_in_government = request.form['work_in_government']
        gov_department = request.form['gov_department']
        gov_position = request.form['gov_position']
        bank_name = request.form['bank_name']
        account_number = request.form['account_number']
        ifsc_code = request.form['ifsc_code']
        account_holder_name = request.form['account_holder_name']
        micr = request.form['micr']
        # Process the fields as needed
        print(request.form)
        print("I am breaking before the if condition")
        if all(value is None for value in existing_record.values()):
            # Save the form data to the database
            print('I am breaking while insertion')
            print('Inserting new record for:' + email)
            if view_mode == 'edit':
                cursor.execute("UPDATE application_page SET father_name = %s, mother_name = %s, work_in_government = %s,"
                               "gov_department = %s, gov_position = %s, bank_name = %s, account_number = %s, "
                               "ifsc_code = %s, account_holder_name = %s, micr = %s WHERE email = %s",
                                (father_name, mother_name, work_in_government, gov_department, gov_position,  bank_name,
                                 account_number, ifsc_code, account_holder_name, micr,
                                email))
                cnx.commit()
            return redirect(url_for('section5'))
        # Select records again to verify the insertion
        cursor.execute("SELECT father_name, mother_name, work_in_government, gov_department, gov_position, bank_name,"
                       "account_number, ifsc_code, account_holder_name, micr FROM application_page WHERE email = %s", (email,))
        existing_record = cursor.fetchone()
        print(existing_record)
    return render_template('AForm_section4.html', existing_record=existing_record, view_mode=view_mode)


@app.route('/section5', methods=['GET', 'POST'])
def section5():
    email = session['email']
    print('I am in section 5:' + email)
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)


    # Check if a record already exists for this user
    cursor.execute("SELECT first_name, last_name, documentfile1, documentfile2, documentfile3, documentfile4, "
                   "documentfile5, documentfile6, documentfile7, documentfile8, signature, documentfile9, documentfile10,"
                   " documentfile11, documentfile12, documentfile13, documentfile14, documentfile15, documentfile16,"
                   "documentfile17, documentfile18, documentfile19, documentfile20, documentfile21, documentfile22,"
                   "documentfile23, documentfile24, documentfile25, documentfile26  FROM application_page WHERE email = %s", (email,))
    existing_record = cursor.fetchone()
    print("existing_record", existing_record)

    if request.method == 'POST':

        cursor.execute("SELECT phd_registration_year, id FROM application_page WHERE email = %s", (email,))
        row = cursor.fetchone()
        unique_id = row['id']
        year = row['phd_registration_year']
        applicant_id = 'TRTI' + '/' + str(year) + '/' + str(unique_id)
        # Get current date
        current_date = datetime.now().strftime('%Y-%m-%d')
        # Get current time
        current_time = datetime.now().strftime('%H:%M:%S')
        print('Got Post request:')
        first_name = existing_record['first_name']
        last_name = existing_record['last_name']
    # Process the form data and files
        document1 = applicant_pdf_upload_section_five(request.files['documentfile1'], first_name, last_name)
        document2 = applicant_pdf_upload_section_five(request.files['documentfile2'], first_name, last_name)
        document3 = applicant_pdf_upload_section_five(request.files['documentfile3'], first_name, last_name)
        document4 = applicant_pdf_upload_section_five(request.files['documentfile4'], first_name, last_name)
        document5 = applicant_pdf_upload_section_five(request.files['documentfile5'], first_name, last_name)
        document6 = applicant_pdf_upload_section_five(request.files['documentfile6'], first_name, last_name)
        document7 = applicant_pdf_upload_section_five(request.files['documentfile7'], first_name, last_name)
        document8 = applicant_pdf_upload_section_five(request.files['documentfile8'], first_name, last_name)
        signature = applicant_pdf_upload_section_five(request.files['signature'], first_name, last_name)
        document9 = applicant_pdf_upload_section_five(request.files['documentfile9'], first_name, last_name)
        document10 = applicant_pdf_upload_section_five(request.files['documentfile10'], first_name, last_name)
        document11 = applicant_pdf_upload_section_five(request.files['documentfile11'], first_name, last_name)
        document12 = applicant_pdf_upload_section_five(request.files['documentfile12'], first_name, last_name)
        document13 = applicant_pdf_upload_section_five(request.files['documentfile13'], first_name, last_name)
        document14 = applicant_pdf_upload_section_five(request.files['documentfile14'], first_name, last_name)
        document15 = applicant_pdf_upload_section_five(request.files['documentfile15'], first_name, last_name)
        document16 = applicant_pdf_upload_section_five(request.files['documentfile16'], first_name, last_name)
        document17 = applicant_pdf_upload_section_five(request.files['documentfile17'], first_name, last_name)
        document18 = applicant_pdf_upload_section_five(request.files['documentfile18'], first_name, last_name)
        document19 = applicant_pdf_upload_section_five(request.files['documentfile19'], first_name, last_name)
        document20 = applicant_pdf_upload_section_five(request.files['documentfile20'], first_name, last_name)
        document21 = applicant_pdf_upload_section_five(request.files['documentfile21'], first_name, last_name)
        document22 = applicant_pdf_upload_section_five(request.files['documentfile22'], first_name, last_name)
        document23 = applicant_pdf_upload_section_five(request.files['documentfile23'], first_name, last_name)
        document24 = applicant_pdf_upload_section_five(request.files['documentfile24'], first_name, last_name)
        # document25 = applicant_pdf_upload_section_five(request.files['documentfile25'], first_name, last_name)
        document26 = applicant_pdf_upload_section_five(request.files['documentfile26'], first_name, last_name)
    else:
        document1 = ''
        document2 = ''
        document3 = ''
        document4 = ''
        document5 = ''
        document6 = ''
        document7 = ''
        document8 = ''
        signature = ''
        document9 = ''
        document10 = ''
        document11 = ''
        document12 = ''
        document13 = ''
        document14 = ''
        document15 = ''
        document16 = ''
        document17 = ''
        document18 = ''
        document19 = ''
        document20 = ''
        document21 = ''
        document22 = ''
        document23 = ''
        document24 = ''
        # document25 = ''
        document26 = ''

    # Check if any of the documents have been uploaded
    if any(doc for doc in
           [document1, document2, document3, document4, document5, document6, document7, document8, signature, document9,
            document10, document11, document12, document13, document14, document15, document16, document17, document18,
            document19, document20, document21, document22, document23, document24, document26]):
        # If any document is uploaded, update or insert into the database
        if all(value is None for value in existing_record.values()):
            print('In the second loop updating the documents')
            # Insert new records
            print('Inserting new record for:' + email)
            cursor.execute(
                "UPDATE application_page SET documentfile1 = %s, documentfile2 = %s, documentfile3 = %s, documentfile4 = %s, "
                "documentfile5 = %s, documentfile6 = %s, documentfile7 = %s, documentfile8 = %s, signature = %s, documentfile9 = %s, "
                "documentfile10 = %s, documentfile11 = %s, documentfile12 = %s, documentfile13 = %s, documentfile14 = %s,"
                "documentfile15 = %s, documentfile16 = %s, documentfile17 = %s, documentfile18 = %s, documentfile19 = %s, "
                "documentfile20 = %s, documentfile21 = %s, documentfile22 = %s, documentfile23 = %s, documentfile24 = %s,"
                "documentfile26 = %s WHERE email = %s",
                (document1, document2, document3, document4, document5, document6, document7, document8, signature, document9,
                 document10, document11, document12, document13, document14, document15, document16, document17,
                 document18, document19, document20, document21, document22, document23, document24, document26, email))
            cnx.commit()     
        else:
            # Update existing records
            print('Updating records for:' + email)
            cursor.execute(
                "UPDATE application_page SET applicant_id = %s, form_filled = 1, documentfile1 = %s, documentfile2 = %s, documentfile3 = %s, documentfile4 = %s, "
                "documentfile5 = %s, documentfile6 = %s, documentfile7 = %s, documentfile8 = %s, signature = %s, documentfile9 = %s,"
                "documentfile10 = %s, documentfile11 = %s, documentfile12 = %s, documentfile13 = %s, documentfile14 = %s,"
                "documentfile15 = %s, documentfile16 = %s, documentfile17 = %s, documentfile18 = %s, documentfile19 = %s, "
                "documentfile20 = %s, documentfile21 = %s, documentfile22 = %s, documentfile23 = %s,  documentfile24 = %s,"
                " documentfile26 = %s, application_date = %s, application_time =%s WHERE email = %s",
                (applicant_id, document1, document2, document3, document4, document5, document6, document7, document8, signature, document9,
                 document10, document11, document12, document13, document14, document15, document16, document17,
                 document18, document19, document20, document21, document22, document23, document24, document26,
                 current_date, current_time, email))
            cnx.commit()
        cursor.execute("SELECT * FROM application_page WHERE email = %s", (email,))
        records = cursor.fetchone()

        return render_template('submit_form.html', records=records)
    else:
        # Handle case where no files were uploaded
        # You can add appropriate logic here (e.g., show an error message)
        print('No files were uploaded')

    cursor.execute("SELECT have_you_qualified FROM application_page WHERE email = %s", (email,))
    mphil_selected = cursor.fetchone()

    return render_template('AForm_section5.html', existing_record=existing_record, mphil_selected=mphil_selected)


def university_college():
    email = session['email']
    cnx = mysql.connector.connect(user='icswebapp', password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    #concerned_university = request.form['concerned_university']
    cursor.execute("SELECT DISTINCT id, u_id, affiliated_universities FROM universities GROUP BY u_id")
    data = cursor.fetchall()
    # print("Data :", data)
    return data


@app.route('/get_colleges', methods=['POST'])
def get_colleges():
    selected_university = request.form['selected_university']
    cnx = mysql.connector.connect(user='icswebapp', password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    # Fetch the colleges for the selected university ID
    cursor.execute("SELECT id, college_name FROM universities WHERE u_id = %s", (selected_university,))
    colleges = cursor.fetchall()
    return jsonify(colleges)


@app.route('/preview', methods=['GET', 'POST'])
def preview():
    email = session['email']
    cnx = mysql.connector.connect(user='icswebapp', password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    # Check if the user's email is in old_users for 2021 or 2022
    cursor.execute("SELECT year FROM signup WHERE email = %s AND year IN ('2020', '2021', '2022')", (email,))
    old_user_data = cursor.fetchone()
    concerned_university = request.get_data('concerned_university')
    cursor.execute("SELECT college_name FROM universities WHERE affiliated_universities = %s ", (concerned_university,))
    data_college = cursor.fetchone()
    cursor.execute('SELECT * FROM districts')
    district_list = cursor.fetchall()
    if old_user_data:
        session['user_type'] = 'old_user' 
        university_data = university_college()
        # User exists in old_users for 2021 or 2022, use this data
        return render_template('preview.html', records=old_user_data, editable=True, university_data=university_data, data_college=data_college, district_list = district_list)
    else:
        session['user_type'] = 'new_user'
        # User does not exist in old_users for 2021 or 2022, fetch from application_page
        cursor.execute("SELECT * FROM application_page WHERE email = %s", (email,))
        application_page_data = cursor.fetchone()
        university_data = university_college()
        return render_template('preview.html', records=application_page_data, editable=True, university_data=university_data, data_college=data_college, district_list = district_list)

    cursor.close()
    cnx.close()
    return render_template('preview.html', records=records, editable=True)


@app.route('/old_user_preview', methods=['GET', 'POST'])
def old_user_preview():
    if session.pop('logged_in_from_login', None):
        flash('Logged in Successfully', 'success')
    email = session['email']
    cnx = mysql.connector.connect(user='icswebapp', password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    # Check if the user's email is in old_users for 2021 or 2022
    cursor.execute("SELECT year FROM signup WHERE email = %s AND year IN ('2020', '2021', '2022')", (email,))
    old_user = cursor.fetchone()
    concerned_university = request.get_data('concerned_university')
    cursor.execute("SELECT college_name FROM universities WHERE affiliated_universities = %s ", (concerned_university,))
    data_college = cursor.fetchone()
    cursor.execute('SELECT * FROM districts')
    district_list = cursor.fetchall()
    if old_user:
        session['user_type'] = 'old_user'
        university_data = university_college()
        cursor.execute("SELECT * FROM old_users WHERE email = %s", (email,))
        old_user_data = cursor.fetchone()
        # User exists in old_users for 2021 or 2022, use this data
        return render_template('old_user_preview.html', records=old_user_data, editable=True, university_data=university_data, data_college=data_college, district_list = district_list)
    else:
        session['user_type'] = 'new_user'
        # User does not exist in old_users for 2021 or 2022, fetch from application_page
        cursor.execute("SELECT * FROM application_page WHERE email = %s", (email,))
        application_page_data = cursor.fetchone()
        university_data = university_college()
        return render_template('old_user_preview.html', records=application_page_data, editable=True, university_data=university_data, data_college=data_college, district_list = district_list)

    cursor.close()
    cnx.close()
    return render_template('old_user_preview.html', records=records, editable=True, adhaar_error=adhaar_error)


def update_database_old_users(data, phd_registration_year, cursor):
    try:
        with mysql.connector.connect(user=user, password=password,
                                     host=host,
                                     database=database) as cnx:
            with cnx.cursor(dictionary=True) as cursor:
                print("phd_registration_year", phd_registration_year)
                # Using the update_data dictionary in the SQL query

                # Build the SQL UPDATE statement
                update_query = f"""
                            UPDATE application_page
                            SET 
                            form_filled = %(form_filled)s,
                            applicant_id = %(applicant_id)s,
                            application_date = %(application_date)s,
                            applicant_photo = %(applicant_photo)s,
                            adhaar_number = %(adhaar_number)s,
                            adhaar_seeding = %(adhaar_seeding)s,
                            first_name = %(first_name)s,
                            middle_name = %(middle_name)s,
                            last_name = %(last_name)s,
                            mobile_number = %(mobile_number)s,
                            date_of_birth = %(date_of_birth)s,
                            gender = %(gender)s,
                            age = %(age)s,
                            marital_status = %(marital_status)s,
                            add_1 = %(add_1)s,
                            add_2 = %(add_2)s,
                            pincode = %(pincode)s,
                            village = %(village)s,
                            city = %(city)s,
                            taluka = %(taluka)s,
                            district = %(district)s,
                            state = %(state)s,
                            ssc_passing_year = %(ssc_passing_year)s,
                            ssc_percentage = %(ssc_percentage)s,
                            ssc_school_name = %(ssc_school_name)s,
                            hsc_passing_year = %(hsc_passing_year)s,
                            hsc_percentage = %(hsc_percentage)s,
                            hsc_school_name = %(hsc_school_name)s,
                            graduation_passing_year = %(graduation_passing_year)s,
                            graduation_percentage = %(graduation_percentage)s,
                            graduation_school_name = %(graduation_school_name)s,
                            phd_passing_year = %(phd_passing_year)s,
                            phd_percentage = %(phd_percentage)s,
                            phd_school_name = %(phd_school_name)s,
                            phd_registration_date = %(phd_registration_date)s,
                            phd_registration_year = %(phd_registration_year)s,
                            concerned_university = %(concerned_university)s,
                            name_of_college = %(name_of_college)s,
                            other_college_name = %(other_college_name)s,
                            department_name = %(department_name)s,
                            topic_of_phd = %(topic_of_phd)s,
                            name_of_guide = %(name_of_guide)s,
                            faculty = %(faculty)s,
                            salaried = %(salaried)s,
                            disability = %(disability)s,
                            family_annual_income = %(family_annual_income)s,
                            income_certificate_number = %(income_certificate_number)s,
                            issuing_authority = %(issuing_authority)s,
                            domicile = %(domicile)s,
                            domicile_certificate = %(domicile_certificate)s,
                            domicile_number = %(domicile_number)s,
                            validity_certificate = %(validity_certificate)s,
                            validity_cert_number = %(validity_cert_number)s,
                            caste_certf = %(caste_certf)s,
                            caste_certf_number = %(caste_certf_number)s,
                            caste = %(caste)s,
                            your_caste = %(your_caste)s,
                            issuing_district = %(issuing_district)s,
                            caste_issuing_authority = %(caste_issuing_authority)s,
                            father_name = %(father_name)s,
                            mother_name = %(mother_name)s,
                            bank_name = %(bank_name)s,
                            account_number = %(account_number)s,
                            ifsc_code = %(ifsc_code)s,
                            account_holder_name = %(account_holder_name)s,
                            signature = %(signature)s,
                            documentfile1 = %(documentfile1)s,
                            documentfile2 = %(documentfile2)s,
                            documentfile3 = %(documentfile3)s,
                            documentfile4 = %(documentfile4)s,
                            documentfile5 = %(documentfile5)s,
                            documentfile6 = %(documentfile6)s,
                            documentfile7 = %(documentfile7)s,
                            documentfile8 = %(documentfile8)s,
                            documentfile9 = %(documentfile9)s,
                            documentfile10 = %(documentfile10)s
                            WHERE email = %(email)s
                        """
                print(update_query)
                # Execute the UPDATE statement with data
                cursor.execute(update_query, data)
                cnx.commit()
                print(f"Rows affected: {cursor.rowcount}")
    except mysql.connector.Error as e:
        # Handle the exception, rollback if necessary
        print(f"Error: {e}")


def update_database(data, phd_registration_year, cursor):
    try:
        with mysql.connector.connect(user=user, password=password,
                                     host=host,
                                     database=database) as cnx:
            with cnx.cursor(dictionary=True) as cursor:
                print("phd_registration_year", phd_registration_year)
                # Using the update_data dictionary in the SQL query

                # Build the SQL UPDATE statement
                update_query = f"""
                            UPDATE application_page
                            SET 
                            form_filled = 1,
                            applicant_id = %(applicant_id)s,
                            application_date = %(application_date)s,
                            applicant_photo = %(applicant_photo)s,
                            adhaar_number = %(adhaar_number)s,
                            adhaar_seeding = %(adhaar_seeding)s,
                            first_name = %(first_name)s,
                            middle_name = %(middle_name)s,
                            last_name = %(last_name)s,
                            mobile_number = %(mobile_number)s,
                            date_of_birth = %(date_of_birth)s,
                            gender = %(gender)s,
                            age = %(age)s,
                            marital_status = %(marital_status)s,
                            add_1 = %(add_1)s,
                            add_2 = %(add_2)s,
                            pincode = %(pincode)s,
                            village = %(village)s,
                            city = %(city)s,
                            taluka = %(taluka)s,
                            district = %(district)s,
                            state = %(state)s,
                            ssc_passing_year = %(ssc_passing_year)s,
                            ssc_percentage = %(ssc_percentage)s,
                            ssc_school_name = %(ssc_school_name)s,
                            hsc_passing_year = %(hsc_passing_year)s,
                            hsc_percentage = %(hsc_percentage)s,
                            hsc_school_name = %(hsc_school_name)s,
                            graduation_passing_year = %(graduation_passing_year)s,
                            graduation_percentage = %(graduation_percentage)s,
                            graduation_school_name = %(graduation_school_name)s,
                            phd_passing_year = %(phd_passing_year)s,
                            phd_percentage = %(phd_percentage)s,
                            phd_school_name = %(phd_school_name)s,
                            phd_registration_date = %(phd_registration_date)s,
                            phd_registration_year = %(phd_registration_year)s,
                            concerned_university = %(concerned_university)s,
                            name_of_college = %(name_of_college)s,
                            other_college_name = %(other_college_name)s,
                            department_name = %(department_name)s,
                            topic_of_phd = %(topic_of_phd)s,
                            name_of_guide = %(name_of_guide)s,
                            faculty = %(faculty)s,
                            salaried = %(salaried)s,
                            disability = %(disability)s,
                            family_annual_income = %(family_annual_income)s,
                            income_certificate_number = %(income_certificate_number)s,
                            issuing_authority = %(issuing_authority)s,
                            domicile = %(domicile)s,
                            domicile_certificate = %(domicile_certificate)s,
                            domicile_number = %(domicile_number)s,
                            validity_certificate = %(validity_certificate)s,
                            validity_cert_number = %(validity_cert_number)s,
                            caste_certf = %(caste_certf)s,
                            caste_certf_number = %(caste_certf_number)s,
                            caste = %(caste)s,
                            your_caste = %(your_caste)s,
                            issuing_district = %(issuing_district)s,
                            caste_issuing_authority = %(caste_issuing_authority)s,
                            father_name = %(father_name)s,
                            mother_name = %(mother_name)s,
                            bank_name = %(bank_name)s,
                            account_number = %(account_number)s,
                            ifsc_code = %(ifsc_code)s,
                            account_holder_name = %(account_holder_name)s,
                            signature = %(signature)s,
                            documentfile1 = %(documentfile1)s,
                            documentfile2 = %(documentfile2)s,
                            documentfile3 = %(documentfile3)s,
                            documentfile4 = %(documentfile4)s,
                            documentfile5 = %(documentfile5)s,
                            documentfile6 = %(documentfile6)s,
                            documentfile7 = %(documentfile7)s,
                            documentfile8 = %(documentfile8)s,
                            documentfile9 = %(documentfile9)s,
                            documentfile10 = %(documentfile10)s
                            WHERE email = %(email)s
                        """
                print('update_query' + update_query)
                # Execute the UPDATE statement with data
                cursor.execute(update_query, data)
                cnx.commit()
                print(f"Rows affected: {cursor.rowcount}")
    except mysql.connector.Error as e:
        # Handle the exception, rollback if necessary
        print(f"Error: {e}")


def update_old_user_in_application_page(data, phd_registration_year, cursor):
    try:
        with mysql.connector.connect(user=user, password=password,
                                     host=host,
                                     database=database) as cnx:
            with cnx.cursor(dictionary=True) as cursor:
                print("phd_registration_year", phd_registration_year)
                # Using the update_data dictionary in the SQL query

                # Build the SQL UPDATE statement
                update_query = f"""
                            UPDATE old_users
                            SET 
                            form_filled = %(form_filled)s,
                            applicant_id = %(applicant_id)s,
                            application_date = %(application_date)s,
                            applicant_photo = %(applicant_photo)s,
                            adhaar_number = %(adhaar_number)s,
                            first_name = %(first_name)s,
                            middle_name = %(middle_name)s,
                            last_name = %(last_name)s,
                            mobile_number = %(mobile_number)s,
                            date_of_birth = %(date_of_birth)s,
                            gender = %(gender)s,
                            age = %(age)s,
                            marital_status = %(marital_status)s,
                            add_1 = %(add_1)s,
                            add_2 = %(add_2)s,
                            pincode = %(pincode)s,
                            village = %(village)s,
                            city = %(city)s,
                            taluka = %(taluka)s,
                            district = %(district)s,
                            state = %(state)s,
                            ssc_passing_year = %(ssc_passing_year)s,
                            ssc_percentage = %(ssc_percentage)s,
                            ssc_school_name = %(ssc_school_name)s,
                            hsc_passing_year = %(hsc_passing_year)s,
                            hsc_percentage = %(hsc_percentage)s,
                            hsc_school_name = %(hsc_school_name)s,
                            graduation_passing_year = %(graduation_passing_year)s,
                            graduation_percentage = %(graduation_percentage)s,
                            graduation_school_name = %(graduation_school_name)s,
                            phd_passing_year = %(phd_passing_year)s,
                            phd_percentage = %(phd_percentage)s,
                            phd_school_name = %(phd_school_name)s,
                            phd_registration_date = %(phd_registration_date)s,
                            phd_registration_year = %(phd_registration_year)s,
                            concerned_university = %(concerned_university)s,
                            name_of_college = %(name_of_college)s,
                            department_name = %(department_name)s,
                            topic_of_phd = %(topic_of_phd)s,
                            name_of_guide = %(name_of_guide)s,
                            faculty = %(faculty)s,
                            salaried = %(salaried)s,
                            disability = %(disability)s,
                            family_annual_income = %(family_annual_income)s,
                            income_certificate_number = %(income_certificate_number)s,
                            issuing_authority = %(issuing_authority)s,
                            domicile = %(domicile)s,
                            domicile_certificate = %(domicile_certificate)s,
                            domicile_number = %(domicile_number)s,
                            caste_certf = %(caste_certf)s,
                            issuing_district = %(issuing_district)s,
                            caste_issuing_authority = %(caste_issuing_authority)s,
                            father_name = %(father_name)s,
                            mother_name = %(mother_name)s,
                            bank_name = %(bank_name)s,
                            account_number = %(account_number)s,
                            ifsc_code = %(ifsc_code)s,
                            account_holder_name = %(account_holder_name)s,
                            signature = %(signature)s,
                            documentfile1 = %(documentfile1)s,
                            documentfile2 = %(documentfile2)s,
                            documentfile3 = %(documentfile3)s,
                            documentfile4 = %(documentfile4)s,
                            documentfile5 = %(documentfile5)s,
                            documentfile6 = %(documentfile6)s,
                            documentfile7 = %(documentfile7)s,
                            documentfile8 = %(documentfile8)s,
                            documentfile9 = %(documentfile9)s,
                            documentfile10 = %(documentfile10)s
                            WHERE email = %(email)s
                        """
                print('update_old_user_in_application_page', update_query)
                # Execute the UPDATE statement with data
                cursor.execute(update_query, data)
                cnx.commit()
                print(f"Rows affected: {cursor.rowcount}")
    except mysql.connector.Error as e:
        # Handle the exception, rollback if necessary
        print(f"Error: {e}")


@app.route('/submit_form', methods=['GET', 'POST'])
def submit_form():
    print('I am in application form submit route')
    email = session['email']
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)

    records = None

    if request.method == 'POST':
        print('Got the POST request')
        today = date.today()
        now = datetime.now()
        dt_string = now.strftime("%d%m%Y%H%M%S")

        # Folder Set
        applicant_photo = request.files['applicant_photo']
        adhaar_number = request.form['adhaar_number']
        adhaar_seeding = request.form['adhaar_seeding']
        first_name = request.form['first_name']
        middle_name = request.form['middle_name']
        last_name = request.form['last_name']
        mobile_number = request.form['mobile_number']
        email = request.form['email']
        date_of_birth = request.form['date_of_birth']
        gender = request.form['gender']
        age = request.form['age']
        marital_status = request.form['marital_status']
        add_1 = request.form['add_1']
        add_2 = request.form['add_2']
        pincode = request.form['pincode']
        village = request.form['village']
        taluka = request.form['taluka']
        city = request.form['city']
        district = request.form['district']
        state = request.form['state']
        ssc_passing_year = request.form['ssc_passing_year']
        ssc_percentage = request.form['ssc_percentage']
        ssc_school_name = request.form['ssc_school_name']
        hsc_passing_year = request.form['hsc_passing_year']
        hsc_percentage = request.form['hsc_percentage']
        hsc_school_name = request.form['hsc_school_name']
        graduation_passing_year = request.form['graduation_passing_year']
        graduation_percentage = request.form['graduation_percentage']
        graduation_school_name = request.form['graduation_school_name']
        phd_passing_year = request.form["phd_passing_year"]
        phd_percentage = request.form["phd_percentage"]
        phd_school_name = request.form["phd_school_name"]
        phd_registration_date = request.form['phd_registration_date']
        phd_registration_year = request.form['phd_registration_year']
        concerned_university = request.form['concerned_university']
        name_of_college = request.form.get('selected_college')
        other_college_name = request.form['other_college_name']
        department_name = request.form['department_name']
        topic_of_phd = request.form['topic_of_phd']
        name_of_guide = request.form['name_of_guide']
        faculty = request.form['faculty']
        salaried = request.form['salaried']
        disability = request.form['disability']
        family_annual_income = request.form['family_annual_income']
        income_certificate_number = request.form['income_certificate_number']
        issuing_authority = request.form['issuing_authority']
        domicile = request.form.get('domicile')
        domicile_certificate = request.form.get('domicile_certificate')
        domicile_number = request.form['domicile_number']
        validity_certificate = request.form['validity_certificate']
        validity_cert_number = request.form['validity_cert_number']
        caste_certf = request.form.get('caste_certf')
        caste_certf_number = request.form['caste_certf_number']
        caste = request.form['caste']
        your_caste = request.form['your_caste']
        issuing_district = request.form['issuing_district']
        caste_issuing_authority = request.form['caste_issuing_authority']
        father_name = request.form['father_name']
        mother_name = request.form['mother_name']
        bank_name = request.form['bank_name']
        account_number = request.form['account_number']
        ifsc_code = request.form['ifsc_code']
        account_holder_name = request.form['account_holder_name']
        signature = request.files['signature']
        print(signature)
        documentfile1 = request.files['documentfile1']
        documentfile2 = request.files['documentfile2']
        documentfile3 = request.files['documentfile3']
        documentfile4 = request.files['documentfile4']
        documentfile5 = request.files['documentfile5']
        documentfile6 = request.files['documentfile6']
        documentfile7 = request.files['documentfile7']
        documentfile8 = request.files['documentfile8']
        documentfile9 = request.files['documentfile9']
        documentfile10 = request.files['documentfile10']


        # Save files
        if signature.filename != '':
            signature_path = save_applicant_photo(signature, first_name, last_name)
        else:
            signature_path = request.form['signature_photo']

        if applicant_photo.filename != '':
            photo_path = save_applicant_photo(applicant_photo, first_name, last_name)
        else:
            photo_path = request.form['user_photo']

        if documentfile1.filename != '':
            documentfile1_path = applicant_pdf_upload_section_five(documentfile1, first_name, last_name)
        else:
            documentfile1_path = request.form['file1']

        if documentfile2.filename != '':
            documentfile2_path = applicant_pdf_upload_section_five(documentfile2, first_name, last_name)
        else:
            documentfile2_path = request.form['file2']

        if documentfile3.filename != '':
            documentfile3_path = applicant_pdf_upload_section_five(documentfile3, first_name, last_name)
        else:
            documentfile3_path = request.form['file3']  

        if documentfile4.filename != '':
            documentfile4_path = applicant_pdf_upload_section_five(documentfile4, first_name, last_name)
        else:
            documentfile4_path = request.form['file4'] 

        if documentfile5.filename != '':
            documentfile5_path = applicant_pdf_upload_section_five(documentfile5, first_name, last_name)
        else:
            documentfile5_path = request.form['file5']   

        if documentfile6.filename != '':
            documentfile6_path = applicant_pdf_upload_section_five(documentfile6, first_name, last_name)
        else:
            documentfile6_path = request.form['file6']   

        if documentfile7.filename != '':
            documentfile7_path = applicant_pdf_upload_section_five(documentfile7, first_name, last_name)
        else:
            documentfile7_path = request.form['file7']  

        if documentfile8.filename != '':
            documentfile8_path = applicant_pdf_upload_section_five(documentfile8, first_name, last_name)
        else:
            documentfile8_path = request.form['file8']  

        if documentfile9.filename != '':
            documentfile9_path = applicant_pdf_upload_section_five(documentfile9, first_name, last_name)
        else:
            documentfile9_path = request.form['file9']

        if documentfile10.filename != '':
            documentfile10_path = applicant_pdf_upload_section_five(documentfile10, first_name, last_name)
        else:
            documentfile10_path = request.form['file10']

        print('Got all the request. fields')

        cursor.execute("SELECT phd_registration_year, id FROM application_page WHERE email = %s", (email,))
        result = cursor.fetchone()

        cursor.execute("SELECT phd_registration_year, id FROM old_users WHERE email = %s", (email,))
        record = cursor.fetchone()
        print("Record", record)

        if result:
            print('I am in result')
            phd_registration_year = result['phd_registration_year']
            id = result['id']
            datatime_now = datetime.now()
            applicant_id = f"TRTI/{phd_registration_year}/{id}"

            print(signature_path)

            update_data = {
                "applicant_id": applicant_id,
                "form_filled": 1,
                "application_date": datetime.now(),
                "applicant_photo": photo_path,
                "adhaar_number": adhaar_number,
                "adhaar_seeding": adhaar_seeding,
                "first_name": first_name,
                "middle_name": middle_name,
                "last_name": last_name,
                "mobile_number": mobile_number,
                "date_of_birth": date_of_birth,
                "gender": gender,
                "age": age,
                "marital_status": marital_status,
                "add_1": add_1,
                "add_2": add_2,
                "pincode": pincode,
                "village": village,
                "city": city,
                "taluka": taluka,
                "district": district,
                "state": state,
                "ssc_passing_year": ssc_passing_year,
                "ssc_percentage": ssc_percentage,
                "ssc_school_name": ssc_school_name,
                "hsc_passing_year": hsc_passing_year,
                "hsc_percentage": hsc_percentage,
                "hsc_school_name": hsc_school_name,
                "graduation_passing_year": graduation_passing_year,
                "graduation_percentage": graduation_percentage,
                "graduation_school_name": graduation_school_name,
                "phd_passing_year": phd_passing_year,
                "phd_percentage": phd_percentage,
                "phd_school_name": phd_school_name,
                "phd_registration_date": phd_registration_date,
                "phd_registration_year": phd_registration_year,
                "concerned_university": concerned_university,
                "name_of_college": name_of_college,
                "other_college_name": other_college_name,
                "department_name": department_name,
                "topic_of_phd": topic_of_phd,
                "name_of_guide": name_of_guide,
                "faculty": faculty,
                "salaried": salaried,
                "disability": disability,
                "family_annual_income": family_annual_income,
                "income_certificate_number": income_certificate_number,
                "issuing_authority": issuing_authority,
                "domicile": domicile,
                "domicile_certificate": domicile_certificate,
                "domicile_number": domicile_number,
                "validity_certificate": validity_certificate,
                "validity_cert_number": validity_cert_number,
                "caste_certf": caste_certf,
                "caste_certf_number": caste_certf_number,
                "caste": caste,
                "your_caste": your_caste,
                "issuing_district": issuing_district,
                "caste_issuing_authority": caste_issuing_authority,
                "father_name": father_name,
                "mother_name": mother_name,
                "bank_name": bank_name,
                "account_number": account_number,
                "ifsc_code": ifsc_code,
                "account_holder_name": account_holder_name,
                "signature": signature_path,
                "documentfile1": documentfile1_path,
                "documentfile2": documentfile2_path,
                "documentfile3": documentfile3_path,
                "documentfile4": documentfile4_path,
                "documentfile5": documentfile5_path,
                "documentfile6": documentfile6_path,
                "documentfile7": documentfile7_path,
                "documentfile8": documentfile8_path,
                "documentfile9": documentfile9_path,
                "documentfile10": documentfile10_path,
                "email": email
            }
            if phd_registration_year is not None and phd_registration_year >= 2023:
                update_database(update_data, phd_registration_year, cursor)
        elif record:
            print('I am in record')
            cursor.execute("INSERT INTO application_page (email) VALUES (%s)", (email,))
            cnx.commit()
            print('I have added the record email in application page')
            cursor.execute("UPDATE application_page SET status='accepted', scrutiny_status='accepted', "
                           "final_approval='accepted', joining_date=(SELECT phd_registration_date FROM old_users WHERE email=%s) WHERE email=%s ", (email,email))
            cnx.commit()
            print('I have updated status and allt threee to accepted')
            cursor.execute("SELECT email FROM award_letter where email = %s", (email,))
            results = cursor.fetchall()
            print('Result of Award Letter', results)
            if email not in results:  # Check if the email is not already in the existing emails
                cursor.execute("INSERT INTO award_letter (email) VALUES (%s) ", (email,))
                cnx.commit()
            phd_registration_year = record['phd_registration_year']
            id = record['id']
            datatime_now = datetime.now()
            applicant_id = f"TRTI/{phd_registration_year}/{id}"

            update_data = {
                "applicant_id": applicant_id,
                "form_filled": 1,
                "application_date": datetime.now(),
                "applicant_photo": photo_path,
                "adhaar_number": adhaar_number,
                "adhaar_seeding": adhaar_seeding,
                "first_name": first_name,
                "middle_name": middle_name,
                "last_name": last_name,
                "mobile_number": mobile_number,
                "date_of_birth": date_of_birth,
                "gender": gender,
                "age": age,
                "marital_status": marital_status,
                "add_1": add_1,
                "add_2": add_2,
                "pincode": pincode,
                "village": village,
                "city": city,
                "taluka": taluka,
                "district": district,
                "state": state,
                "ssc_passing_year": ssc_passing_year,
                "ssc_percentage": ssc_percentage,
                "ssc_school_name": ssc_school_name,
                "hsc_passing_year": hsc_passing_year,
                "hsc_percentage": hsc_percentage,
                "hsc_school_name": hsc_school_name,
                "graduation_passing_year": graduation_passing_year,
                "graduation_percentage": graduation_percentage,
                "graduation_school_name": graduation_school_name,
                "phd_passing_year": phd_passing_year,
                "phd_percentage": phd_percentage,
                "phd_school_name": phd_school_name,
                "phd_registration_date": phd_registration_date,
                "phd_registration_year": phd_registration_year,
                "concerned_university": concerned_university,
                "name_of_college": name_of_college,
                "other_college_name": other_college_name,
                "department_name": department_name,
                "topic_of_phd": topic_of_phd,
                "name_of_guide": name_of_guide,
                "faculty": faculty,
                "salaried": salaried,
                "disability": disability,
                "family_annual_income": family_annual_income,
                "income_certificate_number": income_certificate_number,
                "issuing_authority": issuing_authority,
                "domicile": domicile,
                "domicile_certificate": domicile_certificate,
                "domicile_number": domicile_number,
                "validity_certificate": validity_certificate,
                "validity_cert_number": validity_cert_number,
                "caste_certf": caste_certf,
                "caste_certf_number": caste_certf_number,
                "caste": caste,
                "your_caste": your_caste,
                "issuing_district": issuing_district,
                "caste_issuing_authority": caste_issuing_authority,
                "father_name": father_name,
                "mother_name": mother_name,
                "bank_name": bank_name,
                "account_number": account_number,
                "ifsc_code": ifsc_code,
                "account_holder_name": account_holder_name,
                "signature": signature_path,
                "documentfile1": documentfile1_path,
                "documentfile2": documentfile2_path,
                "documentfile3": documentfile3_path,
                "documentfile4": documentfile4_path,
                "documentfile5": documentfile5_path,
                "documentfile6": documentfile6_path,
                "documentfile7": documentfile7_path,
                "documentfile8": documentfile8_path,
                "documentfile9": documentfile9_path,
                "documentfile10": documentfile10_path,
                "email": email
            }

            if phd_registration_year in (2020, 2021, 2022):
                # update_database_old_users(update_data, phd_registration_year, cursor)
                update_database(update_data, phd_registration_year, cursor)
            cnx.commit()

            cursor.execute("SELECT form_filled='1' FROM old_users WHERE email = %s", (email,))
            records = cursor.fetchone()
            if records:
                update_old_user_in_application_page(update_data, phd_registration_year, cursor)
            print('I have updated the data and commited it successfully')

            # Close the cursor and connection
            cursor.close()
            cnx.close()

        cnx = mysql.connector.connect(user=user, password=password,
                                      host=host,
                                      database=database)
        cursor = cnx.cursor(dictionary=True)
        cursor.execute("SELECT * FROM application_page WHERE email = %s", (email,))
        records = cursor.fetchone()
        cursor.close()
        cnx.close()

    return render_template('submit_form.html', records=records, email=email)


def get_file_extension(filename):
    # Using os.path.splitext
    _, extension = os.path.splitext(filename)
    return extension.lower()


# Function to generate a styled PDF
def generate_pdf_with_styling(data, filename):
    class PDF(FPDF):
        header_added = False  # To track whether the header is added to the first page
        def header(self):
            if not self.header_added:
                # /var/www/fellowship/fellowship/BartiFellowship/BartiFellowship/
                self.set_font("Arial", "B", 12)
                self.cell(0, 10, "Fellowship ", align="C", ln=True)  # Add space by changing the second parameter (e.g., 20)
                # Insert an image (symbol) at the center of the header
                self.image('/var/www/fellowship/fellowship/BartiFellowship/BartiFellowship/static/assets/img/logo/barti.png', 10, 10, 20)  # Replace with the path to your symbol image
                # Insert an image (symbol) at the right of the header
                self.image('/var/www/fellowship/fellowship/BartiFellowship/BartiFellowship/static/Images/satya.png', 155, 10, 20)  # Replace with the path to your small image
                self.image('/var/www/fellowship/fellowship/BartiFellowship/BartiFellowship/static/Images/img.png', 175, 10, 20)  # Replace with the path to your symbol image
                self.cell(0, 10, "Tribal Research & Training Institute, Pune ", align="C", ln=True)
                self.cell(0, 1, "Government of Maharashtra ", align="C", ln=True)
                self.set_font("Arial", "B", size=8)
                self.cell(0,10,
                          "28, Queen's Garden, Bund Garden Rd, near Old Circuit House, Camp, Pune, Maharashtra 411001 ",
                          align="C", ln=True)
                self.set_font("Arial", "B", 13)
                self.cell(0, 10,
                          " Fellowship Application Form 2023 - 2024",
                          align="C", ln=True)
                self.ln(2)  # Adjust this value to control the space after the line
                self.line(10, self.get_y(), 200, self.get_y())  # Draw a line from left (10) to right (200)
                self.header_added = True  # Set to True after adding the header


        def image_and_date(self, data):
            # Date and Applicant ID
            self.set_font("Arial", "B" , size=11)
            current_date = datetime.now().strftime("%Y-%m-%d")  # You can change the date format as needed
            self.cell(50, 10, "Applicant ID: " + data['applicant_id'], ln=False)
            self.cell(108)  # Add space between cells
            # self.cell(50, 10, "Date: " + current_date, ln=True)
            self.set_font("Arial", size=10)
            full_name = data['first_name'] + ' ' + data['middle_name'] + ' ' + data['last_name']
            self.cell(50, 10, "Full Name: " + str(full_name), ln=True)

            if 'applicant_id' and 'application_date' in data:
                data['applicant_id'] = 'TRTI' + '/' + str(data['phd_registration_year']) + '/' + str(data['id'])
                self.cell(50, 10, "Submitted Date: " + str(data['application_date']), ln=True)
                self.cell(50, 10, "Submitted Time: " + str(data['application_time']), ln=True)


            if 'applicant_photo' in data:
                # photo = 'static/Images/trti.jpeg'
                photo = '/var/www/fellowship/fellowship/BartiFellowship/BartiFellowship' + data['applicant_photo']
                print(photo)
                # Insert the applicant photo (adjust the coordinates and size as needed)
                self.image(photo, 165, 65, 30, 35)  # Adjust the Y-coordinate from 60 to 65
                self.rect(165, 65, 30, 35)  # Adjust the Y-coordinate accordingly
                self.ln(10)  # Space between Date/Applicant ID and the main content

        def footer(self):
            # Add a footer
            self.set_y(-15)
            self.set_font("Arial", "B", 8)
            self.cell(0, 10, f" {self.page_no()} ", align="C")

            # Center-align the "TRTI" text
            self.cell(0, 10, " TRTI  |  Fellowship | 2023 - 2024 ", align="R")

    personal_details = {
        "Adhaar Number": data['adhaar_number'],
        "First Name":data['first_name'],
        "Middle Name": data['middle_name'],
        "Last Name": data['last_name'],
        "Mobile Number": data['mobile_number'],
        "Email": data['email'],
        "Date of Birth": data['date_of_birth'],
        "Gender" : data['gender'],
        "Age" : data['age'],
        "Category" : data['caste'],
        "Caste/Tribe ": data['your_caste'],
        "Sub Caste/Tribe": data['subcaste'],
        "Do you belong to PVTG?": data['pvtg'],
        "Which caste/tribe you belong in PVTG?": data['pvtg_caste']

        # Add more fields as needed
    }

    address_details = {
        "Main Address": data['add_1'],
        "Postal Address": data['add_2'],
        "Pincode": data['pincode'],
        "Village": data['village'],
        "Taluka": data['taluka'],
        "District": data['district'],
        "City": data['city'],
        "State": data['state']
    }

    #qualification_details = {
    #SSC
    ssc={
        "SSC Passing Year": data['ssc_passing_year'],
        "SSC School Name": data['ssc_school_name'],
        "SSC Stream": data['ssc_stream'],
        "SSC Attempts": data['ssc_attempts'],
        "SSC Total Marks": data['ssc_total'],
        "SSC Percentage": data['ssc_percentage']
    }

    hsc = {
        "HSC Passing Year": data['hsc_passing_year'],
        "HSC School Name": data['hsc_school_name'],
        "HSC Stream": data['hsc_stream'],
        "HSC Attempts": data['hsc_attempts'],
        "HSC Total Marks": data['hsc_total'],
        "HSC Percentage": data['hsc_percentage']
    }

    grad={
        "Graduation Passing Year": data['graduation_passing_year'],
        "Graduation College Name": data['graduation_school_name'],
        "Graduation Stream": data['grad_stream'],
        "Graduation Attempts": data['grad_attempts'],
        "Graduation Total Marks": data['grad_total'],
        "Graduation Percentage": data['graduation_percentage']
    }

    postgrad = {
        "Post Graduation Passing Year": data['phd_passing_year'],
        "Post Graduation College Name": data['phd_school_name'],
        "Post Graduation Stream": data['pg_stream'],
        "Post Graduation Attempts": data['pg_attempts'],
        "Post Graduation Total Marks": data['pg_total'],
        "Post Graduation Percentage": data['phd_percentage'],

        "What have you Qualified?":data['have_you_qualified']
        # Add more fields as needed
    }

    phd_details = {
        "P.H.D Registration Date": data['phd_registration_date'],
        "P.H.D Registration Year": data['phd_registration_year'],
        "Age at Ph.D. Registration": data['phd_registration_age'],
        "University Name": data['concerned_university'],
        "Name of College": data['name_of_college'],
        "Department Name": data['department_name'],
        "Topic of Ph.D.": data['topic_of_phd'],
        "Name of Guide": data['name_of_guide'],
        "Faculty/Stream": data['faculty']
        # Add more fields as needed
    }

    income_details = {
        "Family Annual Income": data ['family_annual_income'],
        "Income Certificate Number": data['income_certificate_number'],
        "Income Certificate Issuing Authority": data['issuing_authority'],
        "Income Certificate Issuing District": data['income_issuing_district'],
        "Income Certificate Issuing Taluka": data['income_issuing_taluka']
    }

    caste = {
        "Are you Domicile of Maharashtra":data['domicile'],
        "Domicile Certificate":data['domicile_certificate'],
        "Domicile Certificate Number":data['domicile_number'],
        "Do you have Caste/Tribe Certificate":data['caste_certf'],
        "Caste | Tribe":data['your_caste'],
        "Sub Caste/Tribe":data['subcaste'],
        "Caste Certificate Number":data['caste_certf_number'],
        "Caste Certificate Issuing District":data['issuing_district'],
        "Caste Certificate Issuing Authority":data['caste_issuing_authority'],
        "Validity Certificate":data['validity_certificate'],
        "Validity Certificate Number": data['validity_cert_number'],
        "Validity Certificate Issuing District": data['validity_issuing_district'],
        "Validity Certificate Issuing Taluka": data['validity_issuing_taluka'],
        "Validity Certificate Issuing Authority": data['validity_issuing_authority']
    }

    parent_details = {
        "Salaried": data['salaried'],
        "Disability": data['disability'],
        "Type of Disability": data['type_of_disability'],
        "Father Name": data['father_name'],
        "Mother Name": data['mother_name'],
        "Anyone Work in Government": data['work_in_government'],
        "Department in Government":data['gov_department'],
        "Post in Government":data['gov_position']
    }

    bank_details = {
        "IFSC Code": data['ifsc_code'],
        "Account Number": data['account_number'],
        "Bank Name": data['bank_name'],
        "Account Holder Name": data['account_holder_name'],
        "MICR Code":data['micr']
    }


    pdf = PDF(orientation='P', format='A4')
    pdf.add_page()
    pdf.header()
    pdf.image_and_date(data)

    # Personal Details
    pdf.set_font("Arial","B", size=12)
    pdf.cell(0, 10, "Personal Details", ln=True)
    pdf.ln(2)  # Adjust this value to control the space after the line
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())  # Draw a line from left (10) to right (200)
    pdf.header_added = True  # Set to True after adding the header
    pdf.set_font("Arial", size=10)
    for field, value in personal_details.items():
        pdf.cell(70, 10, str(field), border=0)
        pdf.cell(0, 10, str(value), border=0, ln=True)
    pdf.ln(15)  # Adjust this value to control the space after each section


    # Personal Details
    pdf.set_font("Arial", "B", size=12)
    pdf.cell(0, 10, "Address Details", ln=True)
    pdf.ln(2)  # Adjust this value to control the space after the line
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())  # Draw a line from left (10) to right (200)
    pdf.header_added = True  # Set to True after adding the header
    pdf.set_font("Arial", size=10)
    for field, value in address_details.items():
        pdf.cell(70, 10, str(field), border=0)
        pdf.cell(0, 10, str(value), border=0, ln=True)
    pdf.ln(10)  # Adjust this value to control the space after each section

    pdf.set_font("Arial", "B", size=12)
    pdf.cell(0, 10, "Qualification Details", ln=True)

    pdf.ln(10)  # Increase this value to shift the content further down

    # SSC Details
    pdf.set_font("Arial", "B", size=10)
    pdf.cell(0, 10, "S.S.C Details", ln=True)
    pdf.ln(2)  # Adjust this value to control the space after the line
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())  # Draw a line from left (10) to right (200)
    pdf.header_added = True  # Set to True after adding the header
    pdf.set_font("Arial", size=10)
    for field, value in ssc.items():
        pdf.cell(70, 10, str(field), border=0)
        pdf.cell(0, 10, str(value), border=0, ln=True)
    pdf.ln(5)  # Adjust this value to control the space after each section

    # HSC Details
    pdf.set_font("Arial", "B", size=10)
    pdf.cell(0, 10, "H.S.C Details", ln=True)
    pdf.ln(2)  # Adjust this value to control the space after the line
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())  # Draw a line from left (10) to right (200)
    pdf.header_added = True  # Set to True after adding the header
    pdf.set_font("Arial", size=10)
    for field, value in hsc.items():
        pdf.cell(70, 10, str(field), border=0)
        pdf.cell(0, 10, str(value), border=0, ln=True)
    pdf.ln(5)  # Adjust this value to control the space after each section

    # Graduation Details
    pdf.set_font("Arial", "B", size=10)
    pdf.cell(0, 10, "Graduation Details", ln=True)
    pdf.ln(2)  # Adjust this value to control the space after the line
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())  # Draw a line from left (10) to right (200)
    pdf.header_added = True  # Set to True after adding the header
    pdf.set_font("Arial", size=10)
    for field, value in grad.items():
        pdf.cell(70, 10, str(field), border=0)
        pdf.cell(0, 10, str(value), border=0, ln=True)
    pdf.ln(5)  # Adjust this value to control the space after each section

    # Post Graduation Details
    pdf.set_font("Arial", "B", size=10)
    pdf.cell(0, 10, "Post Graduation Details", ln=True)
    pdf.ln(2)  # Adjust this value to control the space after the line
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())  # Draw a line from left (10) to right (200)
    pdf.header_added = True  # Set to True after adding the header
    pdf.set_font("Arial", size=10)
    for field, value in postgrad.items():
        pdf.cell(70, 10, str(field), border=0)
        pdf.cell(0, 10, str(value), border=0, ln=True)
    pdf.ln(5)  # Adjust this value to control the space after each section


    # Personal Details
    pdf.set_font("Arial", "B", size=12)
    pdf.cell(0, 10, "P.H.D Details", ln=True)
    pdf.ln(2)  # Adjust this value to control the space after the line
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())  # Draw a line from left (10) to right (200)
    pdf.header_added = True  # Set to True after adding the header
    pdf.set_font("Arial", size=10)
    for field, value in phd_details.items():
        pdf.cell(70, 10, str(field), border=0)
        pdf.multi_cell(0, 10, str(value), border=0)
    pdf.ln(5)  # Adjust this value to control the space after each section


    # Personal Details
    pdf.set_font("Arial", "B", size=12)
    pdf.cell(0, 10, "Income Details", ln=True)
    pdf.ln(2)  # Adjust this value to control the space after the line
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())  # Draw a line from left (10) to right (200)
    pdf.header_added = True  # Set to True after adding the header
    pdf.set_font("Arial", size=10)
    for field, value in income_details.items():
        pdf.cell(70, 10, str(field), border=0)
        pdf.cell(0, 10, str(value), border=0, ln=True)
    pdf.ln(5)  # Adjust this value to control the space after each section


    # Personal Details
    pdf.set_font("Arial", "B", size=12)
    pdf.cell(0, 10, "Caste/Tribe Details", ln=True)
    pdf.ln(2)  # Adjust this value to control the space after the line
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())  # Draw a line from left (10) to right (200)
    pdf.header_added = True  # Set to True after adding the header
    pdf.set_font("Arial", size=10)
    for field, value in caste.items():
        pdf.cell(70, 10, str(field), border=0)
        pdf.cell(0, 10, str(value), border=0, ln=True)
    pdf.ln(10)  # Adjust this value to control the space after each section

    # Personal Details
    pdf.set_font("Arial", "B", size=12)
    pdf.cell(0, 10, "Parent Details", ln=True)
    pdf.ln(2)  # Adjust this value to control the space after the line
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())  # Draw a line from left (10) to right (200)
    pdf.header_added = True  # Set to True after adding the header
    pdf.set_font("Arial", size=10)
    for field, value in parent_details.items():
        pdf.cell(70, 10, str(field), border=0)
        pdf.cell(0, 10, str(value), border=0, ln=True)
    pdf.ln(10)  # Adjust this value to control the space after each section


    # Personal Details
    pdf.set_font("Arial", "B", size=12)
    pdf.cell(0, 10, "Bank Details", ln=True)
    pdf.ln(2)  # Adjust this value to control the space after the line
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())  # Draw a line from left (10) to right (200)
    pdf.header_added = True  # Set to True after adding the header
    pdf.set_font("Arial", size=10)
    for field, value in bank_details.items():
        pdf.cell(70, 10, str(field), border=0)
        pdf.cell(0, 10, str(value), border=0, ln=True)
    pdf.ln(10)  # Adjust this value to control the space after each section
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())  # Draw a line from left (10) to right (200)
    # Applicant's Signature
    text = (
        "I hereby declare by signing below that the above particulars are true and correct to the best of my knowledge "
        "and belief and nothing has been concealed therein")
    # Define the width for text wrapping
    width = 400  # Adjust this width according to your requirement
    # Draw the text with wrapping
    pdf.cell(width, 10, txt=text, ln=True)
    pdf.set_font("Arial", size=12)
    # Assuming data['signature'] contains the path to the image file
    signature_path = '/var/www/fellowship/fellowship/BartiFellowship/BartiFellowship' + data['signature']
    # Determine the current position
    x = pdf.get_x()
    y = pdf.get_y()
    # Set position for the image
    pdf.set_xy(x + 10, y + 5)  # Adjust position as needed
    # Add the image
    pdf.image(signature_path, x + 50, y + 10, 50)  # Adjust width (50) as needed
    # Move to a new line
    pdf.ln(15)  # Adjust as needed
    pdf.cell(0, 10, "Applicant's Signature:", ln=True)
    pdf.ln(15)  # Adjust this value to control the space after the line
    current_date = datetime.now().strftime("%Y-%m-%d")  # You can change the date format as needed
    pdf.cell(0, 10, "Date:" + ' ' + current_date, ln=True)
    pdf.cell(0, 10, "Place:" + ' ' + data['city'] + ', ' + data['state'], ln=True)

    # Save the PDF to a file
    pdf.output(filename)


@app.route('/generate_pdf', methods=['GET', 'POST'])
def generate_pdf():
    email = session['email']
    output_filename = app.config['PDF_STORAGE_PATH']
    # output_filename = 'static/pdf_application_form/pdfform.pdf'
    cnx = mysql.connector.connect(user=user, password=password, host=host, database=database)
    cursor = cnx.cursor(dictionary=True)

    cursor.execute(" SELECT * FROM signup WHERE year IN ('2020', '2021', '2022', '2023') and email = %s ", (email,))
    output = cursor.fetchall()

    if output:
        cursor.execute(
            "SELECT * FROM application_page WHERE email = %s", (email,))
        old_user_data = cursor.fetchone()
        print(old_user_data)
        # Generate a styled PDF
        print(output_filename)
        generate_pdf_with_styling(old_user_data, output_filename)
    else:
        cursor.execute("SELECT * FROM application_page WHERE email = %s",(email,))
        data = cursor.fetchone()
        print(data)
        # Generate a styled PDF
        generate_pdf_with_styling(data, output_filename)

    # Serve the generated PDF as a response
    with open(output_filename, "rb") as pdf_file:
        response = Response(pdf_file.read(), content_type="application/pdf")
        response.headers['Content-Disposition'] = 'inline; filename=pdfform.pdf'

    return response


def save_file(file, firstname, lastname):
    if file:
        filename = f"{firstname}_{lastname}_{file.filename}"
        file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
        # return os.path.join(app.config['UPLOAD_FOLDER'], filename)
        return '/static/uploads/image_retrive/' + filename
    else:
        return "Save File"


def save_bulk_email_file(file):
    if file:
        filename = f"{file.filename}"
        file.save(os.path.join(app.config['EMAIL_DOCS'], filename))
        return '/static/uploads/sendbulkemails/' + filename
    else:
        return "Save File"


def applicant_pdf_upload_section_five(file, firstname, lastname):
    if file:
        filename = f"{firstname}_{lastname}_{file.filename}"
        file.save(os.path.join(app.config['USER_DOC_SEC_FIVE'], filename))
        # return os.path.join(app.config['USER_DOC_SEC_FIVE'], filename)
        return '/static/uploads/user_doc_secfive/' + filename
    else:
        return "Save File"


def save_applicant_photo(file, firstname, lastname):
    if file:
        filename = f"{firstname}_{lastname}_{file.filename}"
        file.save(os.path.join(app.config['UPLOAD_PHOTO_SECTION1'], filename))
        # return os.path.join(app.config['UPLOAD_PHOTO_SECTION1'], filename)
        return '/static/uploads/image_retrive/' + filename
    else:
        return "Save File"


def save_issue_raised_photo(file):
    if file:
        filename = file.filename
        file.save(os.path.join(app.config['UPLOAD_PHOTO_SECTION1'], filename))
        # return os.path.join(app.config['UPLOAD_PHOTO_SECTION1'], filename)
        return '/static/uploads/image_retrive/' + filename
    else:
        return "Save File"


def save_file_rent_agreement(file, firstname, lastname):
    if file:
        filename = f"{firstname}_{lastname}_{file.filename}"
        file.save(os.path.join(app.config['RENT_AGREEMENT_REPORT'], filename))
        # return os.path.join(app.config['RENT_AGREEMENT_REPORT'], filename)
        return '/static/uploads/rent_agreement/' + filename
    else:
        return "Save File"


def save_file_pdf_cert(file, firstname, lastname):
    if file:
        filename = f"{firstname}_{lastname}_{file.filename}"
        file.save(os.path.join(app.config['PDF_CERTIFICATE'], filename))
        # return os.path.join(app.config['RENT_AGREEMENT_REPORT'], filename)
        return '/static/uploads/phd_certificate/' + filename
    else:
        return "Save File"


def save_file_joining_report(file, firstname, lastname):
    if file:
        filename = f"{firstname}_{lastname}_{file.filename}"
        file.save(os.path.join(app.config['JOINING_REPORT'], filename))
        # return os.path.join(app.config['RENT_AGREEMENT_REPORT'], filename)
        return '/static/uploads/joining_reports/' + filename
    else:
        return "Save File"


def save_file_undertaking_report(file, firstname, lastname):
    if file:
        filename = f"{firstname}_{lastname}_{file.filename}"
        file.save(os.path.join(app.config['UNDERTAKING_REPORT'], filename))
        # return os.path.join(app.config['RENT_AGREEMENT_REPORT'], filename)
        return '/static/uploads/undertaking_doc/' + filename
    else:
        return "Save File"




def save_file_assessment_report(file, firstname, lastname):
    if file:
        filename = f"{firstname}_{lastname}_{file.filename}"
        file.save(os.path.join(app.config['ASSESSMENT_REPORT'], filename))
        # return os.path.join(app.config['RENT_AGREEMENT_REPORT'], filename)
        return '/static/uploads/assessment_report/' + filename
    else:
        return "Save File"


def save_file_half_yearly(file, firstname, lastname):
    if file:
        filename = f"{firstname}_{lastname}_{file.filename}"
        file.save(os.path.join(app.config['HALF_YEARLY_REPORTS'], filename))
        # return os.path.join(app.config['HALF_YEARLY_REPORTS'], filename)
        return '/static/uploads/half_yearly/' + filename
    else:
        return "Save File"


def save_file_uplaod_thesis(file, firstname, lastname):
    if file:
        filename = f"{firstname}_{lastname}_{file.filename}"
        file.save(os.path.join(app.config['UPLOAD_THESIS'], filename))
        # return os.path.join(app.config['HALF_YEARLY_REPORTS'], filename)
        return '/static/uploads/upload_thesis/' + filename
    else:
        return "Save File"


def save_file_presenty_report(file, firstname, lastname):
    if file:
        filename = f"{firstname}_{lastname}_{file.filename}"
        file.save(os.path.join(app.config['PRESENTY_REPORTS'], filename))
        # return os.path.join(app.config['PRESENTY_REPORTS'], filename)
        return '/static/uploads/presenty_reports/' + filename
    else:
        return "Save File"


def save_news(file):
    if file:
        filename = file.filename
        file.save(os.path.join(app.config['SAVE_NEWS'], filename))
        # return os.path.join(app.config['RENT_AGREEMENT_REPORT'], filename)
        return '/static/uploads/save_news/' + filename
    else:
        return "Save File"


def allowed_file(filename):
    ALLOWED_EXTENSIONS = {'pdf', 'txt', 'jpg', 'jpeg', 'png'}
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def insert_into_old_users(email, applicant_id, phd_registration_date, date_of_birth, age):
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    # Update the specific record with the matching email
    with cursor:
        sql = "UPDATE old_users SET applicant_id = %s, phd_registration_date = %s, date_of_birth = %s, age = %s WHERE email = %s"
        cursor.execute(sql, (applicant_id, phd_registration_date, date_of_birth, age, email))
        cnx.commit()
        cursor.close()
        cnx.close()


@app.route('/logout')
def logout():
    # Check if the user is logged in before logging out
    # Check if the user is logged in before logging out
    if 'email' in session:
        # Clear the session variables related to the user
        session.pop('email', None)
        session.pop('user_name', None)
        session.pop('final_approval', None)
        session.clear()
        return redirect(url_for('login'))  # Redirect to the login page after logout
    else:
        return redirect(url_for('admin_login'))

    
# --------------------------------------------------------------------
# -------------      FOOTER FUNCTIONALITY     -----------
# --------------------------------------------------------------------
@app.route('/hyperlink_policy')
def hyperlink_policy():
    if 'language' in session:
        language = session['language']
    else:
        language = 'marathi'
    return render_template('hyperlink_policy.html', multilingual_content=multilingual_content, language=language)


@app.route('/t_and_c')
def t_and_c():
    if 'language' in session:
        language = session['language']
    else:
        language = 'marathi'
    return render_template('t_and_c.html', multilingual_content=multilingual_content, language=language)


@app.route('/privacy_policy')
def privacy_policy():
    if 'language' in session:
        language = session['language']
    else:
        language = 'marathi'
    return render_template('privacy_policy.html', multilingual_content=multilingual_content, language=language)


@app.route('/copyright_policy')
def copyright_policy():
    if 'language' in session:
        language = session['language']
    else:
        language = 'marathi'
    return render_template('copyright_policy.html', multilingual_content=multilingual_content, language=language)


@app.route('/wmp_policy')
def wmp_policy():
    if 'language' in session:
        language = session['language']
    else:
        language = 'marathi'
    return render_template('wmp_policy.html', multilingual_content=multilingual_content, language=language)


@app.route('/sitemap')
def sitemap():
    if 'language' in session:
        language = session['language']
    else:
        language = 'marathi'
    return render_template('sitemap.html', multilingual_content=multilingual_content, language=language)


@app.route('/update_preview_form', methods=['POST'])
def update_preview_form():
    if request.method == 'POST':
        data = request.form.to_dict(flat=False)
        return jsonify(request.form)


@app.route('/faq')
def faq():
    if 'language' in session:
        language = session['language']
    else:
        language = 'marathi'
    return render_template('FAQ.html', multilingual_content=multilingual_content, language=language)


@app.errorhandler(404)
def page_not_found(error):
    return render_template('pages-error-404.html'), 404


@app.errorhandler(500)
def page_not_found_500(error):
    return render_template('pages-error-500.html'), 500


# Get address details using pincode
@app.route('/get_pincode_data', methods=['GET'])
def get_pincode_data():
    pincode_data = request.args.get('pincode')
    api_url = f'https://api.worldpostallocations.com/pincode?postalcode={pincode_data}&countrycode=IN'
    try:
        response = requests.get(api_url)
        response.raise_for_status()  # Raise an exception for bad responses (4xx or 5xx)
        data = response.json()
        return jsonify(data)
    except requests.exceptions.RequestException as e:
        return jsonify({'error': str(e)}), 500
    
@app.route('/get_ifsc_data', methods=['GET'])
def get_ifsc_data():
    ifsc = request.args.get('ifsc')
    api_url = f'https://ifsc.razorpay.com/{ifsc}'
    try:
        response = requests.get(api_url)
        response.raise_for_status()  # Raise an exception for bad responses (4xx or 5xx)
        data = response.json()
        return jsonify(data)
    except requests.exceptions.RequestException as e:
        return jsonify({'error': str(e)}), 500


@app.route('/fellowship_awarded', methods=['GET', 'POST'])
def fellowship_awarded():
    if not session.get('logged_in'):
        # Redirect to the admin login page if the user is not logged in
        return redirect(url_for('admin_login'))
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    sql = """ 
    
            SELECT * 
            FROM application_page 
            WHERE final_approval = 'accepted' 
              AND phd_registration_year >= '2023'
            
            UNION
            
            SELECT * 
            FROM application_page 
            WHERE phd_registration_year > '2020' 
              AND aadesh = 1; 
    
    """
    cursor.execute(sql)
    result = cursor.fetchall()
    print(result)
    cursor.close()
    cnx.close()
    return render_template('fellowship_awarded.html', result=result)


@app.route('/generate_pdf_application/<string:email>', methods=['GET', 'POST'])
def generate_pdf_application(email):
    #email = session['email']
    output_filename = app.config['PDF_STORAGE_PATH']
    # output_filename = 'static/pdf_application_form/pdfform.pdf'
    cnx = mysql.connector.connect(user=user, password=password, host=host, database=database)
    cursor = cnx.cursor(dictionary=True)

    cursor.execute(" SELECT * FROM signup WHERE year IN ('2020', '2021', '2022') and email = %s ", (email,))
    output = cursor.fetchall()

    if output:
        cursor.execute(
            "SELECT * FROM application_page WHERE email = %s", (email,))
        old_user_data = cursor.fetchone()
        print(old_user_data)
        # Generate a styled PDF
        print(output_filename)
        generate_pdf_with_styling(old_user_data, output_filename)
    else:
        cursor.execute("SELECT * FROM application_page WHERE email = %s",(email,))
        data = cursor.fetchone()
        print(data)
        # Generate a styled PDF
        generate_pdf_with_styling(data, output_filename)

    # Serve the generated PDF as a response
    with open(output_filename, "rb") as pdf_file:
        response = Response(pdf_file.read(), content_type="application/pdf")
        response.headers['Content-Disposition'] = 'inline; filename=pdfform.pdf'

    return response


# -------------------------------- ------------------------------------------
# ----------------------- Developer Dashboard -------------------------------
# -------------------------------- ------------------------------------------
@app.route('/developer_dashboard', methods=['GET', 'POST'])
def developer_dashboard():
    return render_template('developer_dashboard.html')


@app.route('/admin_DevDash')
def admin_devdash():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    sql = """ SELECT * FROM admin """
    cursor.execute(sql)
    result = cursor.fetchall()

    cursor.close()
    cnx.close()
    return render_template('admin_devdash.html', result=result)


@app.route('/applicationPage_DevDash')
def applicationPage_DevDash():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    sql = """ SELECT * FROM application_page """
    cursor.execute(sql)
    result = cursor.fetchall()

    cursor.close()
    cnx.close()
    return render_template('applicationPage_DevDash.html', result=result)

# -------------------------------- ------------------------------------------
# ----------------------- Developer Dashboard -------------------------------
# -------------------------------- ------------------------------------------
@app.route('/sendbulkEmails', methods=['GET', 'POST'])
def sendbulkEmails():
    if not session.get('logged_in'):
        # Redirect to the admin login page if the user is not logged in
        return redirect(url_for('admin_login'))
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    record = None  # Initialize record with a default value or None
    email_list = None
    if request.method == 'POST':
        year = request.form['year']
        print(year)
        sql = """ SELECT email FROM signup WHERE year=%s"""
        cursor.execute(sql, (year,))
        record = cursor.fetchall()
        print(record)
        # Process the records as needed
        email_list = [entry['email'] for entry in record]
        print(email_list)
    cursor.close()
    cnx.close()

    return render_template('sendbulkEmails.html', record=record, email_list=email_list)


@app.route('/send_bulk_email', methods=['GET','POST'])
def send_bulk_emails():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)

    cursor = cnx.cursor(dictionary=True)
    if request.method == 'POST':
        print(request.form)
        # Get the email list, message, and subject from the form submission
        email_list = request.form.getlist('email_list[]')

        message = request.form.get('message', '')
        subject = request.form.get('subject', '')
        # attachment = request.files['attachment']
        # attachment_path = save_bulk_email_file(attachment)
        # print("Routing Attachment -", attachment_path)
        print(message, subject)
        # Send emails
        send_bulk_email(message, subject, email_list)

    cursor.close()
    cnx.close()

    return render_template('BulkEmailSent.html')


def send_bulk_email(message, subject, email_list):
    msg = Message(subject=subject, sender='noreply_fellowship@trti-maha.in', recipients=email_list)
    msg_body = message
    msg.html = msg_body
    # Attach file
    # print(attachment_path)
    # with open(attachment_path, 'rb') as f:
    #     file_data = f.read()
    # # Get the file name from the attachment_path
    # filename = os.path.basename(attachment_path)
    # # Determine the MIME type based on the file extension
    # mime_type, _ = mimetypes.guess_type(filename)
    # # If MIME type is not detected, set a default value
    # if not mime_type:
    #     mime_type = 'application/octet-stream'
    # print(attachment_path)
    # Add the attachment to the email message
    #msg.attach(filename, mime_type, file_data)
    mail.send(msg)


@app.route('/send_mail_incomplete_applications')
def send_mail_incomplete_applications():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    sql = """ SELECT email FROM application_page where phd_registration_year>='2023' and form_filled='0' """
    # sql = """ SELECT email FROM application_page where email in ('tupotbhare@gmail.com', 'Tanmay@icspune.com') """
    cursor.execute(sql)
    result = cursor.fetchall()
    # message = 'This is a Test Email Please Ignore'
    # subject = 'Please Ignore - TEST EMAIL for Fellowship'
    message = 'Dear Applicant, This is a reminder to complete your application form for "Fellowship". Please complete and submit the form by 31st May 2024 at 6PM. Thank you for your attention to this matter.'
    subject = 'Regarding Incomplete Form for Fellowship'
    email_list = [row['email'] for row in result]  # Extracting emails from the result
    email_string = ', '.join(email_list)
    send_mail_incomplete(message, subject, email_list=email_list)
    sent_datetime = datetime.now()
    sent_date = sent_datetime.strftime('%Y-%m-%d')
    sent_time = sent_datetime.strftime('%H:%M:%S')
    sent_day = sent_datetime.strftime('%A')

    sql_insert = """INSERT INTO email_record_incompleteform (emails, message, subject, date, time, day) 
                        VALUES (%s, %s, %s, %s, %s, %s)"""
    cursor.execute(sql_insert, (email_string, message, subject, sent_date, sent_time, sent_day))
    # Commit changes to the database
    cnx.commit()
    # Close cursor and database connection
    cursor.close()
    cnx.close()
    return redirect(url_for('fetch_incomplete_form_emails'))


@app.route('/fetch_incomplete_form_emails', methods=['GET', 'POST'])
def fetch_incomplete_form_emails():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    sql = """ SELECT * FROM email_record_incompleteform """
    cursor.execute(sql)
    result = cursor.fetchall()
    cnx.commit()
    # Close cursor and database connection
    cursor.close()
    cnx.close()
    return render_template('fetch_incomplete_form_emails.html', result=result)


def send_mail_incomplete(message, subject, email_list, filename=None, file_data=None):
    msg = Message(subject=subject, sender='noreply_fellowship@trti-maha.in', recipients=email_list)
    msg_body = message
    msg.html = msg_body

    # Attach file if provided
    if filename and file_data:
        # Determine the MIME type based on the file extension
        mime_type, _ = mimetypes.guess_type(filename)
        # If MIME type is not detected, set a default value
        if not mime_type:
            mime_type = 'application/octet-stream'
        print(filename)
        # Add the attachment to the email message
        msg.attach(filename, mime_type, file_data)
    mail.send(msg)


# College name by university data
@app.route('/get_college_data_by_university', methods=['GET','POST'])
def get_college_data_by_university():
    u_id = request.form.get('u_id')
    print(u_id)
    college_obj = universityController(host)
    college_name = college_obj.get_college_name(u_id)
    return jsonify(college_name)


@app.route('/app_form_info')
def app_form_info():
    return render_template('app_form_info.html')


# New Page on Home Page instead of GR
@app.route('/gr_page')
def gr_page():
    if 'language' in session:
        language = session['language']
    else:
        language = 'marathi'
    return render_template('gr_page.html', language=language, multilingual_content=multilingual_content)


@app.route('/student_manage_dashbaord')
def student_manage_dashbaord():
    return render_template('student_manage_dashbaord.html')



@app.route('/admin_issue_dashboard', methods=['GET', 'POST'])
def admin_issue_dashboard():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    sql = """ SELECT * FROM application_page """
    cursor.execute(sql)
    result = cursor.fetchall()
    for record in result:
        id = record['id']
    print(result)
    delete_student = delete_student_management(id)
    return render_template('admin_issue_dashboard.html', result=result, delete_student=delete_student)


@app.route('/issue_resolve/<int:id>', methods=['GET', 'POST'])
def issue_resolve(id):                                                   # -------------- VIEW STUDENT FORM
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    sql = """SELECT * FROM application_page WHERE id = %s"""
    cursor.execute(sql, (id,))
    # Fetch all records matching the query
    records = cursor.fetchall()
    # Close the cursor and database connection
    cursor.close()
    cnx.close()
    return render_template('issue_resolve.html', records=records)


@app.route('/issues_raised_students', methods=['GET', 'POST'])
def issues_raised_students():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    if request.method == 'POST':
        ticket = '#' + str(random.randint(1000, 9999))
        current_datetime = datetime.now()
        date = current_datetime.strftime('%Y-%m-%d')
        time = current_datetime.strftime('%H:%M:%S')

        fullname = request.form['full_name']
        email = request.form['email']
        issue_subject = request.form['issue_subject']
        description = request.form['description']
        document = request.files['document']
        if document.filename != '':
            photo_path = save_issue_raised_photo(document)
        else:
            photo_path = request.form['user_photo']

        cursor.execute(
            "INSERT INTO issue_raised (ticket, fullname, email, issue_subject, description, document, date, time) VALUES (%s, %s, %s,  %s, %s, %s, %s, %s)",
            (ticket, fullname, email, issue_subject, description, photo_path, date, time))
        cnx.commit()

        cnx = mysql.connector.connect(user=user, password=password,
                                      host=host,
                                      database=database)
        cursor = cnx.cursor(dictionary=True)
        sql = """SELECT * FROM issue_raised where email=%s"""
        cursor.execute(sql, (email,))
        # Fetch all records matching the query
        records = cursor.fetchall()
        print(records)
        # Close the cursor and database connection
        cursor.close()
        cnx.close()

        return render_template('submitted_issue_raised.html', records=records)

    return render_template('issues_raised_students.html')


@app.route('/submitted_issue_raised', methods=['GET'])
def submitted_issue_raised():
    email = session['email']
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    sql = """SELECT * FROM issue_raised where email=%s"""
    cursor.execute(sql, (email,))
    # Fetch all records matching the query
    records = cursor.fetchall()
    print(records)
    # Close the cursor and database connection
    cursor.close()
    cnx.close()
    return render_template('submitted_issue_raised.html', records=records)


@app.route('/admin_issue_raised_by_students', methods=['GET'])
def admin_issue_raised_by_students():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    sql = """SELECT * FROM issue_raised """
    cursor.execute(sql)
    # Fetch all records matching the query
    records = cursor.fetchall()
    print(records)
    # Close the cursor and database connection
    cursor.close()
    cnx.close()
    return render_template('admin_issue_raised_by_students.html', records=records)


@app.route('/delete_student_management/<int:id>')
def delete_student_management(id):
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    sql = "DELETE FROM application_page WHERE id = %s"
    cursor.execute(sql, (id,))
    # Close the cursor and database connection
    cursor.close()
    cnx.close()
    return render_template('deleted_student_success.html')


@app.route('/edit_student_admin_management/<int:id>', methods=['GET', 'POST'])
def edit_student_admin_management(id):
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    sql = """SELECT * FROM application_page WHERE id = %s"""
    cursor.execute(sql, (id,))
    # Fetch all records matching the query
    records = cursor.fetchall()
    # Close the cursor and database connection
    cursor.close()
    cnx.close()
    return render_template('edit_student_admin_management.html', records=records)


@app.route('/delete_field/<int:id>/<field_value>', methods=['POST'])
def delete_field(id, field_value):
    try:
        cnx = mysql.connector.connect(user=user, password=password,
                                      host=host,
                                      database=database)
        cursor = cnx.cursor(dictionary=True)
        # Build the SQL query to delete records where email matches the specified value
        sql = f"UPDATE application_page SET {field_value} = NULL WHERE id = %s"
        # Execute the SQL query with the specified email value
        cursor.execute(sql, (id,))
        # Commit the transaction
        cnx.commit()
        # Close the cursor
        cursor.close()
        # Redirect the user to the edit_student_admin_management route with the corresponding id
        return redirect(url_for('edit_student_admin_management', id=id))
    except Exception as e:
        print("Error deleting record:", e)
        return None


@app.route('/update_field/<int:id>/<field_name>', methods=['GET', 'POST'])
def update_field(id, field_name):
    try:
        cnx = mysql.connector.connect(user=user, password=password,
                                      host=host,
                                      database=database)
        cursor = cnx.cursor(dictionary=True)
        new_value = request.form['new_value']
        # Build the SQL query to update the specified field value
        sql = f"UPDATE application_page SET {field_name} = %s WHERE id = %s"
        # Execute the SQL query with the specified field value and id
        cursor.execute(sql, (new_value, id))
        # Commit the transaction
        cnx.commit()
        # Close the cursor
        cursor.close()
        # Redirect the user to the edit_student_admin_management route with the corresponding id
        return redirect(url_for('edit_student_admin_management', id=id))
    except Exception as e:
        print("Error updating record:", e)
        return None


@app.route('/old_user_added_by_admin', methods=['GET', 'POST'])
def old_user_added_by_admin():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    current_date = datetime.now().date()
    sql = """SELECT * FROM signup WHERE added_date = %s"""
    cursor.execute(sql, (current_date,))
    record = cursor.fetchall()
    cnx.commit()
    return render_template('old_user_added_by_admin.html', record=record)


@app.route('/old_user_insertion_by_admin', methods=['GET', 'POST'])
def old_user_insertion_by_admin():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    if request.method == 'POST':
        first_name = request.form['first_name']
        middle_name = request.form['middle_name']
        last_name = request.form['last_name']
        email = request.form['email']
        password = 'Fellowship123'
        confirm_password = 'Fellowship123'
        year = request.form['year']
        phd_registration_date = request.form['phd_registration_date']
        phd_registration_year = request.form['phd_registration_year']
        caste = 'Scheduled Tribes'
        fellowship_withdrawn = 'not_withdrawn'
        form_filled = '0'
        added_by = 'Admin'
        current_datetime = datetime.now()
        added_date = current_datetime.date()
        added_time = current_datetime.time()

        sql = "INSERT INTO signup (first_name, middle_name, last_name, email, password, confirm_password, year, " \
              "fellowship_withdrawn, added_by, added_date, added_time) " \
              "VALUES (%(first_name)s, %(middle_name)s, %(last_name)s, %(email)s, %(password)s, %(confirm_password)s," \
              " %(year)s, %(fellowship_withdrawn)s, %(added_by)s, %(added_date)s, %(added_time)s)"
        signup_data = {
            "first_name": first_name,
            "middle_name": middle_name,
            "last_name": last_name,
            "email": email,
            "password": password,
            "confirm_password": confirm_password,
            "year": year,
            "fellowship_withdrawn": fellowship_withdrawn,
            "added_by": added_by,
            "added_date": added_date,
            "added_time": added_time
        }
        cursor.execute(sql, signup_data)


        sql = "INSERT INTO old_users (form_filled, first_name, middle_name, last_name, email, phd_registration_date, " \
              "phd_registration_year, caste) " \
              "VALUES (%(form_filled)s, %(first_name)s, %(middle_name)s, %(last_name)s, %(email)s, " \
              "%(phd_registration_date)s, %(phd_registration_year)s, %(caste)s)"

        olduser_data = {
            "form_filled": form_filled,
            "first_name": first_name,
            "middle_name": middle_name,
            "last_name": last_name,
            "email": email,
            "phd_registration_date": phd_registration_date,
            "phd_registration_year": phd_registration_year,
            "caste": caste
        }
        cursor.execute(sql, olduser_data)

        cnx.commit()

        cursor.close()
        cnx.close()
        return redirect(url_for('old_user_added_by_admin'))
    return render_template('old_user_insertion_by_admin.html')


@app.route('/start_end_application_form')
def start_end_application_form():
    if 'language' in session:
        language = session['language']
    else:
        language = 'marathi'
    return render_template('startend_login_asper_advertisement.html', language=language, multilingual_content=multilingual_content)


@app.route('/assessment_report_AA', methods=['GET', 'POST'])
def assessment_report_AA():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    records = None
    joining_report = None
    joining_date = None  # Initialize with a default value
    email = session.get('email', None)
    existing_report = []
    # Fetch the joining date from the database

    email = session['email']
    cursor.execute(
        "SELECT first_name, last_name, phd_registration_date, assessment_report FROM application_page WHERE email = %s",
        (email,))
    result = cursor.fetchone()

    if result:
        user = result['first_name'] + ' ' + result['last_name']
    else:
        user = 'Admin'

    if result:
        joining_date = result['phd_registration_date']
        existing_report = result['assessment_report']
        print("existing_report", existing_report)
    if request.method == 'POST':
        print('Inside POST request')
        first_name = result['first_name']
        last_name = result['last_name']
        existing_report = result['assessment_report']
        joining_report = save_file_assessment_report(request.files['assessment_report'], first_name, last_name)
        print(joining_report)
        print(existing_report)

        update_query = "UPDATE application_page SET assessment_report=%s WHERE email = %s"
        cursor.execute(update_query, (joining_report, email))
        cnx.commit()

    cursor.close()
    cnx.close()
    return render_template('assessment_report_AA.html', records=records, joining_date=joining_date,
                           joining_report=joining_report, user=user,
                           result=result, existing_report=existing_report)


@app.route('/old_user_adhaar_seed', methods = ['GET', 'POST'])
def old_user_adhaar_seed():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    records = None
    adhaar_seeding_doc = None
    joining_date = None  # Initialize with a default value
    email = session.get('email', None)
    existing_report = []
    # Fetch the joining date from the database

    email = session['email']
    cursor.execute(
        "SELECT first_name, last_name, phd_registration_date, joining_date, adhaar_seeding, adhaar_seeding_doc, application_date FROM application_page WHERE email = %s",
        (email,))
    result = cursor.fetchone()

    if result:
        user=result['first_name'] + ' ' + result['last_name']
    else:
        user='Admin'

    if result:
        joining_date = result['application_date']
        existing_report = result['adhaar_seeding_doc']
    if request.method == 'POST':
        first_name = result['first_name']
        last_name = result['last_name']
        existing_report = result['adhaar_seeding_doc']
        adhaar_seeding_doc = save_file_joining_report(request.files['adhaar_seeding_doc'], first_name, last_name)
        print(adhaar_seeding_doc)
        print(existing_report)

        update_query = "UPDATE application_page SET adhaar_seeding_doc=%s, joining_date=%s  WHERE email = %s"
        cursor.execute(update_query, (adhaar_seeding_doc, joining_date, email))
        cnx.commit()

    cursor.close()
    cnx.close()
    return render_template('old_user_adhaar_seed.html', records=records, joining_date=joining_date,
                           adhaar_seeding_doc=adhaar_seeding_doc, user=user,
                           result=result, existing_report=existing_report)


@app.route('/news', methods=['GET', 'POST'])
def news():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)

    if request.method == 'POST':
        user = request.form['user']
        title = request.form['title']
        subtitle = request.form['subtitle']
        date = datetime.now()
        time = date.strftime('%H:%M:%S')
        doc = save_news(request.files['doc'])
        sql = """INSERT INTO news_and_updates(user, title, subtitle, date, time, doc) VALUES(%s, %s, %s, %s, %s, %s) """
        data = (user, title, subtitle, date, time, doc)

        # Create cursor
        cursor = cnx.cursor(dictionary=True)
        cursor.execute(sql, data)

        # Commit the changes
        cnx.commit()

        # Close the cursor
        cursor.close()
        return render_template('news.html')
    # Select all records after the insertion
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT * FROM news_and_updates")
    result = cursor.fetchall()
    print('result', result)
    # Close the cursor and connection
    cursor.close()
    cnx.close()

    return render_template('news.html', result=result)


# ---------------------------------------------------------------------------------------
# --------------------------------- LEVEL 1 ADMIN BUTTON ROUTES -------------------------
# ---------------------------------------------------------------------------------------
@app.route('/accepted_students_level1', methods=['GET', 'POST'])
def accepted_students_level1():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    cursor.execute(" SELECT * FROM application_page WHERE phd_registration_year>='2023' and status='accepted' ")
    result = cursor.fetchall()

    return render_template('level_1_templates/accepted_students_level1.html', result=result)


@app.route('/export_accepted_students_level1', methods=['GET', 'POST'])
def export_accepted_students_level1():
    if not session.get('logged_in'):
        # Redirect to the admin login page if the user is not logged in
        return redirect(url_for('admin_login'))
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
                   "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status," 
                    "dependents, state, district, taluka, village, city, add_1, add_2, pincode," 
                    "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total," 
                    "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total," 
                    "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream," 
                    "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
                    "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
                    "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
                    "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income," 
                    "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka," 
                    "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number," 
                    "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
                    "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
                    "type_of_disability, father_name, mother_name, work_in_government, gov_department," 
                    "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
                   " WHERE phd_registration_year>='2023' and status='accepted' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    # ws.append(['applicant_id','email','first_name','last_name','application_date'])

    ws.append(['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
               'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Your Caste', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
               'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
               'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
                'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts', 'HSC Total',
                'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream', 'Graduation Attempts',
                'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts', 'PG Total',
                'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
                'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
                'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
                'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
                'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
                'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District', 'Caste Issuing Authority',
                'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government', 'Government Department',
                'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR' ])

    # Add data to the worksheet
    for row in data:
        ws.append(row)
    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)
    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Level 1 Accepted Students.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response


@app.route('/pending_students_level1', methods=['GET', 'POST'])
def pending_students_level1():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    cursor.execute(" SELECT * FROM application_page WHERE phd_registration_year>='2023' and status='pending' ")
    result = cursor.fetchall()
    print('Pending', result)
    return render_template('level_1_templates/pending_students_level1.html', result=result)


@app.route('/export_pending_students_level1', methods=['GET', 'POST'])
def export_pending_students_level1():
    if not session.get('logged_in'):
        # Redirect to the admin login page if the user is not logged in
        return redirect(url_for('admin_login'))
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
                   "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status," 
                    "dependents, state, district, taluka, village, city, add_1, add_2, pincode," 
                    "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total," 
                    "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total," 
                    "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream," 
                    "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
                    "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
                    "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
                    "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income," 
                    "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka," 
                    "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number," 
                    "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
                    "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
                    "type_of_disability, father_name, mother_name, work_in_government, gov_department," 
                    "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
                   " WHERE phd_registration_year>='2023' and status='pending' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    # ws.append(['applicant_id','email','first_name','last_name','application_date'])

    ws.append(['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
               'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Your Caste', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
               'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
               'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
                'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts', 'HSC Total',
                'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream', 'Graduation Attempts',
                'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts', 'PG Total',
                'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
                'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
                'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
                'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
                'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
                'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District', 'Caste Issuing Authority',
                'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government', 'Government Department',
                'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR' ])

    # Add data to the worksheet
    for row in data:
        ws.append(row)
    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)
    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Level 1 Pending Students.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response


@app.route('/rejected_students_level1')
def rejected_students_level1():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    cursor.execute(" SELECT * FROM application_page WHERE phd_registration_year>='2023' and status='rejected' ")
    result = cursor.fetchall()

    return render_template('level_1_templates/rejected_students_level1.html', result=result)


@app.route('/export_rejected_students_level1', methods=['GET', 'POST'])
def export_rejected_students_level1():
    if not session.get('logged_in'):
        # Redirect to the admin login page if the user is not logged in
        return redirect(url_for('admin_login'))
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
                   "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status," 
                    "dependents, state, district, taluka, village, city, add_1, add_2, pincode," 
                    "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total," 
                    "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total," 
                    "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream," 
                    "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
                    "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
                    "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
                    "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income," 
                    "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka," 
                    "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number," 
                    "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
                    "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
                    "type_of_disability, father_name, mother_name, work_in_government, gov_department," 
                    "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
                   " WHERE phd_registration_year>='2023' and status='rejected' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    # ws.append(['applicant_id','email','first_name','last_name','application_date'])

    ws.append(['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
               'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Your Caste', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
               'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
               'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
                'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts', 'HSC Total',
                'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream', 'Graduation Attempts',
                'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts', 'PG Total',
                'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
                'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
                'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
                'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
                'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
                'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District', 'Caste Issuing Authority',
                'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government', 'Government Department',
                'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR' ])

    # Add data to the worksheet
    for row in data:
        ws.append(row)
    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)
    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Level 1 Rejected Students.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response


@app.route('/pvtg_students_level1')
def pvtg_students_level1():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    cursor.execute(
        " SELECT * FROM application_page WHERE phd_registration_year>='2023' and"
        " your_caste IN ('katkari', 'kolam', 'madia') "
    )
    result = cursor.fetchall()
    return render_template('level_1_templates/pvtg_students_level1.html', result=result)


@app.route('/export_pvtg_students_level1', methods=['GET', 'POST'])
def export_pvtg_students_level1():
    if not session.get('logged_in'):
        # Redirect to the admin login page if the user is not logged in
        return redirect(url_for('admin_login'))
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
                   "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status," 
                    "dependents, state, district, taluka, village, city, add_1, add_2, pincode," 
                    "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total," 
                    "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total," 
                    "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream," 
                    "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
                    "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
                    "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
                    "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income," 
                    "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka," 
                    "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number," 
                    "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
                    "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
                    "type_of_disability, father_name, mother_name, work_in_government, gov_department," 
                    "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
                   " WHERE phd_registration_year>='2023' and your_caste IN ('katkari', 'kolam', 'madia') ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    # ws.append(['applicant_id','email','first_name','last_name','application_date'])

    ws.append(['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
               'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Your Caste', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
               'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
               'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
                'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts', 'HSC Total',
                'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream', 'Graduation Attempts',
                'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts', 'PG Total',
                'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
                'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
                'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
                'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
                'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
                'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District', 'Caste Issuing Authority',
                'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government', 'Government Department',
                'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR' ])

    # Add data to the worksheet
    for row in data:
        ws.append(row)
    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)
    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Level 1 PVTG Students.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response


@app.route('/disabled_students_level1')
def disabled_students_level1():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    cursor.execute(
        " SELECT * FROM application_page WHERE phd_registration_year>='2023' and disability='Yes' "
    )
    result = cursor.fetchall()
    return render_template('level_1_templates/disabled_students_level1.html', result=result)


@app.route('/export_disabled_students_level1', methods=['GET', 'POST'])
def export_disabled_students_level1():
    if not session.get('logged_in'):
        # Redirect to the admin login page if the user is not logged in
        return redirect(url_for('admin_login'))
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
                   "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status," 
                    "dependents, state, district, taluka, village, city, add_1, add_2, pincode," 
                    "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total," 
                    "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total," 
                    "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream," 
                    "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
                    "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
                    "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
                    "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income," 
                    "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka," 
                    "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number," 
                    "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
                    "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
                    "type_of_disability, father_name, mother_name, work_in_government, gov_department," 
                    "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
                   " WHERE phd_registration_year>='2023' and disability='Yes' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    # ws.append(['applicant_id','email','first_name','last_name','application_date'])

    ws.append(['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
               'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Your Caste', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
               'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
               'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
                'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts', 'HSC Total',
                'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream', 'Graduation Attempts',
                'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts', 'PG Total',
                'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
                'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
                'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
                'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
                'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
                'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District', 'Caste Issuing Authority',
                'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government', 'Government Department',
                'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR' ])

    # Add data to the worksheet
    for row in data:
        ws.append(row)
    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)
    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Level 1 Disabled Students.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response
# ---------------------------------------------------------------------------------------
# --------------------------------- LEVEL 2 ADMIN BUTTON ROUTES -------------------------
# ---------------------------------------------------------------------------------------
@app.route('/accepted_students_level2', methods=['GET', 'POST'])
def accepted_students_level2():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    cursor.execute(" SELECT * FROM application_page WHERE phd_registration_year>='2023' and "
                   "scrutiny_status='accepted' and status='accepted' ")
    result = cursor.fetchall()

    return render_template('level_2_templates/accepted_students_level2.html', result=result)


@app.route('/export_accepted_students_level2', methods=['GET', 'POST'])
def export_accepted_students_level2():
    if not session.get('logged_in'):
        # Redirect to the admin login page if the user is not logged in
        return redirect(url_for('admin_login'))
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
                   "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status," 
                    "dependents, state, district, taluka, village, city, add_1, add_2, pincode," 
                    "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total," 
                    "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total," 
                    "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream," 
                    "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
                    "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
                    "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
                    "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income," 
                    "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka," 
                    "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number," 
                    "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
                    "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
                    "type_of_disability, father_name, mother_name, work_in_government, gov_department," 
                    "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
                   " WHERE phd_registration_year>='2023' and scrutiny_status='accepted' and status='accepted' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    # ws.append(['applicant_id','email','first_name','last_name','application_date'])

    ws.append(['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
               'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Your Caste', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
               'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
               'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
                'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts', 'HSC Total',
                'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream', 'Graduation Attempts',
                'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts', 'PG Total',
                'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
                'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
                'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
                'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
                'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
                'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District', 'Caste Issuing Authority',
                'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government', 'Government Department',
                'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR' ])

    # Add data to the worksheet
    for row in data:
        ws.append(row)
    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)
    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Level 2 Accepted Students.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response


@app.route('/pending_students_level2', methods=['GET', 'POST'])
def pending_students_level2():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    cursor.execute(" SELECT * FROM application_page WHERE phd_registration_year>='2023' and scrutiny_status='pending' "
                   "and status='accepted' ")
    result = cursor.fetchall()
    print('Pending', result)
    return render_template('level_2_templates/pending_students_level2.html', result=result)


@app.route('/export_pending_students_level2', methods=['GET', 'POST'])
def export_pending_students_level2():
    if not session.get('logged_in'):
        # Redirect to the admin login page if the user is not logged in
        return redirect(url_for('admin_login'))
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
                   "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status," 
                    "dependents, state, district, taluka, village, city, add_1, add_2, pincode," 
                    "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total," 
                    "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total," 
                    "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream," 
                    "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
                    "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
                    "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
                    "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income," 
                    "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka," 
                    "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number," 
                    "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
                    "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
                    "type_of_disability, father_name, mother_name, work_in_government, gov_department," 
                    "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
                   " WHERE phd_registration_year>='2023' and scrutiny_status='pending' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    # ws.append(['applicant_id','email','first_name','last_name','application_date'])

    ws.append(['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
               'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Your Caste', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
               'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
               'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
                'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts', 'HSC Total',
                'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream', 'Graduation Attempts',
                'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts', 'PG Total',
                'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
                'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
                'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
                'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
                'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
                'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District', 'Caste Issuing Authority',
                'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government', 'Government Department',
                'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR' ])

    # Add data to the worksheet
    for row in data:
        ws.append(row)
    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)
    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Level 2 Pending Students.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response


@app.route('/rejected_students_level2')
def rejected_students_level2():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    cursor.execute(" SELECT * FROM application_page WHERE phd_registration_year>='2023' "
                   "and scrutiny_status='rejected' and status='accepted' ")
    result = cursor.fetchall()

    return render_template('level_2_templates/rejected_students_level2.html', result=result)


@app.route('/export_rejected_students_level2', methods=['GET', 'POST'])
def export_rejected_students_level2():
    if not session.get('logged_in'):
        # Redirect to the admin login page if the user is not logged in
        return redirect(url_for('admin_login'))
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
                   "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status," 
                    "dependents, state, district, taluka, village, city, add_1, add_2, pincode," 
                    "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total," 
                    "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total," 
                    "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream," 
                    "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
                    "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
                    "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
                    "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income," 
                    "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka," 
                    "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number," 
                    "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
                    "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
                    "type_of_disability, father_name, mother_name, work_in_government, gov_department," 
                    "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
                   " WHERE phd_registration_year>='2023' and scrutiny_status='rejected' and status='accepted' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    # ws.append(['applicant_id','email','first_name','last_name','application_date'])

    ws.append(['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
               'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Your Caste', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
               'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
               'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
                'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts', 'HSC Total',
                'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream', 'Graduation Attempts',
                'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts', 'PG Total',
                'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
                'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
                'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
                'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
                'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
                'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District', 'Caste Issuing Authority',
                'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government', 'Government Department',
                'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR' ])

    # Add data to the worksheet
    for row in data:
        ws.append(row)
    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)
    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Level 2 Rejected Students.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response


@app.route('/pvtg_students_level2')
def pvtg_students_level2():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    cursor.execute(
        " SELECT * FROM application_page WHERE phd_registration_year>='2023' and"
        " your_caste IN ('katkari', 'kolam', 'madia') "
    )
    result = cursor.fetchall()
    return render_template('level_2_templates/pvtg_students_level2.html', result=result)


@app.route('/export_pvtg_students_level2', methods=['GET', 'POST'])
def export_pvtg_students_level2():
    if not session.get('logged_in'):
        # Redirect to the admin login page if the user is not logged in
        return redirect(url_for('admin_login'))
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
                   "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status," 
                    "dependents, state, district, taluka, village, city, add_1, add_2, pincode," 
                    "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total," 
                    "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total," 
                    "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream," 
                    "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
                    "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
                    "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
                    "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income," 
                    "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka," 
                    "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number," 
                    "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
                    "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
                    "type_of_disability, father_name, mother_name, work_in_government, gov_department," 
                    "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
                   " WHERE phd_registration_year>='2023' and your_caste IN ('katkari', 'kolam', 'madia') ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    # ws.append(['applicant_id','email','first_name','last_name','application_date'])

    ws.append(['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
               'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Your Caste', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
               'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
               'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
                'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts', 'HSC Total',
                'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream', 'Graduation Attempts',
                'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts', 'PG Total',
                'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
                'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
                'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
                'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
                'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
                'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District', 'Caste Issuing Authority',
                'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government', 'Government Department',
                'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR' ])

    # Add data to the worksheet
    for row in data:
        ws.append(row)
    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)
    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Level 2 PVTG Students.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response


@app.route('/disabled_students_level2')
def disabled_students_level2():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    cursor.execute(
        " SELECT * FROM application_page WHERE phd_registration_year>='2023' and disability='Yes' "
    )
    result = cursor.fetchall()
    return render_template('level_2_templates/disabled_students_level2.html', result=result)


@app.route('/export_disabled_students_level2', methods=['GET', 'POST'])
def export_disabled_students_level2():
    if not session.get('logged_in'):
        # Redirect to the admin login page if the user is not logged in
        return redirect(url_for('admin_login'))
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
                   "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status," 
                    "dependents, state, district, taluka, village, city, add_1, add_2, pincode," 
                    "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total," 
                    "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total," 
                    "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream," 
                    "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
                    "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
                    "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
                    "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income," 
                    "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka," 
                    "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number," 
                    "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
                    "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
                    "type_of_disability, father_name, mother_name, work_in_government, gov_department," 
                    "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
                   " WHERE phd_registration_year>='2023' and disability='Yes' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    # ws.append(['applicant_id','email','first_name','last_name','application_date'])

    ws.append(['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
               'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Your Caste', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
               'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
               'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
                'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts', 'HSC Total',
                'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream', 'Graduation Attempts',
                'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts', 'PG Total',
                'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
                'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
                'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
                'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
                'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
                'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District', 'Caste Issuing Authority',
                'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government', 'Government Department',
                'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR' ])

    # Add data to the worksheet
    for row in data:
        ws.append(row)
    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)
    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Level 2 Disabled Students.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response
# ---------------------------------------------------------------------------------------
# --------------------------------- LEVEL 3 ADMIN BUTTON ROUTES -------------------------
# ---------------------------------------------------------------------------------------
@app.route('/accepted_students_level3', methods=['GET', 'POST'])
def accepted_students_level3():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    cursor.execute(" SELECT * FROM application_page WHERE fellowship_awarded_year='2023' and "
                   "final_approval='accepted' ")
    result = cursor.fetchall()

    return render_template('level_3_templates/accepted_students_level3.html', result=result)


@app.route('/export_accepted_students_level3', methods=['GET', 'POST'])
def export_accepted_students_level3():
    if not session.get('logged_in'):
        # Redirect to the admin login page if the user is not logged in
        return redirect(url_for('admin_login'))
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
                   "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status," 
                    "dependents, state, district, taluka, village, city, add_1, add_2, pincode," 
                    "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total," 
                    "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total," 
                    "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream," 
                    "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
                    "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
                    "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
                    "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income," 
                    "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka," 
                    "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number," 
                    "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
                    "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
                    "type_of_disability, father_name, mother_name, work_in_government, gov_department," 
                    "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
                   " WHERE phd_registration_year>='2023' and final_approval='accepted' and scrutiny_status='accepted' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    # ws.append(['applicant_id','email','first_name','last_name','application_date'])

    ws.append(['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
               'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Your Caste', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
               'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
               'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
                'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts', 'HSC Total',
                'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream', 'Graduation Attempts',
                'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts', 'PG Total',
                'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
                'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
                'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
                'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
                'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
                'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District', 'Caste Issuing Authority',
                'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government', 'Government Department',
                'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR' ])

    # Add data to the worksheet
    for row in data:
        ws.append(row)
    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)
    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Level 3 Accepted Students.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response


@app.route('/pending_students_level3', methods=['GET', 'POST'])
def pending_students_level3():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    cursor.execute(" SELECT * FROM application_page WHERE phd_registration_year>='2023' "
                   "and final_approval='pending' and scrutiny_status='accepted' ")
    result = cursor.fetchall()
    print('Pending', result)
    return render_template('level_3_templates/pending_students_level3.html', result=result)


@app.route('/export_pending_students_level3', methods=['GET', 'POST'])
def export_pending_students_level3():
    if not session.get('logged_in'):
        # Redirect to the admin login page if the user is not logged in
        return redirect(url_for('admin_login'))
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
                   "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status," 
                    "dependents, state, district, taluka, village, city, add_1, add_2, pincode," 
                    "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total," 
                    "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total," 
                    "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream," 
                    "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
                    "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
                    "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
                    "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income," 
                    "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka," 
                    "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number," 
                    "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
                    "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
                    "type_of_disability, father_name, mother_name, work_in_government, gov_department," 
                    "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
                   " WHERE phd_registration_year>='2023' and final_approval='pending' and scrutiny_status='accepted' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    # ws.append(['applicant_id','email','first_name','last_name','application_date'])

    ws.append(['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
               'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Your Caste', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
               'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
               'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
                'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts', 'HSC Total',
                'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream', 'Graduation Attempts',
                'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts', 'PG Total',
                'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
                'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
                'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
                'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
                'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
                'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District', 'Caste Issuing Authority',
                'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government', 'Government Department',
                'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR' ])

    # Add data to the worksheet
    for row in data:
        ws.append(row)
    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)
    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Level 3 Pending Students.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response


@app.route('/rejected_students_level3')
def rejected_students_level3():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    cursor.execute(" SELECT * FROM application_page WHERE phd_registration_year>='2023' "
                   "and final_approval='rejected' and scrutiny_status='accepted' ")
    result = cursor.fetchall()

    return render_template('level_3_templates/rejected_students_level3.html', result=result)


@app.route('/export_rejected_students_level3', methods=['GET', 'POST'])
def export_rejected_students_level3():
    if not session.get('logged_in'):
        # Redirect to the admin login page if the user is not logged in
        return redirect(url_for('admin_login'))
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
                   "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status," 
                    "dependents, state, district, taluka, village, city, add_1, add_2, pincode," 
                    "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total," 
                    "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total," 
                    "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream," 
                    "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
                    "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
                    "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
                    "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income," 
                    "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka," 
                    "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number," 
                    "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
                    "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
                    "type_of_disability, father_name, mother_name, work_in_government, gov_department," 
                    "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
                   " WHERE phd_registration_year>='2023' and final_approval='rejected' and scrutiny_status='accepted' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    # ws.append(['applicant_id','email','first_name','last_name','application_date'])

    ws.append(['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
               'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Your Caste', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
               'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
               'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
                'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts', 'HSC Total',
                'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream', 'Graduation Attempts',
                'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts', 'PG Total',
                'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
                'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
                'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
                'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
                'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
                'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District', 'Caste Issuing Authority',
                'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government', 'Government Department',
                'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR' ])

    # Add data to the worksheet
    for row in data:
        ws.append(row)
    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)
    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Level 3 Rejected Students.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response


@app.route('/pvtg_students_level3')
def pvtg_students_level3():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    cursor.execute(
        " SELECT * FROM application_page WHERE phd_registration_year>='2023' and"
        " your_caste IN ('katkari', 'kolam', 'madia') "
    )
    result = cursor.fetchall()
    return render_template('level_3_templates/pvtg_students_level3.html', result=result)


@app.route('/export_pvtg_students_level3', methods=['GET', 'POST'])
def export_pvtg_students_level3():
    if not session.get('logged_in'):
        # Redirect to the admin login page if the user is not logged in
        return redirect(url_for('admin_login'))
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
                   "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status," 
                    "dependents, state, district, taluka, village, city, add_1, add_2, pincode," 
                    "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total," 
                    "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total," 
                    "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream," 
                    "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
                    "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
                    "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
                    "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income," 
                    "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka," 
                    "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number," 
                    "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
                    "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
                    "type_of_disability, father_name, mother_name, work_in_government, gov_department," 
                    "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
                   " WHERE phd_registration_year>='2023' and your_caste IN ('katkari', 'kolam', 'madia') ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    # ws.append(['applicant_id','email','first_name','last_name','application_date'])

    ws.append(['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
               'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Your Caste', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
               'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
               'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
                'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts', 'HSC Total',
                'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream', 'Graduation Attempts',
                'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts', 'PG Total',
                'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
                'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
                'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
                'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
                'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
                'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District', 'Caste Issuing Authority',
                'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government', 'Government Department',
                'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR' ])

    # Add data to the worksheet
    for row in data:
        ws.append(row)
    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)
    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Level 3 PVTG Students.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response


@app.route('/disabled_students_level3')
def disabled_students_level3():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    cursor.execute(
        " SELECT * FROM application_page WHERE phd_registration_year>='2023' and disability='Yes' "
    )
    result = cursor.fetchall()
    return render_template('level_3_templates/disabled_students_level3.html', result=result)


@app.route('/export_disabled_students_level3', methods=['GET', 'POST'])
def export_disabled_students_level3():
    if not session.get('logged_in'):
        # Redirect to the admin login page if the user is not logged in
        return redirect(url_for('admin_login'))
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute("  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
                   "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status," 
                    "dependents, state, district, taluka, village, city, add_1, add_2, pincode," 
                    "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total," 
                    "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total," 
                    "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream," 
                    "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
                    "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
                    "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
                    "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income," 
                    "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka," 
                    "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number," 
                    "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
                    "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
                    "type_of_disability, father_name, mother_name, work_in_government, gov_department," 
                    "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
                   " WHERE phd_registration_year>='2023' and disability='Yes' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    # ws.append(['applicant_id','email','first_name','last_name','application_date'])

    ws.append(['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
               'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Your Caste', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
               'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
               'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
                'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts', 'HSC Total',
                'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream', 'Graduation Attempts',
                'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts', 'PG Total',
                'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
                'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
                'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
                'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
                'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
                'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District', 'Caste Issuing Authority',
                'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government', 'Government Department',
                'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR' ])

    # Add data to the worksheet
    for row in data:
        ws.append(row)
    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)
    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Level 3 Disabled Students.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response


# --------------------------------------------------------------------------------------------------------
# --------------------------------- Admin Dashboard Generate reports Excel -------------------------------
@app.route('/generate_reports_admin', methods=['GET', 'POST'])
def generate_reports_admin():
    if request.method == 'POST':
        start_date = request.form.get('start_date')
        end_date = request.form.get('end_date')
        year = request.form.get('year')
        filter_option = request.form.get('filter')
        stream = request.form.get('stream')
        query = build_query(start_date, end_date, year, filter_option, stream)
        data = fetch_data(query)

        if 'fetch_result' in request.form:
            # Code to fetch and display results
            return render_template('generate_reports_admin.html', data=data)

        elif 'export_to_excel' in request.form:
            # Code to export data to Excel
            wb = Workbook()
            ws = wb.active

            headers = ['Applicant Id', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
                       'Date Of Birth', 'Gender', 'Age', 'Caste', 'Your Caste', 'Marital Status', 'Dependents', 'Add 1',
                       'Add 2',
                       'Pincode', 'Village', 'Taluka', 'District', 'State', 'Phd Registration Date',
                       'Concerned University', 'Topic Of Phd',
                       'Name Of Guide', 'Name Of College', 'Stream', 'Board University', 'Admission Year',
                       'Passing Year',
                       'Percentage', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
                       'Domicile',
                       'Domicile Certificate', 'Domicile Number', 'Caste Certf', 'Issuing District',
                       'Caste Issuing Authority',
                       'Salaried', 'Disability', 'Father Name', 'Mother Name', 'Work In Government', 'Bank Name',
                       'Account Number',
                       'Ifsc Code', 'Account Holder Name']

            ws.append(headers)

            description = "List of Students"
            ws.insert_rows(1)
            header_row = ws[1]
            header_row[0].value = description
            for cell in header_row:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            for row in data:
                ws.append([row.get(header.lower().replace(' ', '_'), '') for header in headers])

            output = BytesIO()
            wb.save(output)
            output.seek(0)

            filename = f"{filter_option} {year} Students.xlsx" if year else f"{filter_option} Students.xlsx"

            response = make_response(output.getvalue())
            response.headers['Content-Disposition'] = f'attachment; filename={filename}'
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

            return response

    return render_template('generate_reports_admin.html')


def build_query(start_date, end_date, year, filter_option, stream):
    query = "SELECT * FROM application_page"
    conditions = []

    if start_date and end_date:
        conditions.append(f"phd_registration_date BETWEEN '{start_date}' AND '{end_date}'")
    if year:
        conditions.append(f"phd_registration_year = '{year}'")
    if filter_option:
        if filter_option == 'total':
            # No specific filter for total
            pass
        elif filter_option == 'pending':
            conditions.append("status = 'pending'")
        elif filter_option == 'accepted':
            conditions.append("status = 'accepted'")
        elif filter_option == 'rejected':
            conditions.append("status = 'rejected'")
        elif filter_option == 'complete':
            conditions.append("form_filled = '1'")
        elif filter_option == 'incomplete':
            conditions.append("form_filled = '0'")
        elif filter_option == 'male':
            conditions.append("gender = 'male'")
        elif filter_option == 'female':
            conditions.append("gender = 'female'")
        elif filter_option == 'transgender':
            conditions.append("gender = 'transgender'")
        elif filter_option == 'Katari':
            conditions.append("your_caste = 'Katari'")
        elif filter_option == 'Kolam':
            conditions.append("your_caste = 'Kolam'")
        elif filter_option == 'Madia':
            conditions.append("your_caste = 'Madia'")
        elif filter_option == 'disabled':
            conditions.append("disability = 'Yes'")
        elif filter_option == 'notdisabled':
            conditions.append("disability = 'No'")
    if stream:
        if stream == 'science':
            conditions.append("faculty = 'science'")
        if stream == 'arts':
            conditions.append("faculty = 'arts'")
        if stream == 'commerce':
            conditions.append("faculty = 'commerce'")
        if stream == 'other':
            conditions.append("faculty = 'other'")

    if conditions:
        query += " WHERE " + " AND ".join(conditions)

    print('query', query)
    return query


def fetch_data(query):
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    cursor.execute(query)
    results = cursor.fetchall()
    cursor.close()
    cnx.close()
    for result in results:
        for key, value in result.items():
            if isinstance(value, (date, datetime)):
                result[key] = value.isoformat()
    return results


# ------------------  Function Start ----------------------------
# Undertaking Document Uplaod on /mainpage
@app.route('/undertakingDoc', methods=['GET', 'POST'])
def undertakingDoc():
    """
        This function helps in Uploading the Undertaking Document
        from the user and if the document is already uploaded the
        user can view the document.
    """
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    records = None
    undertaking_doc = None
    undertaking_doc_date = None  # Initialize with a default value
    email = session.get('email', None)
    existing_report = []

    # Fetch the joining date from the database
    email = session['email']
    cursor.execute(
        "SELECT first_name, last_name, phd_registration_date, undertaking_doc_date, undertaking_doc FROM application_page WHERE email = %s",
        (email,))
    result = cursor.fetchone()
    print(result)
    if result:
        undertaking_doc_date = result['undertaking_doc_date']
        existing_report = result['undertaking_doc']
    if request.method == 'POST':
        first_name = result['first_name']
        last_name = result['last_name']
        existing_report = result['undertaking_doc']
        undertaking_doc = save_file_undertaking_report(request.files['undertaking_doc'], first_name, last_name)
        undertaking_doc_date = date.today()
        print(undertaking_doc)
        print(existing_report)

        update_query = "UPDATE application_page SET undertaking_doc=%s, undertaking_doc_date=%s WHERE email = %s"
        cursor.execute(update_query, (undertaking_doc, undertaking_doc_date, email))
        cnx.commit()

    cursor.close()
    cnx.close()
    if result:
        user = result['first_name'] + ' ' + result['last_name']
    else:
        user = 'Admin'
    return render_template('undertakingDoc.html', records=records, undertaking_doc_date=undertaking_doc_date,
                           undertaking_doc=undertaking_doc, result=result, existing_report=existing_report,
                           user=user)
# --------------------  Function End ----------------------------


@app.route('/form_for_Old_accepted', methods=['GET', 'POST'])
def old_but_accepted_for_2023():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    email = session['email']
    cursor.execute(
        "SELECT first_name, last_name FROM application_page WHERE email = %s",
        (email,))
    result = cursor.fetchone()
    print(result)
    if result:
        user = result['first_name'] + ' ' + result['last_name']
    else:
        user = 'Admin'
    return render_template('old_but_accepted_for_2023.html', user=user)


@app.route('/submit_form_for_Old_accepted', methods=['GET', 'POST'])
def submit_form_for_Old_accepted():
    email = session['email']

    cursor.execute(
        "SELECT first_name, last_name FROM application_page WHERE email = %s",
        (email,))
    result = cursor.fetchone()
    print(result)
    first_name = result[0]
    last_name = result[1]

    if request.method == 'POST':
        # Retrieve form data
        formfilled_again = '1'
        pvtg = request.form['pvtg']
        pvtg_caste = request.form.get('pvtg_caste', '')
        ssc_stream = request.form.get('ssc_stream', '')
        ssc_attempts = request.form.get('ssc_attempts', '')
        ssc_total = request.form.get('ssc_total', '')
        hsc_stream = request.form['hsc_stream']
        hsc_attempts = request.form['hsc_attempts']
        hsc_total = request.form['hsc_total']
        grad_stream = request.form['grad_stream']
        grad_attempts = request.form['grad_attempts']
        grad_total = request.form['grad_total']
        pg_stream = request.form['pg_stream']
        pg_attempts = request.form['pg_attempts']
        pg_total = request.form['pg_total']
        have_you_qualified = request.form['have_you_qualified']
        income_issuing_district = request.form['income_issuing_district']
        income_issuing_taluka = request.form['income_issuing_taluka']
        signature = applicant_pdf_upload_section_five(request.files['signature'], first_name, last_name)


        # Upload documents
        documents = {}
        for i in range(1, 27):
            document_key = f'documentfile{i}'
            if document_key in request.files:
                documents[document_key] = applicant_pdf_upload_section_five(request.files[document_key], first_name,
                                                                            last_name)
            else:
                documents[document_key] = ''

        # Insert or update your data in the database
        cursor.execute("""
            UPDATE application_page SET 
                formfilled_again = %s, pvtg = %s, pvtg_caste = %s, ssc_stream = %s, ssc_attempts = %s, ssc_total = %s,
                hsc_stream = %s, hsc_attempts = %s, hsc_total = %s, grad_stream = %s, grad_attempts = %s, grad_total = %s,
                pg_stream = %s, pg_attempts = %s, pg_total = %s, have_you_qualified = %s, income_issuing_district = %s, 
                income_issuing_taluka = %s, documentfile1 = %s, documentfile2 = %s, documentfile3 = %s, documentfile4 = %s,
                documentfile5 = %s, documentfile6 = %s, documentfile7 = %s, documentfile8 = %s, signature = %s, documentfile9 = %s,
                documentfile10 = %s, documentfile11 = %s, documentfile12 = %s, documentfile13 = %s, documentfile14 = %s,
                documentfile15 = %s, documentfile16 = %s, documentfile17 = %s, documentfile18 = %s, documentfile19 = %s,
                documentfile20 = %s, documentfile21 = %s, documentfile22 = %s, documentfile23 = %s, documentfile24 = %s, 
                documentfile26 = %s 
            WHERE email = %s
        """, (
            formfilled_again, pvtg, pvtg_caste, ssc_stream, ssc_attempts, ssc_total, hsc_stream, hsc_attempts, hsc_total,
            grad_stream, grad_attempts, grad_total, pg_stream, pg_attempts, pg_total, have_you_qualified,
            income_issuing_district, income_issuing_taluka, documents['documentfile1'], documents['documentfile2'],
            documents['documentfile3'], documents['documentfile4'], documents['documentfile5'],
            documents['documentfile6'],
            documents['documentfile7'], documents['documentfile8'], signature, documents['documentfile9'],
            documents['documentfile10'], documents['documentfile11'], documents['documentfile12'],
            documents['documentfile13'],
            documents['documentfile14'], documents['documentfile15'], documents['documentfile16'],
            documents['documentfile17'],
            documents['documentfile18'], documents['documentfile19'], documents['documentfile20'],
            documents['documentfile21'],
            documents['documentfile22'], documents['documentfile23'], documents['documentfile24'],
            documents['documentfile26'],
            email
        ))

        cnx.commit()
        flash('Details Saved Successfully', 'success')
        return redirect(url_for('main_page'))
    else:
        # Display the form or handle the GET request
        return render_template('old_but_accepted_for_2023.html')


@app.route('/aadesh')
def aadesh():
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)
    query = """
        SELECT * FROM application_page 
        WHERE email IN ('vasavepremsing71@gmail.com', 'yohasavagavit92@gmail.com', 'dvasave775@gmail.com',
                        'paradke.vinod93@gmail.com', 'karanvalvi12@gmail.com', 'vasaved440@gmail.com',
                        'sakharamkhupase@gmail.com', 'manishapardhi33@gmail.com', 'gajanang996@gmail.com', 
                        'sureshbhikjavalvi3@gmail.com', 'pramodkokani999@gmail.com', 'vidyashrihari4.vg@gmail.com',
                        'bdpatil642@gmail.com', 'maneshkokani1@gmail.com', 'ruikey170@gmail.com')
    """
    cursor.execute(query)
    data = cursor.fetchall()  # Use fetchall() to get all matching records
    return render_template('aadesh.html', data=data)  # Pass data to the template

@app.route('/export_aadesh_students')
def export_aadesh_students():
    if not session.get('logged_in'):
        # Redirect to the admin login page if the user is not logged in
        return redirect(url_for('admin_login'))
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor()
    cursor.execute(
        "  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
        "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status,"
        "dependents, state, district, taluka, village, city, add_1, add_2, pincode,"
        "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total,"
        "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total,"
        "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream,"
        "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
        "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
        "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
        "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income,"
        "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka,"
        "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number,"
        "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
        "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
        "type_of_disability, father_name, mother_name, work_in_government, gov_department,"
        "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr "  # Removed trailing comma after 'micr'
        "FROM application_page "  # Added a space before 'FROM'
        "WHERE email IN ('vasavepremsing71@gmail.com', 'yohasavagavit92@gmail.com', 'dvasave775@gmail.com', "
        "'paradke.vinod93@gmail.com', 'karanvalvi12@gmail.com', 'vasaved440@gmail.com', "
        "'sakharamkhupase@gmail.com', 'manishapardhi33@gmail.com', 'gajanang996@gmail.com', "
        "'sureshbhikjavalvi3@gmail.com', 'pramodkokani999@gmail.com', 'vidyashrihari4.vg@gmail.com', "
        "'bdpatil642@gmail.com', 'maneshkokani1@gmail.com', 'ruikey170@gmail.com') "
    )

    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    ws.append(
        ['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
         'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
         'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
         'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
         'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts',
         'HSC Total',
         'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream',
         'Graduation Attempts',
         'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts',
         'PG Total',
         'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
         'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
         'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
         'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
         'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
         'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District',
         'Caste Issuing Authority',
         'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government',
         'Government Department',
         'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR'])

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=2nd List Aadesh (15 Students).xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response


@app.route('/installment_userpage', methods=['GET', 'POST'])
def installment_userpage():
    email = session['email']

    # Database connection
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)

    # Fetch user details
    cursor.execute("SELECT * FROM application_page WHERE email=%s", (email,))
    result = cursor.fetchall()

    if result:
        year = result[0]['phd_registration_year']
        startDate = result[0]['final_approved_date']
        user = result[0]['first_name'] + ' ' + result[0]['last_name']
    else:
        user = 'Admin'
        year = None
        startDate = None  # Initialize startDate to avoid further issues

    # Fetch installment and payment details
    cursor.execute("SELECT * FROM installments WHERE email=%s", (email,))
    installments = cursor.fetchall()

    cursor.execute("SELECT * FROM payment_sheet WHERE email=%s", (email,))
    record = cursor.fetchall()

    # Assuming only one row in record
    today = datetime.today().date()
    installment_list = []  # Initialize the list here
    total_period = 0  # Initialize total_period
    total_balance = 0  # Initialize total_balance

    if record:  # Only proceed if record is not empty
        for row in record:
            total_months = int(row['total_months'])
            start_date = startDate

            # Loop to create 15 installments (5 years * 3 installments per year)
            for i in range(1, 16):
                # Set start and end dates for each installment
                if i == 1:
                    current_start_date = start_date
                else:
                    current_start_date = previous_end_date + timedelta(days=30)

                current_end_date = current_start_date + timedelta(days=90)

                # Create installment dictionary
                installment = {
                    'sr_no': i,
                    'period': total_months,
                    'start_period': current_start_date.strftime('%Y-%m-%d'),
                    'end_period': current_end_date.strftime('%Y-%m-%d'),
                    'due_date': (current_end_date + timedelta(days=60)).strftime('%Y-%m-%d'),
                    'balance': 42000,  # Adjust balance if necessary
                    'installment_number': i,
                    'paid': row.get(f'paid_or_not_installment_{i}', 'Not Available')  # Adjust field accordingly
                }

                installment_list.append(installment)
                previous_end_date = current_end_date

            # Calculate total period and balance for all installments
            total_period += total_months  # Assuming total_months is consistent for each installment
            total_balance += 42000  # Assuming balance remains constant; adjust as necessary

    # Fetch other necessary details
    cursor.execute("SELECT fellowship_withdrawn FROM signup WHERE email=%s", (email,))
    output = cursor.fetchall()

    # print(installments)

    payment_statuses = {}
    latest_paid = 0

    # Loop through installment numbers (from 1 to 15)
    for i in range(1, 16):
        status_key = f'status_paid_{i}'  # e.g., status_paid_1, status_paid_2
        query = f"SELECT {status_key} FROM installments WHERE email=%s"
        cursor.execute(query, (email,))
        result_paid = cursor.fetchone()
        # Check if a result was returned and store the payment status
        if result_paid is not None:
            payment_status = result_paid[status_key]  # Get the payment status from the result tuple
        else:
            payment_status = None  # Handle case where no result is found

        payment_statuses[i] = payment_status  # Map installment number to its status

        if payment_statuses.get(i) == 'Paid':
            latest_paid = i

    # Pass this dictionary to your template context
    context = {
        'payment_statuses': payment_statuses,
        'latest_paid': latest_paid,  # Pass the latest paid installment
    }

    # print('Content', payment_statuses)
    # Fetch installment and payment details
    # cursor.execute("SELECT * FROM installments WHERE email=%s", (email,))
    # installments_latest = cursor.fetchall()

    cnx.commit()
    cursor.close()
    cnx.close()

    print('Installments', installments)

    return render_template('installment_userpage.html',
                           result=result, record=record, output=output,
                           installment_list=installment_list, user=user,
                           total_period=total_period, total_balance=total_balance,
                           today=today, installments=installments, year=year, **context)


@app.route('/pay_installment/<int:inst_no>', methods=['POST'])
def pay_installment(inst_no):
    """
    This function is used on the fellowship_details.html Page.
    The function pays installments to the students by Installment Number.
    """
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)

    if request.method == 'POST':
        email = session.get('email')
        start_period = request.form.get('start_period')
        end_period = request.form.get('end_period')
        received_pay = request.form.get('received_pay')
        installment_number = request.form.get('installment_number')
        # Get the current date and format it
        current_date = datetime.now().date()  # Current date
        received_day = current_date.strftime('%A')  # Current day name (e.g., 'Monday')

        print(
            f"Email: {email}, Start: {start_period}, End: {end_period}, Received Pay: {received_pay}, Received Date: {current_date}")

        # Check if the email already exists
        cursor.execute("SELECT * FROM installments WHERE email = %s", (email,))
        user_record = cursor.fetchone()

        if user_record:  # Email exists
            # Check if the installment already exists
            for i in range(1, 5):  # Check inst_num to inst_num_4
                if user_record[f'inst_num_{i}'] == inst_no:
                    # Installment already exists, exit or handle accordingly
                    return "Installment details already exist.", 400

            # If not found, find the first available installment column to insert
            for i in range(1, 16):
                if user_record[f'inst_num_{i}'] is None:  # Check for None
                    cursor.execute(f"""
                            UPDATE installments
                            SET inst_num_{i} = %s, start_period_{i} = %s, end_period_{i} = %s,
                            recieved_pay_{i} = %s, recieved_date_{i} = %s, received_day_{i} = %s, status_paid_{i} = 'Paid'
                            WHERE email = %s
                        """, (inst_no, start_period, end_period, received_pay, current_date, received_day, email))
                    print(
                        f"Updated: inst_num_{i} with {inst_no}, start_period_{i} with {start_period}")  # Debugging output
                    break
        else:  # Email does not exist, insert a new record

            cursor.execute("""
                    INSERT INTO installments (email, inst_num_1, start_period_1, end_period_1, 
                    recieved_pay_1, recieved_date_1, received_day_1, status_paid_1) 
                    VALUES (%s, %s, %s, %s, %s, %s, %s, 'Paid')
                """, (email, inst_no, start_period, end_period, received_pay, current_date, received_day))

            print(f"Inserted new record for {email} with inst_num: {inst_no}")  # Debugging output

        cnx.commit()  # Commit the changes
        cursor.close()  # Close the cursor
        cnx.close()  # Close the connection

        return "Installment paid successfully.", 200


@app.route('/submit_installments', methods=['GET', 'POST'])
def submit_installments():
    """
    This function is used on the installment_userpage.html Page.
    The function is used for submitting the installments for the Old Users who have PHD Registration year
    before 2023.
    :return:
    """
    cnx = mysql.connector.connect(user=user, password=password,
                                  host=host,
                                  database=database)
    cursor = cnx.cursor(dictionary=True)

    if request.method == 'POST':
        email = session.get('email')  # Use .get() to avoid KeyError if 'email' is not in session
        inst_num = request.form.get('inst_num')
        start_period = request.form.get('start_period')
        end_period = request.form.get('end_period')
        recieved_pay = request.form.get('recieved_pay')
        recieved_date = request.form.get('recieved_date')
        recieved_day = request.form.get('recieved_day')

        try:
            # Check if the email already exists in the installments table
            check_query = "SELECT * FROM installments WHERE email = %s"
            cursor.execute(check_query, (email,))
            result = cursor.fetchone()

            if result:
                # Check if each installment slot is completely filled
                installment_2_filled = (result['start_period_2'] and
                                        result['end_period_2'] and
                                        result['recieved_pay_2'] and
                                        result['recieved_date_2'])
                installment_3_filled = (result['start_period_3'] and
                                        result['end_period_3'] and
                                        result['recieved_pay_3'] and
                                        result['recieved_date_3'])
                installment_4_filled = (result['start_period_4'] and
                                        result['end_period_4'] and
                                        result['recieved_pay_4'] and
                                        result['recieved_date_4'])
                installment_5_filled = (result['start_period_5'] and
                                        result['end_period_5'] and
                                        result['recieved_pay_5'] and
                                        result['recieved_date_5'])

                if not installment_2_filled:
                    update_query = """
                        UPDATE installments
                        SET inst_num_2 = %s, start_period_2 = %s, end_period_2 = %s, 
                            recieved_pay_2 = %s, recieved_date_2 = %s, received_day_2 = %s,  status_paid_2 = %s
                        WHERE email = %s
                    """
                    values = (2, start_period, end_period, recieved_pay, recieved_date, recieved_day, 'Paid', email)
                elif not installment_3_filled:
                    update_query = """
                        UPDATE installments
                        SET inst_num_3 = %s, start_period_3 = %s, end_period_3 = %s, 
                            recieved_pay_3 = %s, recieved_date_3 = %s, received_day_3 = %s,status_paid_3 = %s
                        WHERE email = %s
                    """
                    values = (3, start_period, end_period, recieved_pay, recieved_date, recieved_day, 'Paid', email)
                elif not installment_4_filled:
                    update_query = """
                        UPDATE installments
                        SET inst_num_4 = %s, start_period_4 = %s, end_period_4 = %s, 
                            recieved_pay_4 = %s, recieved_date_4 = %s, received_day_4 = %s, status_paid_4 = %s
                        WHERE email = %s
                    """
                    values = (4, start_period, end_period, recieved_pay, recieved_date, recieved_day, 'Paid', email)
                elif not installment_5_filled:
                    update_query = """
                        UPDATE installments
                        SET inst_num_5 = %s, start_period_5 = %s, end_period_5 = %s, 
                            recieved_pay_5 = %s, recieved_date_5 = %s, received_day_5 = %s, status_paid_5 = %s
                        WHERE email = %s
                    """
                    values = (5, start_period, end_period, recieved_pay, recieved_date, recieved_day, 'Paid', email)
                else:
                    return "Maximum installments reached", 400

                cursor.execute(update_query, values)
                message = "Installment updated successfully"
                flash(message, 'success')
            else:
                print('I am here')
                # If email does not exist, insert a new record
                insert_query = """
                    INSERT INTO installments (email, inst_num_1, start_period_1, end_period_1, 
                                              recieved_pay_1, recieved_date_1, received_day_1, status_paid_1)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                """
                values = (email, inst_num, start_period, end_period, recieved_pay, recieved_date, recieved_day, 'Paid')
                print(values)
                cursor.execute(insert_query, values)
                message = "Installment submitted successfully"
                flash(message, 'success')

            cnx.commit()
            return redirect(url_for('installment_userpage'))
        except Exception as e:
            cnx.rollback()
            return f"An error occurred: {str(e)}", 500
        finally:
            cursor.close()
            cnx.close()

    return redirect(url_for('installment_userpage'))


@app.route('/testapi', methods=['GET', 'POST'])
def testapi():
    return render_template('testapi.html',request=request)


if __name__ == '__main__':
    app.run(debug = True)


# 27th March 2024 (10:04)