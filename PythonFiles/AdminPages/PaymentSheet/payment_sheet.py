from datetime import date, timedelta, datetime
import io
import mysql
from openpyxl.workbook import Workbook
from openpyxl.styles import Font
from classes.database import HostConfig, ConfigPaths, ConnectParam
from fpdf import FPDF
from flask import Blueprint, render_template, session, request, redirect, url_for, flash, make_response, jsonify
from PythonFiles.AdminPages.ExportExcel.export_payment_sheet_column_names import COMMON_COLUMNS, COMMON_HEADERS

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from reportlab.platypus import Table, TableStyle
from io import BytesIO

payment_sheet_blueprint = Blueprint('payment_sheet', __name__)


def payment_sheet_auth(app):
    # ------ HOST Configs are in classes/connection.py
    host = HostConfig.host
    app_paths = ConfigPaths.paths.get(host)
    if app_paths:
        for key, value in app_paths.items():
            app.config[key] = value

    def get_payment_year_quarter(year, quarter):  # Separate function for data fetching
        host = HostConfig.host
        connect_param = ConnectParam(host)
        cnx, cursor = connect_param.connect(use_dict=True)

        if year:
            database = f'payment_sheet_{year}'

        query = f"""
                   SELECT * FROM {database} 
                   WHERE fellowship_awarded_year = %s
                   AND quarters = %s
                """
        cursor.execute(query, (year,quarter))
        data = cursor.fetchall()
        cursor.close()
        cnx.close()

        return data  # Return the data (list of dictionaries)

    @payment_sheet_blueprint.route('/get_payment_sheet_data', methods=['GET', 'POST'])
    def get_payment_sheet_data():
        year = request.args.get('year', '2024')
        quarter = request.args.get('quarter', '1')
        hod_payment_data = get_payment_year_quarter(year, quarter)  # Call the data fetching function
        print("The HOD Payment Data is:")
        for item in hod_payment_data:
            print(item)
        data = {
            'hod_payment_data': hod_payment_data
        }
        return jsonify(data)

    def admin_accept_payment_sheet_status(cnx, cursor, year, sheet_id, status): # Added cnx and cursor
        if year:
            database = f'payment_sheet_{year}'
            admin_action = 'Approved by Admin'
            update_query = f"UPDATE {database} SET admin_approval = %s, admin_action = %s WHERE number = %s"
            try:
                cursor.execute(update_query, (status, admin_action, sheet_id))
                cnx.commit() # Commit the transaction
                print("Status Updated")
            except mysql.connector.Error as err:
                print(f"Error updating status: {err}")
                cnx.rollback() # Rollback on error
                flash('An error occurred while updating status.', 'error') # Flash message

    def admin_reject_payment_sheet_status(cnx, cursor, year, sheet_id, status, rejection_reason): # Added cnx and cursor
        if year:
            database = f'payment_sheet_{year}'
            admin_action = 'Rejected by Admin'
            print('SHEET ID for REJECTION', sheet_id)
            update_query = f"UPDATE {database} SET admin_approval = %s, admin_reject_reason = %s, admin_action = %s WHERE number = %s"
            try:
                cursor.execute(update_query, (status, rejection_reason, admin_action, sheet_id))
                cnx.commit() # Commit the transaction
                print("Status Updated")
            except mysql.connector.Error as err:
                print(f"Error updating status: {err}")
                cnx.rollback() # Rollback on error
                flash('An error occurred while updating status.', 'error') # Flash message


    def admin_hold_payment_sheet_status(cnx, cursor, year, sheet_id, status, hold_reason): # Added cnx and cursor
        if year:
            database = f'payment_sheet_{year}'
            admin_action = 'On Hold by Admin'
            update_query = f"UPDATE {database} SET admin_approval = %s, admin_hold_reason = %s, admin_action = %s WHERE number = %s"
            try:
                cursor.execute(update_query, (status, hold_reason, admin_action, sheet_id))
                cnx.commit() # Commit the transaction
                print("Status Updated")
            except mysql.connector.Error as err:
                print(f"Error updating status: {err}")
                cnx.rollback() # Rollback on error
                flash('An error occurred while updating status.', 'error') # Flash message

    @payment_sheet_blueprint.route('/payment_sheet', methods=['GET', 'POST'])
    def payment_sheet():
        if not session.get('logged_in'):
            # Redirect to the admin login page if the user is not logged in
            return redirect(url_for('admin_login'))

        user = session['user']
        print("The user is " + user)

        user_records = []
        if request.method == 'GET':
            # Establish a database connection
            host = HostConfig.host
            connect_param = ConnectParam(host)
            cnx, cursor = connect_param.connect(use_dict=True)
            # print('I have made connection')

            # Fetch user data based on the email
            cursor.execute("""
                    SELECT *  
                    FROM payment_sheet_2024 where quarters = 'Quarter 1';
            """)
            user_data = cursor.fetchall()  # Use fetchall to retrieve all rows
            print('user data:', user_data)
            
            # Fetch admin details
            cursor.execute("SELECT * FROM admin WHERE username = %s", (user,))
            admin_result = cursor.fetchone()

            # Default year selection
            year_selected = "2023"  

            if admin_result:
                admin_year = admin_result['year']
                admin_username = admin_result['username']
                role = admin_result['role']

                # Extract and format username
                first_name = admin_result.get('first_name', '') or ''
                surname = admin_result.get('surname', '') or ''
                username = first_name + ' ' + surname
                if username.strip() in ('None', ''):
                    username = "Admin"

                print("The username is " + username)

                if role == "Admin":
                    if admin_username == "Admin2021" and admin_year == "BANRF 2021":
                        year_selected = "2021"
                    elif admin_username == "Admin2022" and admin_year == "BANRF 2022":
                        year_selected = "2022"
                    elif admin_username == "Admin2023" and admin_year == "BANRF 2023":
                        year_selected = "2023"
                    elif admin_username == "Admin2024" and admin_year == "BANRF.2024":  # Corrected typo here
                        year_selected = "2024"

            # Set available years
            years = ["2020", "2021", "2022", "2023", "2024"]
            # Set available usernames
            usernames = ["Admin2021", "Admin2022", "Admin2023", "Admin2024"]

            print("I am displaying the Payment Sheet")

            # Get year from request, fallback to admin's assigned year if not provided
            year = request.args.get('year', year_selected)
            print("The year selected is :" +year)


            # Close the database cursor and connection
            cursor.close()
            cnx.close()

        return render_template('AdminPages/PaymentSheet/payment_sheet.html', user_data=user_data,
                year=year,
                years=years,
                username=username,
                year_selected=year_selected,
                admin_username=admin_username,
                usernames=usernames)

    class PDF(FPDF):
        def __init__(self):
            super().__init__(orientation='P', unit='mm', format='A3')  # Set format to A3

        header_added = False  # To track whether the header is added to the first page

        def header(self):
            if not self.header_added:
                var = get_base_url()
                print(var)

                # Adjusted dimensions for A3 format
                self.set_font("Arial", "B", 12)

                # Adjust the X, Y, and image size to fit A3 format
                # for LOCALSERVER
                # self.image('/var/www/icswebapp/icswebapp/static/Images/satya.png', 140, 10,
                #            30)  # Adjusted position and size for A3
                # image_width = 140  # Updated to a size appropriate for A3
                # text_x_position = self.get_x()
                # text_y_position = self.get_y() + 25  # Adjusted for A3 format
                # self.set_xy(text_x_position, text_y_position)
                # # Adjusted positions for A3 format
                # self.image('/var/www/icswebapp/icswebapp/static/Images/newtrtiImage.png', 20, 10,
                #            60)  # Adjust size for larger format
                # self.image('/var/www/icswebapp/icswebapp/static/Images/mahashasn_new.png', 215, 10,
                #            60)  # Adjust size and position for A3

                # For HOSTSERVER
                self.image('static/Images/satya.png', 140, 10, 30)  # Adjusted position and size for A3
                image_width = 140  # Updated to a size appropriate for A3
                text_x_position = self.get_x()
                text_y_position = self.get_y() + 25  # Adjusted for A3 format
                self.set_xy(text_x_position, text_y_position)
                # Adjusted positions for A3 format
                self.image('static/Images/newtrtiImage.png', 20, 10, 60)  # Adjust size for larger format
                self.image('static/Images/mahashasn_new.png', 215, 10, 60)  # Adjust size and position for A3

                # Centered text for A3 format
                self.ln(10)
                self.cell(0, 10, "Government of Maharashtra", align="C", ln=True)
                self.cell(0, 10, "Tribal Research & Training Institute", align="C", ln=True)
                self.cell(0, 10, "28, Queens Garden, Pune - 411001", align="C", ln=True)

                # Adjust the dashed line width for A3
                self.dashed_line(10, self.get_y(), 290, self.get_y(), dash_length=3,
                                 space_length=1)  # Adjust for A3 width

                self.ln(5)  # Adjust space after the line
                self.set_font("Arial", size=10)

                current_date = datetime.now().strftime('%B %Y')
                # Left-aligned and right-aligned text for A3 format
                self.cell(0, 10, "No.: Research-_____/Case.No ____/Desk- __/_____ ", ln=False)
                self.cell(0, 10, f"Date: {current_date}", align="R", ln=True)

                self.set_font("Arial", "B", size=12)
                self.cell(0, 10, "Appendix", align="C", ln=True)

                # Adjust rotation for A3
                # self.ln(2)
                # self.rotate(45)
                # self.set_font('Arial', '', 65)  # Adjust font size for A3
                # self.set_text_color(192, 192, 192)

                # Adjust the position for the watermark on A3
                # self.text(-50, 280, "STRF FELLOWSHIP")  # Adjust position for A3 layout
                # self.rotate(0)

                self.header_added = True  # Set to True after adding the header

        def footer(self):
            self.set_y(-15)
            self.set_font("Arial", 'I', 8)
            self.cell(0, 10, f'Page {self.page_no()}', 0, 0, 'C')

    def get_base_url():
        base_url = request.url_root
        return base_url

     # ----------------------------------------------------------------
    # Common Export to Excel Function
    # ----------------------------------------------------------------
    @payment_sheet_blueprint.route('/export_payment_sheet', methods=['GET'])
    def export_payment_sheet():
        """
            This function is responsible for handling the dynamic exporting of application report data based
            on the selected year.
            Path of AJAX Call: /static/admin.js. (Search the form_types in the JS File)
            Path of HTML can be found in the respective templates.
            {columns_str} will be found in: PythonFiles/AdminPages/ExportExcel/export_payment_sheet_column_names.py
        """
        if not session.get('logged_in'):
            flash('Please enter Email ID and Password', 'error')
            return redirect(url_for('adminlogin.admin_login'))

        host = HostConfig.host
        connect_param = ConnectParam(host)
        cnx, cursor = connect_param.connect(use_dict=True)

        user = session['user']
        print("The EXPORT user is :" , user)

        try:
            # Fetch admin details
            cursor.execute("SELECT * FROM admin WHERE username = %s", (user,))
            admin_result = cursor.fetchone()

            # Default year selection
            year = request.args.get('year', default=2024, type=int)
            quarter = request.args.get('quarter', default="Quarter 1", type=str)

            if admin_result:
                admin_year = admin_result['year']
                admin_username = admin_result['username']
                role = admin_result['role']

                # Extract and format username
                first_name = admin_result.get('first_name', '') or ''
                surname = admin_result.get('surname', '') or ''
                username = first_name + ' ' + surname
                if username.strip() in ('None', ''):
                    username = "Admin"

                print("The username is " + username)

                if role == "Admin":
                    if admin_username == "Admin2021" and admin_year == "BANRF 2021":
                        year = "2021"
                    elif admin_username == "Admin2022" and admin_year == "BANRF 2022":
                        year = "2022"
                    elif admin_username == "Admin2023" and admin_year == "BANRF 2023":
                        year = "2023"
                    elif admin_username == "Admin2024" and admin_year == "BANRF.2024":  # Corrected typo here
                        year = "2024"

                # year = request.args.get('year', default=2023, type=int)
                print("Exporting the Report for YEAR : " , year)
                print("Exporting the Report for Quarter : " , quarter)
                form_type = request.args.get('form_type')  # Get the form type (e.g., "completed_form")

                columns_str = ', '.join(COMMON_COLUMNS)

                if year:
                            table_name = f'payment_sheet_{year}'

                # Dynamically change the SQL query based on form_type
                if form_type == "export_to_excel_payment_sheet":
                    cursor.execute(f"SELECT {columns_str} FROM {table_name} WHERE fellowship_awarded_year = %s AND quarters = %s", (year, quarter,))
                # elif form_type == 'not_disabled_application_records':
                #     cursor.execute(f"SELECT {columns_str} FROM application_page WHERE phd_registration_year = %s and disability='No' ",
                #                 (year,))
                else:
                    # Handle other form types or default case
                    flash('Error fetching Details. Some details are missing.', 'error')

                data = cursor.fetchall()  # Fetch the results

                # Close the connection
                cursor.close()
                cnx.close()

                # Create an Excel workbook
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = f"Data_{year}"

                # Write headers
                if data:
                    # Map database column names to headers
                    headers = [COMMON_HEADERS.get(column, column) for column in
                            data[0].keys()]  # Use COMMON_HEADERS for headers
                    sheet.append(headers)  # Add headers to the first row

                    # Write data rows
                    for row_data in data:
                        sheet.append(
                            [row_data.get(column, '') for column in data[0].keys()])  # Ensure data matches the header order

                # Save the workbook to an in-memory stream
                output = BytesIO()
                workbook.save(output)
                output.seek(0)

                # Return the file as a downloadable response
                response = make_response(output.read())
                response.headers['Content-Disposition'] = f'attachment; filename=export_{form_type}_{year}_{quarter}.xlsx'
                response.headers['Content-type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                return response
        except Exception as e:
            print(f"An error occurred: {e}")
            flash(f"An error occurred: {e}", 'error')
            return redirect(url_for('payment_sheet.payment_sheet')) #replace some_route with a valid route.

    # END Common Export to Excel
    # ----------------------------------------------------------------

    @app.route('/export_payment_sheet_pdf')
    def export_payment_sheet_pdf():
        return render_template('AdminPages/PaymentSheet/payment_sheet_pdf.html')

    @app.route('/generate_pdf', methods=['GET'])
    def generate_pdf():
        """
        Generate Payment Sheet PDF to display in iframe
        """
        if not session.get('logged_in'):
            flash('Please enter Email ID and Password', 'error')
            return redirect(url_for('adminlogin.admin_login'))

        host = HostConfig.host
        connect_param = ConnectParam(host)
        cnx, cursor = connect_param.connect(use_dict=True)

        user = session['user']

        try:
            cursor.execute("SELECT * FROM admin WHERE username = %s", (user,))
            admin_result = cursor.fetchone()

            year = request.args.get('year', default=2023, type=int)
            quarter = request.args.get('quarter', default="Quarter 1", type=str)

            if admin_result:
                role = admin_result['role']
                admin_username = admin_result['username']
                admin_year = admin_result['year']

                if role == "Admin":
                    if admin_username == "Admin2021" and admin_year == "BANRF 2021":
                        year = "2021"
                    elif admin_username == "Admin2022" and admin_year == "BANRF 2022":
                        year = "2022"
                    elif admin_username == "Admin2023" and admin_year == "BANRF 2023":
                        year = "2023"
                    elif admin_username == "Admin2024" and admin_year == "BANRF.2024":
                        year = "2024"

            table_name = f'payment_sheet_{year}'
            cursor.execute(f"SELECT {', '.join(COMMON_COLUMNS)} FROM {table_name} WHERE fellowship_awarded_year = %s AND quarters = %s", (year, quarter))
            data = cursor.fetchall()

            output = BytesIO()
            pdf = canvas.Canvas(output, pagesize=A4)
            pdf.setTitle(f"Payment_Sheet_{year}_{quarter}")

            pdf.drawString(200, 800, f"Payment Sheet for {year} - {quarter}")
            pdf.line(50, 790, 550, 790)

            table_data = [COMMON_HEADERS.values()]  # Headers
            for row in data:
                table_data.append([str(row[column]) for column in COMMON_COLUMNS])

            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 0.25, colors.black),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ]))

            table.wrapOn(pdf, 50, 600)
            table.drawOn(pdf, 50, 600)

            pdf.save()
            output.seek(0)

            response = make_response(output.read())
            response.headers['Content-Type'] = 'application/pdf'
            return response

        except Exception as e:
            print(f"An error occurred: {e}")
            flash(f"An error occurred: {e}", 'error')
            return redirect(url_for('payment_sheet.payment_sheet'))

        finally:
            cursor.close()
            cnx.close()


    @payment_sheet_blueprint.route('/accepted_payment_sheet', methods=['GET', 'POST'])
    def accepted_payment_sheet():
        if not session.get('logged_in'):
            # Redirect to the admin login page if the user is not logged in
            return redirect(url_for('admin_login'))

        user = session['user']
        print("The user is " + user)

        user_records = []
        if request.method == 'GET':
            # Establish a database connection
            host = HostConfig.host
            connect_param = ConnectParam(host)
            cnx, cursor = connect_param.connect(use_dict=True)
            # print('I have made connection')

            # Fetch user data based on the email
            cursor.execute("""
                    SELECT *  
                    FROM payment_sheet_2024 where quarters = 'Quarter 1';
            """)
            user_data = cursor.fetchall()  # Use fetchall to retrieve all rows
            print('user data:', user_data)
            
            # Fetch admin details
            cursor.execute("SELECT * FROM admin WHERE username = %s", (user,))
            admin_result = cursor.fetchone()

            # Default year selection
            year_selected = "2023"  

            if admin_result:
                admin_year = admin_result['year']
                admin_username = admin_result['username']
                role = admin_result['role']

                # Extract and format username
                first_name = admin_result.get('first_name', '') or ''
                surname = admin_result.get('surname', '') or ''
                username = first_name + ' ' + surname
                if username.strip() in ('None', ''):
                    username = "Admin"

                print("The username is " + username)

                if role == "Admin":
                    if admin_username == "Admin2021" and admin_year == "BANRF 2021":
                        year_selected = "2021"
                    elif admin_username == "Admin2022" and admin_year == "BANRF 2022":
                        year_selected = "2022"
                    elif admin_username == "Admin2023" and admin_year == "BANRF 2023":
                        year_selected = "2023"
                    elif admin_username == "Admin2024" and admin_year == "BANRF.2024":  # Corrected typo here
                        year_selected = "2024"

            # Set available years
            years = ["2020", "2021", "2022", "2023", "2024"]
            # Set available usernames
            usernames = ["Admin2021", "Admin2022", "Admin2023", "Admin2024"]

            print("I am displaying the Payment Sheet")

            # Get year from request, fallback to admin's assigned year if not provided
            year = request.args.get('year', year_selected)
            print("The year selected is :" +year)


            # Close the database cursor and connection
            cursor.close()
            cnx.close()

        return render_template('AdminPages/PaymentSheet/accepted_payment_sheet.html', user_data=user_data,
                year=year,
                years=years,
                username=username,
                year_selected=year_selected,
                admin_username=admin_username,
                usernames=usernames)     

    @payment_sheet_blueprint.route('/rejected_payment_sheet', methods=['GET', 'POST'])
    def rejected_payment_sheet():
        if not session.get('logged_in'):
            # Redirect to the admin login page if the user is not logged in
            return redirect(url_for('admin_login'))

        user = session['user']
        print("The user is " + user)

        user_records = []
        if request.method == 'GET':
            # Establish a database connection
            host = HostConfig.host
            connect_param = ConnectParam(host)
            cnx, cursor = connect_param.connect(use_dict=True)
            # print('I have made connection')

            # Fetch user data based on the email
            cursor.execute("""
                    SELECT *  
                    FROM payment_sheet_2024 where quarters = 'Quarter 1';
            """)
            user_data = cursor.fetchall()  # Use fetchall to retrieve all rows
            print('user data:', user_data)
            
            # Fetch admin details
            cursor.execute("SELECT * FROM admin WHERE username = %s", (user,))
            admin_result = cursor.fetchone()

            # Default year selection
            year_selected = "2023"  

            if admin_result:
                admin_year = admin_result['year']
                admin_username = admin_result['username']
                role = admin_result['role']

                # Extract and format username
                first_name = admin_result.get('first_name', '') or ''
                surname = admin_result.get('surname', '') or ''
                username = first_name + ' ' + surname
                if username.strip() in ('None', ''):
                    username = "Admin"

                print("The username is " + username)

                if role == "Admin":
                    if admin_username == "Admin2021" and admin_year == "BANRF 2021":
                        year_selected = "2021"
                    elif admin_username == "Admin2022" and admin_year == "BANRF 2022":
                        year_selected = "2022"
                    elif admin_username == "Admin2023" and admin_year == "BANRF 2023":
                        year_selected = "2023"
                    elif admin_username == "Admin2024" and admin_year == "BANRF.2024":  # Corrected typo here
                        year_selected = "2024"

            # Set available years
            years = ["2020", "2021", "2022", "2023", "2024"]
            # Set available usernames
            usernames = ["Admin2021", "Admin2022", "Admin2023", "Admin2024"]

            print("I am displaying the Payment Sheet")

            # Get year from request, fallback to admin's assigned year if not provided
            year = request.args.get('year', year_selected)
            print("The year selected is :" +year)


            # Close the database cursor and connection
            cursor.close()
            cnx.close()

        return render_template('AdminPages/PaymentSheet/rejected_payment_sheet.html', user_data=user_data,
                year=year,
                years=years,
                username=username,
                year_selected=year_selected,
                admin_username=admin_username,
                usernames=usernames)  

    @payment_sheet_blueprint.route('/onhold_payment_sheet', methods=['GET', 'POST'])
    def onhold_payment_sheet():
        if not session.get('logged_in'):
            # Redirect to the admin login page if the user is not logged in
            return redirect(url_for('admin_login'))

        user = session['user']
        print("The user is " + user)

        user_records = []
        if request.method == 'GET':
            # Establish a database connection
            host = HostConfig.host
            connect_param = ConnectParam(host)
            cnx, cursor = connect_param.connect(use_dict=True)
            # print('I have made connection')

            # Fetch user data based on the email
            cursor.execute("""
                    SELECT *  
                    FROM payment_sheet_2024 where quarters = 'Quarter 1';
            """)
            user_data = cursor.fetchall()  # Use fetchall to retrieve all rows
            print('user data:', user_data)
            
            # Fetch admin details
            cursor.execute("SELECT * FROM admin WHERE username = %s", (user,))
            admin_result = cursor.fetchone()

            # Default year selection
            year_selected = "2023"  

            if admin_result:
                admin_year = admin_result['year']
                admin_username = admin_result['username']
                role = admin_result['role']

                # Extract and format username
                first_name = admin_result.get('first_name', '') or ''
                surname = admin_result.get('surname', '') or ''
                username = first_name + ' ' + surname
                if username.strip() in ('None', ''):
                    username = "Admin"

                print("The username is " + username)

                if role == "Admin":
                    if admin_username == "Admin2021" and admin_year == "BANRF 2021":
                        year_selected = "2021"
                    elif admin_username == "Admin2022" and admin_year == "BANRF 2022":
                        year_selected = "2022"
                    elif admin_username == "Admin2023" and admin_year == "BANRF 2023":
                        year_selected = "2023"
                    elif admin_username == "Admin2024" and admin_year == "BANRF.2024":  # Corrected typo here
                        year_selected = "2024"

            # Set available years
            years = ["2020", "2021", "2022", "2023", "2024"]
            # Set available usernames
            usernames = ["Admin2021", "Admin2022", "Admin2023", "Admin2024"]

            print("I am displaying the Payment Sheet")

            # Get year from request, fallback to admin's assigned year if not provided
            year = request.args.get('year', year_selected)
            print("The year selected is :" +year)


            # Close the database cursor and connection
            cursor.close()
            cnx.close()

        return render_template('AdminPages/PaymentSheet/onhold_payment_sheet.html', user_data=user_data,
                year=year,
                years=years,
                username=username,
                year_selected=year_selected,
                admin_username=admin_username,
                usernames=usernames)     