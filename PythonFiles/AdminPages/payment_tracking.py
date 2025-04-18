from datetime import date, timedelta, datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import mysql.connector
from classes.database import HostConfig, ConfigPaths, ConnectParam
import os
from flask_mail import Mail, Message
from flask import Blueprint, render_template, session, request, redirect, url_for, flash, make_response
from authentication.middleware import auth

payment_tracking_blueprint = Blueprint('payment_tracking', __name__)


def payment_tracking_auth(app):
    # ------ HOST Configs are in classes/connection.py
    host = HostConfig.host
    app_paths = ConfigPaths.paths.get(host)
    if app_paths:
        for key, value in app_paths.items():
            app.config[key] = value

    @payment_tracking_blueprint.route('/payment_tracking', methods=['GET', 'POST'])
    def payment_tracking():
        if not session.get('logged_in'):
            flash('Please enter Email ID and Password', 'error')
            return redirect(url_for('adminlogin.admin_login'))

        user = session['user']
        host = HostConfig.host
        connect_param = ConnectParam(host)
        cnx, cursor = connect_param.connect(use_dict=True)

        try:
            cursor.execute("SELECT * FROM admin WHERE username = %s", (user,))
            admin_result = cursor.fetchone()

            if admin_result:
                admin_year = admin_result['year']
                admin_username = admin_result['username']
                role = admin_result['role']

                first_name = admin_result.get('first_name', '') or ''
                surname = admin_result.get('surname', '') or ''
                username = first_name + ' ' + surname
                if username.strip() in ('None', ''):
                    username = "Admin"

                print("The username is " + username)

                # Determine the admin's year or if they are the HOD
                if role == "Admin" and admin_username.startswith("Admin"):
                    year_selected = admin_username[5:]  # Extract the year from the username
                    admin_type = "year_admin"
                else:
                    year_selected = None  # HOD or other admin
                    admin_type = "hod_admin"

                years = ["2020", "2021", "2022", "2023", "2024"]
                usernames = ["Admin2021", "Admin2022", "Admin2023", "Admin2024"]

                print("Displaying the Report in Track Payments")

                year = request.args.get('year', year_selected) if year_selected else request.args.get('year')
                print("The year selected is :" + (year or "None"))

                records_display = []
                payment_records = []

                print("The year selected for TRACK Payment is :", year_selected)

                if request.method == 'POST':
                    start_date = request.form.get('start_date')
                    end_date = request.form.get('end_date')
                    year = request.form.get('year')
                    month = request.form.get('month')
                    quarters = request.form.get('quarters')
                    print(quarters)

                    conditions = []
                    params = []

                    if start_date and end_date:
                        conditions.append("ps.date BETWEEN %s AND %s")
                        params.extend([start_date, end_date])
                    if year:
                        conditions.append("ps.fellowship_awarded_year = %s")
                        params.append(year)
                    if month:
                        conditions.append("ps.duration_month = %s")
                        params.append(month)
                    if quarters:
                        conditions.append("ps.quarters = %s")
                        params.append(quarters)

                    if admin_type == "year_admin":
                        # Year-specific admin: target their table
                        table_name = f"payment_sheet_{year_selected}"
                        sql = f"""
                            SELECT ap.*, ps.*
                            FROM application_page ap
                            JOIN {table_name} ps ON ap.email = ps.email
                            WHERE ap.final_approval = 'accepted'
                        """
                        if conditions:
                            sql += " AND " + " AND ".join(conditions)
                        cursor.execute(sql, params)
                        records_display = cursor.fetchall()
                        print(sql)

                    elif admin_type == "hod_admin":
                        # HOD admin: query all tables and union the results
                        all_results = []
                        for y in years:
                            table_name = f"payment_sheet_{y}"
                            sql = f"""
                                SELECT ap.*, ps.*
                                FROM application_page ap
                                JOIN {table_name} ps ON ap.email = ps.email
                                WHERE ap.final_approval = 'accepted'
                            """
                            if conditions:
                                sql += " AND " + " AND ".join(conditions)
                            cursor.execute(sql, params)
                            all_results.extend(cursor.fetchall())
                        records_display = all_results
                        print("HOD Admin query executed.")

                    flash('Payment information retrieved successfully', 'success')

                #application_page data
                sql = """
                    SELECT * FROM application_page 
                    WHERE final_approval = 'accepted' 
                        AND phd_registration_year >= '2023'

                    UNION

                    SELECT * FROM application_page 
                    WHERE phd_registration_year > '2020' 
                        AND aadesh = 1;
                """
                cursor.execute(sql)
                records = cursor.fetchall()

                for record in records:
                    email = record['email']
                    sql = "SELECT * FROM payment_sheet WHERE email=%s" #this will be removed in final version.
                    cursor.execute(sql, (email,))
                    result = cursor.fetchall()

                    for payment_record in result:
                        if 'duration_date_from' in payment_record:
                            try:
                                date_obj = datetime.strptime(payment_record['duration_date_from'], '%Y-%m-%d')
                                date_objj = datetime.strptime(payment_record['duration_date_to'], '%Y-%m-%d')
                                payment_record['duration_date_from'] = date_obj.strftime('%d/%m/%Y')
                                payment_record['duration_date_to'] = date_objj.strftime('%d/%m/%Y')
                            except ValueError:
                                pass

                    payment_records.append(result)

                flattened_records = [record for sublist in payment_records for record in sublist]

                return render_template('AdminPages/payment_tracking.html', records_display=records_display, records=records,
                                    payment_records=flattened_records, role=role)
        finally:
            if cursor:
                cursor.close()
            if cnx and cnx.is_connected():
                cnx.close()

    @payment_tracking_blueprint.route('/export_track_payments')
    def export_track_payments():
        host = HostConfig.host
        connect_param = ConnectParam(host)
        cnx, cursor = connect_param.connect(use_dict=True)

        cursor.execute(
            "SELECT number, full_name, email, faculty, jrf_srf, fellowship_awarded_date, date, duration_date_from, duration_date_to, "
            "total_months, fellowship, to_fellowship, rate, amount, months, total_hra, count, pwd, total,"
            "city, bank_name, ifsc_code, account_number FROM payment_sheet")

        data = cursor.fetchall()
        print(data)

        # Create a workbook and add a worksheet
        wb = Workbook()
        ws = wb.active

        header = 4
        date_column_index = 1
        date_column_index_date = 8
        current_date = datetime.now().strftime('%B %Y')
        # Add the bold row before the header
        ws.cell(row=1, column=header, value='Appendix')  # Place "Date:" in column B
        ws.cell(row=2, column=date_column_index, value='Number:')  # Place "Date:" in column B
        ws.cell(row=2, column=date_column_index_date, value=f'Date:{current_date}')  # Place "Date:" in column B
        # ws.cell(row=1, column=date_column_index_date + 1, value=current_date)  # Place the date in column C
        # Add the bold row before the header
        # ws.append(['Date:', current_date])  # Format as "Date: September 2024"
        bold_row = ws[1]  # Get the last added row (the one we just added)
        bold_row_2 = ws[2]  # Get the last added row (the one we just added)
        for cell in bold_row and bold_row_2:
            cell.font = Font(bold=True)  # Make the text bold

        # Add header row
        ws.append(['Sr. No.', 'Name of Student', 'Faculty', 'JRF/SRF', 'Date of PHD Registration',
                   'Fellowship Awarded Date', 'Duration', 'Total Months', 'Fellowship', 'Total Fellowship',
                   'H.R.A Rate',
                   'H.R.A Amount', 'Months', 'Total H.R.A', 'Contingency Yearly', 'PWD', 'Total Amount', 'City'])

        # Add data to the worksheet with formatting
        for index, row in enumerate(data, start=1):
            full_name = row['full_name']

            # Joining date
            joining_date = row['date']
            fellowship_awarded_date = row['fellowship_awarded_date']
            # Duration dates
            duration_date_from = row['duration_date_from']
            duration_date_to = row['duration_date_to']

            # Convert string dates to datetime objects if they are not already
            if isinstance(duration_date_from, str):
                try:
                    duration_date_from = datetime.strptime(duration_date_from, '%Y-%m-%d')  # Adjust format as needed
                except ValueError:
                    duration_date_from = None

            if isinstance(duration_date_to, str):
                try:
                    duration_date_to = datetime.strptime(duration_date_to, '%Y-%m-%d')  # Adjust format as needed
                except ValueError:
                    duration_date_to = None

            if isinstance(duration_date_from, datetime) and isinstance(duration_date_to, datetime):
                duration_date_from_str = duration_date_from.strftime('%d %b %Y')  # Format as "17 Aug 2023"
                duration_date_to_str = duration_date_to.strftime('%d %b %Y')  # Format as "15 Nov 2023"
                duration = f"{duration_date_from_str} to {duration_date_to_str}"
            else:
                duration = "N/A"

            # Other fields
            faculty = row['faculty']
            jrf_srf = row['jrf_srf']
            total_months = row['total_months']
            fellowship = row['fellowship']
            total_felowship = row['to_fellowship']
            hra_rate = row['rate']
            hra_amount = row['amount']
            months = row['months']
            total_hra = row['total_hra']
            count_yearly = row['count']
            pwd = row['pwd']
            total = row['total']
            city = row['city']

            # Append the formatted data
            ws.append([index, full_name, faculty, jrf_srf, joining_date, fellowship_awarded_date, duration,
                       total_months, fellowship, total_felowship, hra_rate, hra_amount, months, total_hra,
                       count_yearly, pwd, total, city])

        # Save the workbook in memory as bytes
        data = BytesIO()
        wb.save(data)
        data.seek(0)

        # Create a response object and attach the workbook as a file
        response = make_response(data.getvalue())
        response.headers['Content-Disposition'] = 'attachment; filename=Track Payment Sheet 2023-2024.xlsx'
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

        return response

    @payment_tracking_blueprint.route('/budget_report/<string:email>', methods=['GET', 'POST'])
    def budget_report(email):
        host = HostConfig.host
        connect_param = ConnectParam(host)
        cnx, cursor = connect_param.connect(use_dict=True)

        cursor.execute("SELECT * FROM application_page where email=%s ", (email,))
        result = cursor.fetchall()
        print("result" + str(result))

        cursor.execute("SELECT * FROM payment_sheet where email=%s ", (email,))
        record = cursor.fetchall()
        print("record" + str(record))
        table_data = []

        for row in record:
            total_months = int(row['total_months'])
            print("Total Months" + str(total_months))
            start_date = datetime.strptime(row['duration_date_from'], '%Y-%m-%d')
            print("Start Date" + str(start_date))
            end_date = datetime.strptime(row['duration_date_to'], '%Y-%m-%d')
            print("End Date" + str(end_date))

            table_data.append({
                'sr_no': 1,
                'period': total_months,
                'start_period': start_date.strftime('%Y-%m-%d'),
                'end_period': end_date.strftime('%Y-%m-%d'),
                'due_date': (end_date + timedelta(days=60)).strftime('%Y-%m-%d'),
                'balance': 31000,
                'installment_number': 1,
                'paid': row['paid_or_not_installment_1']  # Assumes 'paid_or_not_installment_1' is a key in 'row'
            })

            # Generate next two installments
            for i in range(2, 4):
                next_start_date = end_date + timedelta(days=30 * (i - 1))
                next_end_date = next_start_date + timedelta(days=90)
                table_data.append({
                    'sr_no': i,
                    'period': total_months,
                    'start_period': next_start_date.strftime('%Y-%m-%d'),
                    'end_period': next_end_date.strftime('%Y-%m-%d'),
                    'due_date': (next_end_date + timedelta(days=60)).strftime('%Y-%m-%d'),
                    'balance': 31000,
                    'installment_number': i,
                    'paid': row[f'paid_or_not_installment_{i}']  # Adjust this based on your payment status logic
                })

            print('table_data:', table_data)
            total_period = sum(int(item['period']) for item in table_data)
            total_balance = sum(int(item['balance']) for item in table_data)
            print("Total Period:", total_period)

        # approve_pay = approve_payment(email)

        cursor.execute("SELECT * FROM award_letter where email=%s ", (email,))
        solution = cursor.fetchall()
        print("record" + str(solution))

        cursor.execute("SELECT fellowship_withdrawn FROM signup where email=%s ", (email,))
        output = cursor.fetchall()
        print("record" + str(output))

        cnx.commit()
        cursor.close()
        cnx.close()
        return render_template('Admin/budget_report.html', result=result, record=record, output=output, solution=solution,
                               table_data=table_data, total_period=total_period, total_balance=total_balance)
