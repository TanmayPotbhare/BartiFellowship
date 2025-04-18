from io import BytesIO
from openpyxl.workbook import Workbook
from classes.database import HostConfig, ConfigPaths, ConnectParam
from flask import Blueprint, render_template, session, request, redirect, url_for, flash, make_response
from PythonFiles.AdminPages.Dashboard.export_column_names import COMMON_COLUMNS, COMMON_HEADERS


adminlevels_Three_blueprint = Blueprint('adminlevels_Three', __name__)


def adminlevels_Three_auth(app):
    """
        All Export to Excels for Routes on Eligibility Check (Level 2) Routes. (Accepted, Pending, etc)
    """
    # ------ HOST Configs are in classes/connection.py
    host = HostConfig.host
    app_paths = ConfigPaths.paths.get(host)
    if app_paths:
        for key, value in app_paths.items():
            app.config[key] = value

        # ----------------------  EXPORT TO EXCEL Level Three Admin-------------------------------------------

    @adminlevels_Three_blueprint.route('/export_level_three_admin', methods=['GET', 'POST'])
    def export_level_three_admin():
        if not session.get('logged_in'):
            flash('Please enter Email ID and Password', 'error')
            return redirect(url_for('adminlogin.admin_login'))

        host = HostConfig.host
        connect_param = ConnectParam(host)
        cnx, cursor = connect_param.connect(use_dict=True)

        user = session['user']
        print("The EXPORT user is :", user)

        try:
            # Fetch admin details
            cursor.execute("SELECT * FROM admin WHERE username = %s", (user,))
            admin_result = cursor.fetchone()

            # Default year selection
            year = request.args.get('year', default=2024, type=int)

            if admin_result:
                admin_year = admin_result['year']
                admin_username = admin_result['username']
                role = admin_result['role']

                # Extract and format username
                first_name = admin_result.get('first_name', '') or ''
                surname = admin_result.get('surname', '') or ''
                username = first_name + ' ' + surname
                if username.strip() in ('None', ''):
                    username = "Admin"

                print("The username is " + username)

                if role == "Admin":
                    if admin_username == "Admin2021" and admin_year == "BANRF 2021":
                        year = "2021"
                    elif admin_username == "Admin2022" and admin_year == "BANRF 2022":
                        year = "2022"
                    elif admin_username == "Admin2023" and admin_year == "BANRF 2023":
                        year = "2023"
                    elif admin_username == "Admin2024" and admin_year == "BANRF.2024":
                        year = "2024"

            print("Exporting the Report for YEAR : ", year)

            # Assuming you want to export accepted applications
            cursor.execute(
                f"SELECT {', '.join(COMMON_COLUMNS)} FROM application_page WHERE phd_registration_year = %s AND scrutiny_status = 'accepted' and final_approval = 'accepted'",
                (year,)
            )
            data = cursor.fetchall()

            if not data: #Check if data is empty.
                flash(f"No records found for the year {year} in Final Approval Report.", 'info')
                return redirect(url_for('adminlevelthree.level_three_admin')) #return the page.

            # Create a workbook and add a worksheet
            wb = Workbook()
            ws = wb.active

            # Write headers
            if data:
                headers = [COMMON_HEADERS.get(column, column) for column in data[0].keys()]
                ws.append(headers)

                # Write data rows
                for row_data in data:
                    ws.append([row_data.get(column, '') for column in data[0].keys()])

            # Save the workbook in memory as bytes
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            # Create a response object and attach the workbook as a file
            response = make_response(output.read())
            response.headers['Content-Disposition'] = f'attachment; filename=Final Approval (level 3) {year}.xlsx'
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

            return response

        except Exception as e:
            print(f"An error occurred: {e}")
            flash(f"An error occurred: {e}", 'error')
            return redirect(url_for('adminlevelthree.level_three_admin'))  # Redirect to the page where year is selected.

        finally:
            if 'cursor' in locals() and cursor:
                cursor.close()
            if 'cnx' in locals() and cnx:
                cnx.close()

    @adminlevels_Three_blueprint.route('/export_accepted_students_level3', methods=['GET', 'POST'])
    def export_accepted_students_level3():
        if not session.get('logged_in'):
            # Redirect to the admin login page if the user is not logged in
            return redirect(url_for('adminlogin.admin_login'))

        host = HostConfig.host
        connect_param = ConnectParam(host)
        cnx, cursor = connect_param.connect()

        cursor.execute(
            "  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
            "date_of_birth, gender, age, caste, your_caste, marital_status,"
            "state, district, taluka, village, city, add_1, add_2, pincode,"
            "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total,"
            "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total,"
            "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream,"
            "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
            "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
            "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
            "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income,"
            "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka,"
            "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number,"
            "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
            "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
            "type_of_disability, father_name, mother_name, work_in_government, gov_department,"
            "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
            " where phd_registration_year>='2023' and final_approval='accepted' and scrutiny_status='accepted' ")
        data = cursor.fetchall()
        # Create a workbook and add a worksheet
        wb = Workbook()
        ws = wb.active

        # Add headers to the worksheet
        # ws.append(['applicant_id','email','first_name','last_name','application_date'])

        ws.append(
            ['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
             'Date Of Birth', 'Gender', 'Age', 'Caste',
             'Marital Status', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
             'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
             'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts',
             'HSC Total',
             'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream',
             'Graduation Attempts',
             'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts',
             'PG Total',
             'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
             'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
             'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
             'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
             'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
             'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District',
             'Caste Issuing Authority',
             'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government',
             'Government Department',
             'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR'])

        # Add data to the worksheet
        for row in data:
            ws.append(row)

        # Save the workbook in memory as bytes
        data = BytesIO()
        wb.save(data)
        data.seek(0)

        # Create a response object and attach the workbook as a file
        response = make_response(data.getvalue())
        response.headers['Content-Disposition'] = 'attachment; filename=Accepted Students (Level 3).xlsx'
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

        return response

    @adminlevels_Three_blueprint.route('/export_pending_students_level3', methods=['GET', 'POST'])
    def export_pending_students_level3():
        if not session.get('logged_in'):
            # Redirect to the admin login page if the user is not logged in
            return redirect(url_for('adminlogin.admin_login'))

        host = HostConfig.host
        connect_param = ConnectParam(host)
        cnx, cursor = connect_param.connect()

        cursor.execute(
            "SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
            "date_of_birth, gender, age, caste, your_caste, marital_status,"
            "state, district, taluka, village, city, add_1, add_2, pincode,"
            "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total,"
            "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total,"
            "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream,"
            "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
            "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
            "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
            "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income,"
            "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka,"
            "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number,"
            "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
            "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
            "type_of_disability, father_name, mother_name, work_in_government, gov_department,"
            "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
            " where phd_registration_year>='2023' and final_approval='pending' and scrutiny_status='accepted' ")
        data = cursor.fetchall()
        # Create a workbook and add a worksheet
        wb = Workbook()
        ws = wb.active

        # Add headers to the worksheet
        # ws.append(['applicant_id','email','first_name','last_name','application_date'])

        ws.append(
            ['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
             'Date Of Birth', 'Gender', 'Age', 'Caste',
             'Marital Status', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
             'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
             'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts',
             'HSC Total',
             'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream',
             'Graduation Attempts',
             'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts',
             'PG Total',
             'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
             'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
             'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
             'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
             'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
             'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District',
             'Caste Issuing Authority',
             'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government',
             'Government Department',
             'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR'])

        # Add data to the worksheet
        for row in data:
            ws.append(row)

        # Save the workbook in memory as bytes
        data = BytesIO()
        wb.save(data)
        data.seek(0)

        # Create a response object and attach the workbook as a file
        response = make_response(data.getvalue())
        response.headers['Content-Disposition'] = 'attachment; filename=Pending Students (Level 3).xlsx'
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

        return response

    @adminlevels_Three_blueprint.route('/export_rejected_students_level3', methods=['GET', 'POST'])
    def export_rejected_students_level3():
        if not session.get('logged_in'):
            # Redirect to the admin login page if the user is not logged in
            return redirect(url_for('adminlogin.admin_login'))

        host = HostConfig.host
        connect_param = ConnectParam(host)
        cnx, cursor = connect_param.connect()

        cursor.execute(
            "SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
            "date_of_birth, gender, age, caste, your_caste, marital_status,"
            "state, district, taluka, village, city, add_1, add_2, pincode,"
            "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total,"
            "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total,"
            "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream,"
            "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
            "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
            "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
            "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income,"
            "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka,"
            "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number,"
            "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
            "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
            "type_of_disability, father_name, mother_name, work_in_government, gov_department,"
            "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
            " where phd_registration_year>='2023' and final_approval='rejected' and scrutiny_status='accepted' ")
        data = cursor.fetchall()
        # Create a workbook and add a worksheet
        wb = Workbook()
        ws = wb.active

        # Add headers to the worksheet
        # ws.append(['applicant_id','email','first_name','last_name','application_date'])

        ws.append(
            ['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
             'Date Of Birth', 'Gender', 'Age', 'Caste',
             'Marital Status', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
             'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
             'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts',
             'HSC Total',
             'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream',
             'Graduation Attempts',
             'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts',
             'PG Total',
             'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
             'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
             'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
             'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
             'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
             'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District',
             'Caste Issuing Authority',
             'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government',
             'Government Department',
             'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR'])

        # Add data to the worksheet
        for row in data:
            ws.append(row)

        # Save the workbook in memory as bytes
        data = BytesIO()
        wb.save(data)
        data.seek(0)

        # Create a response object and attach the workbook as a file
        response = make_response(data.getvalue())
        response.headers['Content-Disposition'] = 'attachment; filename=Rejected Students (Level 3).xlsx'
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

        return response

    @adminlevels_Three_blueprint.route('/export_pvtg_students_level3', methods=['GET', 'POST'])
    def export_pvtg_students_level3():
        if not session.get('logged_in'):
            # Redirect to the admin login page if the user is not logged in
            return redirect(url_for('adminlogin.admin_login'))

        host = HostConfig.host
        connect_param = ConnectParam(host)
        cnx, cursor = connect_param.connect()

        cursor.execute(
            "  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
            "date_of_birth, gender, age, caste, your_caste, marital_status,"
            "state, district, taluka, village, city, add_1, add_2, pincode,"
            "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total,"
            "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total,"
            "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream,"
            "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
            "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
            "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
            "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income,"
            "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka,"
            "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number,"
            "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
            "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
            "type_of_disability, father_name, mother_name, work_in_government, gov_department,"
            "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
            " where phd_registration_year>='2023' and your_caste IN ('katkari', 'kolam', 'madia') and scrutiny_status='accepted' ")
        data = cursor.fetchall()
        # Create a workbook and add a worksheet
        wb = Workbook()
        ws = wb.active

        # Add headers to the worksheet
        # ws.append(['applicant_id','email','first_name','last_name','application_date'])

        ws.append(
            ['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
             'Date Of Birth', 'Gender', 'Age', 'Caste',
             'Marital Status', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
             'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
             'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts',
             'HSC Total',
             'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream',
             'Graduation Attempts',
             'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts',
             'PG Total',
             'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
             'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
             'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
             'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
             'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
             'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District',
             'Caste Issuing Authority',
             'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government',
             'Government Department',
             'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR'])

        # Add data to the worksheet
        for row in data:
            ws.append(row)

        # Save the workbook in memory as bytes
        data = BytesIO()
        wb.save(data)
        data.seek(0)

        # Create a response object and attach the workbook as a file
        response = make_response(data.getvalue())
        response.headers['Content-Disposition'] = 'attachment; filename=PVTG Students (Level 3).xlsx'
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

        return response

    @adminlevels_Three_blueprint.route('/export_disabled_students_level3', methods=['GET', 'POST'])
    def export_disabled_students_level3():
        if not session.get('logged_in'):
            # Redirect to the admin login page if the user is not logged in
            return redirect(url_for('adminlogin.admin_login'))

        host = HostConfig.host
        connect_param = ConnectParam(host)
        cnx, cursor = connect_param.connect()

        cursor.execute(
            " SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
            "date_of_birth, gender, age, caste, your_caste, marital_status,"
            "state, district, taluka, village, city, add_1, add_2, pincode,"
            "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total,"
            "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total,"
            "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream,"
            "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
            "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
            "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
            "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income,"
            "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka,"
            "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number,"
            "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
            "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
            "type_of_disability, father_name, mother_name, work_in_government, gov_department,"
            "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
            " where phd_registration_year>='2023' and disability='Yes' and scrutiny_status='accepted' ")
        data = cursor.fetchall()
        # Create a workbook and add a worksheet
        wb = Workbook()
        ws = wb.active

        # Add headers to the worksheet
        # ws.append(['applicant_id','email','first_name','last_name','application_date'])

        ws.append(
            ['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
             'Date Of Birth', 'Gender', 'Age', 'Caste',
             'Marital Status', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
             'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
             'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts',
             'HSC Total',
             'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream',
             'Graduation Attempts',
             'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts',
             'PG Total',
             'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
             'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
             'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
             'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
             'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
             'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District',
             'Caste Issuing Authority',
             'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government',
             'Government Department',
             'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR'])

        # Add data to the worksheet
        for row in data:
            ws.append(row)

        # Save the workbook in memory as bytes
        data = BytesIO()
        wb.save(data)
        data.seek(0)

        # Create a response object and attach the workbook as a file
        response = make_response(data.getvalue())
        response.headers['Content-Disposition'] = 'attachment; filename=Disabled Students (Level 3).xlsx'
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

        return response